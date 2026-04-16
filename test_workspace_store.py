import pandas as pd
from openpyxl import Workbook

from review_analyst.connectors import site_config_from_url
from review_analyst.export import autosize_worksheet_openpyxl
from review_analyst.utils import extract_candidate_tokens_from_url, looks_like_sharkninja_uk_eu


def test_sharkninja_uk_host_is_detected():
    assert looks_like_sharkninja_uk_eu("sharkninja.co.uk") is True
    cfg = site_config_from_url(
        "https://sharkninja.co.uk/shark-cryoglow-under-eye-cooling-led-anti-ageing-blemish-repair-mask-fw312uk-blue-frost/FW312UK.html"
    )
    assert cfg is not None
    assert cfg["key"] == "sharkninja_uk_eu"


def test_extract_candidate_tokens_from_new_uk_url_keeps_product_code():
    tokens = extract_candidate_tokens_from_url(
        "https://sharkninja.co.uk/shark-cryoglow-under-eye-cooling-led-anti-ageing-blemish-repair-mask-fw312uk-blue-frost/FW312UK.html"
    )
    lowered = {token.lower() for token in tokens}
    assert "fw312uk" in lowered


def test_openpyxl_autosize_handles_duplicate_columns_and_nested_values():
    df = pd.DataFrame(
        [["ok", {"a": 1}, ["x", "y"]]],
        columns=["review_text", "review_text", "meta"],
    )
    wb = Workbook()
    ws = wb.active
    for col_idx, col_name in enumerate(df.columns, start=1):
        ws.cell(row=1, column=col_idx, value=str(col_name))
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=str(value))

    autosize_worksheet_openpyxl(ws, df)

    assert ws.column_dimensions["A"].width >= 8
    assert ws.column_dimensions["B"].width >= 8
    assert ws.column_dimensions["C"].width >= 8
