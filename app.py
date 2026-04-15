"""
SharkNinja Review Analyst + Symptomizer — Updated and optimized
Updated with:
- Live sidebar filter system with timeframe, stars, core filters, dynamic extra filters, and active filter summary
- Symptom filters shown only when symptom data exists
- Short hoverable Reference tiles (AI Analyst citations only)
- More stable model fallbacks for batch / structured operations
- Symptomizer result cards now show detractors and delighters at the bottom like Review Explorer
"""
from __future__ import annotations

import difflib
import gc
import hashlib
import html
import io
import ipaddress
import json
import math
import os
import random
import re
import textwrap
import time
from collections import Counter
from dataclasses import dataclass
from datetime import date, timedelta
from typing import Any, Dict, List, Optional, Sequence, Set, Tuple
from urllib.parse import parse_qs, parse_qsl, urlencode, urlparse, urlunparse

import numpy as np
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import requests
import streamlit as st
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from plotly.subplots import make_subplots

try:
    from review_analyst.connectors import (
        load_multiple_product_reviews as _package_load_multiple_product_reviews,
        load_product_reviews as _package_load_product_reviews,
        load_uploaded_files as _package_load_uploaded_files,
    )
except Exception:
    _package_load_product_reviews = None
    _package_load_multiple_product_reviews = None
    _package_load_uploaded_files = None

try:
    from review_analyst.tag_quality import (
        UNIVERSAL_DETRACTOR as _UNIVERSAL_DETRACTOR,
        UNIVERSAL_DELIGHTER as _UNIVERSAL_DELIGHTER,
        UNIVERSAL_NEUTRAL_DETRACTORS as _UNIVERSAL_NEUTRAL_DETRACTORS,
        UNIVERSAL_NEUTRAL_DELIGHTERS as _UNIVERSAL_NEUTRAL_DELIGHTERS,
        compute_tag_edit_accuracy as _compute_tag_edit_accuracy,
        ensure_universal_taxonomy as _ensure_universal_taxonomy,
        normalize_tag_list as _normalize_tag_list,
        refine_tag_assignment as _refine_tag_assignment,
        strip_universal_neutral_tags as _strip_universal_neutral_tags,
    )
except Exception:
    _UNIVERSAL_DELIGHTER = "Overall Satisfaction"
    _UNIVERSAL_DETRACTOR = "Overall Dissatisfaction"
    _UNIVERSAL_NEUTRAL_DETRACTORS = [
        "Overall Dissatisfaction",
        "Overpriced",
        "Poor Performance",
        "Poor Quality",
        "Difficult To Use",
        "Unreliable",
        "Hard To Clean",
        "Time Consuming",
    ]
    _UNIVERSAL_NEUTRAL_DELIGHTERS = [
        "Overall Satisfaction",
        "Good Value",
        "Performs Well",
        "High Quality",
        "Easy To Use",
        "Reliable",
        "Easy To Clean",
        "Saves Time",
    ]

    def _normalize_tag_list(values):
        out = []
        seen = set()
        for value in values or []:
            item = re.sub(r"\s+", " ", str(value or "").strip())
            if not item or item.lower() in {"", "na", "n/a", "none", "null", "nan", "<na>", "not mentioned", "unknown"}:
                continue
            item = item.title()
            if item not in seen:
                seen.add(item)
                out.append(item)
        return out

    def _ensure_universal_taxonomy(detractors=None, delighters=None, include_universal_neutral=True):
        det_seed = list(detractors or [])
        del_seed = list(delighters or [])
        if include_universal_neutral:
            det_seed = list(_UNIVERSAL_NEUTRAL_DETRACTORS) + det_seed
            del_seed = list(_UNIVERSAL_NEUTRAL_DELIGHTERS) + del_seed
        dets = _normalize_tag_list(det_seed)
        dels = _normalize_tag_list(del_seed)
        return dets, dels

    def _strip_universal_neutral_tags(detractors=None, delighters=None):
        dets = [tag for tag in _normalize_tag_list(detractors or []) if tag not in set(_UNIVERSAL_NEUTRAL_DETRACTORS + _UNIVERSAL_NEUTRAL_DELIGHTERS)]
        dels = [tag for tag in _normalize_tag_list(delighters or []) if tag not in set(_UNIVERSAL_NEUTRAL_DETRACTORS + _UNIVERSAL_NEUTRAL_DELIGHTERS)]
        return dets, dels

    def _refine_tag_assignment(review_text, detractors, delighters, **kwargs):
        return {
            "dets": _normalize_tag_list(detractors),
            "dels": _normalize_tag_list(delighters),
            "ev_det": {k: list(v or [])[:2] for k, v in (kwargs.get("evidence_det") or {}).items() if k in _normalize_tag_list(detractors)},
            "ev_del": {k: list(v or [])[:2] for k, v in (kwargs.get("evidence_del") or {}).items() if k in _normalize_tag_list(delighters)},
            "support_det": {},
            "support_del": {},
            "added_dets": [],
            "added_dels": [],
            "removed_dets": [],
            "removed_dels": [],
        }

    def _compute_tag_edit_accuracy(baseline_map, current_map):
        baseline_total = 0
        added = 0
        removed = 0
        changed = 0
        for key in set((baseline_map or {}).keys()) | set((current_map or {}).keys()):
            base = baseline_map.get(key, {}) if isinstance(baseline_map, dict) else {}
            cur = current_map.get(key, {}) if isinstance(current_map, dict) else {}
            base_det = set(_normalize_tag_list((base or {}).get("detractors") or []))
            base_del = set(_normalize_tag_list((base or {}).get("delighters") or []))
            cur_det = set(_normalize_tag_list((cur or {}).get("detractors") or []))
            cur_del = set(_normalize_tag_list((cur or {}).get("delighters") or []))
            baseline_total += len(base_det) + len(base_del)
            delta_add = len(cur_det - base_det) + len(cur_del - base_del)
            delta_rem = len(base_det - cur_det) + len(base_del - cur_del)
            added += delta_add
            removed += delta_rem
            if delta_add or delta_rem:
                changed += 1
        total_changes = added + removed
        accuracy = 100.0 if baseline_total <= 0 and total_changes == 0 else max(0.0, 100.0 * (1.0 - total_changes / float(max(baseline_total, 1))))
        return {
            "baseline_total_tags": baseline_total,
            "added_tags": added,
            "removed_tags": removed,
            "total_changes": total_changes,
            "changed_reviews": changed,
            "accuracy_pct": round(accuracy, 1),
        }

try:
    from review_analyst.taxonomy import (
        bucket_symptom_label as _bucket_taxonomy_label,
        build_alias_map_for_labels as _build_taxonomy_alias_map,
        build_structured_taxonomy_rows as _build_structured_taxonomy_rows,
        canonical_theme_name as _canonical_theme_name,
        canonicalize_symptom_catalog as _canonicalize_taxonomy_catalog,
        infer_category as _infer_taxonomy_category,
        infer_l1_theme as _infer_taxonomy_l1_theme,
        merge_alias_maps as _merge_taxonomy_alias_maps,
        prioritize_ai_taxonomy_items as _prioritize_ai_taxonomy_items,
        select_supported_category_pack as _select_supported_category_pack,
        suggest_taxonomy_merges as _suggest_taxonomy_merges,
        taxonomy_prompt_context as _taxonomy_prompt_context,
    )
except Exception:
    def _canonicalize_taxonomy_catalog(delighters=None, detractors=None):
        return _normalize_tag_list(delighters or []), _normalize_tag_list(detractors or []), {}

    def _canonical_theme_name(theme):
        return re.sub(r"\s+", " ", str(theme or "").strip()).title()

    def _infer_taxonomy_l1_theme(label, side=None, family="", theme="", category="general"):
        return _canonical_theme_name(theme or family) or "Product Specific"

    def _build_structured_taxonomy_rows(delighters=None, detractors=None, aliases=None, category="general", preview_items=None):
        rows = []
        for side_key, labels in (("delighter", delighters or []), ("detractor", detractors or [])):
            for raw in labels:
                label = re.sub(r"\s+", " ", str(raw or "").strip()).title()
                if not label:
                    continue
                rows.append({
                    "L1 Theme": "Product Specific",
                    "L2 Symptom": label,
                    "Side": "Delighter" if side_key == "delighter" else "Detractor",
                    "Bucket": _bucket_taxonomy_label(label, side=side_key, category=category),
                    "Aliases": ", ".join((aliases or {}).get(label, [])) if label in (aliases or {}) else "—",
                    "Review Hits": 0,
                    "Support %": 0.0,
                    "Rationale": "",
                    "Example": "",
                    "side_key": side_key,
                    "label": label,
                })
        return rows

    def _suggest_taxonomy_merges(rows, max_suggestions=12):
        return []

    def _build_taxonomy_alias_map(delighters=None, detractors=None, extra_aliases=None):
        merged = {}
        for label in _normalize_tag_list(list(delighters or []) + list(detractors or [])):
            merged.setdefault(label, [])
        for key, values in (extra_aliases or {}).items():
            label = re.sub(r"\s+", " ", str(key or "").strip()).title()
            if not label:
                continue
            bucket = merged.setdefault(label, [])
            for value in values or []:
                alias = re.sub(r"\s+", " ", str(value or "").strip()).title()
                if alias and alias != label and alias not in bucket:
                    bucket.append(alias)
        return merged

    def _merge_taxonomy_alias_maps(*maps):
        merged = {}
        for amap in maps:
            for key, values in (amap or {}).items():
                label = re.sub(r"\s+", " ", str(key or "").strip()).title()
                if not label:
                    continue
                bucket = merged.setdefault(label, [])
                for value in values or []:
                    alias = re.sub(r"\s+", " ", str(value or "").strip()).title()
                    if alias and alias != label and alias not in bucket:
                        bucket.append(alias)
        return merged

    def _infer_taxonomy_category(product_description="", sample_reviews=None):
        return {"category": "general", "confidence": 0.0, "signals": []}

    def _taxonomy_prompt_context(category, include_pack=True, max_labels_per_side=6):
        return ""

    def _select_supported_category_pack(category, sample_reviews, min_hits=1, max_per_side=6):
        return {"delighters": [], "detractors": [], "aliases": {}, "category": str(category or "general")}

    def _bucket_taxonomy_label(label, side=None, category="general"):
        return "Product Specific"

    def _prioritize_ai_taxonomy_items(items, side, sample_reviews, category="general", min_review_hits=1, max_keep=18, exclude_universal=True):
        out = []
        seen = set()
        for raw in items or []:
            if isinstance(raw, dict):
                label = str(raw.get("label", "")).strip()
                aliases = [str(v).strip() for v in (raw.get("aliases") or []) if str(v).strip()]
                family = str(raw.get("family", "")).strip()
                rationale = str(raw.get("rationale", "")).strip()
                bucket = str(raw.get("bucket", "") or "Product Specific")
            else:
                label = str(raw or "").strip()
                aliases = []
                family = ""
                rationale = ""
                bucket = "Product Specific"
            label = re.sub(r"\s+", " ", label).title()
            if not label or label in seen:
                continue
            seen.add(label)
            out.append({
                "label": label,
                "aliases": aliases,
                "family": family,
                "rationale": rationale,
                "bucket": bucket,
                "review_hits": 0,
                "support_ratio": 0.0,
                "score": 0.0,
                "specificity": 0.0,
                "examples": [],
                "seeded": bool(isinstance(raw, dict) and raw.get("seeded")),
            })
            if len(out) >= max_keep:
                break
        return out

try:
    from openai import OpenAI
    _HAS_OPENAI = True
except ImportError:
    OpenAI = None
    _HAS_OPENAI = False

try:
    import tiktoken
    _TIKTOKEN_ENC = tiktoken.get_encoding("cl100k_base")
    _HAS_TIKTOKEN = True
except Exception:
    tiktoken = None
    _TIKTOKEN_ENC = None
    _HAS_TIKTOKEN = False

try:
    from review_analyst.css import APP_CSS as _APP_CSS
except Exception:
    _APP_CSS = ""

try:
    from review_analyst.symptomizer import (
        tag_review_batch as _v3_tag_review_batch,
        gate_polarity as _v3_gate_polarity,
        polarity_confidence_modifier as _v3_polarity_modifier,
        estimate_batch_size as _v3_estimate_batch_size,
        match_label as _v3_match_label,
        validate_evidence as _v3_validate_evidence,
        retry_zero_tag_reviews as _v3_retry_zero_tags,
        calibration_preflight as _v3_calibration_preflight,
        audit_tag_distribution as _v3_audit_distribution,
        LabelTracker as _v3_LabelTracker,
        generate_taxonomy_recommendations as _v3_generate_recommendations,
        _result_cache as _v3_result_cache,
    )
    _HAS_SYMPTOMIZER_V3 = True
except Exception:
    _HAS_SYMPTOMIZER_V3 = False

try:
    from review_analyst.workspace_store import (
        save_workspace_record as _ws_save,
        list_workspace_records as _ws_list,
        load_workspace_record as _ws_load,
        delete_workspace_record as _ws_delete,
        rename_workspace_record as _ws_rename,
        touch_workspace_loaded as _ws_touch,
        count_workspace_records as _ws_count,
    )
    _HAS_WORKSPACE_STORE = True
except Exception:
    _HAS_WORKSPACE_STORE = False

try:
    from review_analyst.logging_config import setup_logging, get_logger
    setup_logging()
    _log = get_logger("app")
except Exception:
    import logging as _logging
    _log = _logging.getLogger("starwalk.app")

try:
    from review_analyst.social_listening import _render_social_listening_tab as _pkg_render_social_listening_tab
    _HAS_SOCIAL_PKG = True
except Exception:
    _HAS_SOCIAL_PKG = False

st.set_page_config(page_title="StarWalk Review Analyst Beta", layout="wide")

if _APP_CSS:
    st.markdown(_APP_CSS, unsafe_allow_html=True)
else:
    st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800;900&display=swap');
:root{--navy:#0f172a;--slate-500:#64748b;--accent:#6366f1;--page-bg:#eef0f4;--surface:#ffffff;--border:#dde1e8;}
html,body,.stApp{font-family:'Inter',system-ui,sans-serif;color:var(--navy);background:var(--page-bg)!important;}
.block-container{max-width:1500px!important;}
</style>
""", unsafe_allow_html=True)

# ═══════════════════════════════════════════════════════════════════════════════
#  CONSTANTS
# ═══════════════════════════════════════════════════════════════════════════════
APP_TITLE           = "StarWalk Review Analyst Beta"
DEFAULT_PASSKEY     = "caC6wVBHos09eVeBkLIniLUTzrNMMH2XMADEhpHe1ewUw"
DEFAULT_DISPLAYCODE = "15973_3_0-en_us"
DEFAULT_API_VERSION = "5.5"
DEFAULT_PAGE_SIZE   = 100
DEFAULT_SORT        = "SubmissionTime:desc"
DEFAULT_CONTENT_LOCALES = (
    "en_US,ar*,zh*,hr*,cs*,da*,nl*,en*,et*,fi*,fr*,de*,el*,he*,hu*,"
    "id*,it*,ja*,ko*,lv*,lt*,ms*,no*,pl*,pt*,ro*,sk*,sl*,es*,sv*,th*,"
    "tr*,vi*,en_AU,en_CA,en_GB"
)
BAZAARVOICE_ENDPOINT = "https://api.bazaarvoice.com/data/reviews.json"
POWERREVIEWS_BASE = "https://display.powerreviews.com"
POWERREVIEWS_MAX_PAGE_SIZE = 25
UK_EU_BV_PASSKEY = "capxzF3xnCmhSCHhkomxF1sQkZmh2zK2fNb8D1VDNl3hY"
COSTCO_BV_PASSKEY = "bai25xto36hkl5erybga10t99"
SEPHORA_BV_PASSKEY = "calXm2DyQVjcCy9agq85vmTJv5ELuuBCF2sdg4BnJzJus"
ULTA_POWERREVIEWS_API_KEY = "daa0f241-c242-4483-afb7-4449942d1a2b"
HOKA_POWERREVIEWS_API_KEY = "ea283fa2-3fdc-4127-863c-b1e2397f7a77"

SITE_REVIEW_CONFIGS = [
    {
        "key": "sharkninja_us",
        "provider": "bazaarvoice",
        "bv_style": "revstats",
        "label": "SharkNinja US",
        "domains": ["sharkninja.com", "www.sharkninja.com", "sharkclean.com", "www.sharkclean.com", "ninjakitchen.com", "www.ninjakitchen.com"],
        "passkey": DEFAULT_PASSKEY,
        "displaycode": DEFAULT_DISPLAYCODE,
        "api_version": DEFAULT_API_VERSION,
        "content_locales": DEFAULT_CONTENT_LOCALES,
        "sort": DEFAULT_SORT,
        "retailer": "SharkNinja US",
    },
    {
        "key": "sharkninja_uk_eu",
        "provider": "bazaarvoice",
        "bv_style": "simple",
        "label": "SharkNinja UK/EU",
        "domains": [
            "sharkninja.co.uk", "www.sharkninja.co.uk",
            "sharkninja.eu", "www.sharkninja.eu",
            "sharkninja.de", "www.sharkninja.de",
            "sharkninja.fr", "www.sharkninja.fr",
            "sharkninja.es", "www.sharkninja.es",
            "sharkninja.it", "www.sharkninja.it",
            "sharkninja.nl", "www.sharkninja.nl",
            "sharkclean.co.uk", "www.sharkclean.co.uk", "ninjakitchen.co.uk", "www.ninjakitchen.co.uk",
            "sharkclean.eu", "www.sharkclean.eu", "ninjakitchen.eu", "www.ninjakitchen.eu",
            "sharkclean.de", "www.sharkclean.de", "ninjakitchen.de", "www.ninjakitchen.de",
            "sharkclean.fr", "www.sharkclean.fr", "ninjakitchen.fr", "www.ninjakitchen.fr",
            "sharkclean.es", "www.sharkclean.es", "ninjakitchen.es", "www.ninjakitchen.es",
            "sharkclean.it", "www.sharkclean.it", "ninjakitchen.it", "www.ninjakitchen.it",
            "sharkclean.nl", "www.sharkclean.nl", "ninjakitchen.nl", "www.ninjakitchen.nl",
        ],
        "passkey": UK_EU_BV_PASSKEY,
        "api_version": "5.4",
        "sort": "SubmissionTime:desc",
        "locale": "en_GB",
        "retailer": "SharkNinja UK/EU",
    },
    {
        "key": "costco",
        "provider": "bazaarvoice",
        "bv_style": "revstats",
        "label": "Costco",
        "domains": ["costco.com", "www.costco.com"],
        "passkey": COSTCO_BV_PASSKEY,
        "displaycode": "2070_2_0-en_us",
        "api_version": "5.5",
        "content_locales": "en_US,ar*,zh*,hr*,cs*,da*,nl*,en*,et*,fi*,fr*,de*,el*,he*,hu*,id*,it*,ja*,ko*,lv*,lt*,ms*,no*,pl*,pt*,ro*,sk*,sl*,es*,sv*,th*,tr*,vi*",
        "sort": "SubmissionTime:desc",
        "retailer": "Costco",
    },
    {
        "key": "sephora",
        "provider": "bazaarvoice",
        "bv_style": "simple",
        "label": "Sephora",
        "domains": ["sephora.com", "www.sephora.com"],
        "passkey": SEPHORA_BV_PASSKEY,
        "api_version": "5.4",
        "sort": "SubmissionTime:desc",
        "locale": "en_US",
        "retailer": "Sephora",
        "extra_filters": ["contentlocale:en*"],
    },
    {
        "key": "ulta",
        "provider": "powerreviews",
        "label": "Ulta",
        "domains": ["ulta.com", "www.ulta.com"],
        "merchant_id": "6406",
        "locale": "en_US",
        "page_locale": "en_US",
        "api_key": ULTA_POWERREVIEWS_API_KEY,
        "sort": "Newest",
        "retailer": "Ulta",
    },
    {
        "key": "hoka",
        "provider": "powerreviews",
        "label": "Hoka",
        "domains": ["hoka.com", "www.hoka.com"],
        "merchant_id": "437772",
        "locale": "en_US",
        "page_locale": "en_US",
        "api_key": HOKA_POWERREVIEWS_API_KEY,
        "sort": "Newest",
        "retailer": "Hoka",
    },
]

POWERREVIEWS_ENDPOINT = "https://display.powerreviews.com"
SHARK_EUUK_BV_PASSKEY = "capxzF3xnCmhSCHhkomxF1sQkZmh2zK2fNb8D1VDNl3hY"
SHARK_EUUK_BV_API_VERSION = "5.4"
SHARK_EUUK_BV_CONFIG = {
    "style": "simple",
    "passkey": SHARK_EUUK_BV_PASSKEY,
    "api_version": SHARK_EUUK_BV_API_VERSION,
    "locale": "en_GB",
    "sort": "SubmissionTime:desc",
    "extra_filters": [],
}
COSTCO_BV_CONFIG = {
    "style": "revstats",
    "passkey": "bai25xto36hkl5erybga10t99",
    "displaycode": "2070_2_0-en_us",
    "api_version": "5.5",
    "sort": "relevancy:a1",
    "content_locales": "en_US,ar*,zh*,hr*,cs*,da*,nl*,en*,et*,fi*,fr*,de*,el*,he*,hu*,id*,it*,ja*,ko*,lv*,lt*,ms*,no*,pl*,pt*,ro*,sk*,sl*,es*,sv*,th*,tr*,vi*",
}
SEPHORA_BV_CONFIG = {
    "style": "simple",
    "passkey": "calXm2DyQVjcCy9agq85vmTJv5ELuuBCF2sdg4BnJzJus",
    "api_version": "5.4",
    "locale": "en_US",
    "sort": "SubmissionTime:desc",
    "extra_filters": ["contentlocale:en*"],
}
ULTA_POWERREVIEWS_CONFIG = {
    "merchant_id": "6406",
    "locale": "en_US",
    "page_locale": "en_US",
    "apikey": "daa0f241-c242-4483-afb7-4449942d1a2b",
    "sort": "Newest",
}
HOKA_POWERREVIEWS_CONFIG = {
    "merchant_id": "437772",
    "locale": "en_US",
    "page_locale": "en_US",
    "apikey": "ea283fa2-3fdc-4127-863c-b1e2397f7a77",
    "sort": "Newest",
}
POWERREVIEWS_ENDPOINT_TEMPLATE = "https://display.powerreviews.com/m/{merchant_id}/l/{locale}/product/{product_id}/reviews"
SHARKNINJA_UK_EU_BV_CONFIG = {
    "passkey": UK_EU_BV_PASSKEY,
    "api_version": "5.4",
    "locale": "en_GB",
    "include": "Products,Comments",
    "sort": "SubmissionTime:desc",
    "content_locale": "en*",
    "retailer": "SharkNinja UK/EU",
}
ULTA_PR_CONFIG = {
    **ULTA_POWERREVIEWS_CONFIG,
    "apikey": ULTA_POWERREVIEWS_CONFIG.get("apikey") or ULTA_POWERREVIEWS_API_KEY,
    "retailer": "Ulta",
}
HOKA_PR_CONFIG = {
    **HOKA_POWERREVIEWS_CONFIG,
    "apikey": HOKA_POWERREVIEWS_CONFIG.get("apikey") or HOKA_POWERREVIEWS_API_KEY,
    "retailer": "Hoka",
}

DEFAULT_PRODUCT_URL = "https://www.sharkninja.com/ninja-air-fryer-pro-xl-6-in-1/AF181.html"
SOURCE_MODE_URL = "Product / review URL"
SOURCE_MODE_FILE = "Uploaded review file"

TAB_DASHBOARD = "📊  Dashboard"
TAB_REVIEW_EXPLORER = "🔍  Review Explorer"
TAB_AI_ANALYST = "🤖  AI Analyst"
TAB_REVIEW_PROMPT = "🏷️  Review Prompt"
TAB_SYMPTOMIZER = "💊  Symptomizer"
TAB_SOCIAL_LISTENING = "📣  Social Listening Beta"
WORKSPACE_TABS = [
    TAB_DASHBOARD,
    TAB_REVIEW_EXPLORER,
    TAB_REVIEW_PROMPT,
    TAB_SYMPTOMIZER,
    TAB_SOCIAL_LISTENING,
]

MODEL_OPTIONS = [
    "gpt-5.4-mini",
    "gpt-5.4",
    "gpt-5.4-pro",
    "gpt-5.4-nano",
    "gpt-5-chat-latest",
    "gpt-5-mini",
    "gpt-5",
    "gpt-5-nano",
    "gpt-4o-mini",
    "gpt-4o",
    "gpt-4.1",
]
DEFAULT_MODEL = "gpt-5.4-mini"
DEFAULT_REASONING = "none"
STRUCTURED_FALLBACK_MODEL = "gpt-5.4-mini"
AI_VISIBLE_CHAT_MESSAGES = 2
AI_CONTEXT_TOKEN_BUDGET = 10_000
NON_VALUES = {"<NA>","NA","N/A","NONE","-","","NAN","NULL"}
STOPWORDS = {
    "a","about","after","again","all","also","am","an","and","any","are","as","at",
    "be","because","been","before","being","best","better","but","by","can","could",
    "did","do","does","don","down","even","every","for","from","get","got","great",
    "had","has","have","he","her","here","hers","him","his","how","i","if","in",
    "into","is","it","its","just","like","love","made","make","many","me","more",
    "most","much","my","new","no","not","now","of","on","one","only","or","other",
    "our","out","over","product","really","so","some","than","that","the","their",
    "them","then","there","these","they","this","to","too","use","used","using",
    "very","was","we","well","were","what","when","which","while","with","would",
    "you","your",
}
PERSONAS: Dict[str, Dict[str, Any]] = {
    "Product Development": {
        "blurb": "Translates reviews into product and feature decisions.",
        "prompt": "Create a report for the product development team. Highlight what customers love, unmet needs, feature gaps, usability friction, and concrete roadmap opportunities. End with the top 5 product actions ranked by impact.",
        "instructions": (
            "You are a senior product strategy analyst specialising in consumer appliances.\n"
            "Structure your response with these exact sections:\n"
            "## What Customers Love\n## Unmet Needs & Feature Gaps\n"
            "## Usability Friction\n## Roadmap Opportunities\n## Top 5 Actions (ranked)\n"
            "Cite review IDs inline as (review_ids: 12345, 67890) for every material claim.\n"
            "Be specific — name exact features, not vague categories.\n"
            "Keep each section to 3-5 bullet points. Total response ≤ 500 words."
        ),
    },
    "Quality Engineer": {
        "blurb": "Focuses on failure modes, defects, durability, and root-cause signals.",
        "prompt": "Create a report for a quality engineer. Identify defect patterns, reliability risks, cleaning issues, performance inconsistencies, and probable root-cause hypotheses. Separate confirmed evidence from inference.",
        "instructions": (
            "You are a senior quality and reliability analyst for consumer appliances.\n"
            "Sections:\n"
            "## Confirmed Defect Patterns\n## Reliability & Durability Risks\n"
            "## Root-Cause Hypotheses\n## Cleaning & Maintenance Issues\n"
            "## Risk Severity Matrix (High/Med/Low)\n"
            "Mark speculative claims as [INFERRED]. Cite review IDs for every confirmed finding.\n"
            "Prioritise by frequency × severity. Total ≤ 500 words."
        ),
    },
    "Consumer Insights": {
        "blurb": "Extracts sentiment drivers, purchase motivations, and voice-of-customer insights.",
        "prompt": "Create a report for the consumer insights team. Summarize key sentiment drivers, barriers to adoption, purchase motivations, key use cases, and how tone changes across star ratings and incentivized vs non-incentivized reviews.",
        "instructions": (
            "You are a consumer insights lead specialising in VoC analysis.\n"
            "Sections:\n"
            "## Top Sentiment Drivers (positive)\n## Top Sentiment Drivers (negative)\n"
            "## Purchase Motivations & Jobs-to-be-Done\n## Barriers to Satisfaction\n"
            "## Organic vs Incentivized Tone Differences\n## Key Verbatim Quotes (3-5)\n"
            "Use plain, executive-ready language. Cite review IDs for quotes. Total ≤ 500 words."
        ),
    },
}
DET_LETTERS  = ["K","L","M","N","O","P","Q","R","S","T"]
DEL_LETTERS  = ["U","V","W","X","Y","Z","AA","AB","AC","AD"]
DET_INDEXES  = [column_index_from_string(c) for c in DET_LETTERS]
DEL_INDEXES  = [column_index_from_string(c) for c in DEL_LETTERS]
META_ORDER   = [("Safety","AE"),("Reliability","AF"),("# of Sessions","AG")]
META_INDEXES = {name: column_index_from_string(col) for name,col in META_ORDER}
AI_DET_HEADERS  = [f"AI Symptom Detractor {i}" for i in range(1,11)]
AI_DEL_HEADERS  = [f"AI Symptom Delighter {i}" for i in range(1,11)]
AI_META_HEADERS = ["AI Safety","AI Reliability","AI # of Sessions"]
SAFETY_ENUM      = ["Not Mentioned","Concern","Positive"]
RELIABILITY_ENUM = ["Not Mentioned","Negative","Neutral","Positive"]
SESSIONS_ENUM    = ["0","1","2–3","4–9","10+","Unknown"]
DEFAULT_PRIORITY_DELIGHTERS = ["Overall Satisfaction","Ease Of Use","Effective Results",
    "Visible Improvement","Time Saver","Comfort","Value","Reliability"]
DEFAULT_PRIORITY_DETRACTORS = ["Overall Dissatisfaction","Poor Results","Ease Of Use","Reliability Issue","High Cost",
    "Irritation","Battery Problem","High Noise","Cleaning Difficulty",
    "Setup Issue","Connectivity Issue","Safety Concern"]
REVIEW_PROMPT_STARTER_ROWS = [
    {"column_name":"perceived_loudness",
     "prompt":"How is product loudness described? Positive, Negative, Neutral, or Not Mentioned.",
     "labels":"Positive, Negative, Neutral, Not Mentioned"},
    {"column_name":"reliability_risk_signal",
     "prompt":"Does the review mention a product reliability or durability risk? Risk Mentioned, Positive Reliability, or Not Mentioned.",
     "labels":"Risk Mentioned, Positive Reliability, Not Mentioned"},
    {"column_name":"product_usage_sessions_if_mentioned",
     "prompt":"If mentioned, how many times has the customer used the product? Choose the closest bucket. Use Unknown if the review does not say.",
     "labels":"0, 1, 2–3, 4–9, 10+, Unknown"},
    {"column_name":"safety_signal_if_mentioned",
     "prompt":"Does the review mention a product safety concern or a positive safety statement? Concern, Positive, or Not Mentioned.",
     "labels":"Concern, Positive, Not Mentioned"},
    {"column_name":"ownership_period_if_mentioned",
     "prompt":"If mentioned, how long has the customer owned or used the product? Choose the closest bucket. Use Unknown if not stated.",
     "labels":"First Use, Under 1 Week, 1–4 Weeks, 1–3 Months, 3+ Months, Unknown"},
]

# ═══════════════════════════════════════════════════════════════════════════════
#  SHARED UTILITIES
# ═══════════════════════════════════════════════════════════════════════════════
class ReviewDownloaderError(Exception):
    pass


@dataclass
class ReviewBatchSummary:
    product_url: str
    product_id: str
    total_reviews: int
    page_size: int
    requests_needed: int
    reviews_downloaded: int


def _safe_text(v, default=""):
    if v is None:
        return default
    if isinstance(v, (list, tuple, set, dict, pd.Series, pd.DataFrame, pd.Index)):
        return default
    try:
        m = pd.isna(v)
    except Exception:
        m = False
    if isinstance(m, bool) and m:
        return default
    t = str(v).strip()
    return default if t.lower() in {"nan", "none", "null", "<na>"} else t


def _safe_int(v, d=0):
    try:
        return int(float(v))
    except Exception:
        return d


def _safe_bool(v, d=False):
    if v is None:
        return d
    if isinstance(v, bool):
        return v
    t = _safe_text(v).lower()
    if t in {"true", "1", "yes", "y", "t"}:
        return True
    if t in {"false", "0", "no", "n", "f", ""}:
        return False
    return d


def _safe_mean(s):
    if s.empty:
        return None
    n = pd.to_numeric(s, errors="coerce").dropna()
    return float(n.mean()) if not n.empty else None


def _safe_pct(num, den):
    return 0.0 if not den else float(num) / float(den)


def _fmt_secs(sec):
    sec = max(0.0, float(sec or 0))
    m = int(sec // 60)
    s = int(round(sec - m * 60))
    return f"{m}:{s:02d}"


def _canon(s):
    return " ".join(str(s).split()).lower().strip()


def _canon_simple(s):
    return "".join(ch for ch in _canon(s) if ch.isalnum())


def _esc(s):
    return html.escape(str(s or ""))


def _chip_html(items):
    if not items:
        return ""
    return "<div class='chip-wrap'>" + "".join(f"<span class='chip {c}'>{_esc(t)}</span>" for t, c in items) + "</div>"


def _is_missing(v):
    if v is None:
        return True
    if isinstance(v, (list, tuple, set, dict, pd.Series, pd.DataFrame, pd.Index)):
        return False
    try:
        m = pd.isna(v)
    except Exception:
        return False
    return bool(m) if isinstance(m, (bool, int)) else False


def _fmt_num(v, d=2):
    if v is None or _is_missing(v):
        return "n/a"
    return f"{v:.{d}f}"


def _fmt_pct(v, d=1):
    if v is None or _is_missing(v):
        return "n/a"
    return f"{100 * float(v):.{d}f}%"


def _trunc(text, max_chars=420):
    text = re.sub(r"\s+", " ", _safe_text(text)).strip()
    return text if len(text) <= max_chars else text[:max_chars - 3].rstrip() + "…"


def _norm_text(text):
    return re.sub(r"\s+", " ", str(text).lower()).strip()


def _tokenize(text):
    return [t for t in re.findall(r"[a-z0-9']+", _norm_text(text)) if len(t) > 2 and t not in STOPWORDS]


def _slugify(text, fallback="custom"):
    c = re.sub(r"[^a-zA-Z0-9]+", "_", _safe_text(text).lower())
    c = re.sub(r"_+", "_", c).strip("_") or fallback
    return ("prompt_" + c if c[0].isdigit() else c)[:64]


def _first_non_empty(series):
    for v in series.astype(str):
        v = _safe_text(v)
        if v and v.lower() != "nan":
            return v
    return ""


def _clean_text(x):
    """Clean review text for AI processing.
    
    Strips HTML tags, normalizes smart quotes/dashes, collapses whitespace,
    and removes embedded URLs. BazaarVoice reviews often contain HTML entities,
    smart punctuation, and URLs that waste AI tokens.
    """
    if pd.isna(x):
        return ""
    s = str(x).strip()
    if not s:
        return ""
    # Strip HTML tags
    s = re.sub(r"<[^>]+>", " ", s)
    # Decode common HTML entities
    s = s.replace("&amp;", "&").replace("&lt;", "<").replace("&gt;", ">")
    s = s.replace("&quot;", '"').replace("&#39;", "'").replace("&nbsp;", " ")
    s = s.replace("&#x27;", "'").replace("&#x2F;", "/")
    # Normalize smart quotes and dashes
    s = s.replace("\u2018", "'").replace("\u2019", "'")  # single smart quotes
    s = s.replace("\u201c", '"').replace("\u201d", '"')  # double smart quotes
    s = s.replace("\u2013", "-").replace("\u2014", " - ")  # en/em dashes
    s = s.replace("\u2026", "...")  # ellipsis
    # Remove URLs (waste tokens, confuse matching)
    s = re.sub(r"https?://\S+", "", s)
    # Collapse whitespace
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _is_filled(val):
    if pd.isna(val):
        return False
    s = str(val).strip()
    return s != "" and s.upper() not in NON_VALUES


def _estimate_tokens(text):
    s = str(text or "")
    if not s:
        return 0
    if _HAS_TIKTOKEN and _TIKTOKEN_ENC is not None:
        try:
            return int(len(_TIKTOKEN_ENC.encode(s)))
        except Exception:
            pass
    return int(max(1, math.ceil(len(s) / 4)))

# ═══════════════════════════════════════════════════════════════════════════════
#  OPENAI
# ═══════════════════════════════════════════════════════════════════════════════
def _get_api_key():
    """Resolve OpenAI API key: secrets → env → manual sidebar input."""
    try:
        if "OPENAI_API_KEY" in st.secrets:
            return str(st.secrets["OPENAI_API_KEY"])
        if "openai" in st.secrets and st.secrets["openai"].get("api_key"):
            return str(st.secrets["openai"]["api_key"])
    except Exception:
        pass
    env_key = os.getenv("OPENAI_API_KEY")
    if env_key:
        return env_key
    manual = (st.session_state.get("sidebar_manual_api_key") or "").strip()
    return manual or None


def _api_key_source() -> str:
    """Return where the API key was found: 'secrets', 'env', 'manual', or 'missing'."""
    try:
        if "OPENAI_API_KEY" in st.secrets: return "secrets"
        if "openai" in st.secrets and st.secrets["openai"].get("api_key"): return "secrets"
    except Exception: pass
    if os.getenv("OPENAI_API_KEY"): return "env"
    if (st.session_state.get("sidebar_manual_api_key") or "").strip(): return "manual"
    return "missing"


@st.cache_resource(show_spinner=False)
def _make_openai_client(api_key: str):
    if not (_HAS_OPENAI and api_key):
        return None
    try:
        return OpenAI(api_key=api_key, timeout=60, max_retries=3)
    except TypeError:
        try:
            return OpenAI(api_key=api_key)
        except Exception:
            return None


def _get_client():
    key = _get_api_key()
    if not (_HAS_OPENAI and key):
        return None
    if _api_key_source() == "manual":
        try: return OpenAI(api_key=key, timeout=60, max_retries=3)
        except TypeError:
            try: return OpenAI(api_key=key)
            except Exception: return None
    return _make_openai_client(key)


def _shared_model():
    return st.session_state.get("shared_model", DEFAULT_MODEL)


def _reasoning_options_for_model(model: str) -> List[str]:
    m = _safe_text(model).lower()
    if not m.startswith("gpt-5"):
        return ["none"]
    if m.startswith("gpt-5.4") or m in {"gpt-5-chat-latest", "gpt-5.2", "gpt-5.2-pro"}:
        return ["none", "low", "medium", "high", "xhigh"]
    if m in {"gpt-5", "gpt-5-mini", "gpt-5-nano"}:
        return ["minimal", "low", "medium", "high"]
    return ["none", "low", "medium", "high"]


def _shared_reasoning():
    current_model = _shared_model()
    allowed = _reasoning_options_for_model(current_model)
    cur = _safe_text(st.session_state.get("shared_reasoning", DEFAULT_REASONING)).lower() or DEFAULT_REASONING
    if cur not in allowed:
        cur = "none" if "none" in allowed else allowed[0]
        st.session_state["shared_reasoning"] = cur
    return cur


def _coerce_ai_target_words(value, default=1200):
    try:
        n = int(value)
    except Exception:
        n = int(default)
    return max(250, min(2400, n))


def _ai_target_token_budget(target_words: int) -> int:
    words = _coerce_ai_target_words(target_words)
    return max(900, min(7000, int(round(words * 2.35))))


def _current_ai_target_words() -> int:
    return _coerce_ai_target_words(st.session_state.get("ai_response_words", 1200))


def _model_supports_reasoning(model: str) -> bool:
    return _safe_text(model).lower().startswith("gpt-5")


def _normalize_reasoning_effort_for_model(model: str, reasoning_effort: Optional[str]) -> Optional[str]:
    if not _model_supports_reasoning(model):
        return None
    allowed = _reasoning_options_for_model(model)
    effort = _safe_text(reasoning_effort).lower()
    if effort in allowed:
        return effort
    if not effort:
        return allowed[0] if allowed else None
    if effort == "none" and "minimal" in allowed:
        return "minimal"
    if effort == "minimal" and "none" in allowed:
        return "none"
    if effort == "xhigh" and "high" in allowed:
        return "high"
    if effort == "high" and "xhigh" in allowed:
        return "high"
    return allowed[0] if allowed else None


def _model_accepts_temperature(model: str, reasoning_effort: Optional[str]) -> bool:
    m = _safe_text(model).lower()
    eff = _safe_text(reasoning_effort).lower()
    if not m.startswith("gpt-5"):
        return True
    if m.startswith("gpt-5.4") or m in {"gpt-5-chat-latest", "gpt-5.2", "gpt-5.2-pro"}:
        return eff in {"", "none"}
    return False


def _split_chat_messages(messages, keep_last=AI_VISIBLE_CHAT_MESSAGES):
    items = list(messages or [])
    keep = max(1, int(keep_last or 1))
    if len(items) <= keep:
        return [], items
    return items[:-keep], items[-keep:]


def _show_thinking(msg):
    ph = st.empty()
    ph.markdown(f"""<div class="thinking-overlay"><div class="thinking-card">
      <div class="thinking-spinner"></div>
      <div class="thinking-title">Working…</div>
      <div class="thinking-sub">{_esc(msg)}</div>
    </div></div>""", unsafe_allow_html=True)
    return ph


def _safe_json_load(s):
    s = (s or "").strip()
    if not s:
        return {}
    try:
        return json.loads(s)
    except Exception:
        pass
    try:
        i = s.find("{")
        j = s.rfind("}")
        if i >= 0 and j > i:
            return json.loads(s[i:j + 1])
    except Exception:
        pass
    return {}


def _prepare_messages_for_model(model: str, messages):
    prepared = []
    use_developer = _safe_text(model).lower().startswith("gpt-5")
    for msg in list(messages or []):
        if not isinstance(msg, dict):
            continue
        item = dict(msg)
        if use_developer and item.get("role") == "system":
            item["role"] = "developer"
        prepared.append(item)
    return prepared


def _build_completion_token_kwargs(max_tokens):
    try:
        limit = int(max_tokens) if max_tokens is not None else None
    except Exception:
        limit = None
    if limit is None or limit <= 0:
        return {}
    return {"max_completion_tokens": limit}


def _chat_complete(client, *, model, messages, temperature=0.0, response_format=None,
                   max_tokens=1200, reasoning_effort=None, _max_retries=3):
    if client is None:
        raise RuntimeError("OpenAI client is not initialized. Check your API key in Settings → OpenAI API Key.")

    effort = _normalize_reasoning_effort_for_model(model, reasoning_effort)
    kwargs = dict(model=model, messages=_prepare_messages_for_model(model, messages))
    kwargs.update(_build_completion_token_kwargs(max_tokens))
    if response_format:
        kwargs["response_format"] = response_format
    if effort:
        kwargs["reasoning_effort"] = effort
    if temperature is not None and _model_accepts_temperature(model, effort):
        kwargs["temperature"] = temperature

    last_exc = None
    reasoning_enabled = "reasoning_effort" in kwargs
    temperature_enabled = "temperature" in kwargs

    for attempt in range(max(1, _max_retries)):
        try:
            resp = client.chat.completions.create(**kwargs)
            return (resp.choices[0].message.content or "").strip()
        except Exception as exc:
            last_exc = exc
            err = str(exc).lower()

            if "max_completion_tokens" in kwargs and any(k in err for k in (
                "unexpected keyword argument 'max_completion_tokens'",
                'unsupported parameter: "max_completion_tokens"',
                "unsupported parameter: 'max_completion_tokens'",
                "unknown parameter: max_completion_tokens",
                "max_completion_tokens is not supported",
            )):
                token_limit = kwargs.pop("max_completion_tokens", None)
                if token_limit is not None:
                    kwargs["max_tokens"] = token_limit
                continue

            if "max_tokens" in kwargs and any(k in err for k in (
                "unexpected keyword argument 'max_tokens'",
                'unsupported parameter: "max_tokens"',
                "unsupported parameter: 'max_tokens'",
                "use 'max_completion_tokens' instead",
                "deprecated in favor of `max_completion_tokens`",
                "not compatible with o-series models",
                "not compatible with reasoning models",
            )):
                token_limit = kwargs.pop("max_tokens", None)
                if token_limit is not None:
                    kwargs["max_completion_tokens"] = token_limit
                continue

            if reasoning_enabled and any(k in err for k in (
                "reasoning_effort",
                "unknown parameter: reasoning_effort",
                'unsupported parameter: "reasoning_effort"',
                "unsupported parameter: 'reasoning_effort'",
                "invalid reasoning",
                "invalid value for reasoning",
                "does not support reasoning effort",
                "not support reasoning effort",
            )):
                kwargs.pop("reasoning_effort", None)
                reasoning_enabled = False
                continue

            if temperature_enabled and any(k in err for k in (
                "temperature",
                "top_p",
                "only supported when using",
                "not supported when reasoning effort",
                "include these fields will raise",
            )):
                kwargs.pop("temperature", None)
                temperature_enabled = False
                continue

            if any(k in err for k in ("rate_limit", "429", "500", "503", "timeout", "overloaded")):
                time.sleep(min((2 ** attempt) + random.uniform(0, 1), 30))
                continue
            raise

    if last_exc:
        raise last_exc
    return ""


def _model_candidates_for_task(selected_model: str, *, structured: bool = False) -> List[str]:
    preferred = _safe_text(selected_model) or DEFAULT_MODEL
    fallbacks = [preferred]
    if structured:
        fallbacks += [STRUCTURED_FALLBACK_MODEL, DEFAULT_MODEL, "gpt-4.1"]
    else:
        fallbacks += [DEFAULT_MODEL]
    out = []
    seen = set()
    for m in fallbacks:
        if m and m not in seen:
            out.append(m)
            seen.add(m)
    return out


def _chat_complete_with_fallback_models(client, *, model, messages, structured=False, **kwargs):
    last_exc = None
    for candidate in _model_candidates_for_task(model, structured=structured):
        try:
            return _chat_complete(client, model=candidate, messages=messages, **kwargs)
        except Exception as exc:
            last_exc = exc
            continue
    if last_exc:
        raise last_exc
    return ""

# ═══════════════════════════════════════════════════════════════════════════════
#  DATA LAYER
# ═══════════════════════════════════════════════════════════════════════════════
def _get_session():
    s = requests.Session()
    s.headers.update({
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,application/json;q=0.9,image/avif,image/webp,*/*;q=0.8",
        "Accept-Language": "en-US,en;q=0.9",
        "Cache-Control": "no-cache",
        "Pragma": "no-cache",
        "Upgrade-Insecure-Requests": "1",
    })
    return s


def _ordered_unique(values):
    out = []
    seen = set()
    for raw in list(values or []):
        val = _safe_text(raw).strip().strip("/")
        val = re.sub(r"\.html?$", "", val, flags=re.IGNORECASE)
        if not val:
            continue
        key = val.lower()
        if key in seen:
            continue
        seen.add(key)
        out.append(val)
    return out


def _domain_matches(host: str, domain: str) -> bool:
    host = _safe_text(host).lower()
    domain = _safe_text(domain).lower()
    return bool(host == domain or host.endswith("." + domain))


def _site_config_from_url(url: str) -> Optional[Dict[str, Any]]:
    host = urlparse(_safe_text(url)).netloc.lower()
    for cfg in SITE_REVIEW_CONFIGS:
        if any(_domain_matches(host, d) for d in cfg.get("domains", [])):
            return dict(cfg)
    return None


def _is_bv_api_url(url: str) -> bool:
    parsed = urlparse(_safe_text(url))
    return "api.bazaarvoice.com" in parsed.netloc.lower() and parsed.path.lower().endswith("/reviews.json")


def _is_powerreviews_api_url(url: str) -> bool:
    parsed = urlparse(_safe_text(url))
    return "display.powerreviews.com" in parsed.netloc.lower() and "/product/" in parsed.path.lower() and parsed.path.lower().endswith("/reviews")


def _extract_pid_from_url(url):
    parsed = urlparse(url.strip())
    path = parsed.path
    m = re.search(r"/([A-Za-z0-9_-]+)\.html(?:$|[?#])", path)
    if m:
        c = m.group(1).strip().upper()
        if re.fullmatch(r"[A-Z0-9_-]{3,}", c):
            return c
    return None


def _extract_pid_from_html(h):
    for pat in [
        r'Item\s*No\.?\s*([A-Z0-9_-]{3,})',
        r'"productId"\s*:\s*"([A-Z0-9_-]{3,})"',
        r'"sku"\s*:\s*"([A-Z0-9_-]{3,})"',
        r'"model"\s*:\s*"([A-Z0-9_-]{3,})"',
    ]:
        m = re.search(pat, h, flags=re.IGNORECASE)
        if m:
            return m.group(1).strip().upper()
    soup = BeautifulSoup(h, "html.parser")
    text = soup.get_text(" ", strip=True)
    for pat in [r"Item\s*No\.?\s*([A-Z0-9_-]{3,})", r"Model\s*:?\s*([A-Z0-9_-]{3,})"]:
        m = re.search(pat, text, flags=re.IGNORECASE)
        if m:
            return m.group(1).strip().upper()
    return None


def _fetch_reviews_page(session, *, product_id, passkey, displaycode, api_version,
                        page_size, offset, sort, content_locales):
    params = dict(
        resource="reviews",
        action="REVIEWS_N_STATS",
        filter=[
            f"productid:eq:{product_id}",
            f"contentlocale:eq:{content_locales}",
            "isratingsonly:eq:false",
        ],
        filter_reviews=f"contentlocale:eq:{content_locales}",
        include="authors,products,comments",
        filteredstats="reviews",
        Stats="Reviews",
        limit=int(page_size),
        offset=int(offset),
        limit_comments=3,
        sort=sort,
        passkey=passkey,
        apiversion=api_version,
        displaycode=displaycode,
    )
    resp = session.get(BAZAARVOICE_ENDPOINT, params=params, timeout=45)
    resp.raise_for_status()
    payload = resp.json()
    if payload.get("HasErrors"):
        raise ReviewDownloaderError(f"BV error: {payload.get('Errors')}")
    return payload


def _fetch_bv_simple_page(session, *, product_id, passkey, api_version,
                          page_size, offset, sort, content_locale="en*",
                          locale="en_US", include="Products,Comments"):
    params: Dict[str, Any] = {
        "apiversion": api_version,
        "passkey": passkey,
        "Include": include,
        "Stats": "Reviews",
        "Limit": int(page_size),
        "Offset": int(offset),
        "Sort": sort,
        "Filter": [f"ProductId:{product_id}"],
    }
    if content_locale:
        params["Filter"].insert(0, f"contentlocale:{content_locale}")
    if locale:
        params["Locale"] = locale
    resp = session.get(BAZAARVOICE_ENDPOINT, params=params, timeout=45)
    resp.raise_for_status()
    payload = resp.json()
    if payload.get("HasErrors"):
        raise ReviewDownloaderError(f"BV error: {payload.get('Errors')}")
    return payload


def _fetch_bazaarvoice_raw_page(session, *, api_url, params):
    resp = session.get(api_url, params=params, timeout=45)
    resp.raise_for_status()
    payload = resp.json()
    if isinstance(payload, dict) and payload.get("HasErrors"):
        raise ReviewDownloaderError(f"BV error: {payload.get('Errors')}")
    return payload


def _fetch_powerreviews_page(session, *, merchant_id, locale, product_id,
                             apikey, paging_from=0, page_size=POWERREVIEWS_MAX_PAGE_SIZE,
                             sort="Newest", filters="", search="",
                             image_only=False):
    endpoint = POWERREVIEWS_ENDPOINT_TEMPLATE.format(
        merchant_id=merchant_id,
        locale=locale,
        product_id=product_id,
    )
    safe_page_size = max(1, min(int(page_size or POWERREVIEWS_MAX_PAGE_SIZE), POWERREVIEWS_MAX_PAGE_SIZE))
    params = {
        "paging.from": int(paging_from),
        "paging.size": safe_page_size,
        "filters": filters or "",
        "search": search or "",
        "sort": sort or "Newest",
        "image_only": "true" if image_only else "false",
        "page_locale": locale,
        "_noconfig": "true",
        "apikey": apikey,
    }
    resp = session.get(endpoint, params=params, timeout=45)
    resp.raise_for_status()
    payload = resp.json()
    if isinstance(payload, dict) and payload.get("errors"):
        raise ReviewDownloaderError(f"PowerReviews error: {payload.get('errors')}")
    return payload


def _find_ci_key(mapping, candidates):
    wanted = {str(c).lower() for c in candidates}
    for k in mapping.keys():
        if str(k).lower() in wanted:
            return k
    return None


def _query_first_ci(mapping, candidates, default=None):
    key = _find_ci_key(mapping, candidates)
    if key is None:
        return default
    val = mapping.get(key)
    if isinstance(val, list):
        return val[0] if val else default
    return val if val not in (None, "") else default


def _set_ci_param(mapping, candidates, value):
    key = _find_ci_key(mapping, candidates) or list(candidates)[0]
    mapping[key] = value
    return key


def _clone_params(mapping):
    out = {}
    for k, v in mapping.items():
        out[k] = list(v) if isinstance(v, list) else v
    return out


def _dedupe_keep_order(values):
    out = []
    seen = set()
    for v in values or []:
        s = str(v).strip()
        if not s or s in seen:
            continue
        seen.add(s)
        out.append(s)
    return out


def _normalize_input_url(product_url: str) -> str:
    product_url = (product_url or "").strip()
    if not product_url:
        return ""
    if not re.match(r"^https?://", product_url, flags=re.IGNORECASE):
        product_url = "https://" + product_url
    parsed = urlparse(product_url)
    if (parsed.scheme or "").lower() not in {"http", "https"}:
        raise ReviewDownloaderError("Only http(s) product/review URLs are supported.")
    host = (parsed.hostname or "").strip().lower()
    if not host:
        raise ReviewDownloaderError("Enter a valid public product/review URL.")
    if host in {"localhost", "0.0.0.0"} or host.endswith(".localhost") or host.endswith(".local"):
        raise ReviewDownloaderError("Local or private URLs are blocked.")
    try:
        ip = ipaddress.ip_address(host)
    except ValueError:
        ip = None
    if ip is not None and (ip.is_private or ip.is_loopback or ip.is_link_local or ip.is_reserved or ip.is_multicast or ip.is_unspecified):
        raise ReviewDownloaderError("Local or private URLs are blocked.")
    return product_url


def _strip_www(host: str) -> str:
    host = (host or "").lower().strip()
    return host[4:] if host.startswith("www.") else host


def _host_matches(host: str, tokens: Sequence[str]) -> bool:
    h = _strip_www(host)
    return any(tok in h for tok in tokens)


def _is_bazaarvoice_api_url(url: str) -> bool:
    parsed = urlparse(url)
    return "api.bazaarvoice.com" in (parsed.netloc or "").lower() and parsed.path.endswith("/reviews.json")


def _is_powerreviews_api_url(url: str) -> bool:
    parsed = urlparse(url)
    return "display.powerreviews.com" in (parsed.netloc or "").lower() and "/product/" in (parsed.path or "") and parsed.path.endswith("/reviews")


def _looks_like_sharkninja_uk_eu(host: str) -> bool:
    h = _strip_www(host)
    if h in {
        "sharkninja.co.uk", "sharkninja.eu", "sharkninja.de", "sharkninja.fr", "sharkninja.es", "sharkninja.it", "sharkninja.nl", "sharkninja.ie",
        "sharkclean.co.uk", "ninjakitchen.co.uk", "sharkclean.eu", "ninjakitchen.eu",
        "sharkclean.de", "ninjakitchen.de", "sharkclean.fr", "ninjakitchen.fr",
        "sharkclean.nl", "ninjakitchen.nl", "sharkclean.ie", "ninjakitchen.ie",
    }:
        return True
    return ("sharkclean" in h or "ninjakitchen" in h or "sharkninja" in h) and not h.endswith(".com")


def _looks_like_sharkninja_us(host: str) -> bool:
    h = _strip_www(host)
    return h in {"sharkclean.com", "www.sharkclean.com", "ninjakitchen.com", "www.ninjakitchen.com", "sharkninja.com", "www.sharkninja.com"} or h.endswith("sharkclean.com") or h.endswith("ninjakitchen.com")


def _safe_candidate_token(raw: Any) -> Optional[str]:
    s = _safe_text(raw)
    s = s.strip().strip("\"' ")
    s = re.sub(r"\.html?$", "", s, flags=re.IGNORECASE)
    s = s.strip()
    if not s or len(s) < 3 or len(s) > 80:
        return None
    if re.search(r"\s", s):
        return None
    if s.lower() in {"product", "products", "reviews", "review", "home", "en", "us", "uk", "gb"}:
        return None
    return s


def _extract_embedded_review_api_urls(html: str) -> List[str]:
    text = (html or "")
    if not text:
        return []
    norm = (
        text.replace(r"\/", "/")
            .replace("&amp;", "&")
            .replace(r"\u0026", "&")
            .replace(r"\x26", "&")
    )
    hits: List[str] = []
    patterns = [
        r"https://api\.bazaarvoice\.com/data/reviews\.json[^\"'<>\s]+",
        r"https://display\.powerreviews\.com/m/\d+/l/[A-Za-z_]+/product/[^\"'<>\s]+/reviews[^\"'<>\s]*",
        r"//display\.powerreviews\.com/m/\d+/l/[A-Za-z_]+/product/[^\"'<>\s]+/reviews[^\"'<>\s]*",
        r"/m/\d+/l/[A-Za-z_]+/product/[^\"'<>\s]+/reviews[^\"'<>\s]*apikey=[^\"'<>\s]+",
    ]
    for pat in patterns:
        for match in re.findall(pat, norm, flags=re.IGNORECASE):
            url = str(match).strip().strip('"\' ,)')
            if url.startswith("//"):
                url = "https:" + url
            elif url.startswith("/m/"):
                url = "https://display.powerreviews.com" + url
            hits.append(url)
    return _dedupe_keep_order(hits)


def _extract_powerreviews_embeds(html: str) -> List[Dict[str, str]]:
    text = (html or "")
    if not text:
        return []
    norm = (
        text.replace(r"\/", "/")
            .replace("&amp;", "&")
            .replace(r"\u0026", "&")
            .replace(r"\x26", "&")
    )
    found: List[Dict[str, str]] = []
    for m in re.finditer(r"(?:https?:)?//display\.powerreviews\.com/m/(\d+)/l/([A-Za-z_]+)/product/([^/?\"'&<>]+)/reviews([^\"'<>]*)", norm, flags=re.IGNORECASE):
        merchant_id, locale, product_id, rest = m.groups()
        api_match = re.search(r"[?&]apikey=([^&#\"']+)", rest or "", flags=re.IGNORECASE)
        if merchant_id and locale and product_id and api_match:
            found.append({
                "merchant_id": merchant_id,
                "locale": locale,
                "product_id": product_id,
                "apikey": api_match.group(1),
            })
    return found


def _extract_bazaarvoice_product_id_from_params(params: Dict[str, Any]) -> Optional[str]:
    direct = _query_first_ci(params, ["productid", "productId", "ProductId", "pid", "id"])
    if direct:
        return _safe_candidate_token(direct)
    for key, vals in params.items():
        if str(key).lower() not in {"filter", "filter_reviews", "filterreviews"}:
            continue
        vals_list = vals if isinstance(vals, list) else [vals]
        for val in vals_list:
            txt = str(val)
            m = re.search(r"productid(?::eq)?:([^,&]+)", txt, flags=re.IGNORECASE)
            if m:
                return _safe_candidate_token(m.group(1))
    return None


def _extract_candidate_tokens_from_url(url: str) -> List[str]:
    parsed = urlparse(url)
    params = parse_qs(parsed.query)
    cands: List[str] = []

    for key, values in params.items():
        lk = key.lower()
        if lk in {"productid", "product_id", "itemid", "item_id", "pid", "id", "sku"}:
            cands.extend(values)
        dw = re.match(r"dwvar_([^_]+)_", key, flags=re.IGNORECASE)
        if dw:
            cands.append(dw.group(1))

    path = parsed.path or ""
    zid_match = re.search(r"[_\-]zid([A-Za-z0-9\-_]+)$", path, flags=re.IGNORECASE)
    if zid_match:
        cands.append(zid_match.group(1))

    segments = [s for s in path.split("/") if s]
    if segments:
        last = re.sub(r"\.html?$", "", segments[-1], flags=re.IGNORECASE)
        cands.append(last)
        for token in re.findall(r"([A-Za-z0-9]{4,30})", last):
            cands.append(token)
        if len(segments) >= 2:
            cands.append(re.sub(r"\.html?$", "", segments[-2], flags=re.IGNORECASE))

    out = []
    for raw in cands:
        tok = _safe_candidate_token(raw)
        if tok:
            out.append(tok)
    return _dedupe_keep_order(out)


def _extract_candidate_tokens_from_html(html: str) -> List[str]:
    if not html:
        return []
    text = html.replace(r"\/", "/")
    soup = BeautifulSoup(html, "html.parser")
    visible = soup.get_text(" ", strip=True)
    cands: List[str] = []

    patterns = [
        r'Item\s*No\.?\s*([A-Za-z0-9_-]{4,40})',
        r'Item\s*([A-Za-z0-9_-]{4,40})',
        r'"productId"\s*[:=]\s*"([A-Za-z0-9_-]{3,40})"',
        r'"product_id"\s*[:=]\s*"([A-Za-z0-9_-]{3,40})"',
        r'"ProductId"\s*[:=]\s*"([A-Za-z0-9_-]{3,40})"',
        r'"page_id"\s*[:=]\s*"([A-Za-z0-9_-]{3,40})"',
        r'"product_page_id"\s*[:=]\s*"([A-Za-z0-9_-]{3,40})"',
        r'"masterId"\s*[:=]\s*"([A-Za-z0-9_-]{3,40})"',
        r'"styleNumber"\s*[:=]\s*"([A-Za-z0-9_-]{3,40})"',
        r'data-product-id=["\']([^"\']+)',
        r'data-sku=["\']([^"\']+)',
        r'display\.powerreviews\.com/m/\d+/l/[A-Za-z_]+/product/([^/\"\'&?<>]+)/reviews',
        r'api\.bazaarvoice\.com/data/reviews\.json[^\"\']*productid(?::eq)?:([^,&\"\']+)',
        r'\b(P\d{5,10})\b',
        r'\b(pimprod\d{5,12})\b',
        r'\b(xlsImpprod\d{5,12})\b',
    ]
    for pat in patterns:
        for match in re.findall(pat, text, flags=re.IGNORECASE):
            cands.append(match)
        for match in re.findall(pat, visible, flags=re.IGNORECASE):
            cands.append(match)

    # Look for numeric or SKU-like ids near review / powerreviews / bazaarvoice references.
    windows = re.findall(r"(?i)(.{0,140}(?:powerreviews|bazaarvoice|reviews|review snapshot).{0,240})", text)
    for window in windows:
        cands.extend(re.findall(r"\b([A-Za-z]*\d[A-Za-z0-9_-]{3,30})\b", window))

    out = []
    for raw in cands:
        tok = _safe_candidate_token(raw)
        if tok:
            out.append(tok)
    return _dedupe_keep_order(out)


def _extract_generic_bv_product_id(url: str, html: str) -> Optional[str]:
    parsed = urlparse(url.strip())
    params = parse_qs(parsed.query)
    for key in ["productId", "product_id", "itemId", "item_id", "pid", "id", "sku"]:
        if key in params and params[key]:
            tok = _safe_candidate_token(params[key][0])
            if tok:
                return tok
    zid_match = re.search(r"[_\-]zid([A-Z0-9\-_]+)$", parsed.path, re.IGNORECASE)
    if zid_match:
        tok = _safe_candidate_token(zid_match.group(1))
        if tok:
            return tok
    html_cands = _extract_candidate_tokens_from_html(html)
    for tok in html_cands:
        if re.fullmatch(r"[A-Za-z0-9_-]{4,40}", tok):
            return tok
    segments = [s for s in parsed.path.split("/") if s]
    if segments:
        last = re.sub(r"\.html?$", "", segments[-1], flags=re.IGNORECASE)
        trailing = re.search(r"([A-Z0-9]{4,20})$", last, re.IGNORECASE)
        if trailing:
            return trailing.group(1)
        tok = _safe_candidate_token(last)
        if tok:
            return tok
    return None


def _is_incentivized(r):
    badges = [str(b).lower() for b in (r.get("BadgesOrder") or [])]
    if any("incentivized" in b for b in badges):
        return True
    ctx = r.get("ContextDataValues") or {}
    if isinstance(ctx, dict):
        for k, v in ctx.items():
            if "incentivized" in str(k).lower():
                flag = str((v.get("Value", "") if isinstance(v, dict) else v)).strip().lower()
                if flag in {"", "true", "1", "yes"}:
                    return True
    return False

def _flatten_review(r):
    photos = r.get("Photos") or []
    urls = []
    for p in photos:
        sz = p.get("Sizes") or {}
        for sn in ["large", "normal", "thumbnail"]:
            u = (sz.get(sn) or {}).get("Url")
            if u:
                urls.append(u)
                break
    syn = r.get("SyndicationSource") or {}
    return dict(
        review_id=r.get("Id"),
        product_id=r.get("ProductId"),
        original_product_name=r.get("OriginalProductName"),
        title=_safe_text(r.get("Title")),
        review_text=_safe_text(r.get("ReviewText")),
        rating=r.get("Rating"),
        is_recommended=r.get("IsRecommended"),
        user_nickname=r.get("UserNickname"),
        author_id=r.get("AuthorId"),
        user_location=r.get("UserLocation"),
        content_locale=r.get("ContentLocale"),
        submission_time=r.get("SubmissionTime"),
        moderation_status=r.get("ModerationStatus"),
        campaign_id=r.get("CampaignId"),
        source_client=r.get("SourceClient"),
        is_featured=r.get("IsFeatured"),
        is_syndicated=r.get("IsSyndicated"),
        syndication_source_name=syn.get("Name"),
        is_ratings_only=r.get("IsRatingsOnly"),
        total_positive_feedback_count=r.get("TotalPositiveFeedbackCount"),
        badges=", ".join(str(x) for x in (r.get("BadgesOrder") or [])),
        context_data_json=json.dumps(r.get("ContextDataValues") or {}, ensure_ascii=False),
        photos_count=len(photos),
        photo_urls=" | ".join(urls),
        incentivized_review=_is_incentivized(r),
        raw_json=json.dumps(r, ensure_ascii=False),
    )

def _ensure_cols(df, cols):
    for c in cols:
        if c not in df.columns:
            df[c] = pd.NA
    return df

def _extract_age_group(val):
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return None
    payload = val
    if isinstance(payload, str):
        stripped = payload.strip()
        if not stripped:
            return None
        try:
            payload = json.loads(stripped)
        except Exception:
            return None
    if not isinstance(payload, dict):
        return None
    for k, raw in payload.items():
        if "age" not in str(k).lower():
            continue
        candidate = raw.get("Value") or raw.get("Label") if isinstance(raw, dict) else raw
        candidate = _safe_text(candidate)
        if candidate and candidate.lower() not in {"nan", "none", "null", "unknown", "prefer not to say"}:
            return candidate
    return None

def _finalize_df(df):
    required = [
        "review_id", "product_id", "base_sku", "sku_item", "product_or_sku",
        "original_product_name", "title", "review_text", "rating", "is_recommended",
        "content_locale", "submission_time", "submission_date", "submission_month",
        "incentivized_review", "is_syndicated", "photos_count", "photo_urls",
        "title_and_text", "retailer", "post_link", "age_group", "user_nickname",
        "user_location", "total_positive_feedback_count", "source_system", "source_file",
    ]
    df = _ensure_cols(df.copy(), required)
    if df.empty:
        for c in ["has_photos", "has_media", "review_length_chars", "review_length_words", "rating_label", "year_month_sort"]:
            if c not in df.columns:
                df[c] = pd.Series(dtype="object")
        return df

    df["review_id"] = df["review_id"].fillna("").astype(str).str.strip()
    missing = df["review_id"].eq("") | df["review_id"].str.lower().isin({"nan", "none", "null"})
    if missing.any():
        df.loc[missing, "review_id"] = [f"review_{i + 1}" for i in range(int(missing.sum()))]
    if "context_data_json" in df.columns:
        df["age_group"] = df["age_group"].fillna(df["context_data_json"].map(_extract_age_group))
    df["rating"] = pd.to_numeric(df["rating"], errors="coerce")
    df["incentivized_review"] = df["incentivized_review"].astype("boolean").fillna(False).astype(bool)
    df["is_syndicated"] = df["is_syndicated"].astype("boolean").fillna(False).astype(bool)
    df["photos_count"] = pd.to_numeric(df["photos_count"], errors="coerce").fillna(0).astype(int)
    df["title"] = df["title"].fillna("").astype(str)
    df["review_text"] = df["review_text"].fillna("").astype(str)
    df["submission_time"] = pd.to_datetime(df["submission_time"], errors="coerce", utc=True).dt.tz_convert(None)
    df["submission_date"] = df["submission_time"].dt.date
    df["submission_month"] = df["submission_time"].dt.to_period("M").astype(str)
    df["content_locale"] = df["content_locale"].fillna("").astype(str).replace({"": pd.NA})
    df["base_sku"] = df.get("base_sku", pd.Series(dtype="str")).fillna("").astype(str).str.strip()
    df["sku_item"] = df.get("sku_item", pd.Series(dtype="str")).fillna("").astype(str).str.strip()
    df["product_id"] = df["product_id"].fillna("").astype(str).str.strip()
    fallback = df["base_sku"].where(df["base_sku"].ne(""), df["product_id"])
    df["product_or_sku"] = df["sku_item"].where(df["sku_item"].ne(""), fallback)
    df["product_or_sku"] = df["product_or_sku"].fillna("").astype(str).str.strip().replace({"": pd.NA})
    df["title_and_text"] = (df["title"].str.strip() + " " + df["review_text"].str.strip()).str.strip()
    df["has_photos"] = df["photos_count"] > 0
    df["has_media"] = df["has_photos"]
    df["review_length_chars"] = df["review_text"].str.len()
    df["review_length_words"] = df["review_text"].str.split().str.len().fillna(0).astype(int)
    df["rating_label"] = df["rating"].map(lambda x: f"{int(x)} star" if pd.notna(x) else "Unknown")
    df["year_month_sort"] = pd.to_datetime(df["submission_month"], format="%Y-%m", errors="coerce")
    sc = [c for c in ["submission_time", "review_id"] if c in df.columns]
    if sc:
        df = df.sort_values(sc, ascending=[False, False], na_position="last").reset_index(drop=True)
    return df

def _pick_col(df, aliases):
    lk = {str(c).strip().lower(): c for c in df.columns}
    for a in aliases:
        c = lk.get(str(a).strip().lower())
        if c:
            return c
    return None


UPLOAD_REVIEW_ID_ALIASES = ["Event Id", "Event ID", "Review ID", "Review Id", "Verbatim Id", "Verbatim ID", "Id", "review_id"]
UPLOAD_REVIEW_TEXT_ALIASES = ["Review Text", "Review", "Verbatim", "Body", "Content", "review_text"]
UPLOAD_TITLE_ALIASES = ["Title", "Review Title", "Review title", "Headline", "title"]
UPLOAD_RATING_ALIASES = ["Rating (num)", "Rating", "Stars", "Star Rating", "rating"]
UPLOAD_DATE_ALIASES = ["Opened date", "Opened Date", "Submission Time", "Review Date", "Date", "submission_time"]
LOCAL_SYMPTOM_META_ALIASES = {
    "AI Safety": ["AI Safety", "Safety"],
    "AI Reliability": ["AI Reliability", "Reliability"],
    "AI # of Sessions": ["AI # of Sessions", "# of Sessions", "Number of Sessions", "Sessions"],
}
SYMPTOM_NON_VALUES = set(NON_VALUES) | {"NOT MENTIONED"}


def _series_alias(df, aliases):
    c = _pick_col(df, aliases)
    if c is None:
        return pd.Series([pd.NA] * len(df), index=df.index)
    return df[c]


def _parse_flag(v, *, pos, neg):
    t = _safe_text(v).lower()
    if t in {"", "nan", "none", "null", "n/a"}:
        return pd.NA
    if any(t == x.lower() for x in neg):
        return False
    if any(t == x.lower() for x in pos):
        return True
    if t.startswith(("not ", "non ")):
        return False
    return True


def _local_symptom_columns(columns):
    out = []
    for col in columns:
        name = str(col).strip()
        lower = name.lower()
        if lower.startswith("ai symptom detractor") or lower.startswith("ai symptom delighter"):
            out.append(name)
            continue
        if re.fullmatch(r"symptom\s+(?:[1-9]|10|1[1-9]|20)", lower):
            out.append(name)
    return out


def _normalize_symptom_series(series):
    text = series.astype("string").fillna("").str.strip()
    valid = (text != "") & (~text.str.upper().isin(SYMPTOM_NON_VALUES)) & (~text.str.startswith("<"))
    return text.where(valid, pd.NA).str.title()


def _score_uploaded_sheet(columns):
    lowered = {str(col).strip().lower() for col in columns}
    score = 0
    if any(alias.lower() in lowered for alias in UPLOAD_REVIEW_TEXT_ALIASES):
        score += 4
    if any(alias.lower() in lowered for alias in UPLOAD_RATING_ALIASES):
        score += 3
    if any(alias.lower() in lowered for alias in UPLOAD_REVIEW_ID_ALIASES):
        score += 3
    if any(alias.lower() in lowered for alias in UPLOAD_TITLE_ALIASES):
        score += 1
    if any(alias.lower() in lowered for alias in UPLOAD_DATE_ALIASES):
        score += 1
    if _local_symptom_columns(list(columns)):
        score += 2
    return score


def _read_best_uploaded_excel_sheet(raw_bytes):
    bio = io.BytesIO(raw_bytes)
    xls = pd.ExcelFile(bio)
    if not xls.sheet_names:
        return pd.DataFrame(), ""
    best_name = xls.sheet_names[0]
    best_score = -1
    for sheet_name in xls.sheet_names:
        try:
            headers = list(pd.read_excel(xls, sheet_name=sheet_name, nrows=0).columns)
        except Exception:
            headers = []
        score = _score_uploaded_sheet(headers)
        if score > best_score:
            best_score = score
            best_name = sheet_name
    return pd.read_excel(xls, sheet_name=best_name), best_name


def _best_uploaded_excel_sheet_name(raw_bytes):
    try:
        _, sheet_name = _read_best_uploaded_excel_sheet(raw_bytes)
        return sheet_name
    except Exception:
        return ""


def _normalize_uploaded_df(raw, *, source_name="", include_local_symptomization=False):
    w = raw.copy()
    w.columns = [str(c).strip() for c in w.columns]
    n = pd.DataFrame(index=w.index)
    n["review_id"] = _series_alias(w, UPLOAD_REVIEW_ID_ALIASES)
    n["product_id"] = _series_alias(w, ["Base SKU", "Model (SKU)", "Model SKU", "Product ID", "Product Id", "ProductId", "BaseSKU"])
    n["base_sku"] = _series_alias(w, ["Base SKU", "Model (SKU)", "Model SKU", "BaseSKU"])
    n["sku_item"] = _series_alias(w, ["SKU Item", "Model (SKU)", "Model SKU", "SKU", "Child SKU", "Variant SKU", "Item Number", "Item No"])
    n["original_product_name"] = _series_alias(w, ["Product Name", "Product", "Name"])
    n["review_text"] = _series_alias(w, UPLOAD_REVIEW_TEXT_ALIASES)
    n["title"] = _series_alias(w, UPLOAD_TITLE_ALIASES)
    n["post_link"] = _series_alias(w, ["Post Link", "Web Link", "URL", "Review URL", "Product URL"])
    n["rating"] = _series_alias(w, UPLOAD_RATING_ALIASES)
    n["submission_time"] = _series_alias(w, UPLOAD_DATE_ALIASES)
    n["content_locale"] = _series_alias(w, ["Content Locale", "Locale", "Reviewer Location", "Location", "Country"])
    n["retailer"] = _series_alias(w, ["Retailer", "Merchant", "Channel", "Source"])
    n["age_group"] = _series_alias(w, ["Age Group", "Age", "Age Range"])
    n["user_location"] = _series_alias(w, ["Reviewer Location", "Location", "Country"])
    n["user_nickname"] = pd.NA
    n["total_positive_feedback_count"] = pd.NA
    n["is_recommended"] = pd.NA
    n["photos_count"] = 0
    n["photo_urls"] = pd.NA
    n["source_file"] = source_name or pd.NA
    n["source_system"] = "Uploaded file"
    seeded = _series_alias(w, ["Seeded Flag", "Seeded", "Incentivized"])
    n["incentivized_review"] = seeded.map(lambda v: _parse_flag(v,
        pos=["seeded", "incentivized", "yes", "true", "1"],
        neg=["not seeded", "not incentivized", "no", "false", "0"]))
    syndicated = _series_alias(w, ["Syndicated Flag", "Syndicated"])
    n["is_syndicated"] = syndicated.map(lambda v: _parse_flag(v,
        pos=["syndicated", "yes", "true", "1"],
        neg=["not syndicated", "no", "false", "0"]))
    if include_local_symptomization:
        for col in _local_symptom_columns(list(w.columns)):
            n[col] = _normalize_symptom_series(w[col])
        for target, aliases in LOCAL_SYMPTOM_META_ALIASES.items():
            source = _pick_col(w, aliases)
            if source is not None:
                n[target] = w[source].astype("string").fillna("").str.strip().replace({"": pd.NA})
    return _finalize_df(n)


def _read_uploaded_file(f, *, include_local_symptomization=False):
    fname = getattr(f, "name", "uploaded_file")
    raw = f.getvalue()
    max_upload_mb = float(os.getenv("STARWALK_MAX_UPLOAD_MB", "40") or 40)
    max_upload_bytes = int(max_upload_mb * 1024 * 1024)
    if max_upload_bytes > 0 and len(raw) > max_upload_bytes:
        raise ReviewDownloaderError(f"{fname} exceeds the upload limit of {max_upload_mb:g} MB.")
    suffix = fname.lower().rsplit(".", 1)[-1] if "." in fname else "csv"
    if suffix == "csv":
        try:
            raw_df = pd.read_csv(io.BytesIO(raw))
        except UnicodeDecodeError:
            raw_df = pd.read_csv(io.BytesIO(raw), encoding="latin-1")
    elif suffix in {"xlsx", "xls", "xlsm"}:
        raw_df, sheet_name = _read_best_uploaded_excel_sheet(raw)
        raw_df.attrs["source_sheet_name"] = sheet_name
    else:
        raise ReviewDownloaderError(f"Unsupported: {fname}")
    if raw_df.empty:
        raise ReviewDownloaderError(f"{fname} is empty.")
    normalized = _normalize_uploaded_df(raw_df, source_name=fname, include_local_symptomization=include_local_symptomization)
    source_sheet_name = raw_df.attrs.get("source_sheet_name")
    if source_sheet_name:
        normalized.attrs["source_sheet_name"] = source_sheet_name
    return normalized


def _load_uploaded_files(files, *, include_local_symptomization=False):
    if not files:
        raise ReviewDownloaderError("Upload at least one file.")
    with st.spinner("Reading files…"):
        frames = [_read_uploaded_file(f, include_local_symptomization=include_local_symptomization) for f in files]
    combined = pd.concat(frames, ignore_index=True)
    combined["review_id"] = combined["review_id"].astype(str)
    combined = combined.drop_duplicates(subset=["review_id"], keep="first").reset_index(drop=True)
    combined = _finalize_df(combined)
    pid = (
        _first_non_empty(combined["base_sku"].fillna("")) or
        _first_non_empty(combined["product_id"].fillna("")) or
        "UPLOADED_REVIEWS"
    )
    names = [getattr(f, "name", "file") for f in files]
    src = names[0] if len(names) == 1 else f"{len(names)} uploaded files"
    summary = ReviewBatchSummary(
        product_url="",
        product_id=pid,
        total_reviews=len(combined),
        page_size=max(len(combined), 1),
        requests_needed=0,
        reviews_downloaded=len(combined),
    )
    source_sheet_name = frames[0].attrs.get("source_sheet_name") if len(frames) == 1 else ""
    return dict(summary=summary, reviews_df=combined, source_type="uploaded", source_label=src, source_sheet_name=source_sheet_name)


def _apply_source_metadata(df: pd.DataFrame, *, retailer: str = "", source_system: str = "", post_link: str = "") -> pd.DataFrame:

    out = df.copy()
    if retailer:
        if "retailer" not in out.columns:
            out["retailer"] = retailer
        else:
            out["retailer"] = out["retailer"].fillna("").astype(str)
            out.loc[out["retailer"].str.strip().eq(""), "retailer"] = retailer
    if source_system:
        out["source_system"] = source_system
    if post_link:
        if "post_link" not in out.columns:
            out["post_link"] = post_link
        else:
            out["post_link"] = out["post_link"].fillna("").astype(str)
            out.loc[out["post_link"].str.strip().eq(""), "post_link"] = post_link
    return out


def _build_bv_dataset(raw_reviews: List[Dict[str, Any]], *, product_url: str, product_id: str,
                      total: int, page_size: int, requests_needed: int,
                      source_label: str, retailer: str = "", source_system: str = "Bazaarvoice"):
    df = _finalize_df(pd.DataFrame([_flatten_review(r) for r in raw_reviews]))
    if not df.empty:
        df["review_id"] = df["review_id"].astype(str)
        df["product_or_sku"] = df.get("product_or_sku", pd.Series(index=df.index, dtype="object")).fillna(product_id)
        df["base_sku"] = df.get("base_sku", pd.Series(index=df.index, dtype="object")).fillna(product_id)
        df["product_id"] = df["product_id"].fillna(product_id)
        df = _apply_source_metadata(df, retailer=retailer, source_system=source_system, post_link=product_url)
    summary = ReviewBatchSummary(
        product_url=product_url,
        product_id=product_id,
        total_reviews=total,
        page_size=page_size,
        requests_needed=requests_needed,
        reviews_downloaded=len(df),
    )
    return dict(summary=summary, reviews_df=df, source_type="bazaarvoice", source_label=source_label or product_url)


def _powerreviews_bool_from_bottom_line(value):
    t = _safe_text(value).lower()
    if t in {"yes", "recommended", "true"}:
        return True
    if t in {"no", "false", "not recommended"}:
        return False
    return pd.NA


def _powerreviews_media_urls(media_items):
    urls = []
    for item in media_items or []:
        if not isinstance(item, dict):
            continue
        for key in [
            "large_url", "normal_url", "thumbnail_url", "url",
            "fullsize_url", "media_url", "src", "link",
        ]:
            val = item.get(key)
            if val:
                urls.append(str(val))
                break
    return urls


def _powerreviews_submission_iso(value):
    ts = pd.to_datetime(value, unit="ms", utc=True, errors="coerce")
    if pd.isna(ts):
        return None
    try:
        return ts.isoformat()
    except Exception:
        return str(ts)


def _flatten_powerreviews_review(review, *, page_id="", product_name="", retailer="", product_url=""):
    details = review.get("details") or {}
    metrics = review.get("metrics") or {}
    media = review.get("media") or []
    photo_urls = _powerreviews_media_urls(media)
    product_id = _safe_text(details.get("product_page_id") or review.get("page_id") or page_id)

    incentivized = False
    badges = review.get("badges") or {}
    for key, val in badges.items():
        lk = str(key).lower()
        if any(tok in lk for tok in ["sampling", "sample", "sweepstakes", "incentivized", "influencer"]):
            if bool(val):
                incentivized = True
                break

    return dict(
        review_id=review.get("review_id") or review.get("ugc_id") or review.get("internal_review_id") or review.get("legacy_id"),
        product_id=product_id,
        base_sku=product_id,
        sku_item=product_id,
        original_product_name=_safe_text(product_name),
        title=_safe_text(details.get("headline") or review.get("headline")),
        review_text=_safe_text(details.get("comments") or review.get("comments")),
        rating=metrics.get("rating") or details.get("rating"),
        is_recommended=_powerreviews_bool_from_bottom_line(details.get("bottom_line") or review.get("bottom_line")),
        user_nickname=_safe_text(details.get("nickname") or review.get("nickname")),
        author_id=review.get("author_id"),
        user_location=_safe_text(details.get("location") or review.get("location")),
        content_locale=_safe_text(details.get("locale") or review.get("locale")),
        submission_time=_powerreviews_submission_iso(details.get("created_date") or review.get("created_date")),
        moderation_status=review.get("status") or pd.NA,
        campaign_id=pd.NA,
        source_client=pd.NA,
        is_featured=pd.NA,
        is_syndicated=bool(review.get("is_syndicated") or badges.get("is_syndicated") or False),
        syndication_source_name=pd.NA,
        is_ratings_only=False,
        total_positive_feedback_count=metrics.get("helpful_votes"),
        badges=", ".join([str(k) for k, v in (badges or {}).items() if bool(v)]),
        context_data_json=json.dumps(details.get("properties") or [], ensure_ascii=False),
        photos_count=len(photo_urls),
        photo_urls=" | ".join(photo_urls),
        incentivized_review=incentivized,
        raw_json=json.dumps(review, ensure_ascii=False),
        retailer=retailer,
        post_link=product_url,
        source_system="PowerReviews",
    )


def _build_powerreviews_dataset(reviews: List[Dict[str, Any]], *, product_url: str, product_id: str,
                                total: int, page_size: int, requests_needed: int,
                                source_label: str, product_name: str = "", retailer: str = "",
                                source_system: str = "PowerReviews"):
    rows = [_flatten_powerreviews_review(r, page_id=product_id, product_name=product_name, retailer=retailer, product_url=product_url) for r in reviews]
    df = _finalize_df(pd.DataFrame(rows))
    if not df.empty:
        df = _apply_source_metadata(df, retailer=retailer, source_system=source_system, post_link=product_url)
    summary = ReviewBatchSummary(
        product_url=product_url,
        product_id=product_id,
        total_reviews=total,
        page_size=page_size,
        requests_needed=requests_needed,
        reviews_downloaded=len(df),
    )
    return dict(summary=summary, reviews_df=df, source_type="powerreviews", source_label=source_label or product_url)


def _load_bazaarvoice_api_url(api_url: str, *, product_url_hint: str = "", retailer_hint: str = ""):
    session = _get_session()
    parsed = urlparse(api_url)
    base_url = f"{parsed.scheme}://{parsed.netloc}{parsed.path}"
    params = parse_qs(parsed.query, keep_blank_values=True)
    page_size = int(_query_first_ci(params, ["limit", "Limit"], default=DEFAULT_PAGE_SIZE) or DEFAULT_PAGE_SIZE)
    product_id = _extract_bazaarvoice_product_id_from_params(params) or "UNKNOWN_PRODUCT"
    first = _fetch_bazaarvoice_raw_page(session, api_url=base_url, params=params)
    total = int(first.get("TotalResults", 0) or 0)
    raw_reviews = list(first.get("Results") or [])
    if total > len(raw_reviews):
        offsets = list(range(len(raw_reviews), total, page_size))
        progress = st.progress(0.0, text="Downloading…")
        for i, offset in enumerate(offsets, 1):
            q = _clone_params(params)
            _set_ci_param(q, ["offset", "Offset"], int(offset))
            _set_ci_param(q, ["limit", "Limit"], int(page_size))
            payload = _fetch_bazaarvoice_raw_page(session, api_url=base_url, params=q)
            raw_reviews.extend(payload.get("Results") or [])
            progress.progress(i / max(len(offsets), 1))
    source_label = product_url_hint or api_url
    return _build_bv_dataset(
        raw_reviews,
        product_url=product_url_hint or api_url,
        product_id=product_id,
        total=total,
        page_size=page_size,
        requests_needed=max(1, math.ceil(total / max(page_size, 1))) if total else 1,
        source_label=source_label,
        retailer=retailer_hint,
        source_system="Bazaarvoice API",
    )


def _load_powerreviews_api_url(api_url: str, *, product_url_hint: str = "", retailer_hint: str = ""):
    session = _get_session()
    parsed = urlparse(api_url)
    m = re.search(r"/m/(\d+)/l/([A-Za-z_]+)/product/([^/]+)/reviews", parsed.path)
    if not m:
        raise ReviewDownloaderError("Could not parse PowerReviews API URL.")
    merchant_id, locale, product_id = m.groups()
    params = parse_qs(parsed.query, keep_blank_values=True)
    apikey = _query_first_ci(params, ["apikey"])
    if not apikey:
        raise ReviewDownloaderError("PowerReviews API URL missing apikey.")
    sort = _query_first_ci(params, ["sort"], default="Newest") or "Newest"
    page_size = min(int(_query_first_ci(params, ["paging.size"], default=POWERREVIEWS_MAX_PAGE_SIZE) or POWERREVIEWS_MAX_PAGE_SIZE), POWERREVIEWS_MAX_PAGE_SIZE)
    first = _fetch_powerreviews_page(
        session,
        merchant_id=merchant_id,
        locale=locale,
        product_id=product_id,
        apikey=apikey,
        paging_from=0,
        page_size=page_size,
        sort=sort,
    )
    paging = first.get("paging") or {}
    total = int(paging.get("total_results", 0) or 0)
    results = first.get("results") or []
    product_name = _safe_text((results[0].get("rollup") or {}).get("name") if results else "")
    all_reviews: List[Dict[str, Any]] = []
    for result in results:
        all_reviews.extend(result.get("reviews") or [])
    if total > len(all_reviews):
        offsets = list(range(len(all_reviews), total, page_size))
        progress = st.progress(0.0, text="Downloading…")
        for i, start in enumerate(offsets, 1):
            payload = _fetch_powerreviews_page(
                session,
                merchant_id=merchant_id,
                locale=locale,
                product_id=product_id,
                apikey=apikey,
                paging_from=int(start),
                page_size=page_size,
                sort=sort,
            )
            for result in payload.get("results") or []:
                all_reviews.extend(result.get("reviews") or [])
            progress.progress(i / max(len(offsets), 1))
    source_label = product_url_hint or api_url
    return _build_powerreviews_dataset(
        all_reviews,
        product_url=product_url_hint or api_url,
        product_id=product_id,
        total=total,
        page_size=page_size,
        requests_needed=max(1, math.ceil(total / max(page_size, 1))) if total else 1,
        source_label=source_label,
        product_name=product_name,
        retailer=retailer_hint,
        source_system="PowerReviews API",
    )


def _probe_bazaarvoice_candidates(session, *, product_url: str, candidates: Sequence[str], cfg: Dict[str, Any]):
    tried = []
    zero_match = None
    for candidate in _dedupe_keep_order(candidates):
        try:
            if cfg.get("kind") == "action":
                payload = _fetch_reviews_page(
                    session,
                    product_id=candidate,
                    passkey=cfg["passkey"],
                    displaycode=cfg["displaycode"],
                    api_version=cfg.get("api_version", DEFAULT_API_VERSION),
                    page_size=1,
                    offset=0,
                    sort=cfg.get("sort", DEFAULT_SORT),
                    content_locales=cfg.get("content_locales", DEFAULT_CONTENT_LOCALES),
                )
            else:
                payload = _fetch_bv_simple_page(
                    session,
                    product_id=candidate,
                    passkey=cfg["passkey"],
                    api_version=cfg.get("api_version", "5.4"),
                    page_size=1,
                    offset=0,
                    sort=cfg.get("sort", "SubmissionTime:desc"),
                    content_locale=cfg.get("content_locale", ""),
                    locale=cfg.get("locale", "en_US"),
                    include=cfg.get("include", "Products,Comments"),
                )
            total = int(payload.get("TotalResults", 0) or 0)
            has_products = bool(((payload.get("Includes") or {}).get("Products") or {}))
            if total > 0 or has_products:
                return candidate, payload
            if zero_match is None:
                zero_match = (candidate, payload)
        except Exception as exc:
            tried.append(f"{candidate}: {exc}")
    if zero_match is not None:
        return zero_match
    raise ReviewDownloaderError("Could not match a Bazaarvoice product ID. Tried: " + "; ".join(tried[:8]))


def _fetch_all_bazaarvoice_for_candidate(session, *, product_url: str, product_id: str, cfg: Dict[str, Any]):
    if cfg.get("kind") == "action":
        first = _fetch_reviews_page(
            session,
            product_id=product_id,
            passkey=cfg["passkey"],
            displaycode=cfg["displaycode"],
            api_version=cfg.get("api_version", DEFAULT_API_VERSION),
            page_size=DEFAULT_PAGE_SIZE,
            offset=0,
            sort=cfg.get("sort", DEFAULT_SORT),
            content_locales=cfg.get("content_locales", DEFAULT_CONTENT_LOCALES),
        )
        total = int(first.get("TotalResults", 0) or 0)
        raw_reviews = list(first.get("Results") or [])
        offsets = list(range(len(raw_reviews), total, DEFAULT_PAGE_SIZE))
        progress = st.progress(0.0, text="Downloading…") if offsets else None
        for i, offset in enumerate(offsets, 1):
            page = _fetch_reviews_page(
                session,
                product_id=product_id,
                passkey=cfg["passkey"],
                displaycode=cfg["displaycode"],
                api_version=cfg.get("api_version", DEFAULT_API_VERSION),
                page_size=DEFAULT_PAGE_SIZE,
                offset=offset,
                sort=cfg.get("sort", DEFAULT_SORT),
                content_locales=cfg.get("content_locales", DEFAULT_CONTENT_LOCALES),
            )
            raw_reviews.extend(page.get("Results") or [])
            if progress is not None:
                progress.progress(i / max(len(offsets), 1))
        return _build_bv_dataset(
            raw_reviews,
            product_url=product_url,
            product_id=product_id,
            total=total,
            page_size=DEFAULT_PAGE_SIZE,
            requests_needed=max(1, math.ceil(total / DEFAULT_PAGE_SIZE)) if total else 1,
            source_label=product_url,
            retailer=cfg.get("retailer", ""),
            source_system=cfg.get("source_system", "Bazaarvoice"),
        )

    first = _fetch_bv_simple_page(
        session,
        product_id=product_id,
        passkey=cfg["passkey"],
        api_version=cfg.get("api_version", "5.4"),
        page_size=DEFAULT_PAGE_SIZE,
        offset=0,
        sort=cfg.get("sort", "SubmissionTime:desc"),
        content_locale=cfg.get("content_locale", ""),
        locale=cfg.get("locale", "en_US"),
        include=cfg.get("include", "Products,Comments"),
    )
    total = int(first.get("TotalResults", 0) or 0)
    raw_reviews = list(first.get("Results") or [])
    offsets = list(range(len(raw_reviews), total, DEFAULT_PAGE_SIZE))
    progress = st.progress(0.0, text="Downloading…") if offsets else None
    for i, offset in enumerate(offsets, 1):
        page = _fetch_bv_simple_page(
            session,
            product_id=product_id,
            passkey=cfg["passkey"],
            api_version=cfg.get("api_version", "5.4"),
            page_size=DEFAULT_PAGE_SIZE,
            offset=offset,
            sort=cfg.get("sort", "SubmissionTime:desc"),
            content_locale=cfg.get("content_locale", ""),
            locale=cfg.get("locale", "en_US"),
            include=cfg.get("include", "Products,Comments"),
        )
        raw_reviews.extend(page.get("Results") or [])
        if progress is not None:
            progress.progress(i / max(len(offsets), 1))
    return _build_bv_dataset(
        raw_reviews,
        product_url=product_url,
        product_id=product_id,
        total=total,
        page_size=DEFAULT_PAGE_SIZE,
        requests_needed=max(1, math.ceil(total / DEFAULT_PAGE_SIZE)) if total else 1,
        source_label=product_url,
        retailer=cfg.get("retailer", ""),
        source_system=cfg.get("source_system", "Bazaarvoice"),
    )


def _probe_powerreviews_candidates(session, *, product_url: str, candidates: Sequence[str], cfg: Dict[str, Any]):
    tried = []
    zero_match = None
    for candidate in _dedupe_keep_order(candidates):
        try:
            payload = _fetch_powerreviews_page(
                session,
                merchant_id=cfg["merchant_id"],
                locale=cfg.get("locale", "en_US"),
                product_id=candidate,
                apikey=cfg["apikey"],
                paging_from=0,
                page_size=5,
                sort=cfg.get("sort", "Newest"),
            )
            paging = payload.get("paging") or {}
            total = int(paging.get("total_results", 0) or 0)
            results = payload.get("results") or []
            if total > 0 or results:
                return candidate, payload
            if zero_match is None:
                zero_match = (candidate, payload)
        except Exception as exc:
            tried.append(f"{candidate}: {exc}")
    if zero_match is not None:
        return zero_match
    raise ReviewDownloaderError("Could not match a PowerReviews product ID. Tried: " + "; ".join(tried[:8]))


def _fetch_all_powerreviews_for_candidate(session, *, product_url: str, product_id: str, cfg: Dict[str, Any]):
    page_size = POWERREVIEWS_MAX_PAGE_SIZE
    first = _fetch_powerreviews_page(
        session,
        merchant_id=cfg["merchant_id"],
        locale=cfg.get("locale", "en_US"),
        product_id=product_id,
        apikey=cfg["apikey"],
        paging_from=0,
        page_size=page_size,
        sort=cfg.get("sort", "Newest"),
    )
    paging = first.get("paging") or {}
    total = int(paging.get("total_results", 0) or 0)
    results = first.get("results") or []
    product_name = _safe_text((results[0].get("rollup") or {}).get("name") if results else "")
    all_reviews: List[Dict[str, Any]] = []
    for result in results:
        all_reviews.extend(result.get("reviews") or [])
    offsets = list(range(len(all_reviews), total, page_size))
    progress = st.progress(0.0, text="Downloading…") if offsets else None
    for i, start in enumerate(offsets, 1):
        payload = _fetch_powerreviews_page(
            session,
            merchant_id=cfg["merchant_id"],
            locale=cfg.get("locale", "en_US"),
            product_id=product_id,
            apikey=cfg["apikey"],
            paging_from=int(start),
            page_size=page_size,
            sort=cfg.get("sort", "Newest"),
        )
        for result in payload.get("results") or []:
            all_reviews.extend(result.get("reviews") or [])
        if progress is not None:
            progress.progress(i / max(len(offsets), 1))
    return _build_powerreviews_dataset(
        all_reviews,
        product_url=product_url,
        product_id=product_id,
        total=total,
        page_size=page_size,
        requests_needed=max(1, math.ceil(total / page_size)) if total else 1,
        source_label=product_url,
        product_name=product_name,
        retailer=cfg.get("retailer", ""),
        source_system=cfg.get("source_system", "PowerReviews"),
    )


def _load_bazaarvoice_product_page(session, *, product_url: str, product_html: str, cfg: Dict[str, Any], extra_candidates: Optional[Sequence[str]] = None):
    embedded_urls = [u for u in _extract_embedded_review_api_urls(product_html) if _is_bazaarvoice_api_url(u)]
    for api_url in embedded_urls:
        try:
            return _load_bazaarvoice_api_url(api_url, product_url_hint=product_url, retailer_hint=cfg.get("retailer", ""))
        except Exception:
            continue

    candidates = []
    candidates.extend(extra_candidates or [])
    candidates.extend(_extract_candidate_tokens_from_url(product_url))
    candidates.extend(_extract_candidate_tokens_from_html(product_html))
    fallback_pid = _extract_generic_bv_product_id(product_url, product_html)
    if fallback_pid:
        candidates.insert(0, fallback_pid)
    product_id, _ = _probe_bazaarvoice_candidates(session, product_url=product_url, candidates=candidates, cfg=cfg)
    return _fetch_all_bazaarvoice_for_candidate(session, product_url=product_url, product_id=product_id, cfg=cfg)


def _load_powerreviews_product_page(session, *, product_url: str, product_html: str, cfg: Dict[str, Any], extra_candidates: Optional[Sequence[str]] = None):
    embedded_urls = [u for u in _extract_embedded_review_api_urls(product_html) if _is_powerreviews_api_url(u)]
    for api_url in embedded_urls:
        try:
            return _load_powerreviews_api_url(api_url, product_url_hint=product_url, retailer_hint=cfg.get("retailer", ""))
        except Exception:
            continue

    embeds = _extract_powerreviews_embeds(product_html)
    for embed in embeds:
        try:
            cfg2 = dict(cfg)
            cfg2.update({k: v for k, v in embed.items() if v})
            return _fetch_all_powerreviews_for_candidate(
                session,
                product_url=product_url,
                product_id=cfg2["product_id"],
                cfg=cfg2,
            )
        except Exception:
            continue

    candidates = []
    candidates.extend(extra_candidates or [])
    candidates.extend(_extract_candidate_tokens_from_html(product_html))
    candidates.extend(_extract_candidate_tokens_from_url(product_url))
    product_id, _ = _probe_powerreviews_candidates(session, product_url=product_url, candidates=candidates, cfg=cfg)
    return _fetch_all_powerreviews_for_candidate(session, product_url=product_url, product_id=product_id, cfg=cfg)


def _load_product_reviews(product_url):
    product_url = _normalize_input_url(product_url)
    parsed = urlparse(product_url)
    host = _strip_www(parsed.netloc)
    retailer_hint = ""
    if "costco.com" in host:
        retailer_hint = "Costco"
    elif "sephora.com" in host:
        retailer_hint = "Sephora"
    elif "ulta.com" in host:
        retailer_hint = "Ulta"
    elif "hoka.com" in host:
        retailer_hint = "Hoka"
    elif _looks_like_sharkninja_uk_eu(host):
        retailer_hint = "SharkNinja UK/EU"
    elif _looks_like_sharkninja_us(host):
        retailer_hint = "SharkNinja"

    if _is_bazaarvoice_api_url(product_url):
        return _load_bazaarvoice_api_url(product_url)
    if _is_powerreviews_api_url(product_url):
        return _load_powerreviews_api_url(product_url)

    session = _get_session()
    product_html = ""
    page_fetch_error = None
    with st.spinner("Loading product page…"):
        try:
            resp = session.get(product_url, timeout=35)
            resp.raise_for_status()
            product_html = resp.text or ""
        except Exception as exc:
            page_fetch_error = exc
            product_html = ""

    # First: if the page source embeds a review API URL, use that directly.
    embedded_urls = _extract_embedded_review_api_urls(product_html) if product_html else []
    if embedded_urls:
        for api_url in embedded_urls:
            try:
                if _is_bazaarvoice_api_url(api_url):
                    return _load_bazaarvoice_api_url(api_url, product_url_hint=product_url, retailer_hint=retailer_hint)
                if _is_powerreviews_api_url(api_url):
                    return _load_powerreviews_api_url(api_url, product_url_hint=product_url, retailer_hint=retailer_hint)
            except Exception:
                continue

    # Site-specific fallbacks that can work even when the retailer blocks page scraping.
    if "costco.com" in host:
        return _load_bazaarvoice_product_page(
            session,
            product_url=product_url,
            product_html=product_html,
            cfg={**COSTCO_BV_CONFIG, "kind": "action", "source_system": "Bazaarvoice", "retailer": "Costco"},
        )

    if "sephora.com" in host:
        sephora_candidates = []
        sephora_candidates.extend(re.findall(r"\b(P\d{5,10})\b", product_url, flags=re.IGNORECASE))
        sephora_candidates.extend(re.findall(r"\b(P\d{5,10})\b", product_html or "", flags=re.IGNORECASE))
        return _load_bazaarvoice_product_page(
            session,
            product_url=product_url,
            product_html=product_html,
            cfg={**SEPHORA_BV_CONFIG, "kind": "simple", "content_locale": "en*", "source_system": "Bazaarvoice", "retailer": "Sephora"},
            extra_candidates=sephora_candidates,
        )

    if "ulta.com" in host:
        ulta_candidates = []
        q = parse_qs(parsed.query)
        ulta_candidates.extend(q.get("sku", []))
        ulta_candidates.extend(re.findall(r"\b(pimprod\d{5,12})\b", product_html or "", flags=re.IGNORECASE))
        ulta_candidates.extend(re.findall(r"\b(pimprod\d{5,12})\b", product_url, flags=re.IGNORECASE))
        ulta_candidates.extend(re.findall(r"\b(xlsImpprod\d{5,12})\b", product_url, flags=re.IGNORECASE))
        return _load_powerreviews_product_page(
            session,
            product_url=product_url,
            product_html=product_html,
            cfg={**ULTA_PR_CONFIG, "source_system": "PowerReviews"},
            extra_candidates=ulta_candidates,
        )

    if "hoka.com" in host:
        hoka_candidates = []
        hoka_candidates.extend(re.findall(r"/(\d+)\.html", parsed.path))
        hoka_candidates.extend(re.findall(r"dwvar_(\d+)_", product_url))
        hoka_candidates.extend(re.findall(r"Item\s*No\.?\s*(\d{5,10})", product_html or "", flags=re.IGNORECASE))
        hoka_candidates.extend(re.findall(r'"product_page_id"\s*[:=]\s*"?(\d{5,10})', product_html or "", flags=re.IGNORECASE))
        hoka_candidates.extend(re.findall(r'"page_id"\s*[:=]\s*"?(\d{5,10})', product_html or "", flags=re.IGNORECASE))
        return _load_powerreviews_product_page(
            session,
            product_url=product_url,
            product_html=product_html,
            cfg={**HOKA_PR_CONFIG, "source_system": "PowerReviews"},
            extra_candidates=hoka_candidates,
        )

    if _looks_like_sharkninja_uk_eu(host):
        uk_pid = _extract_generic_bv_product_id(product_url, product_html)
        return _load_bazaarvoice_product_page(
            session,
            product_url=product_url,
            product_html=product_html,
            cfg={**SHARKNINJA_UK_EU_BV_CONFIG, "kind": "simple", "source_system": "Bazaarvoice"},
            extra_candidates=[uk_pid] if uk_pid else None,
        )

    # Default SharkNinja US / generic Bazaarvoice product pages.
    pid = _extract_pid_from_url(product_url) or _extract_pid_from_html(product_html)
    if pid:
        try:
            return _fetch_all_bazaarvoice_for_candidate(
                session,
                product_url=product_url,
                product_id=pid,
                cfg={
                    "kind": "action",
                    "passkey": DEFAULT_PASSKEY,
                    "displaycode": DEFAULT_DISPLAYCODE,
                    "api_version": DEFAULT_API_VERSION,
                    "sort": DEFAULT_SORT,
                    "content_locales": DEFAULT_CONTENT_LOCALES,
                    "retailer": "SharkNinja",
                    "source_system": "Bazaarvoice",
                },
            )
        except Exception:
            pass

    # Last chance: generic embedded/candidate Bazaarvoice probes.
    generic_candidates = []
    generic_candidates.extend(_extract_candidate_tokens_from_url(product_url))
    generic_candidates.extend(_extract_candidate_tokens_from_html(product_html))
    try:
        return _load_bazaarvoice_product_page(
            session,
            product_url=product_url,
            product_html=product_html,
            cfg={
                "kind": "action",
                "passkey": DEFAULT_PASSKEY,
                "displaycode": DEFAULT_DISPLAYCODE,
                "api_version": DEFAULT_API_VERSION,
                "sort": DEFAULT_SORT,
                "content_locales": DEFAULT_CONTENT_LOCALES,
                "retailer": "SharkNinja",
                "source_system": "Bazaarvoice",
            },
            extra_candidates=generic_candidates,
        )
    except Exception:
        if page_fetch_error is not None:
            if isinstance(page_fetch_error, requests.HTTPError):
                raise
            raise ReviewDownloaderError(f"Could not load product page or match a review feed: {page_fetch_error}")
        raise



def _parse_bulk_product_urls(raw_text: str) -> List[str]:
    urls: List[str] = []
    seen = set()
    for line in re.split(r"[\r\n]+", str(raw_text or "")):
        candidate = re.sub(r"^[\s\-\*\u2022\d\.)]+", "", str(line or "")).strip()
        if not candidate:
            continue
        normalized = _normalize_input_url(candidate)
        key = normalized.lower()
        if key in seen:
            continue
        seen.add(key)
        urls.append(normalized)
    return urls


def _format_multi_source_label(urls: Sequence[str]) -> str:
    hosts = []
    seen = set()
    for url in urls:
        host = _strip_www(urlparse(url).netloc) or _safe_text(url)
        if host and host not in seen:
            seen.add(host)
            hosts.append(host)
    if not hosts:
        return f"{len(urls)} links"
    if len(hosts) <= 3:
        return f"{len(urls)} links · " + ", ".join(hosts)
    return f"{len(urls)} links · " + ", ".join(hosts[:3]) + f" +{len(hosts) - 3}"


def _dedupe_combined_reviews(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df.copy()
    out = df.copy()
    if "review_id" in out.columns:
        exact_key = (
            out["review_id"].fillna("").astype(str).str.strip() + "||" +
            out.get("product_id", pd.Series("", index=out.index)).fillna("").astype(str).str.strip() + "||" +
            out.get("review_text", pd.Series("", index=out.index)).fillna("").astype(str).str.strip()
        )
        out = out.loc[~exact_key.duplicated(keep="first")].copy()
        counts = Counter()
        unique_ids = []
        for rid in out["review_id"].fillna("").astype(str).str.strip().tolist():
            base = rid or f"review_{len(unique_ids) + 1}"
            counts[base] += 1
            unique_ids.append(base if counts[base] == 1 else f"{base} ({counts[base]})")
        out["review_id"] = unique_ids
    return out.reset_index(drop=True)


def _load_multiple_product_reviews(urls: Sequence[str]):
    if isinstance(urls, str):
        url_list = _parse_bulk_product_urls(urls)
    else:
        url_list = _parse_bulk_product_urls("\n".join([str(u) for u in (urls or [])]))
    if not url_list:
        raise ReviewDownloaderError("Add at least one product or review URL.")
    if len(url_list) == 1:
        return _load_product_reviews(url_list[0])

    progress = st.progress(0.0, text="Preparing multi-link load…")
    status = st.empty()
    frames: List[pd.DataFrame] = []
    loaded: List[Dict[str, Any]] = []
    failures: List[Tuple[str, str]] = []

    for i, url in enumerate(url_list, start=1):
        status.info(f"Loading {i}/{len(url_list)} · {url}")
        try:
            ds = _load_product_reviews(url)
            frame = ds["reviews_df"].copy()
            frame["loaded_from_url"] = url
            frame["loaded_from_host"] = _strip_www(urlparse(url).netloc) or "Unknown"
            frame["loaded_from_label"] = _safe_text(ds.get("source_label")) or url
            frame["loaded_from_batch"] = f"Link {i}"
            frames.append(frame)
            loaded.append(ds)
        except Exception as exc:
            failures.append((url, str(exc)))
        progress.progress(i / len(url_list), text=f"Loaded {len(loaded)} of {len(url_list)} links")

    if not frames:
        details = "; ".join(f"{u} → {err}" for u, err in failures[:3]) if failures else "No links loaded."
        raise ReviewDownloaderError(f"Could not load any links. {details}")

    combined = pd.concat(frames, ignore_index=True)
    combined = _dedupe_combined_reviews(combined)
    combined = _finalize_df(combined)

    summary = ReviewBatchSummary(
        product_url="\n".join(url_list),
        product_id=f"MULTI_URL_WORKSPACE_{len(url_list)}",
        total_reviews=len(combined),
        page_size=max(len(combined), 1),
        requests_needed=sum(int(getattr(ds.get("summary"), "requests_needed", 0)) for ds in loaded),
        reviews_downloaded=len(combined),
    )
    status.success(f"Loaded {len(combined):,} reviews from {len(loaded)} link(s).")
    if failures:
        st.warning("Some links could not be loaded: " + " | ".join(f"{u} → {err}" for u, err in failures[:3]))
    return dict(
        summary=summary,
        reviews_df=combined,
        source_type="multi-url",
        source_label=_format_multi_source_label(url_list),
        source_urls=url_list,
        source_failures=failures,
    )


def _use_package_connectors() -> bool:
    """Prefer review_analyst.connectors when available (default: True)."""
    flag = str(os.getenv("STARWALK_USE_PACKAGE_CONNECTORS", "")).strip().lower()
    if flag in {"0", "false", "no", "n", "off"}: return False
    return callable(_package_load_product_reviews)


def _load_product_reviews_dispatch(product_url: str):
    if _use_package_connectors() and callable(_package_load_product_reviews):
        return _package_load_product_reviews(product_url)
    return _load_product_reviews(product_url)


def _load_multiple_product_reviews_dispatch(urls):
    if _use_package_connectors() and callable(_package_load_multiple_product_reviews):
        return _package_load_multiple_product_reviews(urls)
    return _load_multiple_product_reviews(urls)


def _load_uploaded_files_dispatch(files, *, include_local_symptomization=False):
    if _use_package_connectors() and callable(_package_load_uploaded_files):
        return _package_load_uploaded_files(files, include_local_symptomization=include_local_symptomization)
    return _load_uploaded_files(files, include_local_symptomization=include_local_symptomization)


# ═══════════════════════════════════════════════════════════════════════════════
#  ANALYTICS
# ═══════════════════════════════════════════════════════════════════════════════
def _df_cache_key(df):
    cols = [c for c in [
        "review_id", "rating", "incentivized_review", "is_recommended",
        "is_syndicated", "photos_count", "has_photos", "submission_time",
        "title_and_text", "review_length_words", "content_locale", "product_or_sku",
    ] if c in df.columns]
    return df[cols].to_json(orient="split", date_format="iso")


@st.cache_data(show_spinner=False, ttl=300)
def _compute_metrics_cached(df_json):
    df = pd.read_json(io.StringIO(df_json), orient="split")
    return _compute_metrics_direct(df)


def _compute_metrics_direct(df):
    n = len(df)
    if n == 0:
        return dict(
            review_count=0,
            avg_rating=None,
            avg_rating_non_incentivized=None,
            pct_low_star=0.0,
            pct_one_star=0.0,
            pct_two_star=0.0,
            pct_five_star=0.0,
            pct_incentivized=0.0,
            pct_with_photos=0.0,
            pct_syndicated=0.0,
            recommend_rate=None,
            median_review_words=None,
            non_incentivized_count=0,
            low_star_count=0,
        )
    ni = df[~df["incentivized_review"].fillna(False)]
    rb = df[df["is_recommended"].notna()]
    rr = _safe_pct(int(rb["is_recommended"].astype(bool).sum()), len(rb)) if not rb.empty else None
    mw = float(df["review_length_words"].median()) if "review_length_words" in df.columns and not df["review_length_words"].dropna().empty else None
    low = df["rating"].isin([1, 2])
    return dict(
        review_count=n,
        avg_rating=_safe_mean(df["rating"]),
        avg_rating_non_incentivized=_safe_mean(ni["rating"]),
        pct_low_star=_safe_pct(int(low.sum()), n),
        pct_one_star=_safe_pct(int((df["rating"] == 1).sum()), n),
        pct_two_star=_safe_pct(int((df["rating"] == 2).sum()), n),
        pct_five_star=_safe_pct(int((df["rating"] == 5).sum()), n),
        pct_incentivized=_safe_pct(int(df["incentivized_review"].fillna(False).sum()), n),
        pct_with_photos=_safe_pct(int(df["has_photos"].fillna(False).sum()), n),
        pct_syndicated=_safe_pct(int(df["is_syndicated"].fillna(False).sum()), n),
        recommend_rate=rr,
        median_review_words=mw,
        non_incentivized_count=len(ni),
        low_star_count=int(low.sum()),
    )


def _get_metrics(df):
    try:
        return _compute_metrics_cached(_df_cache_key(df))
    except Exception:
        return _compute_metrics_direct(df)


@st.cache_data(show_spinner=False, ttl=300)
def _rating_dist_cached(df_json):
    df = pd.read_json(io.StringIO(df_json), orient="split")
    base = pd.DataFrame({"rating": [1, 2, 3, 4, 5]})
    if df.empty:
        base["review_count"] = 0
        base["share"] = 0.0
        return base
    grouped = (
        df.dropna(subset=["rating"])
        .assign(rating=lambda x: x["rating"].astype(int))
        .groupby("rating", as_index=False)
        .size()
        .rename(columns={"size": "review_count"})
    )
    merged = base.merge(grouped, how="left", on="rating").fillna({"review_count": 0})
    merged["review_count"] = merged["review_count"].astype(int)
    merged["share"] = merged["review_count"] / max(len(df), 1)
    return merged


@st.cache_data(show_spinner=False, ttl=300)
def _rating_dist(df):
    try:
        return _rating_dist_cached(_df_cache_key(df))
    except Exception:
        return pd.DataFrame({"rating": [1, 2, 3, 4, 5], "review_count": [0] * 5, "share": [0.0] * 5})


@st.cache_data(show_spinner=False, ttl=300)
def _monthly_trend_cached(df_json):
    df = pd.read_json(io.StringIO(df_json), orient="split")
    if df.empty:
        return pd.DataFrame(columns=["submission_month", "review_count", "avg_rating", "month_start"])
    df["submission_time"] = pd.to_datetime(df.get("submission_time"), errors="coerce")
    return (
        df.dropna(subset=["submission_time"])
        .assign(month_start=lambda x: x["submission_time"].dt.to_period("M").dt.to_timestamp())
        .groupby("month_start", as_index=False)
        .agg(review_count=("review_id", "count"), avg_rating=("rating", "mean"))
        .assign(submission_month=lambda x: x["month_start"].dt.strftime("%Y-%m"))
        .sort_values("month_start")
    )


def _monthly_trend(df):
    try:
        return _monthly_trend_cached(_df_cache_key(df))
    except Exception:
        return pd.DataFrame(columns=["submission_month", "review_count", "avg_rating", "month_start"])


def _cohort_by_incentivized(df):
    if df.empty:
        return pd.DataFrame()
    w = df.copy()
    w["cohort"] = w["incentivized_review"].fillna(False).map({True: "Incentivized", False: "Organic"})
    w["rating_int"] = pd.to_numeric(w["rating"], errors="coerce")
    w = w.dropna(subset=["rating_int"])
    w["rating_int"] = w["rating_int"].astype(int)
    out = []
    for cohort, grp in w.groupby("cohort"):
        total = max(len(grp), 1)
        for star in [1, 2, 3, 4, 5]:
            cnt = int((grp["rating_int"] == star).sum())
            out.append(dict(cohort=cohort, star=star, count=cnt, pct=cnt / total * 100))
    return pd.DataFrame(out)


def _locale_breakdown(df, top_n=None):
    if df.empty or "content_locale" not in df.columns:
        return pd.DataFrame()
    grp = (
        df.dropna(subset=["content_locale"])
        .groupby("content_locale", as_index=False)
        .agg(count=("review_id", "count"), avg_rating=("rating", "mean"))
        .sort_values("count", ascending=False)
    )
    if top_n not in (None, "All"):
        grp = grp.head(int(top_n))
    grp["pct"] = grp["count"] / max(grp["count"].sum(), 1) * 100
    return grp


def _rolling_velocity(df, window=3):
    md = _monthly_trend(df)
    if md.empty:
        return md
    md = md.copy()
    md["rolling_avg"] = md["review_count"].rolling(window, min_periods=1).mean()
    return md


def _review_length_cohort(df):
    if df.empty or "review_length_words" not in df.columns:
        return pd.DataFrame()
    w = df.dropna(subset=["rating", "review_length_words"]).copy()
    w["review_length_words"] = pd.to_numeric(w["review_length_words"], errors="coerce")
    w = w.dropna(subset=["review_length_words"])
    if len(w) < 8:
        return pd.DataFrame()
    try:
        w["length_bin"] = pd.qcut(
            w["review_length_words"],
            q=4,
            labels=["Short (Q1)", "Medium (Q2)", "Long (Q3)", "Very Long (Q4)"],
            duplicates="drop",
        )
    except Exception:
        return pd.DataFrame()
    return (
        w.groupby("length_bin", as_index=False, observed=True)
        .agg(avg_rating=("rating", "mean"), count=("review_id", "count"), median_words=("review_length_words", "median"))
        .rename(columns={"length_bin": "Length Quartile"})
    )


def _top_locations(df, top_n=None):
    if df.empty or "user_location" not in df.columns:
        return pd.DataFrame()
    grp = (
        df.dropna(subset=["user_location"])
        .groupby("user_location", as_index=False)
        .agg(count=("review_id", "count"), avg_rating=("rating", "mean"))
        .sort_values("count", ascending=False)
    )
    if top_n not in (None, "All"):
        grp = grp.head(int(top_n))
    return grp


def _star_band_trend(df):
    if df.empty:
        return pd.DataFrame()
    md = _monthly_trend(df)
    if md.empty:
        return pd.DataFrame()
    w = df.dropna(subset=["submission_time", "rating"]).copy()
    w["month_start"] = w["submission_time"].dt.to_period("M").dt.to_timestamp()
    w["low"] = w["rating"].isin([1, 2])
    w["high"] = w["rating"].isin([4, 5])
    grp = w.groupby("month_start", as_index=False).agg(total=("review_id", "count"), low_ct=("low", "sum"), high_ct=("high", "sum"))
    grp["pct_low"] = grp["low_ct"] / grp["total"].clip(lower=1) * 100
    grp["pct_high"] = grp["high_ct"] / grp["total"].clip(lower=1) * 100
    return grp.sort_values("month_start")


def _sw_style_fig(fig):
    GRID = "rgba(148,163,184,0.18)"
    trace_count = len(getattr(fig, "data", []) or [])
    if trace_count > 3:
        legend_cfg = dict(
            orientation="v",
            y=1.0,
            x=1.01,
            xanchor="left",
            yanchor="top",
            bgcolor="rgba(255,255,255,0.86)",
            bordercolor="rgba(148,163,184,0.22)",
            borderwidth=1,
            font=dict(size=11),
        )
        margin = dict(l=26, r=108, t=56, b=44)
    else:
        legend_cfg = dict(
            orientation="h",
            y=1.12,
            x=0,
            xanchor="left",
            yanchor="bottom",
            bgcolor="rgba(255,255,255,0.84)",
            bordercolor="rgba(148,163,184,0.18)",
            borderwidth=1,
            font=dict(size=11),
        )
        margin = dict(l=26, r=18, t=64, b=44)
    fig.update_layout(
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
        font=dict(family="Inter, system-ui, sans-serif", size=12),
        margin=margin,
        title=dict(x=0, xanchor="left", font=dict(size=15)),
        legend=legend_cfg,
        hoverlabel=dict(font=dict(family="Inter, system-ui, sans-serif", size=12)),
    )
    fig.update_xaxes(gridcolor=GRID, zerolinecolor=GRID, automargin=True, title_standoff=10)
    fig.update_yaxes(gridcolor=GRID, zerolinecolor=GRID, automargin=True, title_standoff=10)
    return fig


def _show_plotly(fig):
    st.plotly_chart(
        fig,
        use_container_width=True,
        config={
            "displaylogo": False,
            "displayModeBar": False,
            "responsive": True,
            "modeBarButtonsToRemove": ["lasso2d", "select2d", "autoScale2d", "toggleSpikelines"],
        },
    )


def _render_chart_header(title: str, subtitle: str = ""):
    st.markdown(
        f"<div style='padding:2px 2px 10px 2px;'>"
        f"<div style='font-size:13px;font-weight:800;color:var(--navy);line-height:1.25;'>{_esc(title)}</div>"
        + (f"<div style='font-size:11.5px;color:var(--slate-500);margin-top:3px;line-height:1.35;'>{_esc(subtitle)}</div>" if subtitle else "")
        + "</div>",
        unsafe_allow_html=True,
    )


REGION_NAME_MAP = {
    "US": "USA",
    "USA": "USA",
    "GB": "UK",
    "UK": "UK",
    "CA": "Canada",
    "AU": "Australia",
    "DE": "Germany",
    "FR": "France",
    "ES": "Spain",
    "IT": "Italy",
    "JP": "Japan",
    "MX": "Mexico",
    "BR": "Brazil",
    "NL": "Netherlands",
}


def _locale_to_region_label(locale):
    raw = _safe_text(locale).replace("-", "_").strip()
    if not raw:
        return "Unknown"
    parts = [p for p in raw.split("_") if p]
    country = (parts[-1] if parts else raw).upper()
    country = re.sub(r"[^A-Z]", "", country)
    if not country:
        return "Unknown"
    return REGION_NAME_MAP.get(country, country)


def _parse_smoothing_window(label):
    txt = _safe_text(label).lower()
    if txt.startswith("none"):
        return 1
    m = re.search(r"(\d+)", txt)
    return int(m.group(1)) if m else 1


@st.cache_data(show_spinner=False, ttl=300)
def _cumulative_avg_region_trend(df, *, organic_only=False, top_n=None, smoothing_label="7-day"):
    if df.empty or "submission_time" not in df.columns or "rating" not in df.columns:
        return pd.DataFrame(), []

    w = df.copy()
    w["submission_time"] = pd.to_datetime(w["submission_time"], errors="coerce")
    w["rating"] = pd.to_numeric(w["rating"], errors="coerce")
    w = w.dropna(subset=["submission_time", "rating"]).copy()

    if organic_only and "incentivized_review" in w.columns:
        w = w[~w["incentivized_review"].fillna(False)].copy()

    if w.empty:
        return pd.DataFrame(), []

    w["day"] = w["submission_time"].dt.floor("D")
    w["region"] = w.get("content_locale", pd.Series(index=w.index, dtype="object")).map(_locale_to_region_label).fillna("Unknown")

    full_days = pd.date_range(w["day"].min(), w["day"].max(), freq="D")
    base = pd.DataFrame({"day": full_days})

    overall = w.groupby("day", as_index=False).agg(daily_volume=("review_id", "count"), rating_sum=("rating", "sum"))
    trend = base.merge(overall, on="day", how="left").fillna({"daily_volume": 0, "rating_sum": 0})
    trend["daily_volume"] = trend["daily_volume"].astype(int)
    overall_denom = trend["daily_volume"].cumsum()
    trend["overall_cum_avg"] = np.where(overall_denom > 0, trend["rating_sum"].cumsum() / overall_denom, np.nan)

    region_counts = (
        w[w["region"] != "Unknown"]
        .groupby("region")["review_id"]
        .count()
        .sort_values(ascending=False)
    )
    if top_n in (None, "All"):
        regions = region_counts.index.tolist()
    else:
        regions = region_counts.head(int(top_n)).index.tolist()
    if not regions and "Unknown" in set(w["region"]):
        regions = ["Unknown"]

    for region in regions:
        reg = w[w["region"] == region].groupby("day", as_index=False).agg(region_volume=("review_id", "count"), rating_sum=("rating", "sum"))
        reg = base.merge(reg, on="day", how="left").fillna({"region_volume": 0, "rating_sum": 0})
        reg_denom = reg["region_volume"].cumsum()
        trend[f"{region}_cum_avg"] = np.where(reg_denom > 0, reg["rating_sum"].cumsum() / reg_denom, np.nan)

    smoothing_window = _parse_smoothing_window(smoothing_label)
    if smoothing_window > 1:
        for col in [c for c in trend.columns if c.endswith("_cum_avg")]:
            trend[col] = trend[col].rolling(smoothing_window, min_periods=1).mean()

    return trend.sort_values("day").reset_index(drop=True), regions


def _first_non_empty(series):
    for value in series:
        text = _safe_text(value)
        if text:
            return text
    return ""


def _clean_watch_dimension_series(series: pd.Series, *, unknown: str = "Unknown") -> pd.Series:
    cleaned = series.astype("string").fillna("").str.strip()
    cleaned = cleaned.replace({"": unknown, "nan": unknown, "none": unknown, "null": unknown, "<na>": unknown})
    return cleaned.fillna(unknown)


_REGION_TEXT_ALIASES = {
    "UNITED STATES": "USA",
    "UNITED STATES OF AMERICA": "USA",
    "US": "USA",
    "USA": "USA",
    "UNITED KINGDOM": "UK",
    "GREAT BRITAIN": "UK",
    "GB": "UK",
    "UK": "UK",
    "CANADA": "Canada",
    "AUSTRALIA": "Australia",
    "GERMANY": "Germany",
    "FRANCE": "France",
    "SPAIN": "Spain",
    "ITALY": "Italy",
    "JAPAN": "Japan",
    "MEXICO": "Mexico",
    "BRAZIL": "Brazil",
    "NETHERLANDS": "Netherlands",
}


def _region_label_from_value(value: Any) -> str:
    raw = _safe_text(value).strip()
    if not raw:
        return "Unknown"
    upper_words = re.sub(r"[^A-Za-z ]", " ", raw).upper()
    upper_words = re.sub(r"\s+", " ", upper_words).strip()
    if upper_words in _REGION_TEXT_ALIASES:
        return _REGION_TEXT_ALIASES[upper_words]
    if re.search(r"^[A-Za-z]{2,3}(?:[_-][A-Za-z]{2,3})+$", raw):
        return _locale_to_region_label(raw)
    compact = re.sub(r"[^A-Z]", "", upper_words)
    if compact in REGION_NAME_MAP:
        return REGION_NAME_MAP[compact]
    titled = re.sub(r"\s+", " ", raw).strip()
    return titled or "Unknown"


def _region_series_from_column(series: pd.Series) -> pd.Series:
    cleaned = series.astype("string").fillna("").str.strip()
    return cleaned.map(_region_label_from_value).fillna("Unknown")


def _organic_only_mask(series: pd.Series, column_name: str) -> pd.Series:
    norm_name = _normalize_col_key(column_name)
    if pd.api.types.is_bool_dtype(series) or str(series.dtype).lower() == "boolean":
        base = series.astype("boolean")
        return base.fillna(False) if "organic" in norm_name else (~base.fillna(False))

    text = series.astype("string").fillna("").str.strip().str.lower()
    true_like = {"true", "yes", "y", "1", "paid", "sponsored", "incentivized", "gifted", "seeded"}
    organic_like = {"organic", "unpaid", "earned", "consumer", "false", "no", "n", "0", "not seeded", "non-seeded", "not gifted"}
    mapped = pd.Series(pd.NA, index=text.index, dtype="boolean")
    mapped.loc[text.isin(true_like)] = True
    mapped.loc[text.isin(organic_like)] = False
    if "organic" in norm_name:
        return mapped.fillna(False)
    return (~mapped.fillna(False))




def _watch_split_labels(value: Any) -> List[str]:
    """Split a symptom cell into normalized labels."""
    if value is None:
        return []
    if isinstance(value, (list, tuple, set)):
        out: List[str] = []
        for item in value:
            out.extend(_watch_split_labels(item))
        return _normalize_tag_list(out)
    try:
        if isinstance(pd.isna(value), bool) and pd.isna(value):
            return []
    except Exception:
        pass
    text = _safe_text(value)
    if not text:
        return []
    parts = re.split(r"[\n;,|]+", text)
    return _normalize_tag_list(parts)



def _watch_symptom_column_groups(df: pd.DataFrame) -> Tuple[List[str], List[str]]:
    det_cols = [str(c) for c in df.columns if str(c).startswith("AI Symptom Detractor")]
    del_cols = [str(c) for c in df.columns if str(c).startswith("AI Symptom Delighter")]
    for candidate in ["L2 Detractor Condition", "Detractors"]:
        if candidate in df.columns and candidate not in det_cols:
            det_cols.append(candidate)
    for candidate in ["L2 Delighter Condition", "Delighters"]:
        if candidate in df.columns and candidate not in del_cols:
            del_cols.append(candidate)
    return det_cols, del_cols



def _watch_signal_metadata(*, reviews: Any, avg_rating: Any, recent_reviews: Any, delta_30d: Any, gap_vs_region: Any) -> Tuple[str, str]:
    reviews_n = int(pd.to_numeric(reviews, errors="coerce") or 0)
    recent_n = int(pd.to_numeric(recent_reviews, errors="coerce") or 0)
    avg_n = pd.to_numeric(avg_rating, errors="coerce")
    delta_n = pd.to_numeric(delta_30d, errors="coerce")
    gap_n = pd.to_numeric(gap_vs_region, errors="coerce")

    reasons: List[str] = []
    level_rank = 0
    if recent_n >= 3 and pd.notna(delta_n) and float(delta_n) <= -0.35:
        reasons.append("Sharp 30d drop")
        level_rank = max(level_rank, 3)
    elif recent_n >= 3 and pd.notna(delta_n) and float(delta_n) <= -0.20:
        reasons.append("Recent rating down")
        level_rank = max(level_rank, 2)
    if reviews_n >= 5 and pd.notna(gap_n) and float(gap_n) <= -0.25:
        reasons.append("Below region avg")
        level_rank = max(level_rank, 2)
    if reviews_n >= 10 and pd.notna(avg_n) and float(avg_n) <= 3.80:
        reasons.append("Low baseline")
        level_rank = max(level_rank, 1)

    if level_rank <= 0:
        return "", ""
    level = {3: "High", 2: "Watch", 1: "Baseline"}.get(level_rank, "Watch")
    return level, " · ".join(dict.fromkeys(reasons))



def _build_watch_review_rows(df: pd.DataFrame) -> pd.DataFrame:
    title_series = _select_column_series(df, "title") if "title" in df.columns else pd.Series(pd.NA, index=df.index)
    review_series = _select_column_series(df, "review_text") if "review_text" in df.columns else pd.Series(pd.NA, index=df.index)
    review_id_series = _select_column_series(df, "review_id") if "review_id" in df.columns else pd.Series(pd.NA, index=df.index)
    post_link_series = _select_column_series(df, "post_link") if "post_link" in df.columns else pd.Series(pd.NA, index=df.index)
    out = pd.DataFrame(
        {
            "Region": df.get("region", pd.Series("All selected", index=df.index)).astype("string").fillna("Unknown"),
            "Retailer": df.get("retailer_display", pd.Series("Unknown Retailer", index=df.index)).astype("string").fillna("Unknown Retailer"),
            "Date": pd.to_datetime(df.get("_watch_date"), errors="coerce", utc=True).dt.tz_convert("UTC").dt.tz_localize(None),
            "Rating": pd.to_numeric(df.get("rating"), errors="coerce"),
            "Review ID": review_id_series.astype("string").fillna(""),
            "Title": title_series.astype("string").fillna(""),
            "Review": review_series.astype("string").fillna(""),
            "Post Link": post_link_series.astype("string").fillna(""),
            "Recent Window": np.where(df.get("_recent", pd.Series(False, index=df.index)).astype(bool), "Last 30d", "Older"),
            "Symptom Detractors": [", ".join(v or []) for v in df.get("_watch_detractors", pd.Series([[] for _ in range(len(df))], index=df.index))],
            "Symptom Delighters": [", ".join(v or []) for v in df.get("_watch_delighters", pd.Series([[] for _ in range(len(df))], index=df.index))],
        }
    )
    return out



def _build_watch_symptom_summary(df: pd.DataFrame) -> pd.DataFrame:
    records: List[Dict[str, Any]] = []
    if df is None or df.empty:
        return pd.DataFrame(columns=["Region", "Retailer", "Side", "Label", "Mentions", "Scope Reviews", "Share of Reviews"])

    for row in df[["region", "retailer_display", "_watch_detractors", "_watch_delighters"]].itertuples(index=False):
        region = _safe_text(row[0], "Unknown")
        retailer = _safe_text(row[1], "Unknown Retailer")
        for label in row[2] or []:
            records.append({"Region": region, "Retailer": retailer, "Side": "Detractor", "Label": label})
        for label in row[3] or []:
            records.append({"Region": region, "Retailer": retailer, "Side": "Delighter", "Label": label})

    if not records:
        return pd.DataFrame(columns=["Region", "Retailer", "Side", "Label", "Mentions", "Scope Reviews", "Share of Reviews"])

    symptom_long = pd.DataFrame(records)
    scope_reviews = (
        df.groupby(["region", "retailer_display"], dropna=False)
        .size()
        .reset_index(name="Scope Reviews")
        .rename(columns={"region": "Region", "retailer_display": "Retailer"})
    )
    summary = (
        symptom_long.groupby(["Region", "Retailer", "Side", "Label"], dropna=False)
        .size()
        .reset_index(name="Mentions")
    )
    summary = summary.merge(scope_reviews, on=["Region", "Retailer"], how="left")
    summary["Share of Reviews"] = pd.to_numeric(summary["Mentions"], errors="coerce").fillna(0) / pd.to_numeric(summary["Scope Reviews"], errors="coerce").replace(0, np.nan)
    summary["Share of Reviews"] = summary["Share of Reviews"].fillna(0.0)
    summary = summary.sort_values(["Region", "Retailer", "Side", "Mentions", "Label"], ascending=[True, True, True, False, True]).reset_index(drop=True)
    return summary


@st.cache_data(show_spinner=False, ttl=300)
def _prepare_source_rating_watch_payload(
    df,
    *,
    organic_only: bool = False,
    selected_regions: Optional[Sequence[str]] = None,
    combine_regions: bool = True,
    source_col: Optional[str] = None,
    region_col: Optional[str] = None,
    date_col: Optional[str] = None,
):
    if df is None or df.empty or "rating" not in df.columns:
        return {
            "table_df": pd.DataFrame(),
            "region_kpis_df": pd.DataFrame(),
            "known_regions": [],
            "alerts_df": pd.DataFrame(),
            "trend_df": pd.DataFrame(),
            "review_rows_df": pd.DataFrame(),
            "symptom_summary_df": pd.DataFrame(),
            "resolved_source_col": None,
            "resolved_region_col": None,
            "resolved_date_col": None,
            "organic_flag_col": None,
        }

    w = _collapse_duplicate_named_columns(df.copy())
    w["rating"] = pd.to_numeric(w.get("rating"), errors="coerce")
    w = w.dropna(subset=["rating"]).copy()
    if w.empty:
        return {
            "table_df": pd.DataFrame(),
            "region_kpis_df": pd.DataFrame(),
            "known_regions": [],
            "alerts_df": pd.DataFrame(),
            "trend_df": pd.DataFrame(),
            "review_rows_df": pd.DataFrame(),
            "symptom_summary_df": pd.DataFrame(),
            "resolved_source_col": None,
            "resolved_region_col": None,
            "resolved_date_col": None,
            "organic_flag_col": None,
        }

    resolved_source_col = source_col if source_col in w.columns else _resolve_column_alias(w, WATCH_SOURCE_COLUMN_ALIASES)
    resolved_region_col = region_col if region_col in w.columns else _resolve_column_alias(w, WATCH_REGION_COLUMN_ALIASES)
    resolved_date_col = date_col if date_col in w.columns else _resolve_column_alias(w, WATCH_DATE_COLUMN_ALIASES)
    organic_col = _resolve_column_alias(w, WATCH_ORGANIC_COLUMN_ALIASES)

    organic_series = _select_column_series(w, organic_col) if organic_col else None
    if organic_only and organic_col and organic_series is not None:
        w = w[_organic_only_mask(organic_series, organic_col)].copy()
    if w.empty:
        return {
            "table_df": pd.DataFrame(),
            "region_kpis_df": pd.DataFrame(),
            "known_regions": [],
            "alerts_df": pd.DataFrame(),
            "trend_df": pd.DataFrame(),
            "review_rows_df": pd.DataFrame(),
            "symptom_summary_df": pd.DataFrame(),
            "resolved_source_col": resolved_source_col,
            "resolved_region_col": resolved_region_col,
            "resolved_date_col": resolved_date_col,
            "organic_flag_col": organic_col,
        }

    region_series = _select_column_series(w, resolved_region_col) if resolved_region_col else None
    if region_series is not None:
        w["region"] = _region_series_from_column(region_series)
    else:
        w["region"] = "Unknown"
    known_regions = [r for r in sorted(set(w["region"])) if r and r != "Unknown"]

    if known_regions and selected_regions:
        selected = [str(r) for r in selected_regions if str(r) in known_regions]
        if selected:
            w = w[w["region"].isin(selected)].copy()
    if w.empty:
        return {
            "table_df": pd.DataFrame(),
            "region_kpis_df": pd.DataFrame(),
            "known_regions": known_regions,
            "alerts_df": pd.DataFrame(),
            "trend_df": pd.DataFrame(),
            "review_rows_df": pd.DataFrame(),
            "symptom_summary_df": pd.DataFrame(),
            "resolved_source_col": resolved_source_col,
            "resolved_region_col": resolved_region_col,
            "resolved_date_col": resolved_date_col,
            "organic_flag_col": organic_col,
        }

    source_input = _select_column_series(w, resolved_source_col) if resolved_source_col else None
    source_series = _clean_watch_dimension_series(source_input, unknown="") if source_input is not None else pd.Series("", index=w.index, dtype="object")
    source_system_col = _resolve_column_alias(w, ["source_system", "source", "platform", "provider", "review_source"])
    source_system_input = _select_column_series(w, source_system_col) if source_system_col else None
    source_system_series = _clean_watch_dimension_series(source_system_input, unknown="") if source_system_input is not None else pd.Series("", index=w.index, dtype="object")
    w["source_system"] = source_system_series
    w["retailer_display"] = _clean_watch_dimension_series(source_series.mask(source_series.eq(""), source_system_series), unknown="Unknown Retailer")

    date_series = _select_column_series(w, resolved_date_col) if resolved_date_col else None
    if date_series is not None:
        w["_watch_date"] = pd.to_datetime(date_series, errors="coerce", utc=True)
    else:
        w["_watch_date"] = pd.NaT
    anchor_time = w["_watch_date"].dropna().max() if w["_watch_date"].notna().any() else pd.Timestamp.now(tz="UTC")
    recent_cutoff = anchor_time - pd.Timedelta(days=30)
    w["_recent"] = w["_watch_date"].notna() & (w["_watch_date"] >= recent_cutoff)

    det_cols, del_cols = _watch_symptom_column_groups(w)
    if det_cols or del_cols:
        det_lists: List[List[str]] = []
        del_lists: List[List[str]] = []
        symptom_cols = [c for c in det_cols + del_cols if c in w.columns]
        for _, row in w[symptom_cols].iterrows():
            dets: List[str] = []
            dels: List[str] = []
            for col in det_cols:
                dets.extend(_watch_split_labels(row.get(col)))
            for col in del_cols:
                dels.extend(_watch_split_labels(row.get(col)))
            det_lists.append(_normalize_tag_list(dets))
            del_lists.append(_normalize_tag_list(dels))
        w["_watch_detractors"] = det_lists
        w["_watch_delighters"] = del_lists
    else:
        w["_watch_detractors"] = [[] for _ in range(len(w))]
        w["_watch_delighters"] = [[] for _ in range(len(w))]

    group_cols = ["retailer_display"] if combine_regions or not known_regions else ["region", "retailer_display"]
    agg = (
        w.groupby(group_cols, dropna=False)
        .agg(
            source_system=("source_system", _first_non_empty),
            reviews=("rating", "size"),
            avg_rating=("rating", "mean"),
        )
        .reset_index()
    )
    recent = (
        w[w["_recent"]]
        .groupby(group_cols, dropna=False)
        .agg(recent_reviews=("rating", "size"), recent_avg_rating=("rating", "mean"))
        .reset_index()
    )
    table = agg.merge(recent, on=group_cols, how="left")
    table["recent_reviews"] = pd.to_numeric(table.get("recent_reviews"), errors="coerce").fillna(0).astype(int)
    table["recent_avg_rating"] = pd.to_numeric(table.get("recent_avg_rating"), errors="coerce")
    table["delta_30d"] = np.where(table["recent_reviews"] >= 3, table["recent_avg_rating"] - table["avg_rating"], np.nan)
    table["share_of_view"] = pd.to_numeric(table["reviews"], errors="coerce").fillna(0) / max(len(w), 1)

    if known_regions:
        region_kpis = (
            w.groupby("region", dropna=False)
            .agg(reviews=("rating", "size"), avg_rating=("rating", "mean"))
            .reset_index()
        )
        region_recent = (
            w[w["_recent"]]
            .groupby("region", dropna=False)
            .agg(recent_reviews=("rating", "size"), recent_avg_rating=("rating", "mean"))
            .reset_index()
        )
        region_kpis = region_kpis.merge(region_recent, on="region", how="left")
        region_kpis["recent_reviews"] = pd.to_numeric(region_kpis.get("recent_reviews"), errors="coerce").fillna(0).astype(int)
        region_kpis["delta_30d"] = np.where(region_kpis["recent_reviews"] >= 3, region_kpis["recent_avg_rating"] - region_kpis["avg_rating"], np.nan)
        region_kpis = region_kpis.sort_values(["reviews", "avg_rating"], ascending=[False, False]).reset_index(drop=True)
        region_lookup = dict(zip(region_kpis["region"], region_kpis["avg_rating"]))
        if "region" in table.columns:
            table["gap_vs_region_avg"] = table["region"].map(region_lookup)
            table["gap_vs_region_avg"] = table["avg_rating"] - pd.to_numeric(table["gap_vs_region_avg"], errors="coerce")
        region_kpis.rename(columns={"region": "Region"}, inplace=True)
    else:
        region_kpis = pd.DataFrame(columns=["Region", "reviews", "avg_rating", "recent_reviews", "recent_avg_rating", "delta_30d"])
        table["gap_vs_region_avg"] = np.nan

    overall_recent_df = w[w["_recent"]]
    overall_recent_avg = pd.to_numeric(overall_recent_df.get("rating"), errors="coerce").mean() if not overall_recent_df.empty else np.nan
    overall_recent_reviews = int(len(overall_recent_df))
    overall_avg = float(w["rating"].mean()) if not w.empty else np.nan
    overall_delta = overall_recent_avg - overall_avg if overall_recent_reviews >= 3 and pd.notna(overall_recent_avg) else np.nan
    overall_row = pd.DataFrame(
        [{
            "Region": "All selected",
            "reviews": int(len(w)),
            "avg_rating": overall_avg,
            "recent_reviews": overall_recent_reviews,
            "recent_avg_rating": overall_recent_avg,
            "delta_30d": overall_delta,
        }]
    )
    region_kpis = pd.concat([overall_row, region_kpis], ignore_index=True)

    if "region" not in table.columns:
        table["gap_vs_region_avg"] = table["avg_rating"] - overall_avg

    alert_levels: List[str] = []
    alert_signals: List[str] = []
    for row in table.itertuples(index=False):
        level, signal = _watch_signal_metadata(
            reviews=getattr(row, "reviews", 0),
            avg_rating=getattr(row, "avg_rating", np.nan),
            recent_reviews=getattr(row, "recent_reviews", 0),
            delta_30d=getattr(row, "delta_30d", np.nan),
            gap_vs_region=getattr(row, "gap_vs_region_avg", np.nan),
        )
        alert_levels.append(level)
        alert_signals.append(signal)
    table["alert_level"] = alert_levels
    table["signal"] = alert_signals

    gap_label = "gap_vs_region_avg" if "region" in table.columns else "gap_vs_region_avg"
    sort_cols = ["alert_level", "avg_rating"]
    table["_alert_rank"] = table["alert_level"].map({"High": 3, "Watch": 2, "Baseline": 1}).fillna(0)
    if "region" in table.columns:
        sort_cols = ["region", "_alert_rank", "avg_rating"]
        ascending = [True, False, False]
    else:
        ascending = [False, False]
    table = table.sort_values(sort_cols, ascending=ascending).reset_index(drop=True)

    rename_map = {
        "region": "Region",
        "retailer_display": "Retailer",
        "source_system": "Source System",
        "reviews": "Reviews",
        "avg_rating": "Avg Rating",
        "recent_reviews": "Last 30d Reviews",
        "recent_avg_rating": "Last 30d Avg",
        "delta_30d": "30d Delta",
        "share_of_view": "Share of View",
        "gap_vs_region_avg": ("Gap vs Region Avg" if "region" in table.columns else "Gap vs Selected Avg"),
        "alert_level": "Alert Level",
        "signal": "Signal",
    }
    table = _collapse_duplicate_named_columns(table.rename(columns=rename_map)).drop(columns=[c for c in ["_alert_rank"] if c in table.columns], errors="ignore")

    alerts_df = table.copy()
    if "Alert Level" in alerts_df.columns:
        alerts_df = alerts_df[alerts_df["Alert Level"].astype("string").fillna("").str.strip().ne("")].copy()
    if not alerts_df.empty:
        level_rank = {"High": 3, "Watch": 2, "Baseline": 1}
        alerts_df["_rank"] = alerts_df["Alert Level"].map(level_rank).fillna(0)
        sort_cols = ["_rank", "30d Delta", "Avg Rating"]
        ascending = [False, True, True]
        if "Region" in alerts_df.columns:
            sort_cols = ["_rank", "Region", "30d Delta", "Avg Rating"]
            ascending = [False, True, True, True]
        alerts_df = alerts_df.sort_values(sort_cols, ascending=ascending).drop(columns=["_rank"]).reset_index(drop=True)

    if resolved_date_col and w["_watch_date"].notna().any():
        trend_source = w[w["_watch_date"].notna()].copy()
        trend_source["_watch_week"] = trend_source["_watch_date"].dt.tz_convert("UTC").dt.tz_localize(None).dt.to_period("W-SUN").dt.start_time
        trend_group_cols = ["_watch_week", "retailer_display"]
        if known_regions:
            trend_group_cols.insert(1, "region")
        trend_df = (
            trend_source.groupby(trend_group_cols, dropna=False)
            .agg(reviews=("rating", "size"), avg_rating=("rating", "mean"))
            .reset_index()
            .rename(columns={"_watch_week": "Week", "retailer_display": "Retailer", "region": "Region", "reviews": "Reviews", "avg_rating": "Avg Rating"})
        )
        trend_df = trend_df.sort_values([c for c in ["Region", "Retailer", "Week"] if c in trend_df.columns]).reset_index(drop=True)
    else:
        trend_df = pd.DataFrame(columns=["Week", "Region", "Retailer", "Reviews", "Avg Rating"])

    review_rows_df = _build_watch_review_rows(w)
    symptom_summary_df = _build_watch_symptom_summary(w)

    return {
        "table_df": table,
        "region_kpis_df": region_kpis,
        "known_regions": known_regions,
        "alerts_df": alerts_df,
        "trend_df": trend_df,
        "review_rows_df": review_rows_df,
        "symptom_summary_df": symptom_summary_df,
        "resolved_source_col": resolved_source_col,
        "resolved_region_col": resolved_region_col,
        "resolved_date_col": resolved_date_col,
        "organic_flag_col": organic_col,
    }


@st.cache_data(show_spinner=False, ttl=300)
def _prepare_source_rating_watch(
    df,
    *,
    organic_only: bool = False,
    selected_regions: Optional[Sequence[str]] = None,
    combine_regions: bool = True,
    source_col: Optional[str] = None,
    region_col: Optional[str] = None,
    date_col: Optional[str] = None,
):
    payload = _prepare_source_rating_watch_payload(
        df,
        organic_only=organic_only,
        selected_regions=selected_regions,
        combine_regions=combine_regions,
        source_col=source_col,
        region_col=region_col,
        date_col=date_col,
    )
    return payload.get("table_df", pd.DataFrame()), payload.get("region_kpis_df", pd.DataFrame()), payload.get("known_regions", [])



def _source_rating_watch_export_bytes(
    table_df,
    region_kpis_df,
    *,
    alerts_df: Optional[pd.DataFrame] = None,
    trend_df: Optional[pd.DataFrame] = None,
    review_rows_df: Optional[pd.DataFrame] = None,
    symptom_summary_df: Optional[pd.DataFrame] = None,
    split_by_region: bool = False,
    organic_only: bool = False,
    source_col: Optional[str] = None,
    region_col: Optional[str] = None,
    date_col: Optional[str] = None,
    organic_flag_col: Optional[str] = None,
):
    out = io.BytesIO()
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        pd.DataFrame(
            [{
                "Generated At UTC": pd.Timestamp.utcnow().strftime("%Y-%m-%d %H:%M:%S"),
                "Organic only": bool(organic_only),
                "Split by region": bool(split_by_region),
                "Retailer column": source_col or "Retailer (default) / auto-detect",
                "Region column": region_col or "None / auto-unavailable",
                "Date column": date_col or "None / auto-unavailable",
                "Organic flag column": organic_flag_col or "Not detected",
            }]
        ).to_excel(writer, sheet_name="Config", index=False)
        region_kpis_df.to_excel(writer, sheet_name="Region Summary", index=False)
        if split_by_region and "Region" in table_df.columns:
            for region, sub in table_df.groupby("Region", sort=False):
                sub.to_excel(writer, sheet_name=str(region or "Region")[:31], index=False)
        else:
            table_df.to_excel(writer, sheet_name="Retailer Ratings", index=False)
        if alerts_df is not None and not alerts_df.empty:
            alerts_df.to_excel(writer, sheet_name="Smoke Alerts", index=False)
        if trend_df is not None and not trend_df.empty:
            trend_df.to_excel(writer, sheet_name="Weekly Trend", index=False)
        if review_rows_df is not None and not review_rows_df.empty:
            review_rows_df.to_excel(writer, sheet_name="Review Drilldown", index=False)
        if symptom_summary_df is not None and not symptom_summary_df.empty:
            symptom_summary_df.to_excel(writer, sheet_name="Symptom Overlay", index=False)
    out.seek(0)
    return out.getvalue()



def _render_source_rating_watch(df):
    with st.container(border=True):
        st.markdown("<div class='section-title'>🏪 Retailer rating watch</div>", unsafe_allow_html=True)
        st.markdown("<div class='section-sub'>Spot smoke by retailer and region. Track average rating, review count, weekly movement, recent drops, and the review/symptom signals sitting behind those shifts.</div>", unsafe_allow_html=True)

        if df is None or df.empty or "rating" not in df.columns:
            st.info("No rating data available for retailer watch.")
            return

        source_options = _source_rating_watch_column_options(df, allow_none=False)
        region_options = _source_rating_watch_column_options(df, allow_none=True)
        date_options = _source_rating_watch_column_options(df, allow_none=True)
        if st.session_state.get("retailer_watch_source_col") not in source_options:
            st.session_state["retailer_watch_source_col"] = _resolve_column_alias(df, WATCH_SOURCE_COLUMN_ALIASES) or AUTO_COLUMN_SENTINEL
        if st.session_state.get("retailer_watch_region_col") not in region_options:
            st.session_state["retailer_watch_region_col"] = _resolve_column_alias(df, WATCH_REGION_COLUMN_ALIASES) or AUTO_COLUMN_SENTINEL
        if st.session_state.get("retailer_watch_date_col") not in date_options:
            st.session_state["retailer_watch_date_col"] = _resolve_column_alias(df, WATCH_DATE_COLUMN_ALIASES) or AUTO_COLUMN_SENTINEL

        with st.expander("🧭 Data mapping", expanded=False):
            m1, m2, m3 = st.columns(3)
            m1.selectbox(
                "Retailer column",
                options=source_options,
                key="retailer_watch_source_col",
                format_func=_format_column_choice_label,
                help="Defaults to Retailer when it exists. Override it here if the workspace uses a different retailer field.",
            )
            m2.selectbox(
                "Region column",
                options=region_options,
                key="retailer_watch_region_col",
                format_func=_format_column_choice_label,
                help="Defaults to Reviewer Location / Region style columns when present.",
            )
            m3.selectbox(
                "Review date column",
                options=date_options,
                key="retailer_watch_date_col",
                format_func=_format_column_choice_label,
                help="Used for weekly trends and rolling 30-day smoke checks.",
            )

        resolved_source_col = _resolve_optional_column_choice(df, st.session_state.get("retailer_watch_source_col"), WATCH_SOURCE_COLUMN_ALIASES)
        resolved_region_col = _resolve_optional_column_choice(df, st.session_state.get("retailer_watch_region_col"), WATCH_REGION_COLUMN_ALIASES, allow_none=True)
        resolved_date_col = _resolve_optional_column_choice(df, st.session_state.get("retailer_watch_date_col"), WATCH_DATE_COLUMN_ALIASES, allow_none=True)
        organic_flag_col = _resolve_column_alias(df, WATCH_ORGANIC_COLUMN_ALIASES)

        mapping_bits = [f"Retailer: <strong>{_esc(resolved_source_col or 'not found')}</strong>"]
        mapping_bits.append(f"Region: <strong>{_esc(resolved_region_col or 'none')}</strong>")
        mapping_bits.append(f"Date: <strong>{_esc(resolved_date_col or 'none')}</strong>")
        if organic_flag_col:
            mapping_bits.append(f"Organic flag: <strong>{_esc(organic_flag_col)}</strong>")
        st.markdown("<div class='status-note'>" + " · ".join(mapping_bits) + "</div>", unsafe_allow_html=True)

        preview_payload = _prepare_source_rating_watch_payload(
            df,
            organic_only=False,
            selected_regions=None,
            combine_regions=True,
            source_col=resolved_source_col,
            region_col=resolved_region_col,
            date_col=resolved_date_col,
        )
        preview_table = preview_payload.get("table_df", pd.DataFrame())
        known_regions = preview_payload.get("known_regions", [])
        if preview_table.empty:
            st.info("No rating data is available for the current retailer-watch mapping.")
            return

        c0, c1, c2, c3 = st.columns([1, 1.6, 1.05, 1.1])
        organic_only = c0.toggle(
            "Organic only",
            value=False,
            key="retailer_watch_organic_only",
            disabled=(organic_flag_col is None),
            help=("Uses the detected organic / incentivized flag." if organic_flag_col else "No organic / incentivized flag was detected in this workspace."),
        )
        if known_regions:
            current_regions = [r for r in (st.session_state.get("retailer_watch_regions") or known_regions) if r in known_regions]
            if not current_regions:
                current_regions = known_regions
            st.session_state["retailer_watch_regions"] = current_regions
            selected_regions = c1.multiselect("Regions", known_regions, default=current_regions, key="retailer_watch_regions")
            combine_regions = c2.radio("View", ["All combined", "Split by region"], horizontal=True, key="retailer_watch_view") == "All combined"
        else:
            selected_regions = []
            combine_regions = True
            c1.markdown("<div class='status-note'>Region metadata is not available for the current mapping, so the table is shown combined.</div>", unsafe_allow_html=True)
            c2.markdown("<div style='height:1px;'></div>", unsafe_allow_html=True)
        if organic_flag_col is None:
            c0.caption("No organic flag detected in this workspace.")

        payload = _prepare_source_rating_watch_payload(
            df,
            organic_only=organic_only,
            selected_regions=selected_regions,
            combine_regions=combine_regions,
            source_col=resolved_source_col,
            region_col=resolved_region_col,
            date_col=resolved_date_col,
        )
        table_df = payload.get("table_df", pd.DataFrame())
        region_kpis_df = payload.get("region_kpis_df", pd.DataFrame())
        alerts_df = payload.get("alerts_df", pd.DataFrame())
        trend_df = payload.get("trend_df", pd.DataFrame())
        review_rows_df = payload.get("review_rows_df", pd.DataFrame())
        symptom_summary_df = payload.get("symptom_summary_df", pd.DataFrame())
        if table_df.empty:
            st.info("No reviews match the current retailer-watch selection.")
            return

        export_bytes = _source_rating_watch_export_bytes(
            table_df.copy(),
            region_kpis_df.copy(),
            alerts_df=alerts_df.copy(),
            trend_df=trend_df.copy(),
            review_rows_df=review_rows_df.copy(),
            symptom_summary_df=symptom_summary_df.copy(),
            split_by_region=("Region" in table_df.columns),
            organic_only=organic_only,
            source_col=resolved_source_col,
            region_col=resolved_region_col,
            date_col=resolved_date_col,
            organic_flag_col=organic_flag_col,
        )
        c3.download_button(
            "⬇️ Export retailer watch",
            data=export_bytes,
            file_name="retailer_rating_watch.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            key="retailer_rating_watch_export",
        )

        summary_pills = [
            f"<span class='dashboard-pill'><span class='meta'>Retailers</span><strong>{int(table_df['Retailer'].nunique())}</strong></span>",
            f"<span class='dashboard-pill'><span class='meta'>Rows in view</span><strong>{int(pd.to_numeric(table_df['Reviews'], errors='coerce').fillna(0).sum()):,}</strong></span>",
            f"<span class='dashboard-pill'><span class='meta'>Smoke alerts</span><strong>{int(len(alerts_df))}</strong></span>",
        ]
        if resolved_date_col and not trend_df.empty:
            min_week = pd.to_datetime(trend_df['Week'], errors='coerce').min()
            max_week = pd.to_datetime(trend_df['Week'], errors='coerce').max()
            if pd.notna(min_week) and pd.notna(max_week):
                summary_pills.append(f"<span class='dashboard-pill'><span class='meta'>Trend coverage</span><strong>{min_week.strftime('%Y-%m-%d')}</strong><span class='meta'>to {max_week.strftime('%Y-%m-%d')}</span></span>")
        st.markdown("<div class='dashboard-brief'><div class='dashboard-brief-row'>" + "".join(summary_pills) + "</div></div>", unsafe_allow_html=True)

        watch_tabs = st.tabs(["Overview", "Smoke alerts", "Trend line", "Reviews behind shifts", "Symptom overlay"])

        with watch_tabs[0]:
            metric_cards = []
            for _, row in region_kpis_df.iterrows():
                delta = row.get("delta_30d")
                delta_txt = ""
                if pd.notna(delta):
                    direction = "up" if float(delta) > 0 else "down"
                    delta_txt = f" · 30d {direction} {abs(float(delta)):.2f}★"
                avg_value = float(row.get("avg_rating")) if pd.notna(row.get("avg_rating")) else 0.0
                metric_cards.append(
                    f"<span class='dashboard-pill'><span class='meta'>{_esc(row.get('Region'))}</span><strong>{avg_value:.2f}★</strong><span class='meta'>{int(row.get('reviews') or 0):,} reviews{delta_txt}</span></span>"
                )
            st.markdown("<div class='dashboard-brief'><div class='dashboard-brief-title'>Region average rating snapshot</div><div class='dashboard-brief-row'>" + "".join(metric_cards) + "</div></div>", unsafe_allow_html=True)

            chart_df = table_df.copy()
            chart_df["Avg Rating"] = pd.to_numeric(chart_df.get("Avg Rating"), errors="coerce")
            order_source = (
                chart_df.groupby("Retailer", as_index=False)["Avg Rating"]
                .mean()
                .sort_values("Avg Rating", ascending=False)["Retailer"]
                .tolist()
            )
            hover_data = {"Reviews": True, "Last 30d Avg": ":.2f", "30d Delta": ":.2f", "Share of View": ":.1%", "Alert Level": True}
            gap_col = "Gap vs Region Avg" if "Gap vs Region Avg" in chart_df.columns else ("Gap vs Selected Avg" if "Gap vs Selected Avg" in chart_df.columns else None)
            if gap_col:
                hover_data[gap_col] = ":.2f"
            if "Region" in chart_df.columns:
                fig = px.bar(
                    chart_df,
                    x="Retailer",
                    y="Avg Rating",
                    color="Region",
                    barmode="group",
                    category_orders={"Retailer": order_source},
                    hover_data=hover_data,
                )
            else:
                color_field = "Alert Level" if chart_df["Alert Level"].astype("string").fillna("").str.strip().ne("").any() else None
                fig = px.bar(
                    chart_df,
                    x="Retailer",
                    y="Avg Rating",
                    color=color_field,
                    category_orders={"Retailer": order_source},
                    hover_data=hover_data,
                )
                if color_field is None:
                    fig.update_traces(marker_color="#3b82f6")
            overall_avg = float(region_kpis_df.iloc[0]["avg_rating"]) if not region_kpis_df.empty else None
            if overall_avg is not None and not math.isnan(overall_avg):
                fig.add_hline(y=overall_avg, line_dash="dot", line_color="rgba(71,85,105,0.8)", annotation_text=f"Selected avg {overall_avg:.2f}★", annotation_position="top left")
            fig.update_layout(
                title=None,
                margin=dict(l=20, r=18, t=18, b=60),
                xaxis_title="",
                yaxis_title="Average rating ★",
                height=360,
                plot_bgcolor="rgba(0,0,0,0)",
                paper_bgcolor="rgba(0,0,0,0)",
                font_family="Inter",
                legend=dict(orientation="h", y=-0.22, x=0, xanchor="left"),
            )
            fig = _sw_style_fig(fig)
            _show_plotly(fig)

            display_df = table_df.copy()
            base_cols = ["Region", "Retailer", "Reviews", "Avg Rating", "Last 30d Avg", "30d Delta", "Gap vs Region Avg", "Share of View", "Alert Level", "Signal"] if "Region" in display_df.columns else ["Retailer", "Reviews", "Avg Rating", "Last 30d Avg", "30d Delta", "Gap vs Selected Avg", "Share of View", "Alert Level", "Signal"]
            display_cols = [c for c in base_cols if c in display_df.columns]
            if "Source System" in display_df.columns:
                present = display_df["Source System"].astype("string").fillna("").str.strip().replace("", pd.NA).dropna()
                if not present.empty:
                    display_cols.append("Source System")
            display_render_df = display_df[display_cols].copy()
            for col in ["Avg Rating", "Last 30d Avg", "30d Delta", "Gap vs Region Avg", "Gap vs Selected Avg"]:
                if col in display_render_df.columns:
                    display_render_df[col] = pd.to_numeric(display_render_df[col], errors="coerce").round(2)
            if "Share of View" in display_render_df.columns:
                display_render_df["Share of View"] = pd.to_numeric(display_render_df["Share of View"], errors="coerce").map(lambda v: f"{v:.1%}" if pd.notna(v) else "")

            if "Region" in display_render_df.columns:
                for region, sub in display_render_df.groupby("Region", sort=False):
                    st.markdown(f"<div class='section-sub' style='margin-top:.75rem;'><strong>{_esc(region)}</strong></div>", unsafe_allow_html=True)
                    st.dataframe(sub[[c for c in display_cols if c != "Region"]], use_container_width=True, hide_index=True)
            else:
                st.dataframe(display_render_df[display_cols], use_container_width=True, hide_index=True)

        with watch_tabs[1]:
            if alerts_df.empty:
                st.info("No smoke alerts are firing for the current selection.")
            else:
                top_alert = alerts_df.iloc[0]
                headline = f"Top alert: {_esc(top_alert.get('Retailer'))}"
                if "Region" in alerts_df.columns and _safe_text(top_alert.get("Region")):
                    headline += f" · {_esc(top_alert.get('Region'))}"
                detail = f"{_esc(top_alert.get('Alert Level'))} — {_esc(top_alert.get('Signal'))}"
                st.markdown(f"<div class='status-note'><strong>{headline}</strong><br>{detail}</div>", unsafe_allow_html=True)
                alerts_render = alerts_df.copy()
                for col in ["Avg Rating", "Last 30d Avg", "30d Delta", "Gap vs Region Avg", "Gap vs Selected Avg"]:
                    if col in alerts_render.columns:
                        alerts_render[col] = pd.to_numeric(alerts_render[col], errors="coerce").round(2)
                if "Share of View" in alerts_render.columns:
                    alerts_render["Share of View"] = pd.to_numeric(alerts_render["Share of View"], errors="coerce").map(lambda v: f"{v:.1%}" if pd.notna(v) else "")
                st.dataframe(alerts_render, use_container_width=True, hide_index=True)

        default_retailer = None
        default_region = "All selected"
        if not alerts_df.empty:
            default_retailer = _safe_text(alerts_df.iloc[0].get("Retailer"))
            if "Region" in alerts_df.columns:
                default_region = _safe_text(alerts_df.iloc[0].get("Region"), "All selected") or "All selected"
        if not default_retailer and not table_df.empty:
            default_retailer = _safe_text(table_df.iloc[0].get("Retailer"))
        retailer_options = ["All retailers (avg)"] + sorted([_safe_text(v) for v in table_df["Retailer"].dropna().unique() if _safe_text(v)])
        if st.session_state.get("retailer_watch_focus_retailer") not in retailer_options:
            st.session_state["retailer_watch_focus_retailer"] = default_retailer or retailer_options[0]
        region_focus_options = ["All selected"] + known_regions if known_regions else ["All selected"]
        if st.session_state.get("retailer_watch_focus_region") not in region_focus_options:
            st.session_state["retailer_watch_focus_region"] = default_region if default_region in region_focus_options else "All selected"

        with watch_tabs[2]:
            if trend_df.empty:
                st.info("Trend lines need a usable review date column. Select one in Data mapping to unlock weekly tracking.")
            else:
                t1, t2 = st.columns([1.2, 1])
                focus_retailer = t1.selectbox("Trend focus", retailer_options, key="retailer_watch_focus_retailer")
                focus_region = t2.selectbox("Trend region", region_focus_options, key="retailer_watch_focus_region")
                trend_view = trend_df.copy()
                if focus_retailer != "All retailers (avg)":
                    trend_view = trend_view[trend_view["Retailer"] == focus_retailer].copy()
                if focus_region != "All selected" and "Region" in trend_view.columns:
                    trend_view = trend_view[trend_view["Region"] == focus_region].copy()
                if trend_view.empty:
                    st.info("No trend rows match the current focus.")
                else:
                    if focus_retailer == "All retailers (avg)":
                        group_cols = ["Week"] + (["Region"] if ("Region" in trend_view.columns and focus_region == "All selected" and trend_view["Region"].nunique() > 1) else [])
                        trend_plot_df = trend_view.groupby(group_cols, dropna=False).agg(Reviews=("Reviews", "sum"), **{"Avg Rating": ("Avg Rating", "mean")}).reset_index()
                    else:
                        trend_plot_df = trend_view.copy()
                    line_color_col = "Region" if "Region" in trend_plot_df.columns and trend_plot_df["Region"].nunique() > 1 else None
                    fig = make_subplots(specs=[[{"secondary_y": True}]])
                    volume_df = trend_plot_df.groupby("Week", as_index=False)["Reviews"].sum().sort_values("Week")
                    fig.add_bar(x=volume_df["Week"], y=volume_df["Reviews"], name="Reviews", opacity=0.22, secondary_y=True)
                    if line_color_col:
                        for key, sub in trend_plot_df.sort_values("Week").groupby(line_color_col, sort=False):
                            fig.add_scatter(x=sub["Week"], y=sub["Avg Rating"], mode="lines+markers", name=str(key), secondary_y=False)
                    else:
                        sub = trend_plot_df.sort_values("Week")
                        name = focus_retailer if focus_retailer != "All retailers (avg)" else "Selected view"
                        fig.add_scatter(x=sub["Week"], y=sub["Avg Rating"], mode="lines+markers", name=name, secondary_y=False)
                    fig.update_layout(
                        height=360,
                        margin=dict(l=20, r=18, t=18, b=40),
                        plot_bgcolor="rgba(0,0,0,0)",
                        paper_bgcolor="rgba(0,0,0,0)",
                        font_family="Inter",
                        legend=dict(orientation="h", y=-0.22, x=0, xanchor="left"),
                    )
                    fig.update_yaxes(title_text="Average rating ★", range=[0, 5.2], secondary_y=False)
                    fig.update_yaxes(title_text="Reviews/week", secondary_y=True, rangemode="tozero")
                    fig = _sw_style_fig(fig)
                    _show_plotly(fig)
                    latest_row = trend_plot_df.sort_values("Week").iloc[-1]
                    latest_avg = pd.to_numeric(latest_row.get("Avg Rating"), errors="coerce")
                    latest_reviews = int(pd.to_numeric(latest_row.get("Reviews"), errors="coerce") or 0)
                    st.markdown(
                        f"<div class='status-note'><strong>Latest week:</strong> {pd.to_datetime(latest_row.get('Week')).strftime('%Y-%m-%d')} · <strong>{latest_avg:.2f}★</strong> across <strong>{latest_reviews:,}</strong> reviews</div>",
                        unsafe_allow_html=True,
                    )

        with watch_tabs[3]:
            if review_rows_df.empty:
                st.info("No review detail rows are available for the current selection.")
            else:
                r1, r2, r3, r4 = st.columns([1.2, 1.0, 0.85, 0.85])
                review_focus_retailer = r1.selectbox("Retailer focus", retailer_options[1:] or retailer_options, key="retailer_watch_reviews_retailer")
                review_focus_region = r2.selectbox("Region focus", region_focus_options, key="retailer_watch_reviews_region")
                recent_only = r3.toggle("Last 30 days only", value=True, key="retailer_watch_reviews_recent")
                low_only = r4.toggle("Ratings ≤3 only", value=True, key="retailer_watch_reviews_low")
                review_view = review_rows_df.copy()
                if review_focus_retailer and review_focus_retailer != "All retailers (avg)":
                    review_view = review_view[review_view["Retailer"] == review_focus_retailer].copy()
                if review_focus_region != "All selected" and "Region" in review_view.columns:
                    review_view = review_view[review_view["Region"] == review_focus_region].copy()
                if recent_only:
                    review_view = review_view[review_view["Recent Window"] == "Last 30d"].copy()
                if low_only:
                    review_view = review_view[pd.to_numeric(review_view["Rating"], errors="coerce") <= 3].copy()
                review_view = review_view.sort_values(["Rating", "Date"], ascending=[True, False], na_position="last").reset_index(drop=True)
                if review_view.empty:
                    st.info("No review rows match this smoke-investigation focus.")
                else:
                    render = review_view.copy()
                    render["Date"] = pd.to_datetime(render["Date"], errors="coerce").dt.strftime("%Y-%m-%d").fillna("")
                    render["Title"] = render["Title"].map(lambda x: _trunc(x, 100))
                    render["Review"] = render["Review"].map(lambda x: _trunc(x, 220))
                    st.dataframe(render[[c for c in ["Date", "Rating", "Region", "Retailer", "Title", "Review", "Symptom Detractors", "Symptom Delighters", "Post Link"] if c in render.columns]], use_container_width=True, hide_index=True)

        with watch_tabs[4]:
            if symptom_summary_df.empty:
                st.info("No symptomized fields are available in the current view yet. Run the Symptomizer or keep AI symptom columns during upload to unlock this overlay.")
            else:
                s1, s2 = st.columns([1.2, 1.0])
                symptom_focus_retailer = s1.selectbox("Retailer focus", retailer_options[1:] or retailer_options, key="retailer_watch_symptoms_retailer")
                symptom_focus_region = s2.selectbox("Region focus", region_focus_options, key="retailer_watch_symptoms_region")
                sym_view = symptom_summary_df.copy()
                if symptom_focus_retailer and symptom_focus_retailer != "All retailers (avg)":
                    sym_view = sym_view[sym_view["Retailer"] == symptom_focus_retailer].copy()
                if symptom_focus_region != "All selected":
                    sym_view = sym_view[sym_view["Region"] == symptom_focus_region].copy()
                if sym_view.empty:
                    st.info("No symptom summary rows match this focus.")
                else:
                    det_df = sym_view[sym_view["Side"] == "Detractor"].head(12).copy()
                    del_df = sym_view[sym_view["Side"] == "Delighter"].head(12).copy()
                    for frame in [det_df, del_df]:
                        if not frame.empty:
                            frame["Share of Reviews"] = pd.to_numeric(frame["Share of Reviews"], errors="coerce").map(lambda v: f"{v:.1%}" if pd.notna(v) else "")
                    sx1, sx2 = st.columns(2)
                    with sx1:
                        st.markdown("<div class='section-sub'><strong>Top detractors in this slice</strong></div>", unsafe_allow_html=True)
                        st.dataframe(det_df[[c for c in ["Label", "Mentions", "Share of Reviews"] if c in det_df.columns]], use_container_width=True, hide_index=True)
                    with sx2:
                        st.markdown("<div class='section-sub'><strong>Top delighters in this slice</strong></div>", unsafe_allow_html=True)
                        st.dataframe(del_df[[c for c in ["Label", "Mentions", "Share of Reviews"] if c in del_df.columns]], use_container_width=True, hide_index=True)


def _build_volume_bar_series(trend, volume_mode):
    if trend is None or trend.empty:
        return pd.DataFrame(columns=["x", "volume", "width_ms", "label"]), "Reviews/day"

    w = trend[["day", "daily_volume"]].copy()
    w["day"] = pd.to_datetime(w["day"], errors="coerce")
    w["daily_volume"] = pd.to_numeric(w["daily_volume"], errors="coerce").fillna(0)
    w = w.dropna(subset=["day"])
    if w.empty:
        return pd.DataFrame(columns=["x", "volume", "width_ms", "label"]), "Reviews/day"

    mode = _safe_text(volume_mode) or "Reviews/day"
    if mode == "Reviews/week":
        w["bucket_start"] = w["day"].dt.to_period("W-SUN").dt.start_time
        grouped = w.groupby("bucket_start", as_index=False).agg(volume=("daily_volume", "sum"))
        grouped["bucket_days"] = 7.0
        grouped["label"] = grouped["bucket_start"].dt.strftime("Week of %Y-%m-%d")
        axis_title = "Reviews/week"
    elif mode == "Reviews/month":
        w["bucket_start"] = w["day"].dt.to_period("M").dt.start_time
        grouped = w.groupby("bucket_start", as_index=False).agg(volume=("daily_volume", "sum"))
        grouped["bucket_days"] = grouped["bucket_start"].dt.days_in_month.astype(float)
        grouped["label"] = grouped["bucket_start"].dt.strftime("%Y-%m")
        axis_title = "Reviews/month"
    else:
        grouped = w.rename(columns={"day": "bucket_start", "daily_volume": "volume"}).copy()
        grouped["bucket_days"] = 1.0
        grouped["label"] = grouped["bucket_start"].dt.strftime("%Y-%m-%d")
        axis_title = "Reviews/day"

    grouped["x"] = grouped["bucket_start"] + pd.to_timedelta(grouped["bucket_days"] / 2.0, unit="D")
    grouped["width_ms"] = grouped["bucket_days"].map(lambda d: int(pd.Timedelta(days=max(float(d) - 0.15, 0.35)).total_seconds() * 1000))
    grouped["volume"] = pd.to_numeric(grouped["volume"], errors="coerce").fillna(0).astype(int)
    return grouped[["x", "volume", "width_ms", "label"]], axis_title


def _add_axis_break_indicator(fig, *, side="right"):
    if side == "right":
        x0, x1 = 0.988, 0.998
    else:
        x0, x1 = 0.002, 0.012
    y0 = 0.08
    fig.add_shape(type="line", xref="paper", yref="paper", x0=x0, y0=y0, x1=x1, y1=y0 + 0.02, line=dict(color="rgba(71,85,105,0.92)", width=2), layer="above")
    fig.add_shape(type="line", xref="paper", yref="paper", x0=x0, y0=y0 + 0.028, x1=x1, y1=y0 + 0.048, line=dict(color="rgba(71,85,105,0.92)", width=2), layer="above")


def _render_reviews_over_time_chart(df):
    with st.container(border=True):
        st.markdown("<div class='section-title'>📈 Cumulative Avg ★ Over Time by Region (Weighted)</div>", unsafe_allow_html=True)
        st.markdown("<div class='section-sub'>Cumulative average stays on the right axis. Volume bars are optional and use the left axis.</div>", unsafe_allow_html=True)

        r2c0, r2c1, r2c2, r2c3, r2c4, r2c5 = st.columns(6)
        organic_only = r2c0.toggle("Organic only", value=False, key="ot_organic_only")
        show_volume = r2c1.toggle("Show volume", value=bool(st.session_state.get("ot_show_volume", False)), key="ot_show_volume")
        show_overall = r2c2.toggle("Show overall", value=True, key="ot_show_overall")
        smoothing = r2c3.selectbox("Smoothing", ["Off", "7-day", "14-day", "30-day"], index=1, key="ot_smoothing")
        volume_mode = r2c4.selectbox("Volume bars", ["Reviews/day", "Reviews/week", "Reviews/month"], index=0, key="ot_volume_mode")
        y_view = r2c5.radio("Y-axis view", ["Zoomed-in", "Full scale"], horizontal=True, key="ot_y_view")

        top_n_sel = st.selectbox("Regions", ["All", 2, 4, 6, 8, 10], index=0, key="ot_top_n")
        top_n = None if top_n_sel == "All" else int(top_n_sel)

        trend, regions = _cumulative_avg_region_trend(df, organic_only=organic_only, top_n=top_n, smoothing_label=smoothing)
        if trend.empty:
            st.info("No dated reviews available for the over-time chart.")
            return

        volume_bars, volume_axis_title = _build_volume_bar_series(trend, volume_mode)
        fig = make_subplots(specs=[[{"secondary_y": True}]])

        if show_volume and not volume_bars.empty:
            fig.add_trace(
                go.Bar(
                    x=volume_bars["x"],
                    y=volume_bars["volume"],
                    width=volume_bars["width_ms"],
                    name=volume_axis_title,
                    marker_color="rgba(100,116,139,0.46)",
                    marker_line_color="rgba(71,85,105,0.58)",
                    opacity=0.88,
                    customdata=np.stack([volume_bars["label"]], axis=-1),
                    hovertemplate="%{customdata[0]}<br>Reviews: %{y:,}<extra></extra>",
                ),
                secondary_y=False,
            )

        region_colors = [
            "#f97316", "#10b981", "#3b82f6", "#ef4444",
            "#8b5cf6", "#eab308", "#06b6d4", "#ec4899",
            "#84cc16", "#f43f5e", "#14b8a6", "#a855f7",
            "#f59e0b", "#64748b", "#0ea5e9", "#e11d48",
            "#22c55e", "#d97706", "#7c3aed", "#be185d",
        ]
        for idx, region in enumerate(regions):
            col = f"{region}_cum_avg"
            if col not in trend.columns:
                continue
            fig.add_trace(
                go.Scatter(
                    x=trend["day"],
                    y=trend[col],
                    name=region,
                    mode="lines",
                    line=dict(color=region_colors[idx % len(region_colors)], width=2),
                    hovertemplate=f"{region}<br>%{{x|%Y-%m-%d}}<br>Cumulative Avg ★: %{{y:.3f}}<extra></extra>",
                ),
                secondary_y=True,
            )

        if show_overall and "overall_cum_avg" in trend.columns:
            fig.add_trace(
                go.Scatter(
                    x=trend["day"],
                    y=trend["overall_cum_avg"],
                    name="Overall",
                    mode="lines",
                    line=dict(color="#8b5cf6", width=3),
                    hovertemplate="Overall<br>%{x|%Y-%m-%d}<br>Cumulative Avg ★: %{y:.3f}<extra></extra>",
                ),
                secondary_y=True,
            )

        fig.update_layout(
            margin=dict(l=24, r=24, t=16, b=20),
            hovermode="x unified",
            plot_bgcolor="rgba(0,0,0,0)",
            paper_bgcolor="rgba(0,0,0,0)",
            font_family="Inter",
            legend=dict(orientation="h", y=-0.22, x=0, xanchor="left", yanchor="top", bgcolor="rgba(255,255,255,0.8)"),
            bargap=0.06,
            height=430,
        )
        fig.update_xaxes(title_text="", showgrid=False)

        visible_cols = [f"{region}_cum_avg" for region in regions if f"{region}_cum_avg" in trend.columns]
        if show_overall and "overall_cum_avg" in trend.columns:
            visible_cols.append("overall_cum_avg")
        vals = pd.concat([trend[c].dropna() for c in visible_cols], ignore_index=True) if visible_cols else pd.Series(dtype=float)

        if y_view == "Full scale" or vals.empty:
            y_range = [1.0, 5.0]
        else:
            ymin = max(1.0, float(vals.min()) - 0.06)
            ymax = min(5.0, float(vals.max()) + 0.06)
            if ymax - ymin < 0.18:
                mid = (ymin + ymax) / 2
                ymin = max(1.0, mid - 0.09)
                ymax = min(5.0, mid + 0.09)
            y_range = [ymin, ymax]

        fig.update_yaxes(title_text=volume_axis_title if show_volume else "", secondary_y=False, showgrid=False, rangemode="tozero", visible=show_volume)
        fig.update_yaxes(title_text="Cumulative Avg ★", range=y_range, secondary_y=True, showgrid=True, gridcolor="rgba(148,163,184,0.15)")

        if y_view == "Zoomed-in" and y_range[0] > 1.05:
            _add_axis_break_indicator(fig, side="right")

        _show_plotly(fig)

# ═══════════════════════════════════════════════════════════════════════════════
#  SYMPTOM ANALYTICS
# ═══════════════════════════════════════════════════════════════════════════════
def _symptom_col_lists_from_columns(columns):
    cleaned = [str(c).strip() for c in columns]
    man_det = [c for c in cleaned if c.lower() in {f"symptom {i}" for i in range(1, 11)}]
    man_del = [c for c in cleaned if c.lower() in {f"symptom {i}" for i in range(11, 21)}]
    ai_det = [c for c in cleaned if c.lower().startswith("ai symptom detractor")]
    ai_del = [c for c in cleaned if c.lower().startswith("ai symptom delighter")]
    return _dedupe_keep_order(man_det + ai_det), _dedupe_keep_order(man_del + ai_del)


def _get_symptom_col_lists(df):
    return _symptom_col_lists_from_columns(df.columns)


def _detect_symptom_state(df):
    det_cols, del_cols = _get_symptom_col_lists(df)

    def _has(cols):
        for c in cols:
            if c not in df.columns:
                continue
            s = df[c].astype("string").fillna("").str.strip()
            valid = (s != "") & (~s.str.upper().isin(SYMPTOM_NON_VALUES)) & (~s.str.startswith("<"))
            if valid.any():
                return True
        return False

    h_det = _has(det_cols)
    h_del = _has(del_cols)
    if h_det and h_del:
        return "full"
    if h_det or h_del:
        return "partial"
    return "none"


def _empty_symptom_table():
    return pd.DataFrame(columns=["Item", "Avg Star", "Mentions", "% Tagged Reviews", "Avg Tags/Review"])



def _analytics_excluded_symptom_labels() -> set[str]:
    return {
        _safe_text(_UNIVERSAL_NEUTRAL_DETRACTORS[0]),
        _safe_text(_UNIVERSAL_NEUTRAL_DELIGHTERS[0]),
    }



def _prepare_symptom_long(df_in, symptom_cols):
    if df_in is None or df_in.empty:
        return pd.DataFrame(columns=["__row", "symptom", "symptom_count", "review_weight", "star"]), 0
    targets = {str(col).strip() for col in symptom_cols if str(col).strip()}
    if not targets:
        return pd.DataFrame(columns=["__row", "symptom", "symptom_count", "review_weight", "star"]), 0
    col_names = [str(col).strip() for col in df_in.columns]
    positions = [idx for idx, name in enumerate(col_names) if name in targets]
    if not positions:
        return pd.DataFrame(columns=["__row", "symptom", "symptom_count", "review_weight", "star"]), 0

    block = df_in.iloc[:, positions].copy()
    block.columns = [f"__sym_{idx}" for idx in range(block.shape[1])]
    block.insert(0, "__row", np.arange(len(block), dtype=int))
    long = block.melt(id_vars="__row", value_name="symptom", var_name="__col")
    s = long["symptom"].astype("string").fillna("").str.strip()
    mask = (s != "") & (~s.str.upper().isin(SYMPTOM_NON_VALUES)) & (~s.str.startswith("<"))
    long = long.loc[mask, ["__row"]].copy()
    if long.empty:
        return pd.DataFrame(columns=["__row", "symptom", "symptom_count", "review_weight", "star"]), 0

    long["symptom"] = s.loc[mask].str.title()
    long = long.drop_duplicates(subset=["__row", "symptom"])
    if long.empty:
        return pd.DataFrame(columns=["__row", "symptom", "symptom_count", "review_weight", "star"]), 0

    symptomized_reviews = int(long["__row"].nunique())
    long = long.loc[~long["symptom"].isin(_analytics_excluded_symptom_labels())].copy()
    if long.empty:
        return pd.DataFrame(columns=["__row", "symptom", "symptom_count", "review_weight", "star"]), symptomized_reviews

    counts = long.groupby("__row", dropna=False)["symptom"].transform("nunique").astype(float)
    long["symptom_count"] = counts
    long["review_weight"] = (1.0 / counts.replace(0, np.nan)).fillna(0.0)
    if "rating" in df_in.columns:
        stars = pd.to_numeric(df_in.reset_index(drop=True)["rating"], errors="coerce").rename("star")
        long = long.join(stars, on="__row")
    else:
        long["star"] = np.nan
    return long, symptomized_reviews



def analyze_symptoms_fast(df_in, symptom_cols):
    long, symptomized_reviews = _prepare_symptom_long(df_in, symptom_cols)
    if long.empty:
        return _empty_symptom_table()

    grouped = long.groupby("symptom", dropna=False)
    mention_reviews = grouped["__row"].nunique().astype(int)
    avg_tags = grouped["symptom_count"].mean()
    avg_stars = grouped["star"].mean() if "star" in long.columns else pd.Series(index=mention_reviews.index, dtype=float)
    weighted_mentions = grouped["review_weight"].sum()

    out = pd.DataFrame({
        "Item": [str(item).title() for item in mention_reviews.index.tolist()],
        "Mentions": mention_reviews.values.astype(int),
        "% Tagged Reviews": (mention_reviews.values / max(symptomized_reviews, 1) * 100).round(1).astype(str) + "%",
        "Avg Star": [round(float(avg_stars[item]), 1) if item in avg_stars and not pd.isna(avg_stars[item]) else None for item in mention_reviews.index],
        "Avg Tags/Review": np.round(avg_tags.values.astype(float), 2),
        "__Weighted Mentions": weighted_mentions.values.astype(float),
        "__Mention Reviews": mention_reviews.values.astype(int),
        "__Symptomized Reviews": symptomized_reviews,
        "__All Reviews": int(len(df_in)),
    }).sort_values(["Mentions", "__Weighted Mentions", "Item"], ascending=[False, False, True], ignore_index=True)
    out.attrs["symptomized_review_count"] = symptomized_reviews
    out.attrs["all_review_count"] = int(len(df_in))
    return out


def _infer_symptom_total_reviews(tbl):
    if tbl is None or tbl.empty:
        return 0
    if "__All Reviews" in tbl.columns:
        total = int(pd.to_numeric(tbl["__All Reviews"], errors="coerce").fillna(0).max() or 0)
        if total > 0:
            return total
    pct_col = "% Tagged Reviews" if "% Tagged Reviews" in tbl.columns else ("% Total" if "% Total" in tbl.columns else None)
    if pct_col is None:
        return max(int(pd.to_numeric(tbl.get("Mentions"), errors="coerce").fillna(0).max() or 0), 0)
    pct = pd.to_numeric(tbl[pct_col].astype(str).str.replace("%", "", regex=False), errors="coerce")
    mentions = pd.to_numeric(tbl.get("Mentions"), errors="coerce").fillna(0)
    ratios = mentions / (pct / 100.0)
    ratios = ratios[(pct > 0) & ratios.notna() & (ratios > 0)]
    if ratios.empty:
        return max(int(mentions.max() or 0), 0)
    return max(int(round(float(ratios.median()))), 1)



def symptom_table_html(df_in, *, max_height_px=400):
    if df_in is None or df_in.empty:
        return f"<div class='sw-table-wrap' style='max-height:{max_height_px}px;padding:12px;'>No data.</div>"
    cols = [c for c in ["Item", "Mentions", "% Tagged Reviews", "Avg Star", "Avg Tags/Review", "Net Hit", "Forecast Δ★"] if c in df_in.columns]
    th = "".join(f"<th>{html.escape(c)}</th>" for c in cols)
    rows_html = []
    for _, row in df_in[cols].iterrows():
        tds = []
        for c in cols:
            v = row.get(c, "")
            right = "sw-td-right" if c in ("Mentions", "% Tagged Reviews", "Avg Star", "Avg Tags/Review", "Net Hit", "Forecast Δ★") else ""
            if c == "Avg Star":
                try:
                    f = float(v)
                    cls = "sw-star-good" if f >= 4.5 else "sw-star-bad"
                    tds.append(f"<td class='{right} {cls}'>{f:.1f}</td>")
                except Exception:
                    tds.append(f"<td class='{right}'>{html.escape(str(v))}</td>")
            elif c == "Avg Tags/Review":
                try:
                    tds.append(f"<td class='{right}'>{float(v):.2f}</td>")
                except Exception:
                    tds.append(f"<td class='{right}'>{html.escape(str(v))}</td>")
            elif c in {"Net Hit", "Forecast Δ★"}:
                try:
                    tds.append(f"<td class='{right}'>{float(v):.3f}</td>")
                except Exception:
                    tds.append(f"<td class='{right}'>{html.escape(str(v))}</td>")
            else:
                tds.append(f"<td class='{right}'>{html.escape(str(v))}</td>")
        rows_html.append("<tr>" + "".join(tds) + "</tr>")
    body = "".join(rows_html)
    return (
        f"<div class='sw-table-wrap' style='max-height:{max_height_px}px;'>"
        f"<table class='sw-table'><thead><tr>{th}</tr></thead><tbody>{body}</tbody></table></div>"
    )




_SYMPTOM_TABLE_BASE_COLUMNS = ["Item", "L1 Theme", "Mentions", "% Tagged Reviews", "Avg Star", "Avg Tags/Review", "Confidence %", "Net Hit", "Forecast Δ★", "Impact Score"]
_SYMPTOM_TABLE_AUX_COLUMNS = ["Severity Wt", "Impact |Abs|"]


def _symptom_table_for_display(df_in):
    if df_in is None or getattr(df_in, "empty", True):
        return pd.DataFrame(columns=list(_SYMPTOM_TABLE_BASE_COLUMNS))
    out = df_in.copy()
    keep = [c for c in _SYMPTOM_TABLE_BASE_COLUMNS if c in out.columns]
    if "Item" in out.columns:
        out["Item"] = out["Item"].astype(str)
    if "% Tagged Reviews" in out.columns:
        out["% Tagged Reviews"] = pd.to_numeric(out["% Tagged Reviews"].astype(str).str.replace("%", "", regex=False), errors="coerce")
    for col in ["Mentions", "Avg Star", "Avg Tags/Review", "Confidence %", "Net Hit", "Forecast Δ★", "Impact Score", "Severity Wt"]:
        if col in out.columns:
            out[col] = pd.to_numeric(out[col], errors="coerce")
    impact_source = None
    if "Impact Score" in out.columns:
        impact_source = pd.to_numeric(out["Impact Score"], errors="coerce")
    elif "Forecast Δ★" in out.columns:
        impact_source = pd.to_numeric(out["Forecast Δ★"], errors="coerce")
    elif "Net Hit" in out.columns:
        impact_source = pd.to_numeric(out["Net Hit"], errors="coerce")
    if impact_source is not None:
        out["Impact |Abs|"] = impact_source.abs().round(3)
    return out[[c for c in (keep + _SYMPTOM_TABLE_AUX_COLUMNS) if c in out.columns]]


def _symptom_table_column_config(df_in):
    try:
        cfg = {}
        if "Item" in df_in.columns:
            cfg["Item"] = st.column_config.TextColumn("Item")
        if "L1 Theme" in df_in.columns:
            cfg["L1 Theme"] = st.column_config.TextColumn("L1 Theme")
        if "Mentions" in df_in.columns:
            cfg["Mentions"] = st.column_config.NumberColumn("Mentions", format="%d")
        if "% Tagged Reviews" in df_in.columns:
            cfg["% Tagged Reviews"] = st.column_config.NumberColumn("% Tagged Reviews", format="%.1f")
        if "Avg Star" in df_in.columns:
            cfg["Avg Star"] = st.column_config.NumberColumn("Avg Star", format="%.1f")
        if "Avg Tags/Review" in df_in.columns:
            cfg["Avg Tags/Review"] = st.column_config.NumberColumn("Avg Tags/Review", format="%.2f")
        if "Confidence %" in df_in.columns:
            cfg["Confidence %"] = st.column_config.NumberColumn("Confidence %", format="%.1f")
        if "Net Hit" in df_in.columns:
            cfg["Net Hit"] = st.column_config.NumberColumn("Net Hit", format="%.3f")
        if "Forecast Δ★" in df_in.columns:
            cfg["Forecast Δ★"] = st.column_config.NumberColumn("Forecast Δ★", format="%.3f")
        if "Impact Score" in df_in.columns:
            cfg["Impact Score"] = st.column_config.NumberColumn("Impact Score", format="%.3f")
        if "Severity Wt" in df_in.columns:
            cfg["Severity Wt"] = st.column_config.NumberColumn("Severity Wt", format="%.2f")
        if "Impact |Abs|" in df_in.columns:
            cfg["Impact |Abs|"] = st.column_config.NumberColumn("Impact |Abs|", format="%.3f")
        return cfg
    except Exception:
        return {}


def _render_interactive_symptom_table(tbl, *, key_prefix, empty_label):
    display = _symptom_table_for_display(tbl)
    if display.empty:
        st.info(f"No {empty_label.lower()} data.")
        return

    st.caption("Click any column header to sort instantly. The table opens in the cleaner operating view; advanced scoring columns stay available in Table tools when you need them.")
    with st.expander("Table tools", expanded=False):
        r1c1, r1c2, r1c3 = st.columns([2.2, 1.05, 1.05])
        search = r1c1.text_input("Search", key=f"{key_prefix}_search", placeholder="Filter symptom names")
        min_mentions = int(r1c2.number_input("Min mentions", min_value=0, value=0, step=1, key=f"{key_prefix}_min_mentions"))
        min_pct = float(r1c3.number_input("Min % tagged", min_value=0.0, max_value=100.0, value=0.0, step=0.5, key=f"{key_prefix}_min_pct"))
        r2c1, r2c2, r2c3 = st.columns([1.35, 1.0, 2.65])
        sort_candidates = [c for c in (_SYMPTOM_TABLE_BASE_COLUMNS + _SYMPTOM_TABLE_AUX_COLUMNS) if c in display.columns]
        default_sort = "Impact |Abs|" if "Impact |Abs|" in sort_candidates else ("Impact Score" if "Impact Score" in sort_candidates else ("Mentions" if "Mentions" in sort_candidates else sort_candidates[0]))
        sort_by = r2c1.selectbox("Sort by", options=sort_candidates, index=sort_candidates.index(default_sort), key=f"{key_prefix}_sort_by")
        descending_default = sort_by != "Item"
        descending = bool(r2c2.toggle("Descending", value=descending_default, key=f"{key_prefix}_descending"))
        row_options = [25, 50, 100, 250, "All"]
        row_choice = r2c2.selectbox("Rows", options=row_options, index=1, key=f"{key_prefix}_row_limit")
        visible_options = [c for c in (_SYMPTOM_TABLE_BASE_COLUMNS + _SYMPTOM_TABLE_AUX_COLUMNS) if c in display.columns]
        visible_default = [c for c in _SYMPTOM_TABLE_BASE_COLUMNS if c in display.columns and c not in {"Confidence %", "Forecast Δ★", "Impact Score"}]
        visible_cols = r2c3.multiselect("Visible columns", options=visible_options, default=visible_default, key=f"{key_prefix}_visible_cols")

    filtered = display.copy()
    if search:
        mask = filtered["Item"].str.contains(re.escape(str(search).strip()), case=False, na=False)
        if "L1 Theme" in filtered.columns:
            mask = mask | filtered["L1 Theme"].astype(str).str.contains(re.escape(str(search).strip()), case=False, na=False)
        filtered = filtered[mask]
    if "Mentions" in filtered.columns:
        filtered = filtered[filtered["Mentions"].fillna(0) >= min_mentions]
    if "% Tagged Reviews" in filtered.columns:
        filtered = filtered[filtered["% Tagged Reviews"].fillna(0) >= min_pct]
    if sort_by in filtered.columns:
        filtered = filtered.sort_values(sort_by, ascending=not descending, na_position="last", kind="mergesort")
    if row_choice != "All":
        filtered = filtered.head(int(row_choice))
    visible_cols = [c for c in (visible_cols or visible_default) if c in filtered.columns] or [c for c in visible_default if c in filtered.columns]
    st.markdown(_chip_html([
        (f"{len(filtered):,} rows showing", "indigo"),
        (f"{len(display):,} total", "gray"),
        (f"Sorted by {sort_by}", "blue"),
    ]), unsafe_allow_html=True)
    if filtered.empty:
        st.info("No rows match the current symptom table filters.")
        return
    height_px = int(min(max(360, 36 * len(filtered) + 72), 760))
    st.dataframe(filtered[visible_cols], use_container_width=True, hide_index=True, height=height_px, column_config=_symptom_table_column_config(filtered[visible_cols]))


def _compute_detailed_symptom_impact(df_in, symptom_cols, baseline, *, kind):
    long, symptomized_reviews = _prepare_symptom_long(df_in, symptom_cols)
    if long.empty:
        return pd.DataFrame(columns=["Mention Reviews", "Avg Tags/Review", "Avg Star", "Weighted Mentions", "Net Hit Raw"]), 0

    stars = pd.to_numeric(long["star"], errors="coerce")
    if str(kind).lower().startswith("del"):
        gap = (stars - float(baseline)).clip(lower=0)
    else:
        gap = (float(baseline) - stars).clip(lower=0)
    long["gap"] = gap.fillna(0.0)
    long["attributed_gap"] = long["review_weight"].astype(float) * long["gap"].astype(float)

    grouped = long.groupby("symptom", dropna=False)
    out = grouped.agg(**{
        "Mention Reviews": ("__row", "nunique"),
        "Avg Tags/Review": ("symptom_count", "mean"),
        "Avg Star": ("star", "mean"),
        "Weighted Mentions": ("review_weight", "sum"),
        "Net Hit Raw": ("attributed_gap", "sum"),
    })
    return out, symptomized_reviews


_HIGH_SEVERITY_HINTS = {"broke", "broken", "break", "defect", "defective", "danger", "dangerous", "safety", "burn", "burned", "burnt", "smoke", "fire", "leak", "leaks", "leaking", "rash", "itch", "itchy", "irritat", "pain", "painful", "stopped working", "won't work", "doesn't work", "does not work", "shipping damage"}
_MEDIUM_SEVERITY_HINTS = {"poor performance", "unreliable", "hard to clean", "difficult", "connectivity", "battery", "charging", "loud", "wrong size", "instructions", "compatibility", "slow", "time consuming", "poor quality", "overpriced", "cheap", "flimsy", "hot", "overheat"}
_LOW_SEVERITY_HINTS = {"design", "appearance", "attractive", "stylish", "scent", "taste", "texture", "packaging", "price", "value"}
_HIGH_VALUE_DELIGHTS = {"reliable", "high quality", "performs well", "easy to use", "easy to clean", "saves time", "clear instructions", "compatible", "long battery life", "fast charging"}
_LOW_VALUE_DELIGHTS = {"attractive design", "pleasant scent", "great taste", "good texture", "good packaging"}


def _label_severity_weight(label, *, kind):
    norm = str(label or "").strip().lower()
    if not norm:
        return 1.0
    if str(kind).lower().startswith("del"):
        if any(token in norm for token in _HIGH_VALUE_DELIGHTS):
            return 1.08
        if any(token in norm for token in _LOW_VALUE_DELIGHTS):
            return 0.96
        return 1.0
    if any(token in norm for token in _HIGH_SEVERITY_HINTS):
        return 1.35
    if any(token in norm for token in _MEDIUM_SEVERITY_HINTS):
        return 1.18
    if any(token in norm for token in _LOW_SEVERITY_HINTS):
        return 0.96
    return 1.05


def _alignment_confidence(avg_star, baseline, *, kind):
    stars = pd.to_numeric(avg_star, errors="coerce")
    if str(kind).lower().startswith("del"):
        gap = (stars - float(baseline)).clip(lower=0)
    else:
        gap = (float(baseline) - stars).clip(lower=0)
    gap = gap.fillna(0.0)
    scaled = (gap / 1.25).clip(lower=0.0, upper=1.0)
    return 0.45 + 0.55 * scaled


def _add_net_hit(tbl, avg_rating, total_reviews=None, *, kind="detractors", shrink_k=3.0, detail_df=None, symptom_cols=None):
    if tbl is None or tbl.empty:
        return tbl
    d = tbl.copy()
    baseline = float(avg_rating or 0)
    sign = 1.0 if str(kind).lower().startswith("del") else -1.0
    total_reviews = int(total_reviews or _infer_symptom_total_reviews(d) or (len(detail_df) if detail_df is not None else 0) or 0)
    if total_reviews <= 0:
        total_reviews = max(int(pd.to_numeric(d.get("Mentions"), errors="coerce").fillna(0).sum() or 0), 1)

    details = None
    symptomized_reviews = int(d.attrs.get("symptomized_review_count") or 0)
    if detail_df is not None and symptom_cols:
        details, symptomized_reviews = _compute_detailed_symptom_impact(detail_df, symptom_cols, baseline, kind=kind)
        if not details.empty:
            details = details.copy()
            details.index = details.index.to_series().astype(str).str.title()

    d["Item"] = d.get("Item", pd.Series(dtype="string")).astype("string").fillna("").str.strip().str.title()
    d["Mentions"] = pd.to_numeric(d.get("Mentions"), errors="coerce").fillna(0).astype(int)
    d["Avg Star"] = pd.to_numeric(d.get("Avg Star"), errors="coerce")

    if details is not None and not details.empty:
        mention_reviews = d["Item"].map(details["Mention Reviews"]).fillna(d["Mentions"]).astype(float)
        weighted_mentions = d["Item"].map(details["Weighted Mentions"]).fillna(mention_reviews).astype(float)
        d["Mentions"] = mention_reviews.astype(int)
        d["Avg Tags/Review"] = d["Item"].map(details["Avg Tags/Review"]).fillna(pd.to_numeric(d.get("Avg Tags/Review"), errors="coerce"))
        d["Avg Star"] = d["Item"].map(details["Avg Star"]).fillna(d["Avg Star"])
        raw_impact = d["Item"].map(details["Net Hit Raw"]).fillna(0.0).astype(float) / float(max(total_reviews, 1))
        review_conf = mention_reviews / (mention_reviews + float(max(shrink_k, 0.1)))
        weight_conf = weighted_mentions / (weighted_mentions + float(max(shrink_k / 2.0, 0.5)))
        align_conf = _alignment_confidence(d["Avg Star"], baseline, kind=kind).astype(float)
        confidence = np.sqrt(review_conf * weight_conf) * align_conf
        d["Net Hit"] = (sign * raw_impact).round(3)
        d["Forecast Δ★"] = (d["Net Hit"].astype(float) * confidence).round(3)
        if symptomized_reviews <= 0:
            symptomized_reviews = int(details["Mention Reviews"].max() or 0)
    else:
        filled_stars = d["Avg Star"].fillna(baseline)
        if str(kind).lower().startswith("del"):
            rating_gap = (filled_stars - baseline).clip(lower=0)
        else:
            rating_gap = (baseline - filled_stars).clip(lower=0)
        share = d["Mentions"].astype(float) / float(max(total_reviews, 1))
        base_conf = d["Mentions"].astype(float) / (d["Mentions"].astype(float) + float(max(shrink_k, 0.1)))
        align_conf = _alignment_confidence(filled_stars, baseline, kind=kind).astype(float)
        confidence = base_conf * align_conf
        d["Net Hit"] = (sign * share * rating_gap).round(3)
        d["Forecast Δ★"] = (d["Net Hit"].astype(float) * confidence).round(3)
        if "Avg Tags/Review" not in d.columns:
            d["Avg Tags/Review"] = np.nan

    if symptomized_reviews <= 0 and "__Symptomized Reviews" in d.columns:
        symptomized_reviews = int(pd.to_numeric(d["__Symptomized Reviews"], errors="coerce").fillna(0).max() or 0)
    if symptomized_reviews <= 0:
        symptomized_reviews = max(int(d["Mentions"].max() or 0), 1)

    pct_vals = (d["Mentions"].astype(float) / float(max(symptomized_reviews, 1)) * 100).round(1)
    d["% Tagged Reviews"] = pct_vals.astype(str) + "%"
    d["Avg Tags/Review"] = pd.to_numeric(d.get("Avg Tags/Review"), errors="coerce").round(2)
    d["Confidence %"] = (pd.to_numeric(confidence, errors="coerce").fillna(0.0).clip(lower=0.0, upper=1.0) * 100.0).round(1)
    d["Severity Wt"] = d["Item"].map(lambda value: round(float(_label_severity_weight(value, kind=kind)), 2))
    d["Impact Score"] = (d["Forecast Δ★"].astype(float) * d["Severity Wt"].astype(float)).round(3)
    d.attrs["symptomized_review_count"] = symptomized_reviews
    d.attrs["all_review_count"] = total_reviews
    d["_impact_sort"] = np.maximum.reduce([
        d["Impact Score"].astype(float).abs(),
        d["Forecast Δ★"].astype(float).abs(),
        d["Net Hit"].astype(float).abs(),
    ])
    cols = [c for c in ["Item", "Mentions", "% Tagged Reviews", "Avg Star", "Avg Tags/Review", "Confidence %", "Net Hit", "Forecast Δ★", "Impact Score", "Severity Wt"] if c in d.columns]
    return d.sort_values(["_impact_sort", "Mentions", "Item"], ascending=[False, False, True], ignore_index=True)[cols]

def _collect_row_symptom_tags(row, cols):
    tags = []
    seen = set()
    for col in cols:
        raw = row.get(col)
        txt = _safe_text(raw).strip()
        if not txt or txt.upper() in SYMPTOM_NON_VALUES or txt.startswith("<"):
            continue
        label = txt.title()
        if label in seen:
            continue
        seen.add(label)
        tags.append(label)
    return tags


def _opp_scatter(tbl, kind, baseline_avg, *, container_key=""):

    if tbl is None or tbl.empty:
        st.info("No data available.")
        return
    d = tbl.copy()
    d["Mentions"] = pd.to_numeric(d.get("Mentions"), errors="coerce").fillna(0)
    d["Avg Star"] = pd.to_numeric(d.get("Avg Star"), errors="coerce")
    d = d.dropna(subset=["Avg Star"])
    if d.empty:
        st.info("No data available.")
        return
    x = d["Mentions"].astype(float).to_numpy()
    y = d["Avg Star"].astype(float).to_numpy()
    names = d["Item"].astype(str).to_numpy()
    avg_tags = pd.to_numeric(d.get("Avg Tags/Review"), errors="coerce").fillna(np.nan) if "Avg Tags/Review" in d.columns else pd.Series(np.nan, index=d.index)
    confidence_col = pd.to_numeric(d.get("Confidence %"), errors="coerce") if "Confidence %" in d.columns else pd.Series(dtype=float)
    impact_col = pd.to_numeric(d.get("Impact Score"), errors="coerce") if "Impact Score" in d.columns else pd.Series(dtype=float)
    forecast_col = pd.to_numeric(d.get("Forecast Δ★"), errors="coerce") if "Forecast Δ★" in d.columns else pd.Series(dtype=float)
    if not impact_col.empty and impact_col.notna().any():
        score_signed = impact_col.fillna(0).to_numpy()
        score_strength = np.abs(score_signed)
        score_label = "Impact Score"
    elif not forecast_col.empty and forecast_col.notna().any():
        score_signed = forecast_col.fillna(0).to_numpy()
        score_strength = np.abs(score_signed)
        score_label = "Forecast Δ★"
    else:
        score_signed = (x * np.clip(float(baseline_avg) - y, 0, None) * (-1 if kind == "detractors" else 1))
        score_strength = np.abs(score_signed)
        score_label = "Impact |Abs|"
    show_labels = st.toggle("Show labels", value=False, key=f"opp_lbl_{kind}_{container_key}")
    labels_arr = np.array([""] * len(d), dtype=object)
    if show_labels:
        top_idx = np.argsort(-score_strength)[:10]
        labels_arr[top_idx] = names[top_idx]
    mx = max(float(np.nanmax(x)), 1e-9)
    size = (np.sqrt(x) / np.sqrt(mx)) * 24 + 8
    color = "#ef4444" if kind == "detractors" else "#22c55e"
    fig = go.Figure()
    fig.add_trace(go.Scatter(
        x=x,
        y=y,
        mode="markers+text" if show_labels else "markers",
        text=labels_arr,
        textposition="top center",
        textfont=dict(size=10, family="Inter"),
        customdata=np.column_stack([
            names,
            avg_tags.to_numpy(),
            forecast_col.reindex(d.index).fillna(0).to_numpy() if not forecast_col.empty else np.zeros(len(d)),
            impact_col.reindex(d.index).fillna(0).to_numpy() if not impact_col.empty else np.zeros(len(d)),
            confidence_col.reindex(d.index).fillna(0).to_numpy() if not confidence_col.empty else np.zeros(len(d)),
        ]),
        hovertemplate="%{customdata[0]}<br>Mentions=%{x:.0f}<br>Avg ★=%{y:.2f}<br>Avg tags/review=%{customdata[1]:.2f}<br>Forecast Δ★=%{customdata[2]:.3f}<br>Impact Score=%{customdata[3]:.3f}<br>Confidence=%{customdata[4]:.1f}%<extra></extra>",
        marker=dict(size=size, color=color, opacity=0.76, line=dict(width=1, color="rgba(148,163,184,0.38)")),
    ))
    fig.add_hline(y=float(baseline_avg), line_dash="dash", opacity=0.45, annotation_text=f"Avg ★ {baseline_avg:.2f}", annotation_position="right", annotation_font_size=11)
    fig.update_layout(height=420, xaxis_title="Mentions", yaxis_title="Avg ★")
    _sw_style_fig(fig)
    _show_plotly(fig)
    label = "Fix first — high mentions × below-baseline ★" if kind == "detractors" else "Amplify — high mentions × above-baseline ★"
    top15 = d.copy()
    top15[score_label] = score_strength
    top15 = top15.sort_values(score_label, ascending=False).head(15)
    with st.expander(f"📋 {label}", expanded=False):
        keep_cols = [c for c in ["Item", "Mentions", "Avg Star", "Avg Tags/Review", "Confidence %", "Net Hit", "Forecast Δ★", "Impact Score", score_label] if c in top15.columns]
        ds = top15[keep_cols].copy()
        ds["Avg Star"] = ds["Avg Star"].map(lambda v: f"{float(v):.1f}" if pd.notna(v) else "—")
        if "Avg Tags/Review" in ds.columns:
            ds["Avg Tags/Review"] = ds["Avg Tags/Review"].map(lambda v: f"{float(v):.2f}" if pd.notna(v) else "—")
        if "Confidence %" in ds.columns:
            ds["Confidence %"] = ds["Confidence %"].map(lambda v: f"{float(v):.1f}%" if pd.notna(v) else "—")
        for col in ["Net Hit", "Forecast Δ★", "Impact Score", score_label]:
            if col in ds.columns:
                ds[col] = ds[col].map(lambda v: f"{float(v):.3f}" if pd.notna(v) else "—")
        st.dataframe(ds, use_container_width=True, hide_index=True)


def _render_symptom_bar_chart(tbl, title, color, denom, show_pct):
    if tbl is None or tbl.empty:
        st.info(f"No {title.lower()} data.")
        return
    t = tbl.copy()
    t["Mentions"] = pd.to_numeric(t["Mentions"], errors="coerce").fillna(0)
    symptomized_reviews = 0
    if "__Symptomized Reviews" in t.columns:
        symptomized_reviews = int(pd.to_numeric(t["__Symptomized Reviews"], errors="coerce").fillna(0).max() or 0)
    if symptomized_reviews <= 0:
        symptomized_reviews = int(getattr(tbl, "attrs", {}).get("symptomized_review_count") or 0)
    if symptomized_reviews <= 0:
        symptomized_reviews = int(max(denom, 1))
    t["Pct"] = t["Mentions"] / max(symptomized_reviews, 1) * 100
    x_vals = t["Pct"][::-1] if show_pct else t["Mentions"][::-1]
    x_label = "% of symptomized reviews" if show_pct else "Mentions"
    hover = "%{customdata}<br>% of symptomized reviews: %{x:.1f}%<extra></extra>" if show_pct else "%{customdata}<br>Mentions: %{x:.0f}<extra></extra>"
    fig = go.Figure(go.Bar(x=x_vals, y=t["Item"][::-1], orientation="h", marker_color=color, opacity=0.80, customdata=t["Item"][::-1].astype(str).tolist(), hovertemplate=hover))
    fig.update_layout(title=title, height=max(300, 28 * len(t) + 80), xaxis_title=x_label, yaxis_title="", margin=dict(l=160, r=20, t=46, b=30))
    _sw_style_fig(fig)
    _show_plotly(fig)


def _render_symptom_dashboard(filtered_df, overall_df=None):
    od = overall_df if overall_df is not None else filtered_df
    sym_state = _detect_symptom_state(od)
    st.markdown("<hr class='sw-divider'>", unsafe_allow_html=True)
    if sym_state == "none":
        st.markdown("""<div class="sym-state-banner">
          <div class="icon">💊</div><div class="title">No symptoms tagged yet</div>
          <div class="sub">Run the <strong>Symptomizer</strong> tab to AI-tag delighters and detractors,
          then return here for the full analytics.<br>
          If your file already contains <em>Symptom 1–20</em> or <em>AI Symptom</em> columns they'll appear automatically.</div>
        </div>""", unsafe_allow_html=True)
        return
    if sym_state == "partial":
        det_cols, del_cols = _get_symptom_col_lists(od)
        missing = []
        if not _filled_mask(od, det_cols).any():
            missing.append("detractors")
        if not _filled_mask(od, del_cols).any():
            missing.append("delighters")
        if missing:
            st.info(f"Partial tagging — {' and '.join(missing)} not yet labelled.")
    st.markdown("<div class='section-title'>🩺 Detractors & Delighters</div>", unsafe_allow_html=True)
    st.markdown("<div class='section-sub'>% Tagged Reviews uses only symptomized reviews. Net Hit stays signed and observed, Forecast Δ★ keeps shrinkage for steadier forecasting, and Impact Score now layers in confidence and symptom severity so the tables surface the most actionable themes first.</div>", unsafe_allow_html=True)
    with st.expander("How % Tagged Reviews, Net Hit, and Forecast Δ★ work", expanded=False):
        st.markdown("""
- **% Tagged Reviews** = review mentions / reviews in the current view that already have at least one tag on that side (detractor or delighter).
- **Avg Star** = average star rating of reviews that mention the symptom.
- **Avg Tags/Review** shows how many same-side symptoms those reviews usually carry. Higher values mean the review impact is split across more themes.
- **Net Hit** is the observed signed star impact in the filtered review set. Detractors are negative, delighters are positive, and each review's rating gap versus the product average is divided across the same-side symptoms tagged in that review.
- **Confidence %** blends sample size, weighted review share, and rating alignment. Themes mentioned in few reviews or with weak star separation get pushed down.
- **Forecast Δ★** is the shrunk, confidence-adjusted version of Net Hit. It is the steadier estimate to use when forecasting what each symptom is doing to rating.
- **Severity Wt** lightly boosts engineering-critical themes like breakage, defects, leaks, or irritation and lightly down-weights aesthetic-only themes.
- **Impact Score** = Forecast Δ★ × Severity Wt. It is the best single sort for what deserves action first.
- **Impact |Abs|** in Table tools is the magnitude view of Impact Score when you want the strongest movers regardless of sign.
        """)
    det_cols, del_cols = _get_symptom_col_lists(od)
    avg_star = float(_safe_mean(filtered_df["rating"]) or 0)
    total_reviews = len(filtered_df)
    det_base = analyze_symptoms_fast(filtered_df, det_cols)
    del_base = analyze_symptoms_fast(filtered_df, del_cols)
    det_tbl = _add_net_hit(det_base, avg_star, total_reviews=total_reviews, kind="detractors", detail_df=filtered_df, symptom_cols=det_cols)
    del_tbl = _add_net_hit(del_base, avg_star, total_reviews=total_reviews, kind="delighters", detail_df=filtered_df, symptom_cols=del_cols)
    st.caption("Full-width interactive tables below keep every metric visible. Click a header to sort, or open Table tools for saved filters, severity-aware impact ranking, and column visibility.")
    t1, t2 = st.tabs([f"🔴 Detractors ({len(det_tbl):,})", f"🟢 Delighters ({len(del_tbl):,})"])
    with t1:
        with st.container(border=True):
            _render_interactive_symptom_table(det_tbl, key_prefix="sw_det_table", empty_label="detractor")
    with t2:
        with st.container(border=True):
            _render_interactive_symptom_table(del_tbl, key_prefix="sw_del_table", empty_label="delighter")
    try:
        out_xlsx = io.BytesIO()
        with pd.ExcelWriter(out_xlsx, engine="openpyxl") as writer:
            det_tbl.to_excel(writer, sheet_name="Detractors", index=False)
            del_tbl.to_excel(writer, sheet_name="Delighters", index=False)
        ds = st.session_state.get("analysis_dataset") or {}
        pid = _safe_summary_product_slug(ds.get("summary"), ds.get("reviews_df"), default="symptoms")
        st.download_button(
            "⬇️ Download Detractors + Delighters",
            data=out_xlsx.getvalue(),
            file_name=f"{pid}_symptoms.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="sw_sym_dl",
        )
    except Exception:
        pass
    st.markdown("<hr class='sw-divider'>", unsafe_allow_html=True)
    st.markdown("<div class='section-title'>📊 Top Themes</div>", unsafe_allow_html=True)
    bar_ctrl = st.columns([1, 1, 1.2])
    top_n = int(bar_ctrl[0].slider("Top N", 5, 25, 12, 1, key="sw_top_n"))
    org_only = bar_ctrl[1].toggle("Organic only", value=False, key="sw_org_bar")
    show_pct = bar_ctrl[2].toggle("Show %", value=False, key="sw_pct_bar")
    bar_df = filtered_df[~filtered_df["incentivized_review"].fillna(False)] if org_only else filtered_df
    denom = max(1, len(bar_df))
    det_top = analyze_symptoms_fast(bar_df, det_cols).head(top_n)
    del_top = analyze_symptoms_fast(bar_df, del_cols).head(top_n)
    bc1, bc2 = st.columns(2)
    with bc1:
        with st.container(border=True):
            _render_symptom_bar_chart(det_top, "Top Detractors", "#ef4444", denom, show_pct)
    with bc2:
        with st.container(border=True):
            _render_symptom_bar_chart(del_top, "Top Delighters", "#22c55e", denom, show_pct)
    st.markdown("<hr class='sw-divider'>", unsafe_allow_html=True)
    st.markdown("<div class='section-title'>🎯 Opportunity Matrix</div>", unsafe_allow_html=True)
    st.markdown("<div class='section-sub'>Mentions vs Avg ★ · Fix high-mention low-star detractors first · Amplify high-mention high-star delighters.</div>", unsafe_allow_html=True)
    opp_t1, opp_t2 = st.tabs(["🔴 Detractors", "🟢 Delighters"])
    with opp_t1:
        _opp_scatter(det_tbl, "detractors", avg_star, container_key="dash")
    with opp_t2:
        _opp_scatter(del_tbl, "delighters", avg_star, container_key="dash")

# ═══════════════════════════════════════════════════════════════════════════════
#  FILTERS
# ═══════════════════════════════════════════════════════════════════════════════
CORE_REVIEW_FILTER_SPECS = [
    {"key": "product_or_sku", "label": "SKU / Product", "kind": "column", "aliases": ["product_or_sku", "product_name", "product", "product_id", "sku", "item_id", "model", "model_number", "asin"]},
    {"key": "content_locale", "label": "Market / Locale", "kind": "column", "aliases": ["content_locale", "locale", "market", "region", "country", "country_code", "geo", "marketplace"]},
    {"key": "retailer", "label": "Retailer", "kind": "column", "aliases": ["retailer", "merchant", "channel", "store", "seller", "retailer_name", "retail_partner", "site_name"]},
    {"key": "source_system", "label": "Source System", "kind": "column", "aliases": ["source_system", "source", "platform", "provider", "review_source", "connector", "system"]},
    {"key": "loaded_from_host", "label": "Loaded Site", "kind": "column", "aliases": ["loaded_from_host", "loaded_host", "domain", "host", "site", "website"]},
    {"key": "source_file", "label": "Source File", "kind": "column", "aliases": ["source_file", "file_name", "filename", "import_file", "upload_name"]},
    {"key": "age_group", "label": "Age Group", "kind": "column", "aliases": ["age_group", "age_bucket", "age_band", "age"]},
    {"key": "user_location", "label": "Reviewer Location", "kind": "column", "aliases": ["user_location", "reviewer_location", "location", "city", "state", "province"]},
    {"key": "review_type", "label": "Review Type", "kind": "derived"},
    {"key": "recommendation", "label": "Recommendation", "kind": "derived"},
    {"key": "syndication", "label": "Syndication", "kind": "derived"},
    {"key": "media", "label": "Media", "kind": "derived"},
]

WATCH_SOURCE_COLUMN_ALIASES = [
    "retailer", "retailer_name", "merchant", "store", "seller", "retail_partner", "site_name",
    "channel", "source_label", "source", "source_system", "platform", "provider", "loaded_from_host",
]
WATCH_REGION_COLUMN_ALIASES = [
    "reviewer_location", "Reviewer Location", "user_location", "location", "content_locale", "locale", "region", "market", "country", "country_code", "geo", "marketplace",
]
WATCH_DATE_COLUMN_ALIASES = [
    "submission_time", "submission_date", "review_date", "published", "published_at", "created_at", "date",
    "timestamp", "posted_at", "opened_date", "opened", "opened_at",
]
WATCH_ORGANIC_COLUMN_ALIASES = [
    "incentivized_review", "is_incentivized", "incentivized", "sponsored", "is_sponsored", "paid",
    "seeded_flag", "seeded", "gifted", "gifted_flag",
]
AUTO_COLUMN_SENTINEL = "__auto__"
NONE_COLUMN_SENTINEL = "__none__"


def _normalize_col_key(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "_", str(value or "").strip().lower()).strip("_")


def _all_dataframe_columns(df: Optional[pd.DataFrame]) -> List[str]:
    if df is None or not isinstance(df, pd.DataFrame):
        return []
    seen: Set[str] = set()
    ordered: List[str] = []
    for col in df.columns:
        label = str(col)
        if label in seen:
            continue
        seen.add(label)
        ordered.append(label)
    return ordered


def _resolve_column_alias(df: Optional[pd.DataFrame], aliases: Sequence[str]) -> Optional[str]:
    columns = _all_dataframe_columns(df)
    if not columns:
        return None
    norm_to_actual: Dict[str, str] = {}
    for col in columns:
        norm_to_actual.setdefault(_normalize_col_key(col), col)

    alias_list = [str(alias) for alias in aliases or [] if str(alias).strip()]
    for alias in alias_list:
        match = norm_to_actual.get(_normalize_col_key(alias))
        if match:
            return match

    best: Optional[Tuple[Tuple[int, int, int], str]] = None
    for col in columns:
        norm_col = _normalize_col_key(col)
        col_tokens = {tok for tok in norm_col.split("_") if tok}
        for alias in alias_list:
            norm_alias = _normalize_col_key(alias)
            alias_tokens = {tok for tok in norm_alias.split("_") if tok}
            score: Optional[Tuple[int, int, int]] = None
            if norm_alias and norm_alias in norm_col:
                score = (4, len(norm_alias), -len(norm_col))
            elif alias_tokens and alias_tokens.issubset(col_tokens):
                score = (3, len(alias_tokens), -len(norm_col))
            elif len(alias_tokens) >= 2 and len(alias_tokens & col_tokens) >= 2:
                score = (2, len(alias_tokens & col_tokens), -len(norm_col))
            if score and (best is None or score > best[0]):
                best = (score, col)
    return best[1] if best else None


def _resolve_optional_column_choice(
    df: Optional[pd.DataFrame],
    selection: Optional[str],
    aliases: Sequence[str],
    *,
    allow_none: bool = False,
) -> Optional[str]:
    choice = str(selection or AUTO_COLUMN_SENTINEL)
    columns = set(_all_dataframe_columns(df))
    if allow_none and choice == NONE_COLUMN_SENTINEL:
        return None
    if choice not in {AUTO_COLUMN_SENTINEL, NONE_COLUMN_SENTINEL} and choice in columns:
        return choice
    return _resolve_column_alias(df, aliases)


def _select_column_series(df: Optional[pd.DataFrame], column: Optional[str]) -> Optional[pd.Series]:
    """Return a single series for a column label, even if the dataframe has duplicates."""
    if df is None or not isinstance(df, pd.DataFrame) or not column or column not in df.columns:
        return None
    selected = df.loc[:, column]
    if isinstance(selected, pd.Series):
        return selected.copy()
    if isinstance(selected, pd.DataFrame):
        if selected.shape[1] == 1:
            series = selected.iloc[:, 0].copy()
            series.name = column
            return series
        cleaned = selected.copy()
        for idx in range(cleaned.shape[1]):
            ser = cleaned.iloc[:, idx]
            if pd.api.types.is_object_dtype(ser) or pd.api.types.is_string_dtype(ser):
                cleaned.iloc[:, idx] = ser.astype("string").str.strip().replace("", pd.NA)
        collapsed = cleaned.bfill(axis=1).iloc[:, 0]
        collapsed.name = column
        return collapsed
    return pd.Series(selected, index=df.index, name=column)


def _collapse_duplicate_named_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Collapse duplicate column labels by taking the first non-empty value left-to-right."""
    if df is None or not isinstance(df, pd.DataFrame):
        return df
    labels = [str(col) for col in df.columns]
    if len(labels) == len(set(labels)):
        return df
    ordered: List[str] = []
    data: Dict[str, pd.Series] = {}
    for label in labels:
        if label in data:
            continue
        series = _select_column_series(df, label)
        if series is None:
            continue
        ordered.append(label)
        data[label] = series
    return pd.DataFrame({label: data[label] for label in ordered}, index=df.index)


def _source_rating_watch_column_options(df: Optional[pd.DataFrame], *, allow_none: bool = False) -> List[str]:
    options: List[str] = [AUTO_COLUMN_SENTINEL]
    if allow_none:
        options.append(NONE_COLUMN_SENTINEL)
    options.extend(sorted(_all_dataframe_columns(df), key=lambda x: x.lower()))
    return options


def _format_column_choice_label(value: str) -> str:
    if value == AUTO_COLUMN_SENTINEL:
        return "Default / auto-detect"
    if value == NONE_COLUMN_SENTINEL:
        return "None"
    return str(value)


def _series_matches_any(series: pd.Series, values: Sequence[str]) -> pd.Series:
    lookup = {str(v).strip().lower() for v in values if str(v).strip()}
    s = series.astype("string").fillna("").str.strip().str.lower()
    return s.isin(lookup)


def _sanitize_multiselect(key: str, options: Sequence[Any], default: Optional[Sequence[Any]] = None):
    opts = list(options or [])
    default_vals = list(default or ["ALL"])
    cur = st.session_state.get(key, default_vals)
    if not isinstance(cur, list):
        cur = [cur]
    cur = [x for x in cur if x in opts]
    if not cur:
        cur = default_vals
    if "ALL" in cur and len(cur) > 1:
        cur = [x for x in cur if x != "ALL"]
    st.session_state[key] = cur


def _sanitize_multiselect_sym(key: str, options: Sequence[str], default: Optional[Sequence[str]] = None):
    opts = list(options or [])
    default_vals = list(default or ["All"])
    cur = st.session_state.get(key, default_vals)
    if not isinstance(cur, list):
        cur = [cur]
    cur = [x for x in cur if x in opts]
    if not cur:
        cur = default_vals
    if "All" in cur and len(cur) > 1:
        cur = [x for x in cur if x != "All"]
    st.session_state[key] = cur


def _reset_review_filters():
    for key in list(st.session_state.keys()):
        if key.startswith("rf_"):
            st.session_state.pop(key, None)
    st.session_state["review_explorer_page"] = 1
    st.session_state["review_filter_signature"] = None


def _filter_series_for_spec(df: pd.DataFrame, spec: Dict[str, Any]) -> pd.Series:
    key = spec["key"]
    if spec["kind"] == "column":
        s = df[spec["column"]].astype("string").fillna("Unknown").str.strip()
        return s.replace("", "Unknown")
    if key == "review_type":
        raw = df.get("incentivized_review", pd.Series(False, index=df.index)).astype("boolean")
        return raw.map({True: "Incentivized", False: "Organic"}).fillna("Unknown")
    if key == "recommendation":
        raw = df.get("is_recommended", pd.Series(pd.NA, index=df.index)).astype("boolean")
        return raw.map({True: "Recommended", False: "Not Recommended"}).fillna("Unknown")
    if key == "syndication":
        raw = df.get("is_syndicated", pd.Series(pd.NA, index=df.index)).astype("boolean")
        return raw.map({True: "Syndicated", False: "Not Syndicated"}).fillna("Unknown")
    if key == "media":
        raw = df.get("has_photos", pd.Series(False, index=df.index)).astype("boolean")
        return raw.map({True: "With Photos", False: "No Photos"}).fillna("Unknown")
    return pd.Series("Unknown", index=df.index)


def _core_filter_specs_for_df(df: pd.DataFrame) -> List[Dict[str, Any]]:
    specs: List[Dict[str, Any]] = []
    used_columns: Set[str] = set()
    for spec in CORE_REVIEW_FILTER_SPECS:
        resolved_spec = dict(spec)
        if spec["kind"] == "column":
            resolved_col = _resolve_column_alias(df, spec.get("aliases") or ([spec.get("column")] if spec.get("column") else []))
            if not resolved_col or resolved_col in used_columns:
                continue
            resolved_spec["column"] = resolved_col
            used_columns.add(resolved_col)
        s = _filter_series_for_spec(df, resolved_spec)
        opts = [x for x in sorted({str(v).strip() for v in s.dropna().astype(str) if str(v).strip()}, key=lambda x: x.lower()) if x]
        if not opts:
            continue
        if len(opts) == 1 and opts[0] == "Unknown":
            continue
        specs.append({**resolved_spec, "options": ["ALL"] + opts})
    return specs


def _resolved_core_filter_columns(df: pd.DataFrame) -> Set[str]:
    return {str(spec.get("column")) for spec in _core_filter_specs_for_df(df) if spec.get("kind") == "column" and spec.get("column")}


def _looks_like_blob_or_text_column(df: pd.DataFrame, col: str) -> bool:
    if col not in df.columns:
        return False
    norm_name = _normalize_col_key(col)
    if any(tok in norm_name for tok in ["review_text", "title_and_text", "body", "content", "description", "summary", "analysis", "evidence", "prompt", "response", "raw", "json", "html"]):
        return True
    s = df[col].astype("string").fillna("").str.strip()
    non_empty = s[s != ""].head(120)
    if non_empty.empty:
        return False
    avg_len = float(non_empty.str.len().mean())
    return avg_len >= 90


def _looks_like_identifier_or_link(df: pd.DataFrame, col: str) -> bool:
    if col not in df.columns:
        return False
    norm_name = _normalize_col_key(col)
    s = df[col].astype("string").fillna("").str.strip()
    non_empty = s[s != ""]
    if non_empty.empty:
        return False
    nunique = int(non_empty.nunique(dropna=True))
    unique_ratio = nunique / max(len(non_empty), 1)
    if any(tok in norm_name for tok in ["url", "uri", "link"]):
        return unique_ratio >= 0.35
    if any(tok in norm_name for tok in ["uuid", "guid", "hash"]):
        return True
    if norm_name.endswith("_id") or norm_name == "id" or norm_name.startswith("id_"):
        return unique_ratio >= 0.8
    return False


def _extra_filter_column_score(df: pd.DataFrame, col: str) -> Optional[Tuple[int, str]]:
    if col not in df.columns:
        return None
    kind = _infer_extra_filter_kind(df, col)
    s = df[col]
    non_null_ratio = float(s.notna().mean()) if len(s) else 0.0
    if non_null_ratio < 0.05:
        return None
    if kind == "numeric":
        num = pd.to_numeric(s, errors="coerce").dropna()
        if num.nunique(dropna=True) <= 1:
            return None
        return (70, col)
    if kind == "date":
        dt = pd.to_datetime(s, errors="coerce").dropna()
        if dt.nunique(dropna=True) <= 1:
            return None
        return (68, col)
    clean = s.astype("string").fillna("").str.strip().replace("", pd.NA).dropna()
    nunique = int(clean.nunique(dropna=True))
    if nunique <= 1:
        return None
    if _looks_like_blob_or_text_column(df, col) or _looks_like_identifier_or_link(df, col):
        return None
    row_count = max(len(df), 1)
    upper_reasonable = max(50, min(800, int(row_count * 0.8)))
    if nunique > upper_reasonable:
        return None
    if nunique <= 12:
        score = 95
    elif nunique <= 40:
        score = 88
    elif nunique <= 120:
        score = 80
    else:
        score = 72
    return (score, col)


def _col_options(df: pd.DataFrame, col: str, max_vals: Optional[int] = 250) -> List[str]:
    if col not in df.columns:
        return ["ALL"]
    s = df[col]
    vals = s.astype("string").fillna("Unknown").str.strip().replace("", "Unknown").tolist()
    uniq = list(dict.fromkeys(v for v in vals if str(v).strip()))
    uniq = sorted(uniq, key=lambda x: str(x).lower())
    if max_vals is not None:
        uniq = uniq[: int(max_vals)]
    return ["ALL"] + uniq


def _infer_extra_filter_kind(df: pd.DataFrame, col: str) -> str:
    if col not in df.columns:
        return "categorical"
    s = df[col]
    name = str(col).lower()
    try:
        if pd.api.types.is_datetime64_any_dtype(s):
            return "date"
    except Exception:
        pass
    looks_datey = any(tok in name for tok in ["date", "time", "day", "month", "year"])
    if looks_datey and not pd.api.types.is_numeric_dtype(s):
        try:
            as_dt = pd.to_datetime(s, errors="coerce")
            if as_dt.notna().mean() >= 0.75 and as_dt.nunique(dropna=True) > 2:
                return "date"
        except Exception:
            pass
    try:
        num = pd.to_numeric(s, errors="coerce")
        if num.notna().mean() >= 0.9 and num.nunique(dropna=True) > 6:
            return "numeric"
    except Exception:
        pass
    return "categorical"


def _extra_filter_candidates(df: pd.DataFrame) -> List[str]:
    det_cols, del_cols = _get_symptom_col_lists(df)
    excluded = set(_resolved_core_filter_columns(df))
    rating_col = _resolve_column_alias(df, ["rating", "stars", "star_rating", "score"])
    if rating_col:
        excluded.add(rating_col)
    excluded.update(set(det_cols + del_cols))
    excluded.update({str(c) for c in df.columns if str(c).startswith("AI Symptom ") or str(c).startswith("Symptom ")})

    scored: List[Tuple[int, str]] = []
    for raw_col in df.columns:
        col = str(raw_col)
        if col in excluded:
            continue
        score = _extra_filter_column_score(df, col)
        if score is not None:
            scored.append(score)
    scored.sort(key=lambda item: (-item[0], str(item[1]).lower()))
    return [col for _, col in scored]


def _symptom_filter_options(df: pd.DataFrame) -> Tuple[List[str], List[str], List[str], List[str]]:
    det_cols, del_cols = _get_symptom_col_lists(df)
    def _values(cols: Sequence[str]) -> List[str]:
        vals: List[str] = []
        for col in cols:
            if col not in df.columns:
                continue
            s = df[col].astype("string").fillna("").str.strip()
            good = s[(s != "") & (~s.str.upper().isin(SYMPTOM_NON_VALUES)) & (~s.str.startswith("<"))].str.title()
            vals.extend(good.tolist())
        return sorted(list(dict.fromkeys(vals)), key=lambda x: x.lower())
    return _values(det_cols), _values(del_cols), list(det_cols), list(del_cols)


def _collect_active_filter_items(df: pd.DataFrame, *, core_specs: Sequence[Dict[str, Any]], extra_cols: Sequence[str], tf: str, start_date, end_date) -> List[Tuple[str, str]]:
    items: List[Tuple[str, str]] = []
    if tf != "All Time":
        if tf == "Custom Range" and start_date and end_date:
            items.append(("Timeframe", f"{start_date} → {end_date}"))
        else:
            items.append(("Timeframe", tf))

    sr_raw = st.session_state.get("rf_sr", ["All"])
    sr_list = sr_raw if isinstance(sr_raw, list) else [sr_raw]
    sr_sel = [x for x in sr_list if str(x).strip() and str(x).lower() != "all"]
    if sr_sel:
        items.append(("Stars", ", ".join(str(x) for x in sr_sel)))

    for spec in core_specs:
        sel = st.session_state.get(f"rf_{spec['key']}", ["ALL"])
        sel_list = sel if isinstance(sel, list) else [sel]
        sel_clean = [x for x in sel_list if str(x).strip() and str(x).upper() != "ALL"]
        if sel_clean:
            items.append((spec["label"], ", ".join(str(x) for x in sel_clean[:4]) + ("" if len(sel_clean) <= 4 else f" +{len(sel_clean) - 4}")))

    for col in extra_cols:
        if col not in df.columns:
            continue
        kind = _infer_extra_filter_kind(df, col)
        rk = f"rf_{col}_range"
        dk = f"rf_{col}_date_range"
        ck = f"rf_{col}_contains"
        if kind == "numeric":
            num = pd.to_numeric(df[col], errors="coerce").dropna()
            if not num.empty and rk in st.session_state and isinstance(st.session_state.get(rk), (tuple, list)) and len(st.session_state.get(rk)) == 2:
                lo, hi = st.session_state[rk]
                base_lo, base_hi = float(num.min()), float(num.max())
                if round(float(lo), 10) != round(base_lo, 10) or round(float(hi), 10) != round(base_hi, 10):
                    items.append((col, f"{float(lo):g} → {float(hi):g}"))
            continue
        if kind == "date":
            dt = pd.to_datetime(df[col], errors="coerce").dropna()
            if not dt.empty and dk in st.session_state and isinstance(st.session_state.get(dk), (tuple, list)) and len(st.session_state.get(dk)) == 2:
                lo, hi = st.session_state[dk]
                base_lo, base_hi = dt.min().date(), dt.max().date()
                if lo and hi and (lo != base_lo or hi != base_hi):
                    items.append((col, f"{lo} → {hi}"))
            continue
        cv = _safe_text(st.session_state.get(ck))
        if cv:
            items.append((col, f"contains: {cv}"))
            continue
        sel = st.session_state.get(f"rf_{col}", ["ALL"])
        sel_list = sel if isinstance(sel, list) else [sel]
        sel_clean = [x for x in sel_list if str(x).strip() and str(x).upper() != "ALL"]
        if sel_clean:
            items.append((col, ", ".join(str(x) for x in sel_clean[:4]) + ("" if len(sel_clean) <= 4 else f" +{len(sel_clean) - 4}")))

    sel_det = [x for x in (st.session_state.get("rf_sym_detract", ["All"]) or []) if str(x).strip() and str(x).lower() != "all"]
    sel_del = [x for x in (st.session_state.get("rf_sym_delight", ["All"]) or []) if str(x).strip() and str(x).lower() != "all"]
    if sel_det:
        items.append(("Detractors", ", ".join(sel_det[:3]) + ("" if len(sel_det) <= 3 else f" +{len(sel_det) - 3}")))
    if sel_del:
        items.append(("Delighters", ", ".join(sel_del[:3]) + ("" if len(sel_del) <= 3 else f" +{len(sel_del) - 3}")))

    kw = _safe_text(st.session_state.get("rf_kw"))
    if kw:
        items.append(("Keyword", kw))
    return items


def _filter_description_from_items(items: Sequence[Tuple[str, str]]) -> str:
    return "; ".join(f"{k}={v}" for k, v in items) if items else "No active filters"


def _apply_live_review_filters(df_base: pd.DataFrame) -> Dict[str, Any]:
    t0 = time.perf_counter()
    if df_base is None or df_base.empty:
        return {"filtered_df": df_base.copy() if isinstance(df_base, pd.DataFrame) else pd.DataFrame(), "active_items": [], "filter_seconds": 0.0, "description": "No active filters"}
    d0 = df_base
    mask = pd.Series(True, index=d0.index)
    today = date.today()
    tf = st.session_state.get("rf_tf", "All Time")
    start_date = end_date = None
    if tf == "Custom Range":
        rng = st.session_state.get("rf_tf_range", (today - timedelta(days=30), today))
        if isinstance(rng, (tuple, list)) and len(rng) == 2:
            start_date, end_date = rng
    elif tf == "Last Week":
        start_date, end_date = today - timedelta(days=7), today
    elif tf == "Last Month":
        start_date, end_date = today - timedelta(days=30), today
    elif tf == "Last Year":
        start_date, end_date = today - timedelta(days=365), today

    date_col = "submission_date" if "submission_date" in d0.columns else ("submission_time" if "submission_time" in d0.columns else None)
    if start_date and end_date and date_col:
        dt = pd.to_datetime(d0[date_col], errors="coerce")
        end_inclusive = pd.Timestamp(end_date) + pd.Timedelta(days=1) - pd.Timedelta(nanoseconds=1)
        mask &= (dt >= pd.Timestamp(start_date)) & (dt <= end_inclusive)

    sr_raw = st.session_state.get("rf_sr", ["All"])
    sr_list = sr_raw if isinstance(sr_raw, list) else [sr_raw]
    sr_sel = [x for x in sr_list if str(x).strip() and str(x).lower() != "all"]
    if sr_sel and "rating" in d0.columns:
        sr_nums = [int(x) for x in sr_sel if str(x).isdigit()]
        if sr_nums:
            mask &= pd.to_numeric(d0["rating"], errors="coerce").isin(sr_nums)

    core_specs = _core_filter_specs_for_df(d0)
    for spec in core_specs:
        sel = st.session_state.get(f"rf_{spec['key']}", ["ALL"])
        sel_list = sel if isinstance(sel, list) else [sel]
        sel_clean = [x for x in sel_list if str(x).strip() and str(x).upper() != "ALL"]
        if sel_clean:
            mask &= _series_matches_any(_filter_series_for_spec(d0, spec), [str(x) for x in sel_clean])

    extra_cols = [c for c in (st.session_state.get("rf_extra_filter_cols", []) or []) if c in d0.columns]
    for col in extra_cols:
        kind = _infer_extra_filter_kind(d0, col)
        s = d0[col]
        if kind == "numeric":
            rk = f"rf_{col}_range"
            if rk in st.session_state and isinstance(st.session_state.get(rk), (tuple, list)) and len(st.session_state.get(rk)) == 2:
                lo, hi = st.session_state[rk]
                mask &= pd.to_numeric(s, errors="coerce").between(float(lo), float(hi), inclusive="both")
        elif kind == "date":
            dk = f"rf_{col}_date_range"
            if dk in st.session_state and isinstance(st.session_state.get(dk), (tuple, list)) and len(st.session_state.get(dk)) == 2:
                lo, hi = st.session_state[dk]
                if lo and hi:
                    dt = pd.to_datetime(s, errors="coerce")
                    hi_end = pd.Timestamp(hi) + pd.Timedelta(days=1) - pd.Timedelta(nanoseconds=1)
                    mask &= (dt >= pd.Timestamp(lo)) & (dt <= hi_end)
        else:
            ck = f"rf_{col}_contains"
            cv = _safe_text(st.session_state.get(ck)).strip()
            if cv:
                mask &= s.astype("string").fillna("").str.contains(cv, case=False, na=False, regex=False)
            else:
                sel = st.session_state.get(f"rf_{col}", ["ALL"])
                sel_list = sel if isinstance(sel, list) else [sel]
                sel_clean = [x for x in sel_list if str(x).strip() and str(x).upper() != "ALL"]
                if sel_clean:
                    ss = s.astype("string").fillna("")
                    sample = ss.head(200).astype(str)
                    if bool(sample.str.contains(r"\s\|\s", regex=True).any()):
                        toks = [str(x).strip() for x in sel_clean if str(x).strip()]
                        pattern = r"(^|\s*\|\s*)(" + "|".join(re.escape(t) for t in toks) + r")(\s*\|\s*|$)"
                        mask &= ss.str.contains(pattern, case=False, regex=True, na=False)
                    else:
                        mask &= ss.isin([str(x) for x in sel_clean])

    _, _, det_cols, del_cols = _symptom_filter_options(d0)
    sel_det = [str(x).strip().title() for x in (st.session_state.get("rf_sym_detract", ["All"]) or []) if str(x).strip() and str(x).lower() != "all"]
    sel_del = [str(x).strip().title() for x in (st.session_state.get("rf_sym_delight", ["All"]) or []) if str(x).strip() and str(x).lower() != "all"]
    if sel_det and det_cols:
        det_block = d0[det_cols].astype("string").fillna("").apply(lambda col: col.str.strip().str.title())
        mask &= det_block.isin(sel_det).any(axis=1)
    if sel_del and del_cols:
        del_block = d0[del_cols].astype("string").fillna("").apply(lambda col: col.str.strip().str.title())
        mask &= del_block.isin(sel_del).any(axis=1)

    kw = _safe_text(st.session_state.get("rf_kw")).strip()
    search_col = "title_and_text" if "title_and_text" in d0.columns else ("review_text" if "review_text" in d0.columns else None)
    if kw and search_col:
        mask &= d0[search_col].astype("string").fillna("").str.contains(kw, case=False, na=False, regex=False)

    filtered = d0.loc[mask].copy()
    active_items = _collect_active_filter_items(d0, core_specs=core_specs, extra_cols=extra_cols, tf=tf, start_date=start_date, end_date=end_date)
    return {"filtered_df": filtered, "active_items": active_items, "filter_seconds": time.perf_counter() - t0, "description": _filter_description_from_items(active_items)}


def _render_active_filter_summary(filter_state: Dict[str, Any], overall_df: pd.DataFrame):
    active_items = filter_state.get("active_items", [])
    pills = []
    for k, v in active_items[:12]:
        pills.append(f"<div class='pill'><span class='muted'>{_esc(k)}:</span> {_esc(v)}</div>")
    st.markdown(
        f"""
<div class="soft-panel">
  <div><b>Active filters</b> • Showing <b>{len(filter_state.get('filtered_df', [])):,}</b> of <b>{len(overall_df):,}</b> reviews
  <span class="small-muted"> (filter time: {float(filter_state.get('filter_seconds', 0.0)):.3f}s)</span>
  </div>
  <div class="pill-row">{''.join(pills) if pills else '<span class="small-muted">None (All data)</span>'}</div>
</div>
""",
        unsafe_allow_html=True,
    )

def _render_workspace_nav() -> str:
    current = st.session_state.get("workspace_active_tab", TAB_DASHBOARD)
    nav_items = WORKSPACE_TABS
    cols = st.columns(len(nav_items))
    for i, (col, label) in enumerate(zip(cols, nav_items)):
        kwargs = {"use_container_width": True, "key": f"wnav_{i}_{_slugify(label, fallback='tab')}"}
        if current == label:
            kwargs["type"] = "primary"
        display_label = label.split("  ", 1)[1].strip() if "  " in label else re.sub(r"^[^\w]+", "", str(label)).strip()
        if col.button(display_label or str(label).strip(), **kwargs):
            current = label
            st.session_state["workspace_active_tab"] = label
    st.markdown("<div style='height:.25rem'></div>", unsafe_allow_html=True)
    return current



def _summary_attr(summary, field, default=""):
    if summary is None:
        return default
    try:
        value = getattr(summary, field)
        if value not in (None, ""):
            return value
    except Exception:
        pass
    if isinstance(summary, dict):
        value = summary.get(field)
        if value not in (None, ""):
            return value
    if isinstance(summary, str) and field == "product_id":
        value = _safe_text(summary)
        if value:
            return value
    return default


def _safe_summary_product_label(summary, default="Review workspace"):
    value = _safe_text(_summary_attr(summary, "product_id", ""))
    return value or default


def _safe_summary_product_slug(summary, df=None, default="REVIEWS"):
    candidates = [_safe_text(_summary_attr(summary, "product_id", ""))]
    if isinstance(df, pd.DataFrame) and not df.empty:
        for col in ["base_sku", "product_id", "product_or_sku", "original_product_name"]:
            if col in df.columns:
                try:
                    candidates.append(_first_non_empty(df[col].fillna("").astype(str)))
                except Exception:
                    continue
    for raw in candidates:
        raw = _safe_text(raw)
        if not raw:
            continue
        clean = re.sub(r"[^A-Za-z0-9._-]+", "_", raw).strip("._-")
        if clean:
            return clean[:120]
    return default


def _safe_summary_product_url(summary, default=""):
    return _safe_text(_summary_attr(summary, "product_url", default), default)


def _safe_summary_reviews_downloaded(summary, df=None):
    value = _summary_attr(summary, "reviews_downloaded", None)
    if value not in (None, ""):
        try:
            return int(value)
        except Exception:
            pass
    return int(len(df)) if isinstance(df, pd.DataFrame) else 0


def _product_name(summary, df):
    if not df.empty and "original_product_name" in df.columns:
        names = [x for x in dict.fromkeys(df["original_product_name"].fillna("").astype(str).str.strip().tolist()) if x]
        if len(names) == 1:
            return names[0]
        if len(names) > 1:
            return f"Combined review workspace ({len(names)} products)"
    if not df.empty and "product_or_sku" in df.columns:
        prods = [x for x in dict.fromkeys(df["product_or_sku"].fillna("").astype(str).str.strip().tolist()) if x]
        if len(prods) > 1 and str(_safe_summary_product_label(summary, default="")).startswith("MULTI_URL_WORKSPACE"):
            return f"Combined review workspace ({len(prods)} products)"
    return _safe_summary_product_label(summary)

# ═══════════════════════════════════════════════════════════════════════════════
#  AI ANALYST
# ═══════════════════════════════════════════════════════════════════════════════
GENERAL_INSTRUCTIONS = textwrap.dedent("""
    You are SharkNinja Review Analyst — an internal voice-of-customer AI assistant.
    ROLE: Synthesise consumer review data into sharp, actionable insights.
    Prioritise evidence from the supplied dataset over generic assumptions.
    ANSWER FORMAT
    • Use compact markdown. Prefer short bold section labels instead of large headings.
    • Lead with the most important insight — do not bury the lede.
    • Cite review IDs inline: (review_ids: 12345, 67890).
    • For every quantitative claim state the count or percentage from the data.
    • Mark inferences: [INFERRED].
    • Follow the caller-specified response-length target. If no target is given, prefer a detailed answer over a terse one.
    • End every response with a "**Next Steps**" section: 2–3 concrete actions.
    SYMPTOM-AWARE ANALYSIS
    • When symptom_tags data is present, use it to ground your analysis in structured themes rather than guessing from raw text.
    • Reference specific detractor/delighter labels and their frequency when discussing patterns.
    • If symptom data shows a dominant theme (>30%), flag it as a primary driver.
    • If symptom data shows singleton themes (1 occurrence), treat them as anecdotal, not systematic.
    GUARDRAILS
    • Do not invent review IDs, quotes, counts, or trends not in the evidence.
    • If the data is insufficient, say so explicitly.
    • Never hallucinate product specs — only cite what reviews mention.
    • When comparing time periods, state the sample sizes to help the reader gauge significance.
""").strip()


def _persona_instructions(name):
    if not name:
        return GENERAL_INSTRUCTIONS
    return PERSONAS[name]["instructions"]


def _select_relevant(df, question, max_reviews=22):
    if df.empty:
        return df.copy()
    w = df.copy()
    w["blob"] = w["title_and_text"].fillna("").astype(str).map(_norm_text)
    qt = set(_tokenize(question))

    # TF-IDF-inspired scoring: penalize common terms, boost rare matches
    term_doc_freq = {}
    all_blobs = w["blob"].tolist()
    for blob in all_blobs:
        seen_terms = set()
        for tk in qt:
            if tk in blob and tk not in seen_terms:
                term_doc_freq[tk] = term_doc_freq.get(tk, 0) + 1
                seen_terms.add(tk)
    n_docs = max(len(all_blobs), 1)

    def score(row):
        s = 0.0
        t = row["blob"]
        for tk in qt:
            if tk in t:
                # IDF boost: rare terms score higher
                df = term_doc_freq.get(tk, 1)
                idf = max(0.5, 3.0 - (df / n_docs) * 2.5)
                s += idf * (1 + min(t.count(tk), 3))
        r = row.get("rating")
        # Boost low-star reviews for negative queries
        neg_terms = {"defect","broken","issue","problem","bad","fail","broke","terrible","awful","worst","disappointing","leaked","cracked","stopped"}
        pos_terms = {"love","great","perfect","best","amazing","excellent","recommend","easy","quick","comfortable"}
        if qt & neg_terms and pd.notna(r):
            s += max(0, 6 - float(r))
        elif qt & pos_terms and pd.notna(r):
            s += max(0, float(r) - 2)
        # Organic review bonus
        if not _safe_bool(row.get("incentivized_review"), False):
            s += 0.5
        # Length bonus (longer reviews = more evidence)
        if pd.notna(row.get("review_length_words")):
            s += min(float(row.get("review_length_words", 0)) / 60, 2.5)
        # Symptom tag bonus — reviews with symptoms are richer context
        symptom_cols = [c for c in row.index if c.startswith("AI Symptom")]
        for sc in symptom_cols[:6]:
            if _safe_text(row.get(sc)) and _safe_text(row.get(sc)).upper() not in NON_VALUES:
                s += 0.3
                break
        return s

    w["_sc"] = w.apply(score, axis=1)
    ranked = w.sort_values(["_sc", "submission_time"], ascending=[False, False], na_position="last")
    combined = pd.concat([
        ranked.head(max_reviews),
        df[df["rating"].isin([1, 2])].head(max_reviews // 3 or 1),
        df[df["rating"].isin([4, 5])].head(max_reviews // 3 or 1),
    ], ignore_index=True).drop_duplicates(subset=["review_id"])
    return combined.head(max_reviews).drop(columns=["blob", "_sc"], errors="ignore")


@st.cache_data(show_spinner=False, ttl=120)
def _snippet_rows(df, *, max_reviews=22):
    rows = []
    for _, row in df.head(max_reviews).iterrows():
        rows.append(dict(
            review_id=_safe_text(row.get("review_id")),
            rating=_safe_int(row.get("rating"), 0) if pd.notna(row.get("rating")) else None,
            incentivized_review=_safe_bool(row.get("incentivized_review"), False),
            content_locale=_safe_text(row.get("content_locale")),
            submission_date=_safe_text(row.get("submission_date")),
            title=_trunc(row.get("title", ""), 120),
            snippet=_trunc(row.get("review_text", ""), 600),
        ))
    return rows


def _build_ai_context(*, overall_df, filtered_df, summary, filter_description, question):
    om = _get_metrics(overall_df)
    fm = _get_metrics(filtered_df)
    try:
        rd = _rating_dist(filtered_df).to_dict(orient="records")
    except Exception:
        rd = []
    try:
        md = _monthly_trend(filtered_df).tail(12).to_dict(orient="records")
    except Exception:
        md = []
    rel = _select_relevant(filtered_df, question, max_reviews=22)
    rec = filtered_df.sort_values(["submission_time", "review_id"], ascending=[False, False], na_position="last").head(10)
    low = filtered_df[filtered_df["rating"].isin([1, 2])].head(8)
    hi = filtered_df[filtered_df["rating"].isin([4, 5])].head(8)
    ev = pd.concat([rel, rec, low, hi], ignore_index=True).drop_duplicates(subset=["review_id"]).head(32)
    symptom_context = {}
    try:
        det_cols = [c for c in filtered_df.columns if c.startswith("AI Symptom Det")]
        del_cols = [c for c in filtered_df.columns if c.startswith("AI Symptom Del")]
        if det_cols or del_cols:
            from collections import Counter
            det_counts = Counter()
            del_counts = Counter()
            for _, row in filtered_df.iterrows():
                for c in det_cols:
                    v = _safe_text(row.get(c))
                    if v and v.upper() not in NON_VALUES:
                        det_counts[v] += 1
                for c in del_cols:
                    v = _safe_text(row.get(c))
                    if v and v.upper() not in NON_VALUES:
                        del_counts[v] += 1
            if det_counts:
                symptom_context["top_detractors"] = [{"label": k, "count": v, "pct": round(v / max(len(filtered_df), 1) * 100, 1)} for k, v in det_counts.most_common(15)]
            if del_counts:
                symptom_context["top_delighters"] = [{"label": k, "count": v, "pct": round(v / max(len(filtered_df), 1) * 100, 1)} for k, v in del_counts.most_common(15)]
    except Exception:
        pass

    active_tab = st.session_state.get("workspace_active_tab") or TAB_DASHBOARD
    knowledge = _normalize_product_knowledge(st.session_state.get("sym_product_knowledge") or {})
    processed = list(st.session_state.get("sym_processed_rows") or [])
    sym_snapshot = {
        "status": "Not yet symptomized — run the Symptomizer first for richer analysis.",
        "processed_reviews": 0,
        "top_detractors": [],
        "top_delighters": [],
    }
    if processed:
        det_counts = Counter()
        del_counts = Counter()
        zero_tag_reviews = 0
        for rec in processed:
            dets = list(rec.get("wrote_dets") or [])
            dels = list(rec.get("wrote_dels") or [])
            if not dets and not dels:
                zero_tag_reviews += 1
            for label in dets:
                det_counts[label] += 1
            for label in dels:
                del_counts[label] += 1
        n_proc = len(processed)
        sym_snapshot = {
            "status": "Symptomizer results available.",
            "processed_reviews": n_proc,
            "zero_tag_reviews": zero_tag_reviews,
            "top_detractors": [{"label": k, "count": v, "pct": round(v / max(n_proc, 1) * 100, 1)} for k, v in det_counts.most_common(10)],
            "top_delighters": [{"label": k, "count": v, "pct": round(v / max(n_proc, 1) * 100, 1)} for k, v in del_counts.most_common(10)],
        }

    payload = dict(
        workspace=dict(
            workspace_name=_safe_text(st.session_state.get("workspace_name") or ""),
            source_type=_safe_text((st.session_state.get("analysis_dataset") or {}).get("source_type") if isinstance(st.session_state.get("analysis_dataset"), dict) else ""),
            source_label=_safe_text((st.session_state.get("analysis_dataset") or {}).get("source_label") if isinstance(st.session_state.get("analysis_dataset"), dict) else ""),
        ),
        product=dict(
            product_id=_safe_summary_product_label(summary),
            product_url=_safe_summary_product_url(summary),
            product_name=_product_name(summary, overall_df),
        ),
        analysis_scope=dict(
            filter_description=filter_description,
            overall_review_count=len(overall_df),
            filtered_review_count=len(filtered_df),
            current_tab=str(active_tab),
            question=question,
        ),
        active_taxonomy=dict(
            detractors=_normalize_tag_list(st.session_state.get("sym_detractors") or [])[:25],
            delighters=_normalize_tag_list(st.session_state.get("sym_delighters") or [])[:25],
            category=_safe_text(st.session_state.get("sym_taxonomy_category") or "general"),
        ),
        metric_snapshot=dict(
            overall=om,
            filtered=fm,
            rating_distribution_filtered=rd,
            monthly_trend_filtered=md,
        ),
        symptom_tags=symptom_context if symptom_context else "No current tag columns in the filtered view.",
        symptomizer_run_summary=sym_snapshot,
        product_knowledge=knowledge,
        review_text_evidence=_snippet_rows(ev, max_reviews=32),
    )
    full_json = json.dumps(payload, ensure_ascii=False, indent=2, default=str)
    tok = _estimate_tokens(full_json)
    max_ev = 22
    while tok > AI_CONTEXT_TOKEN_BUDGET and max_ev >= 5:
        max_ev -= 4
        payload["review_text_evidence"] = _snippet_rows(ev, max_reviews=max_ev)
        full_json = json.dumps(payload, ensure_ascii=False, indent=2, default=str)
        tok = _estimate_tokens(full_json)
    return full_json


def _call_analyst(*, question, overall_df, filtered_df, summary, filter_description, chat_history, persona_name=None, target_words=1200, include_references=False, freeform_mode=False):
    client = _get_client()
    if client is None:
        raise ReviewDownloaderError("No OpenAI API key configured.")
    target_words = _coerce_ai_target_words(target_words)
    floor_words = max(220, int(round(target_words * 0.8)))
    ceiling_words = min(2600, int(round(target_words * 1.15)))
    max_tok = _ai_target_token_budget(target_words)
    base_instructions = _persona_instructions(persona_name)
    if freeform_mode:
        base_instructions = base_instructions.replace(
            '• End every response with a "**Next Steps**" section: 2–3 concrete actions.',
            '• In freeform mode, answer in the most natural structure for the question. Use a "**Next Steps**" section only when it genuinely helps.',
        )
        length_note = (
            "FREEFORM MODE: answer the user's actual question directly. Use the requested detail level as a ceiling, not a quota. "
            f"For simple questions, answer simply. For broader analytical questions, you may expand up to about {target_words:,} words when useful. "
            "Use compact markdown only when it helps. Do not force a template or fixed section structure."
        )
    else:
        length_note = (
            "RESPONSE LENGTH OVERRIDE: ignore any shorter length caps that may appear earlier in these instructions. "
            f"Aim for about {target_words:,} words, with a practical range of roughly {floor_words:,} to {ceiling_words:,} words. "
            "Use detailed evidence, concrete examples, and a full Next Steps section. Do not compress the answer into a short summary unless the user explicitly asks for that."
        )
    reference_note = (
        "REFERENCE MODE: include inline evidence references using the exact format (review_ids: 12345, 67890) for material claims."
        if include_references else
        "REFERENCE MODE: do NOT include inline citations, do NOT emit (review_ids: ...), and do NOT add reference callouts in the final answer."
    )
    quality_note = (
        "QUALITY BAR: avoid generic consulting language. Tie every major point to a named symptom, repeated consumer phrase, rating pattern, or review cohort from the supplied context. "
        "When you recommend an action, name the likely owner (product, CX, support, content, ops, merchandising, or retention), explain the expected benefit, and state what evidence supports it. "
        "If the data is thin or mixed, say that clearly instead of overcommitting."
    )
    instructions = base_instructions + "\n\n" + length_note + "\n\n" + reference_note + "\n\n" + quality_note
    if freeform_mode:
        instructions += "\n\nFREEFORM ANSWERING: do not force a persona-specific frame. Answer the most natural interpretation of the user's request using the supplied review context. Use structure only when it improves clarity."
    ai_ctx = _build_ai_context(overall_df=overall_df, filtered_df=filtered_df, summary=summary, filter_description=filter_description, question=question)
    msgs = [{"role": m["role"], "content": m["content"]} for m in list(chat_history)[-8:]]
    msgs.append({"role": "user", "content": f"User request:\n{question}\n\nReview dataset context (JSON):\n{ai_ctx}"})
    result = _chat_complete_with_fallback_models(
        client,
        model=_shared_model(),
        structured=False,
        messages=[{"role": "system", "content": instructions}, *msgs],
        temperature=0.0,
        max_tokens=max_tok,
        reasoning_effort=_shared_reasoning(),
    )
    if not result:
        raise ReviewDownloaderError("OpenAI returned empty answer.")
    if not include_references:
        result = _strip_review_citations(result)
    return result

# ═══════════════════════════════════════════════════════════════════════════════
#  REVIEW PROMPT TAGGING
# ═══════════════════════════════════════════════════════════════════════════════
def _default_prompt_df():
    return pd.DataFrame(REVIEW_PROMPT_STARTER_ROWS)


def _normalize_prompt_defs(prompt_df, existing_columns):
    if prompt_df is None or prompt_df.empty:
        return []
    normalized = []
    seen = set()
    existing_set = {str(c) for c in existing_columns}
    for _, row in prompt_df.fillna("").iterrows():
        rp = _safe_text(row.get("prompt"))
        rl = _safe_text(row.get("labels"))
        rc = _safe_text(row.get("column_name"))
        if not rp and not rl and not rc:
            continue
        if not rp:
            raise ReviewDownloaderError("Each prompt row needs a prompt.")
        if not rl:
            raise ReviewDownloaderError("Each prompt row needs labels.")
        labels = [l.strip() for l in rl.split(",") if l.strip()]
        deduped = list(dict.fromkeys(labels))
        if "Not Mentioned" not in deduped and len(deduped) <= 7:
            deduped.append("Not Mentioned")
        if len(deduped) < 2:
            raise ReviewDownloaderError("Each prompt needs at least two labels.")
        col = _slugify(rc or rp)
        if col in existing_set and col not in {"review_id"}:
            col = f"{col}_ai"
        base = col
        suffix = 2
        while col in seen:
            col = f"{base}_{suffix}"
            suffix += 1
        seen.add(col)
        normalized.append(dict(column_name=col, display_name=col.replace("_", " ").title(), prompt=rp, labels=deduped, labels_csv=", ".join(deduped)))
    return normalized


def _build_tagging_schema(prompt_defs):
    item_props = {"review_id": {"type": "string"}}
    required = ["review_id"]
    for p in prompt_defs:
        item_props[p["column_name"]] = {"type": "string", "enum": list(p["labels"])}
        required.append(p["column_name"])
    return {
        "type": "object",
        "additionalProperties": False,
        "properties": {
            "results": {
                "type": "array",
                "items": {
                    "type": "object",
                    "additionalProperties": False,
                    "properties": item_props,
                    "required": required,
                },
            },
        },
        "required": ["results"],
    }


def _classify_chunk(*, client, chunk_df, prompt_defs):
    pc = max(len(prompt_defs), 1)
    max_out = int(max(1600, min(9000, 450 + len(chunk_df) * (18 + 10 * pc))))
    reviews_payload = [
        dict(
            review_id=_safe_text(row.get("review_id")),
            rating=_safe_int(row.get("rating"), 0) if pd.notna(row.get("rating")) else None,
            title=_trunc(row.get("title", ""), 200),
            review_text=_trunc(row.get("review_text", ""), 900),
            incentivized_review=_safe_bool(row.get("incentivized_review"), False),
        )
        for _, row in chunk_df.iterrows()
    ]
    prompt_payload = [dict(column_name=p["column_name"], prompt=p["prompt"], labels=p["labels"]) for p in prompt_defs]
    instructions = "You are a deterministic review-tagging engine. For each review and prompt, return exactly one allowed label. If not mentioned, use Not Mentioned."
    user_content = json.dumps({"prompt_definitions": prompt_payload, "reviews": reviews_payload})
    msgs = [{"role": "system", "content": instructions}, {"role": "user", "content": user_content}]
    structured_rf = {"type": "json_schema", "json_schema": {"name": "review_prompt_tagging", "schema": _build_tagging_schema(prompt_defs), "strict": True}}
    result_text = ""
    try:
        result_text = _chat_complete_with_fallback_models(
            client,
            model=_shared_model(),
            structured=True,
            messages=msgs,
            temperature=0.0,
            response_format=structured_rf,
            max_tokens=max_out,
            reasoning_effort=_shared_reasoning(),
        )
    except Exception as exc:
        col_hints = ", ".join(f'{p["column_name"]}: one of [{", ".join(p["labels"])}]' for p in prompt_defs)
        fallback_instructions = (
            "You are a deterministic review-tagging engine. Return ONLY a JSON object with key 'results' containing an array. "
            "Each element must have: review_id (string), " + col_hints + ". Include every review_id from the input. Use 'Not Mentioned' if not applicable."
        )
        result_text = _chat_complete_with_fallback_models(
            client,
            model=_shared_model(),
            structured=True,
            messages=[{"role": "system", "content": fallback_instructions}, {"role": "user", "content": user_content}],
            temperature=0.0,
            response_format={"type": "json_object"},
            max_tokens=max_out,
            reasoning_effort=_shared_reasoning(),
        )
    if not result_text:
        raise ReviewDownloaderError("OpenAI returned an empty response. Check your API key and model selection.")
    data = _safe_json_load(result_text)
    output_rows = data.get("results") or []
    out_df = pd.DataFrame(output_rows)
    if out_df.empty:
        raise ReviewDownloaderError(f"OpenAI returned no tagged rows. Raw response snippet: {result_text[:300]}")
    out_df["review_id"] = out_df["review_id"].astype(str)
    expected = set(chunk_df["review_id"].astype(str))
    returned = set(out_df["review_id"].astype(str))
    if expected != returned:
        miss = sorted(expected - returned)
        if miss:
            import warnings
            warnings.warn(f"Batch partial: missing {miss[:5]}")
        out_df = out_df[out_df["review_id"].isin(expected)]
    return out_df


def _run_review_prompt_tagging(*, client, source_df, prompt_defs, chunk_size):
    if source_df.empty:
        raise ReviewDownloaderError("No reviews in scope.")
    chunks = list(range(0, len(source_df), chunk_size))
    prog = st.progress(0.0, text="Preparing…")
    status = st.empty()
    outputs = []
    errors = []
    for i, start in enumerate(chunks, 1):
        chunk_df = source_df.iloc[start:start + chunk_size].copy()
        status.info(f"Classifying {start + 1}–{min(start + chunk_size, len(source_df))} of {len(source_df)}")
        try:
            outputs.append(_classify_chunk(client=client, chunk_df=chunk_df, prompt_defs=prompt_defs))
        except Exception as exc:
            errors.append(f"Batch {i}: {exc}")
            status.warning(f"Batch {i} failed — {exc}")
        prog.progress(i / len(chunks))
        gc.collect()
    if not outputs:
        err_detail = "; ".join(errors[:3]) if errors else "unknown error"
        raise ReviewDownloaderError(f"All batches failed. First error: {err_detail}")
    if errors:
        status.warning(f"{len(errors)} of {len(chunks)} batch(es) failed — partial results saved.")
    else:
        status.success(f"Finished tagging {len(source_df):,} reviews.")
    return pd.concat(outputs, ignore_index=True).drop_duplicates(subset=["review_id"], keep="last")


def _merge_prompt_results(overall_df, prompt_results_df, prompt_defs):
    updated = overall_df.copy()
    rids = updated["review_id"].astype(str)
    lk = prompt_results_df.set_index("review_id")
    for p in prompt_defs:
        col = p["column_name"]
        if col not in updated.columns:
            updated[col] = pd.NA
        mapping = lk[col].to_dict()
        nv = rids.map(mapping)
        updated[col] = nv.where(nv.notna(), updated[col])
    return updated


def _summarize_prompt_results(prompt_results_df, prompt_defs, source_df=None):
    merged = prompt_results_df.copy()
    merged["review_id"] = merged["review_id"].astype(str)
    if source_df is not None and not source_df.empty and "review_id" in source_df.columns:
        lk = source_df[[c for c in ["review_id", "rating"] if c in source_df.columns]].copy()
        lk["review_id"] = lk["review_id"].astype(str)
        merged = merged.merge(lk, on="review_id", how="left")
    rows = []
    total = max(len(prompt_results_df), 1)
    for p in prompt_defs:
        col = p["column_name"]
        for label in p["labels"]:
            sub = merged[merged[col] == label]
            rows.append(dict(column_name=col, display_name=p["display_name"], label=str(label), review_count=len(sub), share=_safe_pct(len(sub), total), avg_rating=_safe_mean(sub["rating"]) if "rating" in sub.columns else None))
    return pd.DataFrame(rows)

# ═══════════════════════════════════════════════════════════════════════════════
#  EXPORT
# ═══════════════════════════════════════════════════════════════════════════════
def _autosize_cell_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, (pd.Series, pd.DataFrame, pd.Index, list, tuple, set, dict)):
        try:
            return json.dumps(value, ensure_ascii=False, default=str)
        except Exception:
            return str(value)
    try:
        missing = pd.isna(value)
    except Exception:
        missing = False
    if isinstance(missing, bool) and missing:
        return ""
    return str(value)


def _autosize_ws(ws, df):
    ws.freeze_panes = "A2"
    if df is None or not isinstance(df, pd.DataFrame) or df.empty and len(df.columns) == 0:
        return
    for idx, col in enumerate(list(df.columns), 1):
        try:
            series = df.iloc[:, idx - 1]
        except Exception:
            series = pd.Series(dtype="object")
        if isinstance(series, pd.DataFrame):
            values = []
            for nested_idx in range(series.shape[1]):
                values.extend(series.iloc[:, nested_idx].head(250).tolist())
        else:
            values = series.head(250).tolist()
        max_len = max([len(str(col))] + [len(_autosize_cell_text(v)) for v in values] + [0])
        ws.column_dimensions[get_column_letter(idx)].width = min(max_len + 2, 48)


def _filter_criteria_df(active_items=None, *, filter_description="", export_scope_label="", total_loaded_reviews=None, export_review_count=None):
    rows = [
        {"field": "Export scope", "value": export_scope_label or "Current view"},
        {"field": "Export review count", "value": export_review_count if export_review_count is not None else ""},
        {"field": "Total loaded reviews", "value": total_loaded_reviews if total_loaded_reviews is not None else ""},
        {"field": "Filter description", "value": filter_description or "No active filters"},
    ]
    if active_items:
        for label, value in active_items:
            rows.append({"field": label, "value": value})
    else:
        rows.append({"field": "Filters", "value": "No active filters"})
    return pd.DataFrame(rows)


def _build_master_excel(summary, reviews_df, *, prompt_defs=None, prompt_summary_df=None, prompt_scope="", active_items=None, filter_description="", export_scope_label="Current view", total_loaded_reviews=None):
    metrics = _get_metrics(reviews_df)
    summary_pid = _safe_summary_product_slug(summary, reviews_df, default="reviews")
    summary_label = _safe_summary_product_label(summary)
    summary_url = _safe_summary_product_url(summary)
    summary_reviews_downloaded = _safe_summary_reviews_downloaded(summary, reviews_df)
    try:
        rd = _rating_dist(reviews_df)
        md = _monthly_trend(reviews_df)
    except Exception:
        rd = pd.DataFrame()
        md = pd.DataFrame()
    summary_df = pd.DataFrame([dict(
        product_name=_product_name(summary, reviews_df),
        product_id=summary_label,
        product_url=summary_url,
        reviews_downloaded=summary_reviews_downloaded,
        export_review_count=len(reviews_df),
        total_loaded_reviews=total_loaded_reviews if total_loaded_reviews is not None else len(reviews_df),
        export_scope=export_scope_label,
        filter_description=filter_description or "No active filters",
        avg_rating=metrics.get("avg_rating"),
        avg_rating_non_incentivized=metrics.get("avg_rating_non_incentivized"),
        pct_low_star=metrics.get("pct_low_star"),
        pct_incentivized=metrics.get("pct_incentivized"),
        generated_utc=pd.Timestamp.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC"),
    )])
    priority_cols = ["review_id", "product_id", "rating", "incentivized_review", "is_recommended", "submission_time", "content_locale", "title", "review_text"]
    pc = [p["column_name"] for p in (prompt_defs or []) if p["column_name"] in reviews_df.columns]
    ordered = [c for c in priority_cols + pc if c in reviews_df.columns]
    remaining = [c for c in reviews_df.columns if c not in ordered]
    exp_reviews = reviews_df[ordered + remaining]
    filter_df = _filter_criteria_df(active_items, filter_description=filter_description, export_scope_label=export_scope_label, total_loaded_reviews=total_loaded_reviews, export_review_count=len(reviews_df))
    out = io.BytesIO()
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        # ── Executive Summary sheet ──────────────────────────────────────
        try:
            exec_data = [
                {"Metric": "Product", "Value": summary_label},
                {"Metric": "Total Reviews", "Value": len(reviews_df)},
                {"Metric": "Avg Rating", "Value": round(pd.to_numeric(reviews_df.get("rating"), errors="coerce").mean(), 2) if not reviews_df.empty else "N/A"},
                {"Metric": "Organic Count", "Value": int((~reviews_df.get("incentivized_review", pd.Series(False)).fillna(False)).sum()) if not reviews_df.empty else 0},
                {"Metric": "Export Scope", "Value": export_scope_label},
                {"Metric": "Filters", "Value": filter_description or "None"},
            ]
            # Add rating breakdown
            for star in [5, 4, 3, 2, 1]:
                count = int((pd.to_numeric(reviews_df.get("rating"), errors="coerce") == star).sum()) if not reviews_df.empty else 0
                pct = round(count / max(len(reviews_df), 1) * 100, 1)
                exec_data.append({"Metric": f"{star}★ Reviews", "Value": f"{count} ({pct}%)"})
            # Add top symptom tags
            det_cols = [c for c in reviews_df.columns if c.startswith("AI Symptom Det")]
            del_cols = [c for c in reviews_df.columns if c.startswith("AI Symptom Del")]
            if det_cols or del_cols:
                from collections import Counter
                exec_data.append({"Metric": "", "Value": ""})
                det_c = Counter()
                for c in det_cols:
                    for v in reviews_df[c].dropna().astype(str):
                        v = v.strip()
                        if v and v.upper() not in {"","NA","N/A","NONE","NULL","NAN","<NA>","NOT MENTIONED"}: det_c[v] += 1
                for label, cnt in det_c.most_common(8):
                    exec_data.append({"Metric": f"Top Detractor: {label}", "Value": f"{cnt} ({round(cnt/max(len(reviews_df),1)*100,1)}%)"})
                del_c = Counter()
                for c in del_cols:
                    for v in reviews_df[c].dropna().astype(str):
                        v = v.strip()
                        if v and v.upper() not in {"","NA","N/A","NONE","NULL","NAN","<NA>","NOT MENTIONED"}: del_c[v] += 1
                for label, cnt in del_c.most_common(8):
                    exec_data.append({"Metric": f"Top Delighter: {label}", "Value": f"{cnt} ({round(cnt/max(len(reviews_df),1)*100,1)}%)"})
            exec_df = pd.DataFrame(exec_data)
            exec_df.to_excel(writer, sheet_name="Executive Summary", index=False)
            _autosize_ws(writer.sheets["Executive Summary"], exec_df)
        except Exception:
            pass
        sheets = {"Summary": summary_df, "FilterCriteria": filter_df, "Reviews": exp_reviews, "RatingDistribution": rd, "ReviewVolume": md}
        if prompt_defs:
            sheets["ReviewPromptDefinitions"] = pd.DataFrame([dict(column_name=p["column_name"], display_name=p["display_name"], prompt=p["prompt"], labels=", ".join(p["labels"]), scope=prompt_scope) for p in prompt_defs])
        if prompt_summary_df is not None and not prompt_summary_df.empty:
            sheets["ReviewPromptSummary"] = prompt_summary_df
        for sname, df_ in sheets.items():
            if df_ is None or df_.empty:
                continue
            df_.to_excel(writer, sheet_name=sname, index=False)
            _autosize_ws(writer.sheets[sname], df_)
    out.seek(0)
    return out.getvalue()


def _get_master_bundle(summary, reviews_df, prompt_artifacts, *, active_items=None, filter_description="", export_scope_label="Current view", total_loaded_reviews=None):
    pd_ = (prompt_artifacts or {}).get("definitions") or []
    summary_pid = _safe_summary_product_slug(summary, reviews_df, default="reviews")
    psd = (prompt_artifacts or {}).get("summary_df")
    ps = (prompt_artifacts or {}).get("scope_label", "")
    key = json.dumps(dict(
        pid=summary_pid,
        n=len(reviews_df),
        cols=sorted(str(c) for c in reviews_df.columns),
        psig=(prompt_artifacts or {}).get("definition_signature"),
        filters=list(active_items or []),
        filter_description=filter_description or "",
        export_scope_label=export_scope_label or "",
        total_loaded_reviews=total_loaded_reviews,
    ), sort_keys=True)
    b = st.session_state.get("master_export_bundle")
    if b and b.get("key") == key:
        return b
    xlsx = _build_master_excel(
        summary,
        reviews_df,
        prompt_defs=pd_,
        prompt_summary_df=psd,
        prompt_scope=ps,
        active_items=active_items,
        filter_description=filter_description,
        export_scope_label=export_scope_label,
        total_loaded_reviews=total_loaded_reviews,
    )
    ts = pd.Timestamp.utcnow().strftime("%Y%m%d_%H%M%S")
    scope_slug = "filtered_view" if active_items else "review_workspace"
    b = dict(key=key, excel_bytes=xlsx, excel_name=f"{summary_pid}_{scope_slug}_{ts}.xlsx")
    st.session_state["master_export_bundle"] = b
    return b


def _safe_get_master_bundle(summary, reviews_df, prompt_artifacts, **kwargs):
    try:
        return _get_master_bundle(summary, reviews_df, prompt_artifacts, **kwargs), None
    except Exception as exc:
        return None, exc

# ═══════════════════════════════════════════════════════════════════════════════
#  SYMPTOMIZER HELPERS
# ═══════════════════════════════════════════════════════════════════════════════
def _get_symptom_whitelists(file_bytes):
    bio = io.BytesIO(file_bytes)
    df_sym = None
    try:
        df_sym = pd.read_excel(bio, sheet_name="Symptoms")
    except Exception:
        df_sym = None
    if df_sym is not None and not df_sym.empty:
        df_sym.columns = [str(c).strip() for c in df_sym.columns]
        lc = {c.lower(): c for c in df_sym.columns}
        alias_col = next((lc[c] for c in ["aliases", "alias"] if c in lc), None)
        label_col = next((lc[c] for c in ["symptom", "label", "name", "item"] if c in lc), None)
        type_col = next((lc[c] for c in ["type", "polarity", "category", "side"] if c in lc), None)
        pos_tags = {"delighter", "delighters", "positive", "pos", "pros"}
        neg_tags = {"detractor", "detractors", "negative", "neg", "cons"}

        def _clean(s):
            vals = s.dropna().astype(str).str.strip()
            out = []
            seen = set()
            for v in vals:
                if v and v not in seen:
                    seen.add(v)
                    out.append(v)
            return out

        delighters, detractors, alias_map = [], [], {}
        if label_col and type_col:
            df_sym[type_col] = df_sym[type_col].astype(str).str.lower().str.strip()
            delighters = _clean(df_sym.loc[df_sym[type_col].isin(pos_tags), label_col])
            detractors = _clean(df_sym.loc[df_sym[type_col].isin(neg_tags), label_col])
            if alias_col:
                for _, row in df_sym.iterrows():
                    lbl = str(row.get(label_col, "")).strip()
                    als = str(row.get(alias_col, "")).strip()
                    if lbl:
                        alias_map[lbl] = [p.strip() for p in als.replace(",", "|").split("|") if p.strip()] if als else []
        else:
            for lck, orig in lc.items():
                if "delight" in lck or "positive" in lck or lck == "pros":
                    delighters.extend(_clean(df_sym[orig]))
                if "detract" in lck or "negative" in lck or lck == "cons":
                    detractors.extend(_clean(df_sym[orig]))
            delighters = list(dict.fromkeys(delighters))
            detractors = list(dict.fromkeys(detractors))
        if delighters or detractors:
            delighters, detractors = _canonical_symptom_catalog(delighters, detractors)
        return delighters, detractors, alias_map
    try:
        review_df, _sheet_name = _read_best_uploaded_excel_sheet(file_bytes)
    except Exception:
        return [], [], {}
    if review_df is None or review_df.empty:
        return [], [], {}
    det_vals, del_vals, _, _ = _symptom_filter_options(review_df)
    if del_vals or det_vals:
        del_vals, det_vals = _canonical_symptom_catalog(list(del_vals), list(det_vals))
        return del_vals, det_vals, {}
    return [], [], {}


def _ensure_ai_cols(df):
    for h in AI_DET_HEADERS + AI_DEL_HEADERS + AI_META_HEADERS:
        if h not in df.columns:
            df[h] = None
    return df


def _detect_sym_cols(df):
    det_cols, del_cols = _symptom_col_lists_from_columns(df.columns)
    return dict(
        manual_detractors=[c for c in det_cols if c.lower() in {f"symptom {i}" for i in range(1, 11)}],
        manual_delighters=[c for c in del_cols if c.lower() in {f"symptom {i}" for i in range(11, 21)}],
        ai_detractors=[c for c in det_cols if c.lower().startswith("ai symptom detractor")],
        ai_delighters=[c for c in del_cols if c.lower().startswith("ai symptom delighter")],
    )


def _filled_mask(df, cols):
    if not cols:
        return pd.Series(False, index=df.index)
    mask = pd.Series(False, index=df.index)
    for c in cols:
        if c not in df.columns:
            continue
        s = df[c].astype("string").fillna("").str.strip()
        mask |= (s != "") & (~s.str.upper().isin(SYMPTOM_NON_VALUES)) & (~s.str.startswith("<"))
    return mask


def _detect_missing(df, colmap):
    out = df.copy()
    det_cols = colmap["manual_detractors"] + colmap["ai_detractors"]
    del_cols = colmap["manual_delighters"] + colmap["ai_delighters"]
    out["Has_Detractors"] = _filled_mask(out, det_cols)
    out["Has_Delighters"] = _filled_mask(out, del_cols)
    out["Needs_Detractors"] = ~out["Has_Detractors"]
    out["Needs_Delighters"] = ~out["Has_Delighters"]
    out["Needs_Symptomization"] = out["Needs_Detractors"] & out["Needs_Delighters"]
    return out


def _match_label(raw, allowed, aliases=None, cutoff=0.76):
    if _HAS_SYMPTOMIZER_V3:
        return _v3_match_label(raw, allowed, aliases=aliases, cutoff=0.72)
    if not raw or not allowed:
        return None
    raw_s = raw.strip()
    exact = {_canon_simple(x): x for x in allowed}
    lbl = exact.get(_canon_simple(raw_s))
    if lbl:
        return lbl
    if aliases:
        for canonical, als in (aliases or {}).items():
            if canonical not in allowed:
                continue
            for a in (als or []):
                if _canon_simple(raw_s) == _canon_simple(a):
                    return canonical
    m = difflib.get_close_matches(raw_s, allowed, n=1, cutoff=cutoff)
    if m:
        return m[0]
    raw_lower = raw_s.lower()
    for label in allowed:
        if raw_lower in label.lower() or label.lower() in raw_lower:
            return label
    return None


def _validate_evidence(evidence_list, review_text, max_ev_chars=120):
    if _HAS_SYMPTOMIZER_V3:
        return _v3_validate_evidence(evidence_list, review_text, max_ev_chars)
    if not evidence_list or not review_text:
        return []
    rv = re.sub(r"\s+", " ", review_text.lower())
    _neg_patt = re.compile(r"\b(not|no|never|don't|doesn't|won't|can't|isn't|wasn't|without|hardly|barely)\s+", re.I)
    out = []
    for e in evidence_list:
        e = str(e).strip()[:max_ev_chars]
        # Minimum quality: 6+ chars, 2+ words (matches v3)
        if len(e) < 6 or len(e.split()) < 2:
            continue
        e_norm = re.sub(r"\s+", " ", e.lower())
        if e_norm not in rv:
            continue
        # Negation check: reject short evidence preceded by negation
        idx = rv.find(e_norm)
        if idx > 0 and len(e_norm.split()) <= 4:
            preceding = rv[max(0, idx - 30):idx]
            if _neg_patt.search(preceding):
                continue
        out.append(e)
    return out[:2]


def _canonical_index_key(value):
    text = str(value or "").strip()
    if not text:
        return ""
    if re.fullmatch(r"-?\d+(?:\.0+)?", text):
        try:
            return str(int(float(text)))
        except Exception:
            return text
    return text


def _build_symptom_baseline_map(processed_rows):
    baseline = {}
    for rec in processed_rows or []:
        idx = _canonical_index_key(rec.get("idx", ""))
        if not idx:
            continue
        baseline[idx] = {
            "detractors": _normalize_tag_list(rec.get("wrote_dets", [])),
            "delighters": _normalize_tag_list(rec.get("wrote_dels", [])),
        }
    return baseline


def _collect_ai_tag_map(df, row_ids=None):
    if df is None or df.empty:
        return {}
    out = {}
    ids = [_canonical_index_key(rid) for rid in row_ids] if row_ids is not None else [_canonical_index_key(idx) for idx in df.index]
    for rid in ids:
        if not rid:
            continue
        try:
            row = df.loc[int(float(rid))]
        except Exception:
            continue
        out[rid] = {
            "detractors": _normalize_tag_list(_collect_row_symptom_tags(row, AI_DET_HEADERS)),
            "delighters": _normalize_tag_list(_collect_row_symptom_tags(row, AI_DEL_HEADERS)),
        }
    return out


def _write_ai_symptom_row(df, idx, *, dets=None, dels=None, safety=None, reliability=None, sessions=None):
    df = _ensure_ai_cols(df)
    if dets is None:
        det_values = _normalize_tag_list(_collect_row_symptom_tags(df.loc[idx], AI_DET_HEADERS))
    else:
        det_values = _normalize_tag_list(dets or [])[:10]
    if dels is None:
        del_values = _normalize_tag_list(_collect_row_symptom_tags(df.loc[idx], AI_DEL_HEADERS))
    else:
        del_values = _normalize_tag_list(dels or [])[:10]

    for header in AI_DET_HEADERS:
        df.loc[idx, header] = None
    for header in AI_DEL_HEADERS:
        df.loc[idx, header] = None
    for j, label in enumerate(det_values, start=1):
        df.loc[idx, f"AI Symptom Detractor {j}"] = label
    for j, label in enumerate(del_values, start=1):
        df.loc[idx, f"AI Symptom Delighter {j}"] = label
    if safety is not None:
        df.loc[idx, "AI Safety"] = safety
    if reliability is not None:
        df.loc[idx, "AI Reliability"] = reliability
    if sessions is not None:
        df.loc[idx, "AI # of Sessions"] = sessions
    return df


def _qa_review_option_label(row):
    rating = _safe_int(row.get("rating"), 0) if pd.notna(row.get("rating")) else 0
    title = _safe_text(row.get("title"), "No title") or "No title"
    review_id = _safe_text(row.get("review_id"))
    suffix = f" · {review_id}" if review_id else ""
    return f"{rating}/5 · {title[:72]}{suffix}"


def _qa_accuracy_metrics(reviews_df):
    baseline = st.session_state.get("sym_qa_baseline_map") or {}
    baseline_total = sum(len(v.get("detractors", []) or []) + len(v.get("delighters", []) or []) for v in baseline.values())
    if not baseline:
        metrics = {
            "baseline_total_tags": 0,
            "added_tags": 0,
            "removed_tags": 0,
            "total_changes": 0,
            "changed_reviews": 0,
            "accuracy_pct": 100.0,
        }
        st.session_state["sym_qa_accuracy"] = metrics
        return metrics
    if not bool(st.session_state.get("sym_qa_user_edited", False)):
        metrics = {
            "baseline_total_tags": baseline_total,
            "added_tags": 0,
            "removed_tags": 0,
            "total_changes": 0,
            "changed_reviews": 0,
            "accuracy_pct": 100.0,
        }
        st.session_state["sym_qa_accuracy"] = metrics
        return metrics
    current = _collect_ai_tag_map(reviews_df, row_ids=list(baseline.keys()))
    metrics = _compute_tag_edit_accuracy(baseline, current)
    st.session_state["sym_qa_accuracy"] = metrics
    return metrics


def _parse_manual_tag_entries(raw_text):
    return _normalize_tag_list([part.strip() for part in re.split(r"[\n,;|]+", str(raw_text or "")) if part.strip()])


def _standardize_symptom_lists(delighters=None, detractors=None):
    try:
        return _canonicalize_taxonomy_catalog(delighters or [], detractors or [])
    except Exception:
        return _normalize_tag_list(delighters or []), _normalize_tag_list(detractors or []), {}


def _alias_map_for_catalog(delighters=None, detractors=None, *, extra_aliases=None, existing_aliases=None):
    labels = set(_normalize_tag_list(list(delighters or []) + list(detractors or [])))
    base = _build_taxonomy_alias_map(delighters or [], detractors or [], extra_aliases=extra_aliases or {})
    filtered_existing = {}
    for key, values in (existing_aliases or {}).items():
        label = re.sub(r"\s+", " ", str(key or "").strip()).title()
        if not label or label not in labels:
            continue
        filtered_existing[label] = [re.sub(r"\s+", " ", str(v or "").strip()).title() for v in (values or []) if str(v or "").strip()]
    merged = _merge_taxonomy_alias_maps(filtered_existing, base)
    return {key: vals for key, vals in merged.items() if key in labels and vals}


def _custom_universal_catalog():
    custom_dels, custom_dets, _ = _standardize_symptom_lists(
        st.session_state.get("sym_custom_universal_delighters") or [],
        st.session_state.get("sym_custom_universal_detractors") or [],
    )
    built_in_dels = set(_UNIVERSAL_NEUTRAL_DELIGHTERS)
    built_in_dets = set(_UNIVERSAL_NEUTRAL_DETRACTORS)
    return [label for label in custom_dels if label not in built_in_dels], [label for label in custom_dets if label not in built_in_dets]


def _universal_neutral_catalog():
    custom_dels, custom_dets = _custom_universal_catalog()
    return _normalize_tag_list(list(_UNIVERSAL_NEUTRAL_DELIGHTERS) + list(custom_dels)), _normalize_tag_list(list(_UNIVERSAL_NEUTRAL_DETRACTORS) + list(custom_dets))


def _all_universal_neutral_labels():
    neutral_dels, neutral_dets = _universal_neutral_catalog()
    return set(neutral_dels + neutral_dets)


def _save_custom_universal_catalog(delighters=None, detractors=None, *, merge=False):
    current_dels, current_dets = _custom_universal_catalog() if merge else ([], [])
    raw_dels = list(current_dels) + list(delighters or [])
    raw_dets = list(current_dets) + list(detractors or [])
    saved_dels, saved_dets, extra_aliases = _standardize_symptom_lists(raw_dels, raw_dets)
    built_in_dels = set(_UNIVERSAL_NEUTRAL_DELIGHTERS)
    built_in_dets = set(_UNIVERSAL_NEUTRAL_DETRACTORS)
    saved_dels = [label for label in saved_dels if label not in built_in_dels]
    saved_dets = [label for label in saved_dets if label not in built_in_dets]
    st.session_state["sym_custom_universal_delighters"] = saved_dels
    st.session_state["sym_custom_universal_detractors"] = saved_dets
    active_dels, active_dets = _canonical_symptom_catalog(st.session_state.get("sym_delighters") or [], st.session_state.get("sym_detractors") or [])
    st.session_state["sym_delighters"] = active_dels
    st.session_state["sym_detractors"] = active_dets
    st.session_state["sym_aliases"] = _alias_map_for_catalog(
        active_dels,
        active_dets,
        extra_aliases=extra_aliases,
        existing_aliases=st.session_state.get("sym_aliases", {}),
    )
    st.session_state["sym_qa_user_edited"] = True
    return saved_dels, saved_dets


def _promote_labels_to_custom_universal(labels, *, side):
    cleaned = _normalize_tag_list(labels or [])
    if not cleaned:
        return [], []
    if str(side).lower().startswith("del"):
        return _save_custom_universal_catalog(delighters=cleaned, detractors=[], merge=True)
    return _save_custom_universal_catalog(delighters=[], detractors=cleaned, merge=True)


def _canonical_symptom_catalog(delighters=None, detractors=None, include_universal_neutral=None):
    if include_universal_neutral is None:
        include_universal_neutral = bool(st.session_state.get("sym_include_universal_neutral", True))
    neutral_dels, neutral_dets = _universal_neutral_catalog()
    all_universal = set(neutral_dels + neutral_dets)
    del_seed, det_seed, _ = _standardize_symptom_lists(delighters or [], detractors or [])
    del_seed = [label for label in del_seed if label not in all_universal]
    det_seed = [label for label in det_seed if label not in all_universal]
    if include_universal_neutral:
        del_seed = list(neutral_dels) + del_seed
        det_seed = list(neutral_dets) + det_seed
    dels, dets, _ = _standardize_symptom_lists(del_seed, det_seed)
    return list(dels), list(dets)


def _scrub_universal_neutral_from_processed_rows(processed_rows):
    universal_labels = _all_universal_neutral_labels()
    cleaned = []
    for rec in processed_rows or []:
        rec2 = dict(rec)
        rec2["wrote_dets"] = [tag for tag in _normalize_tag_list(rec2.get("wrote_dets") or []) if tag not in universal_labels][:10]
        rec2["wrote_dels"] = [tag for tag in _normalize_tag_list(rec2.get("wrote_dels") or []) if tag not in universal_labels][:10]
        rec2["ev_det"] = {k: list(v or [])[:2] for k, v in (rec2.get("ev_det") or {}).items() if k in rec2["wrote_dets"]}
        rec2["ev_del"] = {k: list(v or [])[:2] for k, v in (rec2.get("ev_del") or {}).items() if k in rec2["wrote_dels"]}
        cleaned.append(rec2)
    return cleaned


def _scrub_universal_neutral_from_reviews_df(reviews_df):
    if reviews_df is None or getattr(reviews_df, "empty", True):
        return reviews_df, 0
    universal_labels = _all_universal_neutral_labels()
    edited = reviews_df.copy()
    changed_rows = 0
    for idx, row in reviews_df.iterrows():
        current_dets = _normalize_tag_list(_collect_row_symptom_tags(row, AI_DET_HEADERS))
        current_dels = _normalize_tag_list(_collect_row_symptom_tags(row, AI_DEL_HEADERS))
        next_dets = [tag for tag in current_dets if tag not in universal_labels]
        next_dels = [tag for tag in current_dels if tag not in universal_labels]
        if next_dets == current_dets and next_dels == current_dels:
            continue
        edited = _write_ai_symptom_row(
            edited,
            int(idx),
            dets=next_dets,
            dels=next_dels,
            safety=row.get("AI Safety"),
            reliability=row.get("AI Reliability"),
            sessions=row.get("AI # of Sessions"),
        )
        changed_rows += 1
    return edited, changed_rows


def _apply_universal_neutral_toggle(include_enabled):
    current_dels = list(st.session_state.get("sym_delighters") or [])
    current_dets = list(st.session_state.get("sym_detractors") or [])
    new_dels, new_dets = _canonical_symptom_catalog(current_dels, current_dets, include_universal_neutral=include_enabled)
    st.session_state["sym_delighters"] = new_dels
    st.session_state["sym_detractors"] = new_dets
    st.session_state["sym_aliases"] = _alias_map_for_catalog(new_dels, new_dets, existing_aliases=st.session_state.get("sym_aliases", {}))
    if include_enabled:
        st.session_state["sym_run_notice"] = "Universal Neutral Symptoms added back into the catalog for future tagging and inline edits."
        return

    dataset_edit = dict(st.session_state.get("analysis_dataset") or {})
    reviews_df = dataset_edit.get("reviews_df")
    reviews_df, changed_rows = _scrub_universal_neutral_from_reviews_df(reviews_df)
    if reviews_df is not None:
        dataset_edit["reviews_df"] = reviews_df
        st.session_state["analysis_dataset"] = dataset_edit
    processed_rows = st.session_state.get("sym_processed_rows") or []
    if processed_rows:
        st.session_state["sym_processed_rows"] = _scrub_universal_neutral_from_processed_rows(processed_rows)
    if reviews_df is not None:
        try:
            original_bytes = st.session_state.get("_uploaded_raw_bytes")
            summary_obj = dataset_edit.get("summary") or (st.session_state.get("analysis_dataset") or {}).get("summary")
            if original_bytes:
                st.session_state["sym_export_bytes"] = _gen_symptomized_workbook(original_bytes, reviews_df)
            elif summary_obj is not None:
                st.session_state["sym_export_bytes"] = _build_master_excel(summary_obj, reviews_df)
            else:
                st.session_state["sym_export_bytes"] = None
        except Exception:
            st.session_state["sym_export_bytes"] = None
        _qa_accuracy_metrics(reviews_df)
    st.session_state["sym_qa_user_edited"] = True
    st.session_state["sym_run_notice"] = f"Universal Neutral Symptoms removed from the catalog and stripped from {changed_rows:,} currently tagged review(s)."


def _sync_symptom_catalog_session(*, default_source=None):
    current_dels = list(st.session_state.get("sym_delighters") or [])
    current_dets = list(st.session_state.get("sym_detractors") or [])
    dels, dets = _canonical_symptom_catalog(current_dels, current_dets)
    changed = dels != current_dels or dets != current_dets
    if changed:
        st.session_state["sym_delighters"] = dels
        st.session_state["sym_detractors"] = dets
    if default_source and st.session_state.get("sym_symptoms_source", "none") in {"", "none"} and (dels or dets):
        st.session_state["sym_symptoms_source"] = default_source
    return dels, dets, changed


def _apply_pending_symptomizer_ui_state():
    pending_desc = st.session_state.pop("_sym_pdesc_pending", None)
    if pending_desc is not None:
        value = _safe_text(pending_desc)
        st.session_state["sym_pdesc"] = value
        st.session_state["sym_product_profile"] = value
    pending_inline = st.session_state.pop("_sym_inline_defaults_pending", None) or {}
    for row_id, payload in pending_inline.items():
        rid = _canonical_index_key(row_id)
        if not rid:
            continue
        st.session_state[f"sym_inline_det_select_{rid}"] = _normalize_tag_list((payload or {}).get("dets") or [])
        st.session_state[f"sym_inline_del_select_{rid}"] = _normalize_tag_list((payload or {}).get("dels") or [])
        st.session_state[f"sym_inline_det_new_{rid}"] = _safe_text((payload or {}).get("new_det"))
        st.session_state[f"sym_inline_del_new_{rid}"] = _safe_text((payload or {}).get("new_del"))


def _queue_inline_editor_defaults(row_id, *, dets=None, dels=None, new_det="", new_del=""):
    rid = str(row_id).strip()
    if not rid:
        return
    pending = dict(st.session_state.get("_sym_inline_defaults_pending") or {})
    pending[rid] = {
        "dets": _normalize_tag_list(dets or []),
        "dels": _normalize_tag_list(dels or []),
        "new_det": _safe_text(new_det),
        "new_del": _safe_text(new_del),
    }
    st.session_state["_sym_inline_defaults_pending"] = pending


def _symptomizer_review_text(row):
    if row is None:
        return ""
    preferred_fields = [
        "title",
        "review_text",
        "pros",
        "cons",
        "headline",
        "body",
        "comments",
        "reviewer_comments",
        "usage_context",
        "routine",
        "variant",
        "size",
        "color",
        "flavor",
        "scent",
        "style",
        "fit",
        "material",
        "attachments",
        "accessories",
    ]
    excluded_markers = (
        "symptom",
        "ai ",
        "needs_",
        "has_",
        "rating",
        "review_id",
        "submission",
        "locale",
        "product_id",
        "sku",
        "token",
        "score",
    )

    values = []
    seen = set()
    title_and_text = _safe_text(row.get("title_and_text"))
    if title_and_text:
        values.append(title_and_text)
        seen.add(_clean_text(title_and_text).lower())
        # Also mark individual title and review_text as seen to prevent double-counting
        # since title_and_text = title + " " + review_text
        _t = _clean_text(_safe_text(row.get("title"))).lower()
        _r = _clean_text(_safe_text(row.get("review_text"))).lower()
        if _t: seen.add(_t)
        if _r: seen.add(_r)

    def _append_field(value, prefix=""):
        text = _safe_text(value)
        cleaned = _clean_text(text)
        if not cleaned:
            return
        key = cleaned.lower()
        if key in seen:
            return
        seen.add(key)
        values.append(f"{prefix}{cleaned}" if prefix else cleaned)

    for field in preferred_fields:
        if field in getattr(row, "index", []) or (hasattr(row, "get") and row.get(field) is not None):
            _append_field(row.get(field), prefix=f"{field.replace('_', ' ').title()}: ")

    for key in getattr(row, "index", []):
        key_text = str(key or "").strip()
        key_lower = key_text.lower()
        if not key_text or key_lower == "title_and_text":
            continue
        if any(marker in key_lower for marker in excluded_markers):
            continue
        if key_lower in preferred_fields:
            continue
        value = row.get(key)
        if value is None:
            continue
        text = _safe_text(value)
        if len(text.strip()) < 3:
            continue
        if len(text) > 400:
            text = text[:400]
        _append_field(text, prefix=f"{key_text.replace('_', ' ').title()}: ")

    assembled = _clean_text(" \n ".join(values))
    # Intelligent truncation: cap at 1500 chars to prevent one long review
    # from consuming the entire batch's token budget
    if len(assembled) > 1500:
        # Keep first 600 chars (opening sentiment) + last 400 chars (conclusion/summary)
        # + keyword-dense middle section
        first = assembled[:600]
        last = assembled[-400:]
        middle = assembled[600:-400]
        # Find the most signal-dense 500-char window in the middle
        if len(middle) > 500:
            _signal_kw = re.compile(r"\b(love|hate|broke|broken|great|terrible|loud|quiet|easy|hard|damage|perfect|worst|best|recommend|return|refund|issue|problem)\b", re.I)
            best_start, best_score = 0, 0
            for i in range(0, len(middle) - 500, 50):
                window = middle[i:i + 500]
                score = len(_signal_kw.findall(window))
                if score > best_score:
                    best_start, best_score = i, score
            middle = middle[best_start:best_start + 500]
        assembled = first + " [...] " + middle + " [...] " + last
    return assembled


def _sample_reviews_for_symptomizer(df, sample_n):
    if df is None or getattr(df, "empty", True):
        return []
    n = max(0, int(sample_n or 0))
    if n <= 0:
        return []
    w = df.copy()
    text_series = w.get("title_and_text", w.get("review_text", pd.Series("", index=w.index))).fillna("").astype(str).str.strip()
    w = w.loc[text_series != ""].copy()
    if w.empty:
        return []
    text_series = text_series.loc[w.index]
    rating = pd.to_numeric(w.get("rating", pd.Series(np.nan, index=w.index)), errors="coerce").fillna(3).round().clip(lower=1, upper=5).astype(int)
    lengths = pd.to_numeric(w.get("review_length_words", text_series.str.split().str.len()), errors="coerce").fillna(text_series.str.split().str.len()).fillna(0)
    review_norm = text_series.str.lower()
    submitted = pd.to_datetime(w.get("submission_time", pd.Series(pd.NaT, index=w.index)), errors="coerce", utc=True)
    prioritized = _prioritize_for_symptomization(w.assign(_sample_rating_bucket=rating, _sample_length=lengths, _sample_submitted=submitted))

    selected_idx = []
    seen = set()

    def _extend(index_values, limit):
        added = 0
        for idx in list(index_values):
            if idx in seen:
                continue
            review_text = _safe_text(text_series.get(idx))
            if not review_text:
                continue
            selected_idx.append(idx)
            seen.add(idx)
            added += 1
            if len(selected_idx) >= n or added >= limit:
                break

    available_buckets = [bucket for bucket in [1, 2, 3, 4, 5] if (rating == bucket).any()]
    per_bucket = max(1, int(math.ceil((n * 0.34) / float(max(len(available_buckets), 1)))))
    for bucket in [1, 2, 3, 4, 5]:
        bucket_df = prioritized.loc[rating.loc[prioritized.index] == bucket]
        _extend(bucket_df.index.tolist(), per_bucket)
        if len(selected_idx) >= n:
            break

    long_threshold = float(lengths.quantile(0.75)) if len(lengths) >= 4 else float(lengths.max() or 0)
    signal_masks = [
        ("explicit defects", review_norm.str.contains(r"\b(leak|broken|crack|defect|defective|stopped working|won't|wouldn't|doesn't|does not|hard to|difficult|messy|loud|noisy|disconnect|pair|charge|battery|falls flat|frizz|flyaways|curl|hold|wears off|itch|rash|shrink|sag|suction)\b", regex=True, na=False), prioritized),
        ("mixed sentiment", review_norm.str.contains(r"\b(but|however|although|except|wish|if only|yet|while)\b", regex=True, na=False), prioritized),
        ("maintenance or workflow", review_norm.str.contains(r"\b(clean|cleanup|wash|descale|filter|residue|lint|maintenance|attachment|attachments|switch|swap|refill|empty|setup|assembly|pairing)\b", regex=True, na=False), prioritized),
        ("strong delight", review_norm.str.contains(r"\b(love|favorite|easy|quiet|comfortable|fast|quick|helpful|works great|worth it|smooth|volume|hydrat|clear sound|good flavor|supportive)\b", regex=True, na=False), prioritized),
        ("long reviews", lengths >= long_threshold, prioritized.sort_values(["_sample_length"], ascending=[False], na_position="last", kind="mergesort")),
    ]
    if submitted.notna().any():
        recent_sorted = prioritized.sort_values(["_sample_submitted", "_sample_length"], ascending=[False, False], na_position="last", kind="mergesort")
        signal_masks.insert(3, ("recent reviews", submitted.notna(), recent_sorted))

    signal_quota = max(1, int(math.ceil((n * 0.5) / float(max(len(signal_masks), 1)))))
    for _label, mask, base_df in signal_masks:
        scoped = base_df.loc[mask.loc[base_df.index]]
        _extend(scoped.index.tolist(), signal_quota)
        if len(selected_idx) >= n:
            break

    if len(selected_idx) < n:
        _extend(prioritized.index.tolist(), n - len(selected_idx))

    reviews = []
    for _, row in prioritized.loc[selected_idx].iterrows():
        review_text = _symptomizer_review_text(row)
        if review_text:
            reviews.append(review_text)
            if len(reviews) >= n:
                break
    return reviews


def _coerce_product_knowledge_list(values, *, max_items=8):
    out = []
    seen = set()
    for raw in values or []:
        item = re.sub(r"\s+", " ", str(raw or "").strip().strip("•-"))
        if not item:
            continue
        key = item.lower()
        if key in seen:
            continue
        seen.add(key)
        out.append(item)
        if len(out) >= max_items:
            break
    return out


def _normalize_product_knowledge(knowledge):
    data = dict(knowledge or {}) if isinstance(knowledge, dict) else {}
    return {
        "product_archetype": _safe_text(data.get("product_archetype")),
        "product_areas": _coerce_product_knowledge_list(data.get("product_areas")),
        "use_cases": _coerce_product_knowledge_list(data.get("use_cases")),
        "desired_outcomes": _coerce_product_knowledge_list(data.get("desired_outcomes")),
        "comparison_set": _coerce_product_knowledge_list(data.get("comparison_set")),
        "workflow_steps": _coerce_product_knowledge_list(data.get("workflow_steps")),
        "user_contexts": _coerce_product_knowledge_list(data.get("user_contexts")),
        "csat_drivers": _coerce_product_knowledge_list(data.get("csat_drivers")),
        "likely_failure_modes": _coerce_product_knowledge_list(data.get("likely_failure_modes")),
        "likely_themes": _coerce_product_knowledge_list(data.get("likely_themes")),
        "likely_delighter_themes": _coerce_product_knowledge_list(data.get("likely_delighter_themes")),
        "likely_detractor_themes": _coerce_product_knowledge_list(data.get("likely_detractor_themes")),
        "watchouts": _coerce_product_knowledge_list(data.get("watchouts")),
        "confidence_note": _safe_text(data.get("confidence_note")),
    }


def _product_knowledge_context_text(knowledge, *, limit_per_section=5):
    info = _normalize_product_knowledge(knowledge)
    lines = []
    if info.get("product_archetype"):
        lines.append(f"Product archetype: {info['product_archetype']}")
    for heading, key in [
        ("Product areas", "product_areas"),
        ("Use cases", "use_cases"),
        ("Desired outcomes", "desired_outcomes"),
        ("Workflow steps", "workflow_steps"),
        ("Comparison set", "comparison_set"),
        ("User contexts", "user_contexts"),
        ("CSAT drivers", "csat_drivers"),
        ("Likely failure modes", "likely_failure_modes"),
        ("Likely themes", "likely_themes"),
        ("Likely delighter themes", "likely_delighter_themes"),
        ("Likely detractor themes", "likely_detractor_themes"),
        ("Watchouts", "watchouts"),
    ]:
        values = [str(v).strip() for v in list(info.get(key) or []) if str(v).strip()][:max(int(limit_per_section or 0), 1)]
        if values:
            lines.append(f"{heading}: " + ", ".join(values))
    note = _safe_text(info.get("confidence_note"))
    if note:
        lines.append(f"Knowledge note: {note}")
    return "\n".join(lines)


def _enrich_product_knowledge_from_run(processed_rows, *, min_mentions=3, max_per_field=20):
    """Post-run feedback loop: enrich product knowledge from tagging results.
    
    Analyzes the completed run to discover:
    - New failure modes from frequent detractor tags
    - New delighter themes from frequent delighter tags  
    - Unlisted candidates that appeared often enough to be real themes
    - Watchouts from safety/reliability findings
    
    Updates session state product knowledge so subsequent runs and
    taxonomy rebuilds benefit from what the tagger learned.
    """
    if not processed_rows or len(processed_rows) < 5:
        return {}
    
    knowledge = _normalize_product_knowledge(st.session_state.get("sym_product_knowledge") or {})
    existing_failure_modes = set(str(x).lower() for x in (knowledge.get("likely_failure_modes") or []))
    existing_del_themes = set(str(x).lower() for x in (knowledge.get("likely_delighter_themes") or []))
    existing_det_themes = set(str(x).lower() for x in (knowledge.get("likely_detractor_themes") or []))
    existing_watchouts = set(str(x).lower() for x in (knowledge.get("watchouts") or []))
    
    # Count tag frequencies
    det_freq = {}
    del_freq = {}
    unl_det_freq = {}
    unl_del_freq = {}
    safety_issues = 0
    reliability_failures = 0
    
    for rec in processed_rows:
        for tag in (rec.get("wrote_dets") or []):
            det_freq[tag] = det_freq.get(tag, 0) + 1
        for tag in (rec.get("wrote_dels") or []):
            del_freq[tag] = del_freq.get(tag, 0) + 1
        for tag in (rec.get("unl_dets") or []):
            t = str(tag).strip()
            if t:
                unl_det_freq[t] = unl_det_freq.get(t, 0) + 1
        for tag in (rec.get("unl_dels") or []):
            t = str(tag).strip()
            if t:
                unl_del_freq[t] = unl_del_freq.get(t, 0) + 1
        if str(rec.get("safety", "")).strip() in ("Safety Issue", "Minor Concern"):
            safety_issues += 1
        if str(rec.get("reliability", "")).strip() == "Failure":
            reliability_failures += 1
    
    n = max(len(processed_rows), 1)
    enrichments = {"failure_modes": [], "delighter_themes": [], "detractor_themes": [], "watchouts": []}
    
    # Discover new failure modes from high-frequency detractors
    for tag, count in sorted(det_freq.items(), key=lambda x: -x[1]):
        if count >= min_mentions and tag.lower() not in existing_failure_modes and tag.lower() not in existing_det_themes:
            enrichments["detractor_themes"].append(tag)
            # Tags appearing in >15% of reviews are likely failure modes
            if count / n >= 0.15:
                enrichments["failure_modes"].append(tag)
    
    # Discover new delighter themes
    for tag, count in sorted(del_freq.items(), key=lambda x: -x[1]):
        if count >= min_mentions and tag.lower() not in existing_del_themes:
            enrichments["delighter_themes"].append(tag)
    
    # Promote frequent unlisted candidates
    for tag, count in sorted(unl_det_freq.items(), key=lambda x: -x[1]):
        if count >= max(min_mentions, 3) and tag.lower() not in existing_det_themes and tag.lower() not in existing_failure_modes:
            enrichments["detractor_themes"].append(tag)
    for tag, count in sorted(unl_del_freq.items(), key=lambda x: -x[1]):
        if count >= max(min_mentions, 3) and tag.lower() not in existing_del_themes:
            enrichments["delighter_themes"].append(tag)
    
    # Safety/reliability findings → watchouts
    if safety_issues >= 2 and "safety concerns reported" not in existing_watchouts:
        enrichments["watchouts"].append(f"Safety concerns in {safety_issues}/{n} reviews")
    if reliability_failures >= max(3, n * 0.1) and "reliability failures reported" not in existing_watchouts:
        enrichments["watchouts"].append(f"Reliability failures in {reliability_failures}/{n} reviews")
    
    # Apply enrichments to session state
    updated = dict(knowledge)
    changed = False
    for field, session_key in [
        ("failure_modes", "likely_failure_modes"),
        ("delighter_themes", "likely_delighter_themes"),
        ("detractor_themes", "likely_detractor_themes"),
        ("watchouts", "watchouts"),
    ]:
        new_items = enrichments.get(field, [])
        if new_items:
            current = list(updated.get(session_key) or [])
            for item in new_items[:max_per_field - len(current)]:
                if item not in current:
                    current.append(item)
                    changed = True
            updated[session_key] = current[:max_per_field]
    
    if changed:
        st.session_state["sym_product_knowledge"] = updated
        _log.info("Product knowledge enriched: %s", {k: len(v) for k, v in enrichments.items() if v})
    
    return enrichments


def _auto_learn_aliases_from_run(processed_rows):
    """Post-run alias learning: discover consumer language variants.
    
    Checks unlisted candidates from the run against the existing catalog.
    Near-misses (fuzzy match score 0.55-0.72) are auto-added as aliases
    for the matched catalog label. This teaches the system consumer language
    like "makes a racket" → alias for "Loud Noise".
    """
    delighters = list(st.session_state.get("sym_delighters") or [])
    detractors = list(st.session_state.get("sym_detractors") or [])
    aliases = dict(st.session_state.get("sym_aliases") or {})
    all_catalog = delighters + detractors
    if not all_catalog or not processed_rows:
        return 0

    # Collect all unlisted candidates with their frequencies
    unl_freq: Dict[str, int] = {}
    for rec in processed_rows:
        for tag in (rec.get("unl_dets") or []) + (rec.get("unl_dels") or []):
            t = str(tag).strip()
            if t and len(t) >= 4:
                unl_freq[t] = unl_freq.get(t, 0) + 1

    added = 0
    for candidate, count in unl_freq.items():
        if count < 2:
            continue  # Need at least 2 mentions to be worth aliasing
        # Check if this candidate is a near-miss to an existing label
        # Use a lower cutoff than normal matching to catch fuzzy variants
        matched = _v3_match_label(candidate, all_catalog, aliases=aliases, cutoff=0.55) if _HAS_SYMPTOMIZER_V3 else None
        if matched and matched != candidate:
            # It's a near-miss — add as alias if not already there
            existing_aliases = list(aliases.get(matched) or [])
            if candidate not in existing_aliases and len(existing_aliases) < 10:
                existing_aliases.append(candidate)
                aliases[matched] = existing_aliases
                added += 1
                _log.info("Auto-alias: '%s' → '%s' (%d mentions)", candidate, matched, count)

    if added:
        st.session_state["sym_aliases"] = aliases
        _log.info("Auto-learned %d new aliases from unlisted candidates", added)
    return added


_PRODUCT_KNOWLEDGE_VISIBLE_KEYS = [
    "product_archetype", "product_areas", "use_cases", "desired_outcomes", "comparison_set",
    "workflow_steps", "user_contexts", "csat_drivers", "likely_failure_modes",
    "likely_themes", "likely_delighter_themes", "likely_detractor_themes", "watchouts",
]


def _has_visible_product_knowledge(knowledge):
    knowledge = _normalize_product_knowledge(knowledge)
    return any(knowledge.get(key) for key in _PRODUCT_KNOWLEDGE_VISIBLE_KEYS)


def _render_product_knowledge_panel(knowledge):
    knowledge = _normalize_product_knowledge(knowledge)
    if not _has_visible_product_knowledge(knowledge):
        return
    with st.container(border=True):
        st.caption("Built from the review sample to make the first-pass symptom list broader, sharper, and more useful across CSAT, consumer insights, CX, product, and quality teams.")
        if knowledge.get("product_archetype"):
            st.markdown(_chip_html([(knowledge.get("product_archetype").replace("_", " ").title(), "blue")]), unsafe_allow_html=True)
        for heading, key, color in [
            ("Product areas", "product_areas", "blue"),
            ("Use cases", "use_cases", "indigo"),
            ("Desired outcomes", "desired_outcomes", "green"),
            ("Workflow steps", "workflow_steps", "gray"),
            ("Comparison set", "comparison_set", "indigo"),
            ("User contexts", "user_contexts", "blue"),
            ("CSAT drivers", "csat_drivers", "green"),
            ("Likely failure modes", "likely_failure_modes", "red"),
            ("Likely delighter themes", "likely_delighter_themes", "green"),
            ("Likely detractor themes", "likely_detractor_themes", "red"),
            ("Watchouts", "watchouts", "red"),
        ]:
            values = list(knowledge.get(key) or [])
            if values:
                st.markdown(f"**{heading}**")
                st.markdown(_chip_html([(item, color) for item in values]), unsafe_allow_html=True)
        if knowledge.get("likely_themes"):
            st.markdown("**Likely first-pass themes**")
            st.markdown(_chip_html([(item, "gray") for item in knowledge.get("likely_themes")[:12]]), unsafe_allow_html=True)
        note = _safe_text(knowledge.get("confidence_note"))
        if note:
            st.caption(note)


_GENERIC_ARCHETYPE_KEYWORDS = {
    "wireless_audio": ("earbuds", "headphones", "bluetooth", "pairing", "noise cancelling", "charging case", "anc"),
    "vacuum_floorcare": ("vacuum", "robot vacuum", "suction", "pet hair", "brushroll", "mop", "dock", "dust bin"),
    "coffee_espresso": ("coffee", "espresso", "frother", "steam wand", "portafilter", "brew", "carafe"),
    "air_fryer_oven": ("air fryer", "basket", "preheat", "crisp", "countertop oven", "tray"),
    "apparel": ("fabric", "fit", "runs small", "runs large", "shirt", "dress", "jacket", "leggings"),
    "footwear": ("shoe", "sneaker", "boot", "arch support", "insole", "toe box", "laces"),
    "mattress_bedding": ("mattress", "pillow", "cooling", "firm", "soft", "motion isolation", "off gassing", "support"),
    "drinkware": ("water bottle", "tumbler", "cup holder", "keeps cold", "keeps hot", "straw", "lid", "insulated"),
    "skincare_topical": ("serum", "cream", "moisturizer", "cleanser", "retinol", "breakout", "hydrating"),
    "oral_care": ("toothbrush", "water flosser", "reservoir", "brush head", "gums", "teeth"),
}


_GENERIC_DRIVER_LIBRARY = {
    "results": {
        "signals": ("result", "results", "performance", "coverage", "flavor", "taste", "suction", "support", "hydrat", "crisp", "volume"),
        "delighter": {"label": "Delivers Strong Results", "aliases": ["results are strong", "gets the job done"], "theme": "Performance", "family": "Results & Outcome"},
        "detractor": {"label": "Falls Short On Results", "aliases": ["results are weak", "doesn't deliver the outcome"], "theme": "Performance", "family": "Results & Outcome"},
    },
    "longevity": {
        "signals": ("hold", "last", "lasting", "staying power", "runtime", "wear time", "keeps cold", "keeps hot"),
        "delighter": {"label": "Results Last", "aliases": ["lasting results", "holds up over time"], "theme": "Reliability", "family": "Outcome Longevity"},
        "detractor": {"label": "Results Fade Quickly", "aliases": ["doesn't last", "wears off fast", "drops quickly"], "theme": "Reliability", "family": "Outcome Longevity"},
    },
    "learning_curve": {
        "signals": ("learn", "learning curve", "practice", "master", "figure out", "technique"),
        "delighter": {"label": "Easy To Learn", "aliases": ["quick to figure out", "easy learning curve"], "theme": "Ease Of Use", "family": "Learning Curve & Technique"},
        "detractor": {"label": "Hard To Learn", "aliases": ["steep learning curve", "takes practice"], "theme": "Ease Of Use", "family": "Learning Curve & Technique"},
    },
    "workflow": {
        "signals": ("workflow", "routine", "session", "attachment", "attachments", "switch", "swap", "step", "tips", "refill", "empty"),
        "delighter": {"label": "Workflow Feels Smooth", "aliases": ["easy routine", "flows well in use"], "theme": "Ease Of Use", "family": "Workflow & Attachments"},
        "detractor": {"label": "Workflow Feels Awkward", "aliases": ["routine feels clunky", "steps are awkward"], "theme": "Ease Of Use", "family": "Workflow & Attachments"},
    },
    "maintenance": {
        "signals": ("clean", "cleanup", "maintenance", "filter", "wash", "descale", "care"),
        "delighter": {"label": "Easy Routine Maintenance", "aliases": ["easy upkeep", "simple maintenance"], "theme": "Cleaning & Maintenance", "family": "Cleaning & Maintenance"},
        "detractor": {"label": "Maintenance Feels Confusing", "aliases": ["upkeep is confusing", "maintenance steps are unclear"], "theme": "Cleaning & Maintenance", "family": "Cleaning & Maintenance"},
    },
    "quality": {
        "signals": ("quality", "durable", "break", "broken", "leak", "crack", "wear", "flimsy"),
        "delighter": {"label": "Built To Last", "aliases": ["holds up well", "feels durable"], "theme": "Quality & Durability", "family": "Quality & Durability"},
        "detractor": {"label": "Breaks Too Easily", "aliases": ["doesn't hold up", "wears out too fast"], "theme": "Quality & Durability", "family": "Quality & Durability"},
    },
    "comfort_fit": {
        "signals": ("comfort", "comfortable", "support", "fit", "size", "runs small", "runs large", "heavy", "lightweight", "cup holder"),
        "delighter": {"label": "Comfortable To Use", "aliases": ["feels comfortable", "comfortable during use"], "theme": "Comfort", "family": "Comfort"},
        "detractor": {"label": "Uncomfortable To Use", "aliases": ["feels uncomfortable", "causes discomfort"], "theme": "Comfort", "family": "Comfort"},
    },
    "value": {
        "signals": ("value", "price", "worth", "expensive", "overpriced", "compare", "alternative", "premium"),
        "delighter": {"label": "Feels Worth The Price", "aliases": ["worth the price", "good value"], "theme": "Value", "family": "Value & Comparison"},
        "detractor": {"label": "Does Not Feel Worth The Price", "aliases": ["not worth the price", "overpriced for what it does"], "theme": "Value", "family": "Value & Comparison"},
    },
    "battery_connectivity": {
        "signals": ("battery", "charge", "charging", "power", "app", "wifi", "bluetooth", "pairing", "connect"),
        "delighter": {"label": "Battery And Connectivity Work Reliably", "aliases": ["battery lasts and connectivity is stable"], "theme": "Compatibility & Connectivity", "family": "Power & Connectivity"},
        "detractor": {"label": "Battery Or Connectivity Cause Friction", "aliases": ["battery drains or connection is unstable"], "theme": "Compatibility & Connectivity", "family": "Power & Connectivity"},
    },
}


def _infer_generic_archetype(product_description, product_knowledge):
    corpus = "\n".join([_safe_text(product_description)] + [str(v) for v in _coerce_product_knowledge_list(_product_knowledge_context_text(product_knowledge).splitlines(), max_items=30)])
    corpus_norm = corpus.lower()
    best = "general"
    best_score = 0.0
    for archetype, keywords in _GENERIC_ARCHETYPE_KEYWORDS.items():
        score = 0.0
        for keyword in keywords:
            if str(keyword).lower() in corpus_norm:
                score += 1.0 if " " in str(keyword) else 0.65
        if score > best_score:
            best = archetype
            best_score = score
    return best


def _titleize_theme_label(value, *, default=""):
    text = re.sub(r"\s+", " ", str(value or "").strip().strip("•-"))
    if not text:
        return default
    words = [w for w in re.split(r"[^A-Za-z0-9']+", text) if w]
    if not words:
        return default
    return " ".join(word.capitalize() if not word.isupper() else word for word in words[:6])


def _specific_generic_pairs(product_knowledge, product_description, category):
    info = _normalize_product_knowledge(product_knowledge)
    items = []
    seeds = list(info.get("desired_outcomes") or []) + list(info.get("workflow_steps") or []) + list(info.get("csat_drivers") or []) + list(info.get("likely_failure_modes") or [])
    corpus = " ".join([_safe_text(product_description)] + [str(v) for v in seeds]).lower()

    def add_pair(pos_label, neg_label, *, signals, pos_aliases=None, neg_aliases=None, theme="Performance", family="CSAT Driver"):
        if not any(sig in corpus for sig in signals):
            return
        items.append({"side": "delighter", "label": pos_label, "aliases": pos_aliases or [], "theme": theme, "family": family, "bucket": "Category Driver", "seeded": True})
        items.append({"side": "detractor", "label": neg_label, "aliases": neg_aliases or [], "theme": theme, "family": family, "bucket": "Category Driver", "seeded": True})

    add_pair("Keeps Drinks Cold", "Loses Cold Quickly", signals=["cold", "keeps cold", "temperature retention", "insulated"], pos_aliases=["stays cold all day"], neg_aliases=["doesn't stay cold", "ice melts quickly"], theme="Reliability", family="Outcome Longevity")
    add_pair("Fits Cup Holder", "Does Not Fit Cup Holder", signals=["cup holder", "car"], pos_aliases=["fits my cup holder"], neg_aliases=["too big for cup holder"], theme="Size & Fit", family="Size & Fit")
    add_pair("Hydrates Well", "Not Hydrating Enough", signals=["hydrate", "hydrating", "moistur"], pos_aliases=["skin feels hydrated"], neg_aliases=["skin feels dry", "not moisturizing enough"], theme="Performance", family="Results & Outcome")
    add_pair("Strong Suction", "Weak Suction", signals=["suction", "pet hair", "pickup"], pos_aliases=["picks up everything"], neg_aliases=["doesn't pick up enough"], theme="Performance", family="Results & Outcome")
    add_pair("Rich Coffee Flavor", "Coffee Lacks Flavor", signals=["coffee", "espresso", "flavor", "taste"], pos_aliases=["great tasting coffee"], neg_aliases=["coffee tastes weak"], theme="Performance", family="Results & Outcome")
    add_pair("Fast Preheat", "Long Preheat Time", signals=["preheat", "heats up", "heat up"], pos_aliases=["preheats quickly"], neg_aliases=["slow to preheat"], theme="Time Efficiency", family="Time Efficiency")
    add_pair("Fits As Expected", "Fit Does Not Match Expectation", signals=["fit", "size", "runs small", "runs large", "true to size"], pos_aliases=["true to size"], neg_aliases=["runs small", "runs large"], theme="Size & Fit", family="Size & Fit")
    add_pair("Long Battery Life", "Battery Dies Fast", signals=["battery", "charge", "charging", "runtime"], pos_aliases=["battery lasts a long time"], neg_aliases=["battery drains quickly"], theme="Reliability", family="Power & Connectivity")
    return items


def _derive_csat_seed_candidates(product_description, sample_reviews, product_knowledge, category):
    archetype = _infer_generic_archetype(product_description, product_knowledge)
    corpus = "\n".join([_safe_text(product_description)] + [str(r) for r in (sample_reviews or [])[:60]] + [str(v) for v in (_product_knowledge_context_text(product_knowledge).splitlines())]).lower()
    delighters, detractors = [], []
    aliases = {}

    def add_item(side_key, label, *, aliases_list=None, theme="Performance", family="CSAT Driver"):
        clean_label = _titleize_theme_label(label)
        if not clean_label:
            return
        target = delighters if side_key == "delighter" else detractors
        if clean_label not in target:
            target.append(clean_label)
        aliases[clean_label] = _coerce_product_knowledge_list(aliases_list or [], max_items=8)

    for driver_name, payload in _GENERIC_DRIVER_LIBRARY.items():
        if any(sig in corpus for sig in payload.get("signals", ())):
            add_item("delighter", payload["delighter"]["label"], aliases_list=payload["delighter"].get("aliases"), theme=payload["delighter"].get("theme", "Performance"), family=payload["delighter"].get("family", "CSAT Driver"))
            add_item("detractor", payload["detractor"]["label"], aliases_list=payload["detractor"].get("aliases"), theme=payload["detractor"].get("theme", "Performance"), family=payload["detractor"].get("family", "CSAT Driver"))

    for item in _specific_generic_pairs(product_knowledge, product_description, category):
        add_item("delighter" if item.get("side") == "delighter" else "detractor", item.get("label"), aliases_list=item.get("aliases") or [], theme=item.get("theme", "Performance"), family=item.get("family", "CSAT Driver"))

    for raw in _coerce_product_knowledge_list((_normalize_product_knowledge(product_knowledge).get("likely_delighter_themes") or []), max_items=12):
        add_item("delighter", raw, aliases_list=[], theme="Performance", family="Product Knowledge Theme")
    for raw in _coerce_product_knowledge_list((_normalize_product_knowledge(product_knowledge).get("likely_detractor_themes") or []) + (_normalize_product_knowledge(product_knowledge).get("likely_failure_modes") or []), max_items=14):
        add_item("detractor", raw, aliases_list=[], theme="Performance", family="Product Knowledge Theme")

    archetype_pack = {
        "wireless_audio": {
            "delighters": ["Clear Sound Quality", "Comfortable Fit", "Long Battery Life"],
            "detractors": ["Battery Dies Fast", "Bluetooth Connection Drops", "Uncomfortable Fit"],
        },
        "vacuum_floorcare": {
            "delighters": ["Strong Suction", "Easy Bin Emptying", "Smart Navigation"],
            "detractors": ["Weak Suction", "Gets Stuck Often", "Bin Hard To Empty"],
        },
        "coffee_espresso": {
            "delighters": ["Rich Coffee Flavor", "Fast Heat Up", "Consistent Brew Results"],
            "detractors": ["Coffee Lacks Flavor", "Takes Too Long To Heat Up", "Inconsistent Brew Results"],
        },
        "air_fryer_oven": {
            "delighters": ["Crisps Food Well", "Fast Preheat", "Easy Basket Cleanup"],
            "detractors": ["Food Cooks Unevenly", "Long Preheat Time", "Basket Hard To Clean"],
        },
        "apparel": {
            "delighters": ["Fits As Expected", "Comfortable To Wear", "Fabric Feels Premium"],
            "detractors": ["Fit Does Not Match Expectation", "Feels Uncomfortable In Use", "Fabric Feels Cheap"],
        },
        "footwear": {
            "delighters": ["Comfortable To Wear", "Fits As Expected", "Good Arch Support"],
            "detractors": ["Fit Does Not Match Expectation", "Hurts After Long Wear", "Lacks Arch Support"],
        },
        "mattress_bedding": {
            "delighters": ["Comfortable Support", "Sleeps Cooler", "Good Motion Isolation"],
            "detractors": ["Too Firm", "Too Soft", "Sleeps Hot"],
        },
        "drinkware": {
            "delighters": ["Keeps Drinks Cold", "Leakproof Lid", "Fits Cup Holder"],
            "detractors": ["Loses Cold Quickly", "Leaks In Bag", "Does Not Fit Cup Holder"],
        },
        "skincare_topical": {
            "delighters": ["Hydrates Well", "Gentle On Skin", "Absorbs Quickly"],
            "detractors": ["Not Hydrating Enough", "Causes Irritation", "Feels Sticky"],
        },
        "oral_care": {
            "delighters": ["Leaves Teeth Feeling Clean", "Long Battery Life", "Gentle Yet Effective"],
            "detractors": ["Reservoir Leaks", "Battery Dies Fast", "Pressure Feels Too Strong"],
        },
    }.get(archetype, {})
    for label in archetype_pack.get("delighters", []):
        add_item("delighter", label)
    for label in archetype_pack.get("detractors", []):
        add_item("detractor", label)

    return {
        "delighters": _normalize_tag_list(delighters)[:18],
        "detractors": _normalize_tag_list(detractors)[:22],
        "aliases": aliases,
        "archetype": archetype,
    }


def _knowledge_driven_taxonomy_candidates(product_knowledge, product_description="", sample_reviews=None, category="general"):
    """Systematically derive symptom candidates from structured product knowledge.

    Instead of dumping knowledge as text context, this function exploits each
    knowledge field to generate specific, actionable symptom pairs:

    - product_areas    → Component-level detractor/delighter pairs
    - desired_outcomes → Outcome achievement / failure pairs
    - likely_failure_modes → Direct detractor labels
    - workflow_steps   → Usability symptom pairs
    - use_cases        → Context-specific symptom candidates
    - csat_drivers     → Category-level satisfaction drivers
    - watchouts        → Pre-flagged risk areas as detractors

    Returns a dict with 'delighters', 'detractors', 'aliases', and 'generation_log'.
    """
    info = _normalize_product_knowledge(product_knowledge)
    delighters, detractors = [], []
    aliases = {}
    generation_log = []
    corpus = "\n".join([_safe_text(product_description)] + [str(r) for r in (sample_reviews or [])[:40]]).lower()

    def _add(side, label, *, from_field="", aliases_list=None):
        label = _titleize_theme_label(label)
        if not label: return
        target = delighters if side == "del" else detractors
        if label not in target:
            target.append(label)
            generation_log.append({"label": label, "side": "delighter" if side == "del" else "detractor", "source": from_field})
        if aliases_list:
            aliases[label] = [str(a).strip() for a in aliases_list if str(a).strip()][:6]

    def _has_support(keywords):
        return any(kw.lower() in corpus for kw in keywords if kw)

    # ── 1. Product Areas → Component symptom pairs ──────────────────────
    component_templates = {
        "filter":      ("Easy Filter Cleaning", "Filter Hard To Clean", ["filter maintenance", "clogged filter"]),
        "motor":       ("Powerful Motor", "Motor Issues", ["motor noise", "motor failure"]),
        "battery":     ("Long Battery Life", "Battery Dies Fast", ["charge", "battery drain"]),
        "cord":        ("Cord Length Adequate", "Cord Too Short", ["cord", "power cord"]),
        "heater":      ("Heats Up Fast", "Slow To Heat", ["heat up time", "warm up"]),
        "brush":       ("Brush Works Well", "Brush Issues", ["bristle", "brush head"]),
        "nozzle":      ("Nozzle Design", "Nozzle Issues", ["nozzle", "concentrator"]),
        "attachment":  ("Attachments Work Well", "Attachment Issues", ["attachment", "accessory"]),
        "display":     ("Clear Display", "Display Hard To Read", ["screen", "display"]),
        "lid":         ("Secure Lid", "Lid Leaks", ["lid", "seal", "leak"]),
        "handle":      ("Comfortable Grip", "Handle Uncomfortable", ["grip", "handle", "ergonomic"]),
        "blade":       ("Sharp Blade", "Blade Dulls Quickly", ["blade", "cutting"]),
        "pump":        ("Pump Works Well", "Pump Issues", ["pump", "dispenser"]),
        "wheel":       ("Wheels Roll Smoothly", "Wheel Issues", ["wheel", "caster", "roll"]),
        "sensor":      ("Sensor Accuracy", "Sensor Issues", ["sensor", "detect"]),
        "app":         ("App Works Well", "App Issues", ["app", "bluetooth", "wifi", "connect"]),
    }
    for area in (info.get("product_areas") or []):
        area_lower = area.lower()
        for key, (del_label, det_label, signals) in component_templates.items():
            if key in area_lower or any(s in area_lower for s in signals):
                if _has_support(signals + [key]):
                    _add("del", del_label, from_field=f"product_areas:{area}")
                    _add("det", det_label, from_field=f"product_areas:{area}")
                break
        else:
            # Generic component pair
            clean = _titleize_theme_label(area)
            if clean:
                _add("del", f"{clean} Works Well", from_field=f"product_areas:{area}")
                _add("det", f"{clean} Issues", from_field=f"product_areas:{area}")

    # ── 2. Desired Outcomes → Achievement / failure pairs ───────────────
    for outcome in (info.get("desired_outcomes") or []):
        clean = _titleize_theme_label(outcome)
        if not clean: continue
        # Generate the positive and negative version
        _add("del", clean, from_field=f"desired_outcomes:{outcome}", aliases_list=[outcome])
        # Generate the negative (failure to achieve)
        neg = clean
        for pos_word, neg_word in [("Good","Poor"),("Fast","Slow"),("Easy","Difficult"),("Quiet","Loud"),
                                     ("Strong","Weak"),("Smooth","Rough"),("Long","Short"),("Clean","Dirty"),
                                     ("Clear","Unclear"),("Comfortable","Uncomfortable")]:
            if pos_word.lower() in clean.lower():
                neg = clean.replace(pos_word, neg_word).replace(pos_word.lower(), neg_word.lower())
                break
        else:
            neg = f"Poor {clean}" if not clean.startswith("Poor") else clean
        if neg != clean:
            _add("det", neg, from_field=f"desired_outcomes:{outcome}")

    # ── 3. Failure Modes → Direct detractors ────────────────────────────
    for mode in (info.get("likely_failure_modes") or []):
        _add("det", mode, from_field=f"likely_failure_modes:{mode}")

    # ── 4. Workflow Steps → Usability pairs ─────────────────────────────
    for step in (info.get("workflow_steps") or []):
        clean = _titleize_theme_label(step)
        if not clean: continue
        _add("del", f"Easy {clean}", from_field=f"workflow_steps:{step}")
        _add("det", f"Difficult {clean}", from_field=f"workflow_steps:{step}")

    # ── 5. CSAT Drivers → High-level satisfaction labels ────────────────
    for driver in (info.get("csat_drivers") or []):
        _add("del", driver, from_field=f"csat_drivers:{driver}")

    # ── 6. Watchouts → Pre-flagged detractors ───────────────────────────
    for watchout in (info.get("watchouts") or []):
        _add("det", watchout, from_field=f"watchouts:{watchout}")

    # ── 7. Explicit theme lists from knowledge ──────────────────────────
    for theme in (info.get("likely_delighter_themes") or []):
        _add("del", theme, from_field="likely_delighter_themes")
    for theme in (info.get("likely_detractor_themes") or []):
        _add("det", theme, from_field="likely_detractor_themes")

    # ── 8. Validate against reviews — flag unsupported candidates ───────
    validated_dels, validated_dets = [], []
    for label in delighters:
        tokens = _tokenize(label)
        if tokens and any(t in corpus for t in tokens):
            validated_dels.append(label)
        elif len(delighters) < 12:  # Keep if catalog is small
            validated_dels.append(label)
    for label in detractors:
        tokens = _tokenize(label)
        if tokens and any(t in corpus for t in tokens):
            validated_dets.append(label)
        elif len(detractors) < 14:
            validated_dets.append(label)

    return {
        "delighters": _normalize_tag_list(validated_dels)[:20],
        "detractors": _normalize_tag_list(validated_dets)[:25],
        "aliases": aliases,
        "generation_log": generation_log,
        "total_candidates": len(delighters) + len(detractors),
        "validated_count": len(validated_dels) + len(validated_dets),
    }


def _build_inline_tag_suggestions(row, current_dets, current_dels, detractors, delighters):
    review_text = _symptomizer_review_text(row)
    if not review_text:
        return {"missing_detractors": [], "missing_delighters": []}
    aliases = st.session_state.get("sym_aliases", {}) or {}
    custom_universal_dels, custom_universal_dets = _custom_universal_catalog()
    refined = _refine_tag_assignment(
        review_text,
        current_dets,
        current_dels,
        allowed_detractors=_normalize_tag_list(detractors or []),
        allowed_delighters=_normalize_tag_list(delighters or []),
        evidence_det={},
        evidence_del={},
        aliases=aliases,
        max_per_side=10,
        include_universal_neutral=bool(st.session_state.get("sym_include_universal_neutral", True)),
        rating=row.get("rating"),
        extra_universal_detractors=custom_universal_dets,
        extra_universal_delighters=custom_universal_dels,
    )
    return {
        "missing_detractors": [label for label in refined.get("added_dets", []) if label not in current_dets],
        "missing_delighters": [label for label in refined.get("added_dels", []) if label not in current_dels],
    }


def _upsert_processed_symptom_record(processed_rows, row_id, dets, dels, *, row_meta=None, ev_det=None, ev_del=None):
    rid = _canonical_index_key(row_id)
    updated = []
    found = False
    for rec in processed_rows or []:
        if str(rec.get("idx", "")).strip() == rid:
            rec2 = dict(rec)
            rec2["wrote_dets"] = _normalize_tag_list(dets or [])[:10]
            rec2["wrote_dels"] = _normalize_tag_list(dels or [])[:10]
            rec2["ev_det"] = {k: list(v or [])[:2] for k, v in (ev_det or {}).items() if k in rec2["wrote_dets"]}
            rec2["ev_del"] = {k: list(v or [])[:2] for k, v in (ev_del or {}).items() if k in rec2["wrote_dels"]}
            if row_meta is not None:
                rec2["safety"] = row_meta.get("AI Safety", rec2.get("safety", "Not Mentioned"))
                rec2["reliability"] = row_meta.get("AI Reliability", rec2.get("reliability", "Not Mentioned"))
                rec2["sessions"] = row_meta.get("AI # of Sessions", rec2.get("sessions", "Unknown"))
            found = True
            updated.append(rec2)
        else:
            updated.append(rec)
    if not found:
        row_meta = row_meta or {}
        updated.append(dict(
            idx=int(rid),
            wrote_dets=_normalize_tag_list(dets or [])[:10],
            wrote_dels=_normalize_tag_list(dels or [])[:10],
            safety=row_meta.get("AI Safety", "Not Mentioned"),
            reliability=row_meta.get("AI Reliability", "Not Mentioned"),
            sessions=row_meta.get("AI # of Sessions", "Unknown"),
            ev_det={k: list(v or [])[:2] for k, v in (ev_det or {}).items() if k in _normalize_tag_list(dets or [])[:10]},
            ev_del={k: list(v or [])[:2] for k, v in (ev_del or {}).items() if k in _normalize_tag_list(dels or [])[:10]},
            unl_dels=[],
            unl_dets=[],
        ))
    return updated


def _save_inline_symptom_edit(row_id, row, final_dets, final_dels, *, updated_reviews, processed_rows, detractors, delighters, notice_prefix="Updated"):
    rid = str(row_id).strip()
    final_dels, final_dets, inline_aliases = _standardize_symptom_lists(final_dels or [], final_dets or [])
    final_dets = list(final_dets)[:10]
    final_dels = list(final_dels)[:10]
    if not bool(st.session_state.get("sym_include_universal_neutral", True)):
        final_dets, final_dels = _strip_universal_neutral_tags(final_dets, final_dels)
        final_dets = final_dets[:10]
        final_dels = final_dels[:10]
    overlap = sorted(set(final_dets) & set(final_dels))
    if overlap:
        return False, f"The same symptom cannot be saved as both a detractor and a delighter in the same review: {', '.join(overlap)}"

    session_dels, session_dets = _canonical_symptom_catalog(list(delighters or []) + final_dels, list(detractors or []) + final_dets)
    st.session_state["sym_detractors"] = session_dets
    st.session_state["sym_delighters"] = session_dels
    st.session_state["sym_aliases"] = _alias_map_for_catalog(session_dels, session_dets, extra_aliases=inline_aliases, existing_aliases=st.session_state.get("sym_aliases", {}))

    edited_reviews = updated_reviews.copy()
    edited_reviews = _write_ai_symptom_row(
        edited_reviews,
        int(rid),
        dets=final_dets,
        dels=final_dels,
        safety=row.get("AI Safety"),
        reliability=row.get("AI Reliability"),
        sessions=row.get("AI # of Sessions"),
    )
    custom_universal_dels, custom_universal_dets = _custom_universal_catalog()
    support = _refine_tag_assignment(
        _symptomizer_review_text(row),
        final_dets,
        final_dels,
        allowed_detractors=final_dets,
        allowed_delighters=final_dels,
        evidence_det={},
        evidence_del={},
        aliases=st.session_state.get("sym_aliases", {}) or {},
        max_per_side=10,
        include_universal_neutral=bool(st.session_state.get("sym_include_universal_neutral", True)),
        rating=row.get("rating"),
        extra_universal_detractors=custom_universal_dets,
        extra_universal_delighters=custom_universal_dels,
    )
    support_det = support.get("support_det", {}) or {}
    support_del = support.get("support_del", {}) or {}
    ev_det = {label: list((support_det.get(label, {}) or {}).get("snippets") or [])[:2] for label in final_dets if list((support_det.get(label, {}) or {}).get("snippets") or [])[:2]}
    ev_del = {label: list((support_del.get(label, {}) or {}).get("snippets") or [])[:2] for label in final_dels if list((support_del.get(label, {}) or {}).get("snippets") or [])[:2]}

    dataset_edit = dict(st.session_state.get("analysis_dataset") or {})
    dataset_edit["reviews_df"] = edited_reviews
    st.session_state["analysis_dataset"] = dataset_edit
    st.session_state["sym_processed_rows"] = _upsert_processed_symptom_record(
        processed_rows,
        rid,
        final_dets,
        final_dels,
        row_meta=row,
        ev_det=ev_det,
        ev_del=ev_del,
    )
    export_suffix = " Export will regenerate when you click Prepare export."
    try:
        original_bytes = st.session_state.get("_uploaded_raw_bytes")
        summary_obj = dataset_edit.get("summary") or (st.session_state.get("analysis_dataset") or {}).get("summary")
        if original_bytes:
            st.session_state["sym_export_bytes"] = _gen_symptomized_workbook(original_bytes, edited_reviews)
            export_suffix = " Export refreshed."
        elif summary_obj is not None:
            st.session_state["sym_export_bytes"] = _build_master_excel(summary_obj, edited_reviews)
            export_suffix = " Export refreshed."
        else:
            st.session_state["sym_export_bytes"] = None
    except Exception:
        st.session_state["sym_export_bytes"] = None
    _queue_inline_editor_defaults(rid, dets=final_dets, dels=final_dels, new_det="", new_del="")
    st.session_state["sym_qa_user_edited"] = True
    metrics_now = _qa_accuracy_metrics(edited_reviews)
    st.session_state["sym_qa_notice"] = f"{notice_prefix} row {rid}. Accuracy is now {metrics_now.get('accuracy_pct', 100.0):.1f}%.{export_suffix}"
    return True, ""


def _prioritize_for_symptomization(df):
    if df.empty:
        return df
    w = df.copy()
    text = w.get("title_and_text", w.get("review_text", pd.Series("", index=w.index))).fillna("").astype(str)
    rating = pd.to_numeric(w.get("rating", pd.Series(dtype=float)), errors="coerce").fillna(3)
    lengths = pd.to_numeric(w.get("review_length_words", text.str.split().str.len()), errors="coerce").fillna(text.str.split().str.len()).fillna(0)
    neg_kw = r"\b(broke|broken|fail|failed|defect|issue|problem|stopped|won't|doesn't|difficult|hard|confusing|loud|noise|leak|burned|smoke|stuck|cracked|itchy|rash|damaged|expensive|overpriced|slow)\b"
    pos_kw = r"\b(love|perfect|amazing|excellent|great|easy|simple|quiet|durable|worth|recommend|best|happy|works well|works great|comfortable|stylish|quick)\b"
    mixed_kw = r"\b(but|however|although|except|yet|while|wish)\b"
    w["_prio"] = (
        (rating <= 2).astype(int) * 4.0 +
        (rating >= 4.5).astype(int) * 2.5 +
        (rating.between(2.5, 3.5)).astype(int) * 1.2 +
        (lengths.clip(upper=500) / 120.0)
    )
    w["_prio"] += text.str.lower().str.count(neg_kw, flags=re.IGNORECASE).clip(upper=4) * 0.9
    w["_prio"] += text.str.lower().str.count(pos_kw, flags=re.IGNORECASE).clip(upper=4) * 0.45
    w["_prio"] += text.str.lower().str.count(mixed_kw, flags=re.IGNORECASE).clip(upper=2) * 0.9
    w["_prio"] += text.str.len().clip(upper=1200) / 1500.0
    return w.sort_values(["_prio"], ascending=False).drop(columns=["_prio"])


def _historical_symptom_counts_from_workspace(
    *,
    allowed_detractors: Optional[Sequence[str]] = None,
    allowed_delighters: Optional[Sequence[str]] = None,
) -> Tuple[Dict[str, int], Dict[str, int]]:
    """Build lightweight historical label counts from the current workspace."""

    det_allow = set(_normalize_tag_list(list(allowed_detractors or [])))
    del_allow = set(_normalize_tag_list(list(allowed_delighters or [])))
    det_counts: Counter[str] = Counter()
    del_counts: Counter[str] = Counter()

    processed = list(st.session_state.get("sym_processed_rows") or [])
    for rec in processed:
        for label in _normalize_tag_list(rec.get("wrote_dets") or []):
            if not det_allow or label in det_allow:
                det_counts[label] += 1
        for label in _normalize_tag_list(rec.get("wrote_dels") or []):
            if not del_allow or label in del_allow:
                del_counts[label] += 1

    dataset = st.session_state.get("analysis_dataset") or {}
    reviews_df = dataset.get("reviews_df") if isinstance(dataset, dict) else None
    if isinstance(reviews_df, pd.DataFrame) and not reviews_df.empty:
        det_cols = [col for col in reviews_df.columns if col.startswith("AI Symptom Det")]
        del_cols = [col for col in reviews_df.columns if col.startswith("AI Symptom Del")]
        for col in det_cols:
            for label in _normalize_tag_list(reviews_df[col].fillna("").astype(str).tolist()):
                if not det_allow or label in det_allow:
                    det_counts[label] += 1
        for col in del_cols:
            for label in _normalize_tag_list(reviews_df[col].fillna("").astype(str).tolist()):
                if not del_allow or label in del_allow:
                    del_counts[label] += 1

    return dict(det_counts), dict(del_counts)


def _symptomizer_cache_signature(
    *,
    allowed_detractors: Sequence[str],
    allowed_delighters: Sequence[str],
    aliases: Optional[Dict[str, List[str]]] = None,
    detractor_specs: Optional[Sequence[Any]] = None,
    delighter_specs: Optional[Sequence[Any]] = None,
    product_profile: str = "",
    product_knowledge: Any = None,
    max_ev_chars: int = 120,
    include_universal_neutral: bool = True,
    taxonomy_category: str = "general",
    v4_enabled: bool = True,
) -> str:
    """Build a stable fingerprint of the active symptomizer prompt context."""

    def _default(value: Any) -> Any:
        if isinstance(value, set):
            return sorted(str(v) for v in value)
        if hasattr(value, "__dict__"):
            return value.__dict__
        return str(value)

    payload = {
        "allowed_detractors": list(_normalize_tag_list(list(allowed_detractors or []))),
        "allowed_delighters": list(_normalize_tag_list(list(allowed_delighters or []))),
        "aliases": aliases or {},
        "detractor_specs": list(detractor_specs or []),
        "delighter_specs": list(delighter_specs or []),
        "product_profile": _safe_text(product_profile),
        "product_knowledge": product_knowledge or {},
        "max_ev_chars": int(max_ev_chars or 0),
        "include_universal_neutral": bool(include_universal_neutral),
        "taxonomy_category": _safe_text(taxonomy_category) or "general",
        "v4_enabled": bool(v4_enabled),
    }
    raw = json.dumps(payload, sort_keys=True, ensure_ascii=False, default=_default, separators=(",", ":"))
    return hashlib.sha256(raw.encode("utf-8", errors="replace")).hexdigest()[:24]


def _invalidate_symptomizer_cache_if_needed(
    *,
    allowed_detractors: Sequence[str],
    allowed_delighters: Sequence[str],
    aliases: Optional[Dict[str, List[str]]] = None,
    detractor_specs: Optional[Sequence[Any]] = None,
    delighter_specs: Optional[Sequence[Any]] = None,
    product_profile: str = "",
    product_knowledge: Any = None,
    max_ev_chars: int = 120,
    include_universal_neutral: bool = True,
    taxonomy_category: str = "general",
    v4_enabled: bool = True,
) -> None:
    """Clear the session symptomizer cache when the active prompt context changes."""

    if not _HAS_SYMPTOMIZER_V3:
        return

    signature = _symptomizer_cache_signature(
        allowed_detractors=allowed_detractors,
        allowed_delighters=allowed_delighters,
        aliases=aliases,
        detractor_specs=detractor_specs,
        delighter_specs=delighter_specs,
        product_profile=product_profile,
        product_knowledge=product_knowledge,
        max_ev_chars=max_ev_chars,
        include_universal_neutral=include_universal_neutral,
        taxonomy_category=taxonomy_category,
        v4_enabled=v4_enabled,
    )
    previous = st.session_state.get("_sym_cache_signature")
    if previous != signature:
        _v3_result_cache.clear()
        st.session_state["_sym_cache_signature"] = signature
        if previous:
            _log.info("Symptomizer cache cleared after prompt context change")

def _call_symptomizer_batch(*, client, items, allowed_delighters, allowed_detractors,
                             product_profile="", product_knowledge=None, max_ev_chars=120, aliases=None, include_universal_neutral=True):
    if not items:
        return {}
    _det_specs = st.session_state.get("sym_detractor_specs", [])
    _del_specs = st.session_state.get("sym_delighter_specs", [])
    _use_v4 = bool(st.session_state.get("sym_v4_pipeline", True))
    _invalidate_symptomizer_cache_if_needed(
        allowed_detractors=allowed_detractors,
        allowed_delighters=allowed_delighters,
        aliases=aliases,
        detractor_specs=_det_specs or None,
        delighter_specs=_del_specs or None,
        product_profile=product_profile,
        product_knowledge=product_knowledge,
        max_ev_chars=max_ev_chars,
        include_universal_neutral=include_universal_neutral,
        taxonomy_category=st.session_state.get("sym_taxonomy_category", "general"),
        v4_enabled=_use_v4,
    )
    _hist_det_counts, _hist_del_counts = _historical_symptom_counts_from_workspace(
        allowed_detractors=allowed_detractors,
        allowed_delighters=allowed_delighters,
    )

    if _HAS_SYMPTOMIZER_V3 and _use_v4:
        from review_analyst.symptomizer import tag_review_batch_v4
        return tag_review_batch_v4(
            client=client,
            items=items,
            allowed_delighters=allowed_delighters,
            allowed_detractors=allowed_detractors,
            detractor_specs=_det_specs or None,
            delighter_specs=_del_specs or None,
            historical_detractor_counts=_hist_det_counts,
            historical_delighter_counts=_hist_del_counts,
            product_profile=product_profile,
            product_knowledge=product_knowledge,
            max_ev_chars=max_ev_chars,
            aliases=aliases,
            include_universal_neutral=include_universal_neutral,
            pre_category=st.session_state.get("sym_taxonomy_category", ""),
            chat_complete_fn=_chat_complete_with_fallback_models,
            safe_json_load_fn=_safe_json_load,
            refine_fn=_refine_tag_assignment,
            model_fn=_shared_model,
            reasoning_fn=_shared_reasoning,
            infer_category_fn=_infer_taxonomy_category,
            taxonomy_context_fn=_taxonomy_prompt_context,
            product_knowledge_text_fn=_product_knowledge_context_text,
            custom_universal_fn=_custom_universal_catalog,
            standardize_fn=lambda d, t: _standardize_symptom_lists(d, t),
        )
    # ── v3 staged pipeline (extract → classify → verify) ────────────────
    # Auto-activates for long reviews (>100 avg words) where the quality gain
    # from taxonomy-free claim extraction justifies the extra API call.
    _user_staged = bool(st.session_state.get("sym_staged_pipeline"))
    _avg_review_words = sum(len(str(it.get("review", "")).split()) for it in items) / max(len(items), 1)
    _auto_staged = _avg_review_words > 200 and len(items) <= 5  # Only for small batches of very long reviews
    if _HAS_SYMPTOMIZER_V3 and (_user_staged or _auto_staged):
        from review_analyst.symptomizer import extract_claims, map_claims_to_taxonomy
        _SAFETY_KW = re.compile(r"\b(burn|fire|shock|electr|hazard|danger|injur|hospital|smoke|smoking|melted|caught fire|sparks?)\b", re.I)
        _RELIABILITY_KW_NEG = re.compile(r"\b(broke|broken|fail|died|stopped|malfunction|defective|DOA|dead on arrival)\b", re.I)
        _RELIABILITY_KW_MIX = re.compile(r"\b(intermittent|sometimes|occasional|inconsistent|hit.or.miss)\b", re.I)
        out_staged = {}
        for it in items:
            idx = int(it["idx"])
            review_text = it.get("review", "")
            rating = it.get("rating")
            claims = extract_claims(client=client, review_text=review_text, rating=rating,
                chat_fn=_chat_complete_with_fallback_models, json_fn=_safe_json_load,
                model_fn=_shared_model, reasoning_fn=_shared_reasoning)
            if claims:
                dets, dels, ev_det, ev_del = map_claims_to_taxonomy(
                    claims, allowed_detractors, allowed_delighters, aliases=aliases)
                # Run refinement pass
                if dets or dels:
                    custom_dels, custom_dets = _custom_universal_catalog()
                    refined = _refine_tag_assignment(review_text, dets, dels,
                        allowed_detractors=allowed_detractors, allowed_delighters=allowed_delighters,
                        evidence_det=ev_det, evidence_del=ev_del, aliases=aliases, max_per_side=10,
                        include_universal_neutral=bool(include_universal_neutral), rating=rating,
                        extra_universal_detractors=custom_dets, extra_universal_delighters=custom_dels)
                    dets = list(refined.get("dets", []))[:10]
                    dels = list(refined.get("dels", []))[:10]
                    ev_det = dict(refined.get("ev_det", {}) or {})
                    ev_del = dict(refined.get("ev_del", {}) or {})
                # Infer safety/reliability/sessions from claims + review text
                _rt = review_text.lower()
                safety = "Safety Issue" if _SAFETY_KW.search(_rt) else "Not Mentioned"
                if safety == "Not Mentioned" and any("safe" in c.get("aspect","").lower() for c in claims):
                    safety = "Safe"
                reliability = "Not Mentioned"
                if _RELIABILITY_KW_NEG.search(_rt): reliability = "Failure"
                elif _RELIABILITY_KW_MIX.search(_rt): reliability = "Intermittent Issue"
                elif not dets and dels: reliability = "Reliable"
                sessions = "Unknown"
                # Extract unlisted candidates from claims that didn't match taxonomy
                all_matched = set(dets + dels)
                unl_dets, unl_dels = [], []
                for claim in claims:
                    aspect = claim.get("aspect", "").strip()
                    polarity = claim.get("polarity", "neutral")
                    if not aspect or len(aspect) < 4 or len(aspect.split()) > 6:
                        continue
                    # Check if this claim's aspect already matched a catalog label
                    from review_analyst.symptomizer import match_label as _ml
                    if _ml(aspect, list(allowed_detractors) + list(allowed_delighters), aliases=aliases):
                        continue
                    aspect_title = " ".join(w.capitalize() for w in aspect.split())
                    if polarity in ("negative", "mixed") and aspect_title not in unl_dets:
                        unl_dets.append(aspect_title)
                    elif polarity in ("positive",) and aspect_title not in unl_dels:
                        unl_dels.append(aspect_title)
                out_staged[idx] = dict(dels=dels, dets=dets, ev_del=ev_del, ev_det=ev_det,
                    unl_dels=unl_dels[:5], unl_dets=unl_dets[:5], safety=safety, reliability=reliability, sessions=sessions)
            else:
                out_staged[idx] = dict(dels=[], dets=[], ev_del={}, ev_det={},
                    unl_dels=[], unl_dets=[], safety="Not Mentioned", reliability="Not Mentioned", sessions="Unknown")
        return out_staged
    # ── v3 single-pass engine (default) ──────────────────────────────────
    if _HAS_SYMPTOMIZER_V3:
        return _v3_tag_review_batch(
            client=client, items=items, allowed_delighters=allowed_delighters,
            allowed_detractors=allowed_detractors, product_profile=product_profile,
            detractor_specs=_det_specs or None,
            delighter_specs=_del_specs or None,
            product_knowledge=product_knowledge, max_ev_chars=max_ev_chars,
            aliases=aliases, include_universal_neutral=include_universal_neutral,
            pre_category=st.session_state.get("sym_taxonomy_category", ""),
            chat_complete_fn=_chat_complete_with_fallback_models,
            safe_json_load_fn=_safe_json_load, refine_fn=_refine_tag_assignment,
            model_fn=_shared_model, reasoning_fn=_shared_reasoning,
            infer_category_fn=_infer_taxonomy_category,
            taxonomy_context_fn=_taxonomy_prompt_context,
            product_knowledge_text_fn=_product_knowledge_context_text,
            custom_universal_fn=_custom_universal_catalog,
            standardize_fn=lambda d, t: _standardize_symptom_lists(d, t),
        )
    # ── Fallback: original inline tagger ─────────────────────────────────
    out_by_idx = {}
    category_info = _infer_taxonomy_category(product_profile, [it.get("review", "") for it in items[:12]])
    category = category_info.get("category", "general")
    product_knowledge_text = _product_knowledge_context_text(product_knowledge, limit_per_section=4)
    taxonomy_context = _taxonomy_prompt_context(category)
    # Try to use the v3 prompt builder for consistency even in fallback mode
    try:
        from review_analyst.symptomizer import _build_system_prompt as _v3_build_prompt
        system_prompt = _v3_build_prompt(
            allowed_detractors=allowed_detractors, allowed_delighters=allowed_delighters,
            product_profile=product_profile, product_knowledge_text=product_knowledge_text,
            taxonomy_context=taxonomy_context, category=category, max_ev_chars=max_ev_chars,
        )
    except Exception:
        # Ultimate fallback: minimal prompt
        det_list = "\n".join(f"  - {l}" for l in allowed_detractors) or "  (none defined)"
        del_list = "\n".join(f"  - {l}" for l in allowed_delighters) or "  (none defined)"
        system_prompt = f"""You are an expert consumer product review analyst.
Tag reviews against this symptom taxonomy. Evidence required for every tag.

{f"Product: {product_profile[:500]}" if product_profile else ""}

DETRACTORS: {det_list}
DELIGHTERS: {del_list}

Rules: exact labels only, verbatim evidence (4-{max_ev_chars} chars), tag both sides regardless of rating,
watch for negation ("didn't break" = positive), no inference.

Output JSON: {{"items":[{{"id":"<id>","detractors":[{{"label":"<exact>","evidence":["<verbatim>"]}}],
"delighters":[{{"label":"<exact>","evidence":["<verbatim>"]}}],
"unlisted_detractors":[],"unlisted_delighters":[],
"safety":"<Safe|Minor Concern|Safety Issue|Not Mentioned>",
"reliability":"<Reliable|Intermittent Issue|Failure|Not Mentioned>",
"sessions":"<1-5|6-20|21-50|50+|Unknown>"}}]}}"""
    payload = dict(items=[dict(id=str(it["idx"]), review=it["review"], rating=it.get("rating")) for it in items])
    catalog_size = len(allowed_detractors) + len(allowed_delighters)
    catalog_mult = 1.0 + min(0.5, max(0, catalog_size - 20) / 80)
    max_out = min(6500, max(1200, int((200 * len(items) + 400) * catalog_mult)))
    result_text = _chat_complete_with_fallback_models(
        client,
        model=_shared_model(),
        structured=True,
        messages=[{"role": "system", "content": system_prompt}, {"role": "user", "content": json.dumps(payload)}],
        temperature=0.0,
        response_format={"type": "json_object"},
        max_tokens=max_out,
        reasoning_effort=_shared_reasoning(),
    )
    data = _safe_json_load(result_text)
    items_out = data.get("items") or (data if isinstance(data, list) else [])
    by_id = {str(o.get("id")): o for o in items_out if isinstance(o, dict) and "id" in o}
    for it in items:
        idx = int(it["idx"])
        review_text = it.get("review", "")
        obj = by_id.get(str(idx)) or {}

        def _extract_side(objs, allowed):
            labels = []
            ev_map = {}
            for obj2 in (objs or []):
                if not isinstance(obj2, dict):
                    continue
                raw = str(obj2.get("label", "")).strip()
                lbl = _match_label(raw, allowed, aliases=aliases)
                if not lbl:
                    continue
                raw_evs = [str(e) for e in (obj2.get("evidence") or []) if isinstance(e, str)]
                validated = _validate_evidence(raw_evs, review_text, max_ev_chars)
                if not validated:
                    validated = [str(e).strip()[:max_ev_chars] for e in raw_evs if str(e).strip()][:1]
                if lbl not in labels:
                    labels.append(lbl)
                    ev_map[lbl] = validated[:2]
                if len(labels) >= 10:
                    break
            return labels, ev_map

        dels, ev_del = _extract_side(obj.get("delighters", []), allowed_delighters)
        dets, ev_det = _extract_side(obj.get("detractors", []), allowed_detractors)
        custom_universal_dels, custom_universal_dets = _custom_universal_catalog()
        refined = _refine_tag_assignment(
            review_text,
            dets,
            dels,
            allowed_detractors=allowed_detractors,
            allowed_delighters=allowed_delighters,
            evidence_det=ev_det,
            evidence_del=ev_del,
            aliases=aliases,
            max_per_side=10,
            include_universal_neutral=bool(include_universal_neutral),
            rating=it.get("rating"),
            extra_universal_detractors=custom_universal_dets,
            extra_universal_delighters=custom_universal_dels,
        )
        dets = list(refined.get("dets", []))[:10]
        dels = list(refined.get("dels", []))[:10]
        ev_det = dict(refined.get("ev_det", {}) or {})
        ev_del = dict(refined.get("ev_del", {}) or {})
        safety = str(obj.get("safety", "Not Mentioned")).strip()
        reliability = str(obj.get("reliability", "Not Mentioned")).strip()
        sessions = str(obj.get("sessions", "Unknown")).strip()
        safety = safety if safety in SAFETY_ENUM else "Not Mentioned"
        reliability = reliability if reliability in RELIABILITY_ENUM else "Not Mentioned"
        sessions = sessions if sessions in SESSIONS_ENUM else "Unknown"
        canon_unl_dels, _, _ = _standardize_symptom_lists([str(x).strip() for x in (obj.get("unlisted_delighters") or []) if str(x).strip()][:10], [])
        _, canon_unl_dets, _ = _standardize_symptom_lists([], [str(x).strip() for x in (obj.get("unlisted_detractors") or []) if str(x).strip()][:10])
        out_by_idx[idx] = dict(
            dels=dels,
            dets=dets,
            ev_del=ev_del,
            ev_det=ev_det,
            unl_dels=list(canon_unl_dels)[:10],
            unl_dets=list(canon_unl_dets)[:10],
            safety=safety,
            reliability=reliability,
            sessions=sessions,
        )
    return out_by_idx


def _ai_build_symptom_list(*, client, product_description, sample_reviews, product_knowledge=None):
    category_info = _infer_taxonomy_category(product_description, sample_reviews)
    category = category_info.get("category", "general")
    sample_reviews = [_trunc(str(review or "").strip(), 420) for review in list(sample_reviews or [])[:60] if str(review or "").strip()]
    min_hits = 1 if len(sample_reviews) <= 15 else 2
    product_knowledge = _normalize_product_knowledge(product_knowledge)
    supported_pack = _select_supported_category_pack(category, sample_reviews, min_hits=min_hits, max_per_side=10)
    generic_seed_pack = _derive_csat_seed_candidates(product_description, sample_reviews, product_knowledge, category)
    # Augment with knowledge-driven candidates
    try:
        knowledge_candidates = _knowledge_driven_taxonomy_candidates(product_knowledge, product_description, sample_reviews, category)
        # Merge knowledge candidates into the seed pack
        for label in knowledge_candidates.get("delighters", []):
            if label not in generic_seed_pack.get("delighters", []):
                generic_seed_pack.setdefault("delighters", []).append(label)
        for label in knowledge_candidates.get("detractors", []):
            if label not in generic_seed_pack.get("detractors", []):
                generic_seed_pack.setdefault("detractors", []).append(label)
        for k, v in knowledge_candidates.get("aliases", {}).items():
            if k not in generic_seed_pack.get("aliases", {}):
                generic_seed_pack.setdefault("aliases", {})[k] = v
    except Exception:
        pass
    neutral_dels, neutral_dets = _universal_neutral_catalog()
    excluded_universal = {
        "delighters": list(neutral_dels),
        "detractors": list(neutral_dets),
    }
    knowledge_context = _product_knowledge_context_text(product_knowledge, limit_per_section=6)
    sys = textwrap.dedent(f"""
        You are a consumer insights expert building a reusable first-pass symptom taxonomy for product review analysis.
        The taxonomy must work for ANY consumer product category and stay useful to consumer insights, quality engineers, product developers, CX, and brand teams.
        DELIGHTERS = praised features, strengths, positive outcomes.
        DETRACTORS = problems, failures, frustrations, negative outcomes.

        {_taxonomy_prompt_context(category)}

        PRODUCT KNOWLEDGE RULES:
        - The GENERIC SEED CANDIDATES below were systematically derived from structured product knowledge:
          • Each product_area generated component-level symptom pairs (e.g., Filter → "Easy Filter Cleaning" / "Filter Hard To Clean")
          • Each desired_outcome generated achievement/failure pairs
          • Each likely_failure_mode became a direct detractor candidate
          • Each workflow_step generated usability pairs
        - These seeds are your starting point — validate each against the sample reviews and KEEP labels that have review support.
        - ADD new labels for patterns you see in reviews that the seeds missed.
        - REMOVE seeds that have zero review support and don't match any review pattern.
        - When the product clearly has important outcomes, create paired labels when supported, such as Fits As Expected vs Fit Does Not Match Expectation, Long Battery Life vs Battery Dies Fast, Strong Suction vs Weak Suction, Hydrates Well vs Not Hydrating Enough, or Keeps Drinks Cold vs Loses Cold Quickly.
        - Prefer concrete customer-facing labels that are still useful to cross-functional teams.
        - When in doubt, prefer component + mode labels such as Filter Door Hard To Open, Lid Leaks In Bag, App Mapping Is Confusing, or Steam Wand Hard To Clean over vague catch-alls.

        SYSTEMATIC LABELING RULES:
        - Do not call the same concept two different things. Collapse near-duplicates to one canonical label.
        - Use concise Title Case labels, usually 2-6 words.
        - Universal Neutral Symptoms are managed elsewhere, so do NOT spend slots returning these labels: {', '.join(excluded_universal['delighters'] + excluded_universal['detractors'])}.
        - Separate your output into Category Drivers vs Product Specific labels.
        - Category Drivers should represent the major reasons customers are satisfied or dissatisfied in this category.
        - Product Specific labels should be concrete, engineer-usable recurring issues or strengths tied to components, workflow, setup, cleaning, fit, packaging, formula, or failure modes.
        - Keep the catalog MECE-ish in practice: avoid overlapping labels that describe the same issue at different levels of abstraction when one sharper label will do.
        - Keep each label cross-functional and actionable for consumer insights, quality, product, CX, and brand teams.
        - Avoid vague duplicates like Great Product, Good Features, Bad Experience, Poor Item.
        - Include aliases for common alternate phrasings reviewers use.
        - Keep the final list clean, deduplicated, and useful for downstream analytics tables.

        PRODUCT KNOWLEDGE SNAPSHOT:
        {knowledge_context or 'No structured product knowledge available.'}

        GENERIC FIRST-CUT SEED CANDIDATES TO CONSIDER:
        Delighters: {', '.join(generic_seed_pack.get('delighters', [])[:18]) or 'None'}
        Detractors: {', '.join(generic_seed_pack.get('detractors', [])[:22]) or 'None'}

        OUTPUT — strict JSON only:
        {{
          "category":"<short category>",
          "delighters":{{
            "category_drivers":[{{"label":"<Title Case 2-6 words>","theme":"<L1 theme>","aliases":["<alternate phrase>"],"family":"<short family>","rationale":"<why>"}}],
            "product_specific":[{{"label":"<Title Case 2-6 words>","theme":"<L1 theme>","aliases":["<alternate phrase>"],"family":"<short family>","rationale":"<why>"}}]
          }},
          "detractors":{{
            "category_drivers":[{{"label":"<Title Case 2-6 words>","theme":"<L1 theme>","aliases":["<alternate phrase>"],"family":"<short family>","rationale":"<why>"}}],
            "product_specific":[{{"label":"<Title Case 2-6 words>","theme":"<L1 theme>","aliases":["<alternate phrase>"],"family":"<short family>","rationale":"<why>"}}]
          }},
          "notes":"<short note>"
        }}
        Aim for roughly 6-10 category drivers and 12-24 product-specific labels per side based on actual review patterns. Favor broader coverage only when labels stay concrete, non-overlapping, and review-backed.
    """).strip()
    payload = dict(
        product_description=product_description or "General consumer product",
        product_knowledge=product_knowledge,
        category_hint=category,
        category_signals=category_info.get("signals", []),
        sample_reviews=sample_reviews,
        supported_category_drivers={
            "delighters": supported_pack.get("delighters", []),
            "detractors": supported_pack.get("detractors", []),
        },
        generic_seed_candidates={
            "delighters": generic_seed_pack.get("delighters", []),
            "detractors": generic_seed_pack.get("detractors", []),
        },
        universal_neutral_managed_separately=excluded_universal,
    )
    result_text = _chat_complete_with_fallback_models(
        client,
        model=_shared_model(),
        structured=True,
        messages=[{"role": "system", "content": sys}, {"role": "user", "content": json.dumps(payload)}],
        temperature=0.0,
        response_format={"type": "json_object"},
        max_tokens=4200,
        reasoning_effort=_shared_reasoning(),
    )
    data = _safe_json_load(result_text)

    def _coerce_taxonomy_item(obj, *, bucket, seeded=False):
        if isinstance(obj, dict):
            label = str(obj.get("label", "")).strip()
            aliases = [str(a).strip() for a in (obj.get("aliases") or []) if str(a).strip()]
            family = _safe_text(obj.get("family"))
            theme = _safe_text(obj.get("theme") or obj.get("l1_theme") or obj.get("l1"))
            rationale = _safe_text(obj.get("rationale"))
        else:
            label = str(obj or "").strip()
            aliases = []
            family = ""
            theme = ""
            rationale = ""
        if not label:
            return None
        return {
            "label": label,
            "aliases": aliases,
            "family": family,
            "theme": theme,
            "rationale": rationale,
            "bucket": bucket,
            "seeded": bool(seeded),
        }

    def _parse_side(section):
        parsed = []
        if isinstance(section, dict):
            for bucket_key, bucket_name in (("category_drivers", "Category Driver"), ("product_specific", "Product Specific")):
                for raw in (section.get(bucket_key) or []):
                    item = _coerce_taxonomy_item(raw, bucket=bucket_name)
                    if item:
                        parsed.append(item)
            if not parsed:
                for raw in (section.get("items") or []):
                    item = _coerce_taxonomy_item(raw, bucket="Product Specific")
                    if item:
                        parsed.append(item)
        else:
            for raw in (section or []):
                item = _coerce_taxonomy_item(raw, bucket="Product Specific")
                if item:
                    parsed.append(item)
        return parsed

    seeded_delighters = [
        {
            "label": label,
            "aliases": list((supported_pack.get("aliases") or {}).get(label, [])) + list((generic_seed_pack.get("aliases") or {}).get(label, [])),
            "family": "",
            "theme": _infer_taxonomy_l1_theme(label, side="delighter", category=category),
            "rationale": "Supported CSAT or category-general pattern from product knowledge and the review sample.",
            "bucket": "Category Driver",
            "seeded": True,
        }
        for label in _dedupe_keep_order(list(supported_pack.get("delighters") or []) + list(generic_seed_pack.get("delighters") or []))
    ]
    seeded_detractors = [
        {
            "label": label,
            "aliases": list((supported_pack.get("aliases") or {}).get(label, [])) + list((generic_seed_pack.get("aliases") or {}).get(label, [])),
            "family": "",
            "theme": _infer_taxonomy_l1_theme(label, side="detractor", category=category),
            "rationale": "Supported CSAT or category-general pattern from product knowledge and the review sample.",
            "bucket": "Category Driver",
            "seeded": True,
        }
        for label in _dedupe_keep_order(list(supported_pack.get("detractors") or []) + list(generic_seed_pack.get("detractors") or []))
    ]

    ai_del_items = _parse_side(data.get("delighters") or [])
    ai_det_items = _parse_side(data.get("detractors") or [])

    prioritized_dels = _prioritize_ai_taxonomy_items(
        seeded_delighters + ai_del_items,
        side="delighter",
        sample_reviews=sample_reviews,
        category=category,
        min_review_hits=min_hits,
        max_keep=36,
        exclude_universal=True,
    )
    prioritized_dets = _prioritize_ai_taxonomy_items(
        seeded_detractors + ai_det_items,
        side="detractor",
        sample_reviews=sample_reviews,
        category=category,
        min_review_hits=min_hits,
        max_keep=42,
        exclude_universal=True,
    )

    merged_aliases = {}
    for item in prioritized_dels + prioritized_dets:
        label = _safe_text(item.get("label"))
        aliases = [str(v).strip() for v in (item.get("aliases") or []) if str(v).strip()]
        if label and aliases:
            merged_aliases[label] = aliases

    canon_dels, canon_dets, canon_aliases = _standardize_symptom_lists(
        [item.get("label") for item in prioritized_dels],
        [item.get("label") for item in prioritized_dets],
    )
    alias_map = _alias_map_for_catalog(
        canon_dels,
        canon_dets,
        extra_aliases=_merge_taxonomy_alias_maps(
            canon_aliases,
            merged_aliases,
            supported_pack.get("aliases", {}),
            generic_seed_pack.get("aliases", {}),
        ),
    )

    def _finalize_preview(items, *, side):
        lookup = {str(item.get("label") or ""): dict(item) for item in items}
        ordered = canon_dels if side == "delighter" else canon_dets
        preview_rows = []
        for label in ordered:
            base = dict(lookup.get(label, {}))
            base.setdefault("label", label)
            base["bucket"] = base.get("bucket") or _bucket_taxonomy_label(label, side=side, category=category)
            base["aliases"] = list(alias_map.get(label, base.get("aliases") or []))
            base.setdefault("family", "")
            base["l1_theme"] = _infer_taxonomy_l1_theme(label, side=side, family=base.get("family"), theme=base.get("theme") or base.get("l1_theme"), category=category)
            base["side"] = side
            base.setdefault("rationale", "")
            base.setdefault("review_hits", 0)
            base.setdefault("support_ratio", 0.0)
            base.setdefault("score", 0.0)
            base.setdefault("specificity", 0.0)
            base.setdefault("examples", [])
            preview_rows.append(base)
        return preview_rows[:36]

    preview_dels = _finalize_preview(prioritized_dels, side="delighter")
    preview_dets = _finalize_preview(prioritized_dets, side="detractor")

    note = str(data.get("notes", "")).strip()
    if not note:
        note = (
            f"Detected category: {category}. Universal neutral labels were held out, CSAT and outcome drivers were seeded from product knowledge plus review evidence, "
            f"and product-specific labels were ranked by review support to keep the catalog systematic and useful."
        )
    return dict(
        delighters=list(canon_dels)[:36],
        detractors=list(canon_dets)[:42],
        aliases=alias_map,
        category=category,
        category_confidence=category_info.get("confidence", 0.0),
        taxonomy_note=note,
        preview_delighters=preview_dels,
        preview_detractors=preview_dets,
        product_knowledge=product_knowledge,
    )


def _ai_generate_product_description(*, client, sample_reviews, existing_description=""):
    _log.info("Generating product description from %d sample reviews (desc length: %d)", len(sample_reviews or []), len(existing_description or ""))
    if not client:
        raise RuntimeError("OpenAI client is None — check API key configuration")
    if not sample_reviews:
        raise ValueError("No sample reviews provided for product description generation")
    sys = textwrap.dedent("""
        You are a product marketing and consumer-insights analyst writing a concise product description from customer reviews.
        Use only facts and recurring capabilities clearly supported by the review sample.
        Do not invent capacities, dimensions, accessories, or retailer-specific claims not grounded in the reviews.
        Write 2-4 concise sentences describing what the product is, what it is mainly used for, and the most repeated strengths or caveats.
        Also extract structured product knowledge that will help downstream symptom generation create a sharper, more CSAT-driven first-pass taxonomy for ANY consumer product.
        Focus on the outcomes customers are trying to achieve, the workflow steps that matter, likely failure modes, comparison benchmarks, and user contexts.
        OUTPUT — strict JSON only:
        {
          "description":"<2-4 sentence product description>",
          "confidence_note":"<short note on confidence>",
          "product_archetype":"<short archetype such as wireless audio, vacuum floorcare, coffee espresso, drinkware, skincare topical, footwear>",
          "product_areas":["<component, workflow area, or product part>"],
          "use_cases":["<main use case or job to be done>"],
          "desired_outcomes":["<customer outcome the product should deliver>"],
          "comparison_set":["<brand, premium alternative, or comparison set>"],
          "workflow_steps":["<important step in setup, use, maintenance, or switching workflow>"],
          "user_contexts":["<skill level, environment, hair type, skin type, family size, commute context, etc.>"],
          "csat_drivers":["<core reason a customer would be satisfied or dissatisfied>"],
          "likely_failure_modes":["<important failure mode or friction point>"],
          "likely_themes":["<probable first-pass theme>"],
          "likely_delighter_themes":["<positive theme>"],
          "likely_detractor_themes":["<negative theme or failure mode>"],
          "watchouts":["<important caution, limitation, or risk>" ]
        }
    """).strip()
    compact_reviews = [_trunc(str(review or "").strip(), 420) for review in list(sample_reviews or [])[:60] if str(review or "").strip()]
    payload = dict(existing_description=existing_description or "", sample_reviews=compact_reviews)
    result_text = _chat_complete_with_fallback_models(
        client,
        model=_shared_model(),
        structured=True,
        messages=[{"role": "system", "content": sys}, {"role": "user", "content": json.dumps(payload)}],
        temperature=0.0,
        response_format={"type": "json_object"},
        max_tokens=1800,
        reasoning_effort=_shared_reasoning(),
    )
    data = _safe_json_load(result_text)
    product_knowledge = _normalize_product_knowledge(data)
    return {
        "description": _safe_text(data.get("description")),
        "confidence_note": _safe_text(data.get("confidence_note")),
        "product_knowledge": product_knowledge,
    }


def _gen_symptomized_workbook(original_bytes, updated_df):
    wb = load_workbook(io.BytesIO(original_bytes))
    sheet_name = _best_uploaded_excel_sheet_name(original_bytes)
    if not sheet_name or sheet_name not in wb.sheetnames:
        sheet_name = "Star Walk scrubbed verbatims" if "Star Walk scrubbed verbatims" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]
    df2 = _ensure_ai_cols(updated_df.copy())

    def _canon_review_key(value):
        if value is None or (isinstance(value, float) and pd.isna(value)):
            return ""
        if isinstance(value, (int, np.integer)):
            return str(int(value))
        if isinstance(value, (float, np.floating)) and not pd.isna(value) and float(value).is_integer():
            return str(int(value))
        text = str(value).strip()
        if re.fullmatch(r"-?\d+\.0+", text):
            text = text.split(".", 1)[0]
        return text

    def _header_lookup(sheet):
        out = {}
        for col_idx in range(1, sheet.max_column + 1):
            header = _safe_text(sheet.cell(row=1, column=col_idx).value).strip()
            if header and header.lower() not in out:
                out[header.lower()] = col_idx
        return out

    header_map = _header_lookup(ws)

    def _ensure_ai_column(header_name, *, fallback_index=None):
        lower = header_name.lower()
        if lower in header_map:
            return header_map[lower]
        if fallback_index is not None:
            existing = _safe_text(ws.cell(row=1, column=fallback_index).value).strip()
            if not existing or existing.lower() == lower:
                ws.cell(row=1, column=fallback_index, value=header_name)
                header_map[lower] = fallback_index
                return fallback_index
        new_idx = max(ws.max_column, 0) + 1
        ws.cell(row=1, column=new_idx, value=header_name)
        header_map[lower] = new_idx
        return new_idx

    det_write_cols = [
        _ensure_ai_column(header, fallback_index=DET_INDEXES[j - 1])
        for j, header in enumerate(AI_DET_HEADERS, start=1)
    ]
    del_write_cols = [
        _ensure_ai_column(header, fallback_index=DEL_INDEXES[j - 1])
        for j, header in enumerate(AI_DEL_HEADERS, start=1)
    ]
    meta_write_cols = {
        "AI Safety": _ensure_ai_column("AI Safety", fallback_index=META_INDEXES["Safety"]),
        "AI Reliability": _ensure_ai_column("AI Reliability", fallback_index=META_INDEXES["Reliability"]),
        "AI # of Sessions": _ensure_ai_column("AI # of Sessions", fallback_index=META_INDEXES["# of Sessions"]),
    }

    review_id_col = next((header_map[a.lower()] for a in UPLOAD_REVIEW_ID_ALIASES if a.lower() in header_map), None)
    worksheet_rows_by_review_id = {}
    if review_id_col is not None:
        for row_idx in range(2, ws.max_row + 1):
            key = _canon_review_key(ws.cell(row=row_idx, column=review_id_col).value)
            if key and key not in worksheet_rows_by_review_id:
                worksheet_rows_by_review_id[key] = row_idx

    fg = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fr = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    fy = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    fb = PatternFill(start_color="CFE2F3", end_color="CFE2F3", fill_type="solid")
    fp = PatternFill(start_color="EAD1DC", end_color="EAD1DC", fill_type="solid")

    def _write_cell(row_idx, col_idx, value, fill):
        clean = None if (pd.isna(value) or str(value).strip() == "") else str(value).strip()
        cell = ws.cell(row=row_idx, column=col_idx, value=clean)
        cell.fill = fill if clean else PatternFill()

    for seq_row_idx, (_, row) in enumerate(df2.iterrows(), start=2):
        target_row = None
        if review_id_col is not None:
            review_key = _canon_review_key(row.get("review_id"))
            if review_key and review_key in worksheet_rows_by_review_id:
                target_row = worksheet_rows_by_review_id[review_key]
            elif review_key:
                target_row = ws.max_row + 1
                worksheet_rows_by_review_id[review_key] = target_row
                ws.cell(row=target_row, column=review_id_col, value=review_key)
        if target_row is None:
            target_row = seq_row_idx
        for j, ci in enumerate(det_write_cols, 1):
            _write_cell(target_row, ci, row.get(f"AI Symptom Detractor {j}"), fr)
        for j, ci in enumerate(del_write_cols, 1):
            _write_cell(target_row, ci, row.get(f"AI Symptom Delighter {j}"), fg)
        _write_cell(target_row, meta_write_cols["AI Safety"], row.get("AI Safety"), fy)
        _write_cell(target_row, meta_write_cols["AI Reliability"], row.get("AI Reliability"), fb)
        _write_cell(target_row, meta_write_cols["AI # of Sessions"], row.get("AI # of Sessions"), fp)
    for c in det_write_cols + del_write_cols + list(meta_write_cols.values()):
        try:
            ws.column_dimensions[get_column_letter(c)].width = 28
        except Exception:
            pass
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def _dedup_candidates(raw):
    def _norm(s):
        s = s.strip().lower()
        s = re.sub(r"^(not\s+too\s+|not\s+very\s+|not\s+overly\s+|not\s+)", "", s)
        s = re.sub(r"[^a-z0-9 ]", " ", s)
        return re.sub(r"\s+", " ", s).strip()

    labels = sorted(raw.keys(), key=lambda l: -int(raw[l].get("count", 0)))
    merged = {}
    used = set()
    for a in labels:
        if a in used:
            continue
        merged[a] = dict(raw[a])
        used.add(a)
        na = _norm(a)
        for b in labels:
            if b in used or b == a:
                continue
            nb = _norm(b)
            if difflib.SequenceMatcher(None, na, nb).ratio() >= 0.72 or na in nb or nb in na:
                merged[a]["count"] = int(merged[a].get("count", 0)) + int(raw[b].get("count", 0))
                refs = list(merged[a].get("refs", []))
                for r in raw[b].get("refs", []):
                    if r not in refs and len(refs) < 50:
                        refs.append(r)
                merged[a]["refs"] = refs
                merged[a].setdefault("_merged_from", []).append(b)
                used.add(b)
    return merged


def _record_new_symptom_candidate(label, *, idx, side):
    lab = _safe_text(label).strip()
    if not lab:
        return
    bucket = st.session_state.setdefault("sym_new_candidates", {}).setdefault(
        lab,
        {"count": 0, "refs": [], "delighter_count": 0, "detractor_count": 0, "delighter_refs": [], "detractor_refs": []},
    )
    bucket["count"] = int(bucket.get("count", 0)) + 1
    if idx not in bucket.get("refs", []) and len(bucket.get("refs", [])) < 50:
        bucket.setdefault("refs", []).append(idx)
    side_key = "delighter" if str(side).lower().startswith("del") else "detractor"
    count_key = f"{side_key}_count"
    refs_key = f"{side_key}_refs"
    bucket[count_key] = int(bucket.get(count_key, 0)) + 1
    if idx not in bucket.get(refs_key, []) and len(bucket.get(refs_key, [])) < 50:
        bucket.setdefault(refs_key, []).append(idx)


def _candidate_rows_for_side(raw_candidates, *, side):
    side_key = "delighter" if str(side).lower().startswith("del") else "detractor"
    other_key = "detractor" if side_key == "delighter" else "delighter"
    category = str(st.session_state.get("sym_taxonomy_category") or (st.session_state.get("sym_ai_build_result") or {}).get("category") or "general")
    side_raw = {}
    for label, payload in (raw_candidates or {}).items():
        clean_label = _safe_text(label).strip()
        if not clean_label:
            continue
        if side_key == "delighter":
            canon_labels, _, _ = _standardize_symptom_lists([clean_label], [])
            canonical_label = canon_labels[0] if canon_labels else clean_label.title()
        else:
            _, canon_labels, _ = _standardize_symptom_lists([], [clean_label])
            canonical_label = canon_labels[0] if canon_labels else clean_label.title()
        if f"{side_key}_count" in payload or f"{other_key}_count" in payload:
            count = int(payload.get(f"{side_key}_count", 0) or 0)
            refs = list(payload.get(f"{side_key}_refs", []) or [])
            other_count = int(payload.get(f"{other_key}_count", 0) or 0)
        else:
            count = int(payload.get("count", 0) or 0)
            refs = list(payload.get("refs", []) or [])
            other_count = 0
        if count <= 0:
            continue
        bucket = side_raw.setdefault(canonical_label, {"count": 0, "refs": [], "other_count": 0, "raw_labels": []})
        bucket["count"] = int(bucket.get("count", 0)) + count
        bucket["other_count"] = int(bucket.get("other_count", 0)) + other_count
        bucket["raw_labels"] = _normalize_tag_list(list(bucket.get("raw_labels", [])) + [clean_label])
        existing_refs = list(bucket.get("refs", []))
        for ref in refs:
            if ref not in existing_refs and len(existing_refs) < 50:
                existing_refs.append(ref)
        bucket["refs"] = existing_refs
    deduped = _dedup_candidates(side_raw) if side_raw else {}
    rows = []
    for label, payload in sorted(deduped.items(), key=lambda kv: (-int(kv[1].get("count", 0)), kv[0])):
        merged_from = payload.get("_merged_from", []) or []
        notes = []
        raw_labels = [lab for lab in payload.get("raw_labels", []) or [] if lab and lab != label]
        if int(payload.get("other_count", 0) or 0) > 0:
            notes.append(f"also seen as {other_key}: {int(payload.get('other_count', 0))}")
        if raw_labels:
            notes.append("normalized from: " + ", ".join(raw_labels[:3]))
        if merged_from:
            notes.append("merged from: " + ", ".join(merged_from[:3]))
        rows.append({
            "L1 Theme": _infer_taxonomy_l1_theme(label, side=side_key, category=category),
            "Label": label,
            "Bucket": _bucket_taxonomy_label(label, side=side_key, category=category),
            "Count": int(payload.get("count", 0) or 0),
            "Refs": ", ".join(str(r) for r in list(payload.get("refs", []) or [])[:5]),
            "Notes": " · ".join(notes),
        })
    return rows


def _try_load_symptoms_from_file():
    raw = st.session_state.get("_uploaded_raw_bytes")
    if not raw:
        return False
    d, t, a = _get_symptom_whitelists(raw)
    if d or t:
        loaded_dels, loaded_dets = _canonical_symptom_catalog(d, t)
        st.session_state.update(sym_delighters=loaded_dels, sym_detractors=loaded_dets, sym_aliases=_alias_map_for_catalog(loaded_dels, loaded_dets, extra_aliases=a, existing_aliases=st.session_state.get("sym_aliases", {})), sym_symptoms_source="file", sym_taxonomy_preview_items=[], sym_taxonomy_category="general")
        return True
    return False


def _local_symptom_catalog(df):
    if df is None or df.empty:
        return [], []
    det_vals, del_vals, _, _ = _symptom_filter_options(df)
    if not del_vals and not det_vals:
        return [], []
    del_vals, det_vals = _canonical_symptom_catalog(list(del_vals), list(det_vals))
    return list(del_vals), list(det_vals)

# ═══════════════════════════════════════════════════════════════════════════════
#  SESSION STATE
# ═══════════════════════════════════════════════════════════════════════════════
def _auto_discover_product(dataset, *, max_sample=20):
    """Phase 1: Auto-discover product description + knowledge from reviews.
    Runs automatically after workspace build — no user action needed."""
    client = _get_client()
    if not client:
        return
    reviews_df = dataset.get("reviews_df", pd.DataFrame())
    if reviews_df.empty:
        return
    try:
        # Sample reviews stratified by rating
        sample = _sample_reviews_for_symptomizer(reviews_df, max_sample)
        if not sample:
            return
        # Auto-generate product description
        existing_desc = st.session_state.get("sym_product_profile", "")
        if not existing_desc:
            result = _ai_generate_product_description(
                client=client, sample_reviews=sample, existing_description=""
            )
            if result and result.get("description"):
                st.session_state["sym_product_profile"] = result["description"]
                st.session_state["sym_product_profile_ai_note"] = result.get("confidence_note", "Auto-generated on workspace build.")
                _log.info("Auto-discovered product description: %s", result["description"][:80])
        # Auto-generate product knowledge
        desc = st.session_state.get("sym_product_profile", "")
        if desc and not st.session_state.get("sym_product_knowledge"):
            category_info = _infer_taxonomy_category(desc, sample[:12])
            category = category_info.get("category", "general")
            st.session_state["sym_taxonomy_category"] = category
            # Build knowledge from the description + sample
            knowledge = {
                "product_archetype": _infer_generic_archetype(desc, {}),
                "confidence_note": "Auto-generated from review sample on workspace build.",
            }
            # Use the AI to fill in structured knowledge
            try:
                knowledge_result = _ai_generate_product_description(
                    client=client, sample_reviews=sample, existing_description=desc
                )
                if knowledge_result:
                    for key in ["product_areas", "desired_outcomes", "likely_failure_modes",
                                "workflow_steps", "use_cases", "comparison_set", "user_contexts",
                                "csat_drivers", "watchouts", "likely_delighter_themes", "likely_detractor_themes"]:
                        if knowledge_result.get(key):
                            knowledge[key] = knowledge_result[key]
            except Exception:
                pass
            st.session_state["sym_product_knowledge"] = _normalize_product_knowledge(knowledge)
            _log.info("Auto-discovered product knowledge for category: %s", category)
    except Exception as exc:
        _log.warning("Auto-discovery failed: %s", exc)


def _batch_html(*blocks):
    """Render multiple HTML blocks in a single st.markdown call to reduce Streamlit rerenders."""
    combined = "\n".join(str(b) for b in blocks if b)
    if combined.strip():
        st.markdown(combined, unsafe_allow_html=True)


def _init_state():
    defaults = dict(
        analysis_dataset=None,
        chat_messages=[],
        master_export_bundle=None,
        prompt_definitions_df=_default_prompt_df(),
        prompt_builder_suggestion=None,
        prompt_run_artifacts=None,
        prompt_run_notice=None,
        chat_scope_signature=None,
        chat_scope_notice=None,
        review_explorer_page=1,
        review_explorer_per_page=20,
        review_explorer_sort="Newest",  # Options: Newest, Oldest, Lowest rated, Highest rated, Longest, Most tagged, Least tagged
        review_filter_signature=None,
        shared_model=DEFAULT_MODEL,
        shared_reasoning=DEFAULT_REASONING,
        ai_response_preset="Large (1200 words)",
        ai_response_words=1200,
        ai_include_references=False,
        ot_show_volume=False,
        workspace_source_mode=SOURCE_MODE_URL,
        workspace_product_url=DEFAULT_PRODUCT_URL,
        workspace_product_urls_bulk="",
        workspace_file_uploader_nonce=0,
        workspace_include_local_symptomization=True,
        workspace_active_tab=TAB_DASHBOARD,
        workspace_tab_request=None,
        ai_scroll_to_top=False,
        sym_delighters=[],
        sym_detractors=[],
        sym_custom_universal_delighters=[],
        sym_custom_universal_detractors=[],
        sym_aliases={},
        sym_taxonomy_preview_items=[],
        sym_taxonomy_category="general",
        sym_symptoms_source="none",
        sym_processed_rows=[],
        sym_new_candidates={},
        sym_product_profile="",
        sym_product_knowledge={},
        sym_include_universal_neutral=True,
        sym_scope_choice="Missing both",
        sym_n_to_process=10,
        sym_batch_size=8,
        sym_max_ev_chars=120,
        sym_review_log_limit=50,
        sym_run_notice=None,
        sym_qa_baseline_map={},
        sym_qa_accuracy={},
        sym_qa_user_edited=False,
        sym_qa_row_ids=[],
        sym_qa_selected_row=None,
        sym_qa_notice=None,
        sym_product_profile_ai_note="",
        sym_calibration_result=None,
        sym_staged_pipeline=False,
        sym_v4_pipeline=True,
        sym_last_run_stats=None,
        sidebar_manual_api_key="",
        workspace_name="",
        workspace_id=None,
        _ws_show_rename=False,
        _sym_inline_defaults_pending={},
        _prompt_defs_cache={},
        _prompt_bundle_ready=False,
    )
    for k, v in defaults.items():
        st.session_state.setdefault(k, v)


def _reset_workspace_state(*, reset_source=True):
    st.session_state["analysis_dataset"] = None
    st.session_state["chat_messages"] = []
    st.session_state["chat_scope_signature"] = None
    st.session_state["chat_scope_notice"] = None
    st.session_state["master_export_bundle"] = None
    st.session_state["prompt_run_artifacts"] = None
    st.session_state["prompt_run_notice"] = None
    st.session_state["review_explorer_page"] = 1
    st.session_state["workspace_active_tab"] = TAB_DASHBOARD
    st.session_state["workspace_tab_request"] = None
    st.session_state["ai_scroll_to_top"] = False
    st.session_state["sym_processed_rows"] = []
    st.session_state["sym_new_candidates"] = {}
    st.session_state["sym_product_profile"] = ""
    st.session_state["sym_product_knowledge"] = {}
    st.session_state["sym_pdesc"] = ""
    st.session_state["sym_include_universal_neutral"] = True
    st.session_state["sym_run_notice"] = None
    st.session_state["sym_review_log_limit"] = 50
    st.session_state["sym_symptoms_source"] = "none"
    st.session_state["sym_delighters"] = []
    st.session_state["sym_detractors"] = []
    st.session_state["sym_aliases"] = {}
    st.session_state["sym_taxonomy_preview_items"] = []
    st.session_state["sym_taxonomy_category"] = "general"
    st.session_state["sym_qa_baseline_map"] = {}
    st.session_state["sym_qa_accuracy"] = {}
    st.session_state["sym_qa_user_edited"] = False
    st.session_state["sym_qa_row_ids"] = []
    st.session_state["sym_qa_selected_row"] = None
    st.session_state["sym_qa_notice"] = None
    st.session_state["sym_product_profile_ai_note"] = ""
    st.session_state.pop("_sym_pdesc_pending", None)
    st.session_state.pop("_sym_inline_defaults_pending", None)
    st.session_state["_uploaded_raw_bytes"] = None
    st.session_state["sym_export_bytes"] = None
    st.session_state["_prompt_bundle_ready"] = False
    st.session_state["ai_include_references"] = False
    st.session_state["ot_show_volume"] = False
    st.session_state.pop("sym_ai_build_result", None)
    _reset_review_filters()
    if reset_source:
        st.session_state["workspace_source_mode"] = SOURCE_MODE_URL
        st.session_state["workspace_product_url"] = DEFAULT_PRODUCT_URL
        st.session_state["workspace_product_urls_bulk"] = ""
        st.session_state["workspace_file_uploader_nonce"] = int(st.session_state.get("workspace_file_uploader_nonce", 0)) + 1


_init_state()
# ═══════════════════════════════════════════════════════════════════════════════
#  SIDEBAR
# ═══════════════════════════════════════════════════════════════════════════════
def _render_sidebar(df: Optional[pd.DataFrame]):
    api_key = _get_api_key()
    filter_state = {"filtered_df": df.copy() if isinstance(df, pd.DataFrame) else pd.DataFrame(), "active_items": [], "filter_seconds": 0.0, "description": "No active filters"}
    with st.sidebar:
        current_tab = st.session_state.get("workspace_active_tab", TAB_DASHBOARD)
        st.markdown("### 🔍 Review Filters")
        st.caption("Applies live to every workspace tab.")
        if st.button("🧹 Clear all filters", use_container_width=True, key="rf_clear_btn"):
            _reset_review_filters()
            st.rerun()
        if df is None:
            st.info("Build a workspace to unlock filters.")
        else:
            core_specs = _core_filter_specs_for_df(df)
            det_opts, del_opts, _, _ = _symptom_filter_options(df)
            st.markdown(
                f"<div class='sidebar-scope-card'><div class='sidebar-scope-title'>Loaded workspace</div><div class='sidebar-scope-value'>{len(df):,} reviews · {len(core_specs):,} core filters{' · symptoms ready' if (det_opts or del_opts) else ''}</div></div>",
                unsafe_allow_html=True,
            )
            with st.expander("🗓️ Timeframe", expanded=False):
                tf_opts = ["All Time", "Last Week", "Last Month", "Last Year", "Custom Range"]
                if st.session_state.get("rf_tf") not in tf_opts:
                    st.session_state["rf_tf"] = "All Time"
                st.selectbox("Select timeframe", options=tf_opts, key="rf_tf")
                if st.session_state.get("rf_tf") == "Custom Range":
                    today = date.today()
                    rng = st.session_state.get("rf_tf_range", (today - timedelta(days=30), today))
                    if not (isinstance(rng, (tuple, list)) and len(rng) == 2):
                        rng = (today - timedelta(days=30), today)
                    st.session_state["rf_tf_range"] = tuple(rng)
                    st.date_input("Start / end", value=st.session_state["rf_tf_range"], key="rf_tf_range")
            with st.expander("⭐ Star rating", expanded=False):
                sr_opts = ["All", 5, 4, 3, 2, 1]
                cur = st.session_state.get("rf_sr", ["All"])
                if not isinstance(cur, list):
                    cur = [cur]
                cur = [v for v in cur if v in sr_opts]
                if not cur:
                    cur = ["All"]
                if "All" in cur and len(cur) > 1:
                    cur = [v for v in cur if v != "All"]
                st.session_state["rf_sr"] = cur
                st.multiselect("Select stars", options=sr_opts, default=st.session_state["rf_sr"], key="rf_sr")
            with st.expander("🧭 Review filters", expanded=False):
                st.caption("")
                for spec in core_specs:
                    key = f"rf_{spec['key']}"
                    _sanitize_multiselect(key, spec["options"], ["ALL"])
                    st.multiselect(spec["label"], options=spec["options"], default=st.session_state[key], key=key)
            if det_opts or del_opts:
                with st.expander("🩺 Symptom filters", expanded=False):
                    st.caption("Only shown when symptom tags are present in the workspace.")
                    if det_opts:
                        det_all = ["All"] + det_opts
                        _sanitize_multiselect_sym("rf_sym_detract", det_all, ["All"])
                        st.multiselect("Detractors", options=det_all, default=st.session_state["rf_sym_detract"], key="rf_sym_detract")
                    if del_opts:
                        del_all = ["All"] + del_opts
                        _sanitize_multiselect_sym("rf_sym_delight", del_all, ["All"])
                        st.multiselect("Delighters", options=del_all, default=st.session_state["rf_sym_delight"], key="rf_sym_delight")
            with st.expander("🔎 Keyword", expanded=False):
                st.text_input("Search in title + review text", value=st.session_state.get("rf_kw", ""), key="rf_kw", placeholder="e.g. loud noise, filter cleaning, cord length")
            extra_candidates = _extra_filter_candidates(df)
            current_extra = [c for c in (st.session_state.get("rf_extra_filter_cols", []) or []) if c in extra_candidates]
            st.session_state["rf_extra_filter_cols"] = current_extra
            with st.expander("➕ Add Filters", expanded=False):
                st.caption("Suggested from the current workspace schema so this list adapts to each uploaded file.")
                st.multiselect("Available columns from this workspace", options=extra_candidates, default=current_extra, key="rf_extra_filter_cols")
            extra_cols = st.session_state.get("rf_extra_filter_cols", []) or []
            if extra_cols:
                with st.expander("🧩 Extra filters", expanded=False):
                    for col in extra_cols:
                        if col not in df.columns:
                            continue
                        kind = _infer_extra_filter_kind(df, col)
                        s = df[col]
                        if kind == "numeric":
                            num = pd.to_numeric(s, errors="coerce").dropna()
                            if num.empty:
                                continue
                            lo, hi = float(num.min()), float(num.max())
                            if lo == hi:
                                st.caption(f"{col}: {lo:g} (constant)")
                                continue
                            key = f"rf_{col}_range"
                            default = st.session_state.get(key, (lo, hi))
                            if not (isinstance(default, (tuple, list)) and len(default) == 2):
                                default = (lo, hi)
                            st.session_state[key] = (float(default[0]), float(default[1]))
                            st.slider(col, min_value=lo, max_value=hi, value=st.session_state[key], key=key)
                        elif kind == "date":
                            dt = pd.to_datetime(s, errors="coerce").dropna()
                            if dt.empty:
                                continue
                            lo, hi = dt.min().date(), dt.max().date()
                            key = f"rf_{col}_date_range"
                            default = st.session_state.get(key, (lo, hi))
                            if not (isinstance(default, (tuple, list)) and len(default) == 2):
                                default = (lo, hi)
                            st.session_state[key] = tuple(default)
                            st.date_input(col, value=st.session_state[key], min_value=lo, max_value=hi, key=key)
                        else:
                            try:
                                nunique = int(s.astype("string").replace({"": pd.NA}).nunique(dropna=True))
                            except Exception:
                                nunique = 0
                            if nunique > 600:
                                st.text_input(f"{col} contains", value=str(st.session_state.get(f"rf_{col}_contains") or ""), key=f"rf_{col}_contains", help="High-cardinality column — using a contains filter for speed.")
                            else:
                                opts = _col_options(df, col, max_vals=None)
                                _sanitize_multiselect(f"rf_{col}", opts, ["ALL"])
                                st.multiselect(col, options=opts, default=st.session_state[f"rf_{col}"], key=f"rf_{col}")
            filter_state = _apply_live_review_filters(df)
        with st.expander("🤖 AI Model & Symptomizer", expanded=False):
            st.caption("Use higher reasoning for ELT-ready analysis and more nuanced summaries.")
            cur_model = st.session_state.get("shared_model", DEFAULT_MODEL)
            if cur_model not in MODEL_OPTIONS:
                cur_model = DEFAULT_MODEL
                st.session_state["shared_model"] = cur_model
            st.selectbox("Model", options=MODEL_OPTIONS, index=MODEL_OPTIONS.index(cur_model), key="shared_model", help="Used by AI Analyst, Review Prompt, and Symptomizer.")
            effort_options = _reasoning_options_for_model(st.session_state.get("shared_model", DEFAULT_MODEL))
            cur_reasoning = _safe_text(st.session_state.get("shared_reasoning", DEFAULT_REASONING)).lower() or DEFAULT_REASONING
            if cur_reasoning not in effort_options:
                cur_reasoning = "none" if "none" in effort_options else effort_options[0]
                st.session_state["shared_reasoning"] = cur_reasoning
            st.selectbox("Reasoning effort", options=effort_options, index=effort_options.index(cur_reasoning), key="shared_reasoning", help="Applied to GPT-5 family models. Raising this usually improves quality on nuanced and long-form analysis, but may be a bit slower.")
            key_source = _api_key_source()
            if key_source in ("secrets", "env"):
                st.markdown("<div class='helper-chip-row'><span class='helper-chip' style='background:rgba(5,150,105,.10);color:#059669;border-color:rgba(5,150,105,.25);'>✅ Key loaded</span><span class='helper-chip'>Higher reasoning = higher quality</span></div>", unsafe_allow_html=True)
            elif key_source == "manual":
                st.markdown("<div class='helper-chip-row'><span class='helper-chip' style='background:rgba(217,119,6,.10);color:#d97706;border-color:rgba(217,119,6,.25);'>🔑 Using manual key</span></div>", unsafe_allow_html=True)
            else:
                st.warning("No API key detected. Paste one below or set OPENAI_API_KEY in secrets.")
            if key_source in ("missing", "manual"):
                st.text_input("OpenAI API key", type="password", placeholder="sk-proj-...", key="sidebar_manual_api_key", help="Paste your OpenAI API key. Only stored in your browser session.")
                api_key = _get_api_key()
            st.markdown("<div style='height:.25rem'></div>", unsafe_allow_html=True)
            st.slider("Symptomizer batch size", 1, 12, key="sym_batch_size")
            st.slider("Symptomizer max evidence chars", 60, 200, step=10, key="sym_max_ev_chars")
            if _HAS_SYMPTOMIZER_V3:
                st.toggle(
                    "Use Symptomizer v4 pipeline",
                    value=bool(st.session_state.get("sym_v4_pipeline", True)),
                    key="sym_v4_pipeline",
                    help="Default on. Uses rich taxonomy guidance, long-review claim extraction, polarity audits, taxonomy hygiene, and targeted verification for riskier results.",
                )
        st.divider()
        st.markdown("""<div class='sidebar-scope-card sidebar-scope-card--feature'>
          <div class='sidebar-scope-title'>Beta feature</div>
          <div class='sidebar-scope-value'>Open the Social Listening beta route when you want to preview the placeholder Meltwater-style workflow and five-module analysis experience.</div>
        </div>""", unsafe_allow_html=True)
        social_btn_kwargs = {"use_container_width": True, "key": "sidebar_open_social_beta"}
        if current_tab == TAB_SOCIAL_LISTENING:
            social_btn_kwargs["type"] = "primary"
        if st.button("📣 Open Social Listening Beta", **social_btn_kwargs):
            st.session_state["workspace_active_tab"] = TAB_SOCIAL_LISTENING
            st.session_state["workspace_tab_request"] = TAB_SOCIAL_LISTENING
            st.rerun()
    return {"api_key": api_key, "review_filters": filter_state}

# ═══════════════════════════════════════════════════════════════════════════════
#  RENDER HELPERS
# ═══════════════════════════════════════════════════════════════════════════════
def _render_metric_card(label, value, subtext, accent=False):
    cls = "metric-card accent" if accent else "metric-card"
    st.markdown(f"""<div class="{cls}">
      <div class="metric-label">{label}</div>
      <div class="metric-value">{value}</div>
      <div class="metric-sub">{subtext}</div>
    </div>""", unsafe_allow_html=True)


def _render_workspace_header(summary, overall_df, filtered_df, prompt_artifacts, *, source_type, source_label, filter_description, active_items):
    export_df = filtered_df.copy()
    bundle, bundle_error = _safe_get_master_bundle(
        summary,
        export_df,
        prompt_artifacts,
        active_items=active_items,
        filter_description=filter_description,
        export_scope_label="Filtered current view" if active_items else "Current view (all reviews)",
        total_loaded_reviews=len(overall_df),
    )
    product_name = _product_name(summary, overall_df)
    summary_label = _safe_summary_product_label(summary)
    organic = int((~overall_df["incentivized_review"].fillna(False)).sum()) if not overall_df.empty else 0
    n = len(overall_df)
    view_count = len(filtered_df)
    if source_type == "uploaded":
        src_chip = f"Uploaded · {source_label}"
    elif source_type == "multi-url":
        src_chip = f"Multi-link batch · {source_label}"
    elif source_type == "powerreviews":
        src_chip = f"{(source_label or 'PowerReviews')} · {summary_label}"
    elif source_type == "bazaarvoice":
        src_chip = f"{(source_label or 'Bazaarvoice')} · {summary_label}"
    else:
        src_chip = f"{(source_label or str(source_type).title())} · {summary_label}"
    st.markdown(f"""<div class="hero-card">
      <div style="display:flex;justify-content:space-between;align-items:flex-start;gap:16px;flex-wrap:wrap;">
        <div>
          <div class="hero-kicker">Review workspace · Beta</div>
          <div class="hero-title">{_esc(product_name)}</div>
        </div>
        <div class="badge-row">
          <span class="chip gray">{_esc(src_chip)}</span>
          <span class="chip yellow">Beta</span>
          <span class="chip blue">{'Filtered view' if active_items else 'All reviews'}</span>
          <span class="chip indigo">{n:,} reviews</span>
          <span class="chip green">{organic:,} organic</span>
        </div>
      </div>
    </div>""", unsafe_allow_html=True)
    a0, a1, a2 = st.columns([1.35, 1.15, 4])
    if bundle is not None:
        a0.download_button("⬇️ Download current view", data=bundle["excel_bytes"], file_name=bundle["excel_name"], mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
    else:
        a0.button("⬇️ Download current view", use_container_width=True, disabled=True, key="ws_dl_disabled")
    if a1.button("🔄 Reset workspace", use_container_width=True):
        _reset_workspace_state(reset_source=True)
        st.rerun()
    if bundle_error is not None:
        a2.caption(f"Download export is temporarily unavailable: {bundle_error}")
    else:
        a2.caption(f"Download includes {view_count:,} row(s) from the current view, a FilterCriteria sheet, Rating Distribution, Volume trend, and any AI prompt or Symptomizer columns.")


def _render_top_metrics(overall_df, filtered_df):
    m = _get_metrics(filtered_df)
    recommend_value = _fmt_pct(m.get("recommend_rate")) if m.get("recommend_rate") is not None else "n/a"
    recommend_sub = "Share of reviews with recommendation data" if m.get("recommend_rate") is not None else "No recommendation field available"
    organic_share = max(0.0, 1.0 - float(m.get("pct_incentivized") or 0.0))
    cards = [
        ("Reviews in view", f"{m['review_count']:,}", f"of {len(overall_df):,} loaded", False),
        ("Avg rating", _fmt_num(m["avg_rating"]), f"Organic avg {_fmt_num(m['avg_rating_non_incentivized'])}", False),
        ("Recommend rate", recommend_value, recommend_sub, False),
        ("% 1-2 star", _fmt_pct(m["pct_low_star"]), f"{m['low_star_count']:,} low-star reviews", True),
        ("% organic", _fmt_pct(organic_share), f"{m['non_incentivized_count']:,} organic reviews", False),
    ]
    cols = st.columns(len(cards))
    for col, (label, value, sub, acc) in zip(cols, cards):
        with col:
            _render_metric_card(label, value, sub, accent=acc)


_REVIEW_REF_PATTERN = re.compile(r"\(review_ids?\s*:\s*([^)]+)\)", flags=re.IGNORECASE)


def _reference_preview_rows(review_ids: Sequence[str], df: pd.DataFrame, max_items: int = 4) -> List[Dict[str, str]]:
    if df is None or df.empty:
        return []

    def _clean_rid(value: Any) -> str:
        rid = str(value or "").strip()
        rid = re.sub(r'^[`\'"\s]+', '', rid)
        rid = re.sub(r'[`\'"\s.,;:()\[\]{}]+$', '', rid)
        return rid.strip()

    lookup = df.copy()
    lookup["review_id"] = lookup["review_id"].astype(str)
    lookup["__rid_norm"] = lookup["review_id"].astype(str).str.strip().str.lower()
    lookup["__rid_simple"] = lookup["__rid_norm"].str.replace(r"[^a-z0-9]+", "", regex=True)
    out = []
    used = set()
    for rid in review_ids:
        cleaned = _clean_rid(rid)
        if not cleaned:
            continue
        rid_norm = cleaned.lower()
        if rid_norm in used:
            continue
        used.add(rid_norm)
        hit = lookup[lookup["__rid_norm"] == rid_norm]
        if hit.empty:
            rid_simple = re.sub(r"[^a-z0-9]+", "", rid_norm)
            if rid_simple:
                hit = lookup[lookup["__rid_simple"] == rid_simple]
        if hit.empty:
            hit = lookup[lookup["review_id"].astype(str).str.contains(re.escape(cleaned), case=False, na=False)].head(1)
        if hit.empty:
            continue
        row = hit.iloc[0]
        title = _safe_text(row.get("title"), "Untitled review") or "Untitled review"
        snippet = _trunc(_safe_text(row.get("review_text")) or _safe_text(row.get("title_and_text")), 220)
        meta = []
        if pd.notna(row.get("rating")):
            meta.append(f"★{_safe_int(row.get('rating'), 0)}")
        if _safe_text(row.get("submission_date")):
            meta.append(_safe_text(row.get("submission_date")))
        if _safe_text(row.get("content_locale")):
            meta.append(_safe_text(row.get("content_locale")))
        out.append({"meta": " · ".join(meta), "title": title, "snippet": snippet})
        if len(out) >= max_items:
            break
    return out


def _reference_tile_html_from_ids(review_ids: Sequence[str], df: pd.DataFrame, *, label: str = "Reference") -> str:
    ids = [str(x).strip() for x in review_ids if str(x).strip()]
    previews = _reference_preview_rows(ids, df)
    if not previews:
        raw_ids = ", ".join(ids[:4]) + ("…" if len(ids) > 4 else "")
        tip = f"<div class='ref-empty'>Referenced review preview not available in the loaded dataset. {_esc(raw_ids) if raw_ids else ''}</div>"
    else:
        bits = []
        for item in previews:
            bits.append(
                "<div class='ref-item'>"
                + (f"<div class='ref-meta'>{_esc(item['meta'])}</div>" if item.get("meta") else "")
                + f"<div class='ref-title'>{_esc(item['title'])}</div>"
                + f"<div class='ref-snippet'>{_esc(item['snippet'])}</div>"
                + "</div>"
            )
        extra = max(0, len(ids) - len(previews))
        if extra:
            bits.append(f"<div class='ref-item'><div class='ref-empty'>+{extra} more referenced review(s)</div></div>")
        tip = "".join(bits)
    return f"<span class='ref-wrap' tabindex='0' role='button' aria-label='{_esc(label)} review reference'><span class='ref-tile'>{_esc(label)}</span><span class='ref-tip'><span class='ref-tip-inner'>{tip}</span></span></span>"


def _normalize_ai_answer_display(text: str) -> str:
    raw = str(text or "")
    if not raw:
        return ""
    lines = []
    for line in raw.splitlines():
        if re.match(r"^\s{0,3}#{1,6}\s+", line):
            section = re.sub(r"^\s{0,3}#{1,6}\s+", "", line).strip()
            section = section.rstrip(":")
            lines.append(f"**{section}**" if section else "")
        else:
            lines.append(line)
    cleaned = "\n".join(lines)
    cleaned = re.sub(r"\n{3,}", "\n\n", cleaned)
    return cleaned.strip()


def _reference_tile_html_for_row(row) -> str:
    rid = _safe_text(row.get("review_id"))
    title = _safe_text(row.get("title"), "Untitled review") or "Untitled review"
    snippet = _trunc(_safe_text(row.get("review_text")) or _safe_text(row.get("title_and_text")), 220)
    meta = []
    if pd.notna(row.get("rating")):
        meta.append(f"★{_safe_int(row.get('rating'), 0)}")
    if _safe_text(row.get("submission_date")):
        meta.append(_safe_text(row.get("submission_date")))
    if _safe_text(row.get("content_locale")):
        meta.append(_safe_text(row.get("content_locale")))
    if rid:
        meta.append("Loaded review")
    tip = (
        "<div class='ref-item'>"
        + (f"<div class='ref-meta'>{_esc(' · '.join(meta))}</div>" if meta else "")
        + f"<div class='ref-title'>{_esc(title)}</div>"
        + f"<div class='ref-snippet'>{_esc(snippet)}</div>"
        + "</div>"
    )
    return f"<span class='ref-wrap'><span class='ref-tile'>Reference</span><span class='ref-tip'><span class='ref-tip-inner'>{tip}</span></span></span>"


def _replace_review_citations_with_reference_tiles(text: str, df: pd.DataFrame) -> str:
    normalized = _normalize_ai_answer_display(text)
    safe = html.escape(normalized, quote=False)
    def repl(match):
        raw = match.group(1)
        ids = [p.strip() for p in re.split(r"[,;\n|]+", raw) if p.strip()]
        if len(ids) <= 1:
            token_ids = [
                tok for tok in re.findall(r"[A-Za-z0-9_-]{3,}", raw)
                if tok.lower() not in {"review", "reviews", "reviewid", "reviewids", "review_id", "review_ids", "id", "ids"}
            ]
            if token_ids:
                ids = token_ids
        return _reference_tile_html_from_ids(ids, df, label="Reference")
    return _REVIEW_REF_PATTERN.sub(repl, safe)


def _render_markdown_with_reference_tiles(text: str, df: pd.DataFrame):
    processed = _replace_review_citations_with_reference_tiles(text, df)
    st.markdown(processed, unsafe_allow_html=True)


def _strip_review_citations(text: str) -> str:
    raw = _normalize_ai_answer_display(text)
    cleaned = _REVIEW_REF_PATTERN.sub("", raw)
    cleaned = re.sub(r"\s+([,.;:])", r"\1", cleaned)
    cleaned = re.sub(r"\n{3,}", "\n\n", cleaned)
    return cleaned.strip()


def _render_ai_response(text: str, df: pd.DataFrame, *, include_references: bool):
    if include_references:
        _render_markdown_with_reference_tiles(text, df)
    else:
        st.markdown(_strip_review_citations(text))


def _sort_reviews(df, sort_mode):
    w = df.copy()
    if sort_mode == "Newest":
        return w.sort_values(["submission_time", "review_id"], ascending=[False, False], na_position="last")
    if sort_mode == "Oldest":
        return w.sort_values(["submission_time", "review_id"], ascending=[True, True], na_position="last")
    if sort_mode == "Highest rating":
        return w.sort_values(["rating", "submission_time"], ascending=[False, False], na_position="last")
    if sort_mode == "Lowest rating":
        return w.sort_values(["rating", "submission_time"], ascending=[True, False], na_position="last")
    if sort_mode == "Longest":
        return w.sort_values(["review_length_words", "submission_time"], ascending=[False, False], na_position="last")
    if sort_mode in ("Most tagged", "Least tagged"):
        sym_cols = [c for c in w.columns if c.startswith("AI Symptom")]
        if sym_cols:
            w["_tag_n"] = sum((w[c].notna() & (w[c].astype(str).str.strip() != "") & (~w[c].astype(str).str.upper().isin({"","NA","N/A","NONE","NULL","NAN","<NA>","NOT MENTIONED"}))) for c in sym_cols)
            asc = sort_mode == "Least tagged"
            return w.sort_values(["_tag_n", "submission_time"], ascending=[asc, False], na_position="last").drop(columns=["_tag_n"])
    return w


def _highlight_keywords_in_text(text, keywords):
    """Highlight search keywords in review text with a subtle background."""
    if not keywords or not text:
        return str(text)
    result = str(text)
    for kw in keywords:
        kw = kw.strip()
        if len(kw) < 2: continue
        pattern = re.compile(re.escape(kw), re.IGNORECASE)
        result = pattern.sub(lambda m: f"<mark style='background:rgba(99,102,241,.15);padding:1px 3px;border-radius:3px;'>{_esc(m.group())}</mark>", result)
    return result


def _normalize_highlight_text(text):
    replacements = str.maketrans({
        "‘": "'", "’": "'", "‛": "'",
        "“": '"', "”": '"',
        "–": '-', "—": '-', "−": '-',
        " ": ' ',
    })
    return str(text or "").translate(replacements)



def _normalized_text_with_map(text):
    raw = _normalize_highlight_text(text)
    chars = []
    idx_map = []
    prev_space = True
    for idx, ch in enumerate(raw):
        if ch.isspace():
            if prev_space:
                continue
            chars.append(" ")
            idx_map.append(idx)
            prev_space = True
            continue
        chars.append(ch.lower())
        idx_map.append(idx)
        prev_space = False
    return "".join(chars), idx_map



def _evidence_content_tokens(text):
    stop = {"the", "a", "an", "and", "to", "for", "of", "in", "on", "with", "is", "it", "very", "really", "so"}
    return [w for w in re.findall(r"[a-z0-9']+", str(text or "").lower()) if len(w) > 2 and w not in stop]



def _find_evidence_hits(text, evidence_items):
    text_str = str(text or "")
    if not evidence_items or not text_str.strip():
        return []

    norm_text, idx_map = _normalized_text_with_map(text_str)
    if not norm_text:
        return []

    hits = []
    for item in evidence_items:
        ev_text = item[0] if len(item) >= 1 else ""
        tag_label = item[1] if len(item) >= 2 else ""
        side = item[2] if len(item) >= 3 else ""
        ev_clean = re.sub(r"\s+", " ", _normalize_highlight_text(ev_text)).strip()
        if not ev_clean:
            continue
        ev_norm, _ = _normalized_text_with_map(ev_clean)
        ev_norm = re.sub(r"\s+", " ", ev_norm).strip()
        if not ev_norm:
            continue

        found_any = False
        # Tier 1: normalized exact substring match (handles curly quotes/dashes/spaces)
        for match in re.finditer(re.escape(ev_norm), norm_text, flags=re.IGNORECASE):
            start_n, end_n = match.span()
            if end_n <= start_n:
                continue
            start_o = idx_map[start_n]
            end_o = idx_map[end_n - 1] + 1
            hits.append((start_o, end_o, tag_label, text_str[start_o:end_o], side))
            found_any = True

        # Tier 2: flexible token chain match allowing punctuation gaps
        if not found_any:
            tokens = _evidence_content_tokens(ev_norm)
            if tokens:
                token_pattern = re.compile(r"\b" + r"\W+".join(re.escape(tok) for tok in tokens) + r"\b", re.IGNORECASE)
                match = token_pattern.search(norm_text)
                if match:
                    start_n, end_n = match.span()
                    start_o = idx_map[start_n]
                    end_o = idx_map[end_n - 1] + 1
                    hits.append((start_o, end_o, tag_label, text_str[start_o:end_o], side))
                    found_any = True

        # Tier 3: shortest window covering all content tokens
        if not found_any:
            tokens = _evidence_content_tokens(ev_norm)
            if len(tokens) >= 2:
                best = None
                for token in tokens:
                    pos = norm_text.find(token)
                    while pos >= 0:
                        window_start = max(0, pos - 18)
                        window_end = min(len(norm_text), pos + max(len(ev_norm) + 42, 120))
                        window = norm_text[window_start:window_end]
                        if all(tok in window for tok in tokens):
                            positions = []
                            for tok in tokens:
                                tok_pos = window.find(tok)
                                if tok_pos >= 0:
                                    positions.append((window_start + tok_pos, window_start + tok_pos + len(tok)))
                            if positions:
                                start_n = min(p[0] for p in positions)
                                end_n = max(p[1] for p in positions)
                                span_len = end_n - start_n
                                if span_len < min(max(len(ev_norm) + 36, 80), 220):
                                    if best is None or span_len < best[0]:
                                        best = (span_len, start_n, end_n)
                        pos = norm_text.find(token, pos + 1)
                if best is not None:
                    _, start_n, end_n = best
                    start_o = idx_map[start_n]
                    end_o = idx_map[end_n - 1] + 1
                    hits.append((start_o, end_o, tag_label, text_str[start_o:end_o], side))

    if not hits:
        return []

    hits.sort(key=lambda item: (item[0], -(item[1] - item[0])))
    deduped = []
    seen = set()
    cursor = -1
    for start_o, end_o, tag_label, matched, side in hits:
        key = (start_o, end_o, str(tag_label).strip().lower(), str(side).strip().lower())
        if key in seen:
            continue
        seen.add(key)
        if start_o < cursor:
            continue
        deduped.append((start_o, end_o, tag_label, matched, side))
        cursor = end_o
    return deduped



def _highlight_evidence(text, evidence_items, *, hits=None):
    text_str = str(text or "")
    if not text_str.strip():
        return f"<div class='review-body'>{html.escape(text_str)}</div>"
    hits = list(hits or _find_evidence_hits(text_str, evidence_items))
    if not hits:
        return f"<div class='review-body'>{html.escape(text_str)}</div>"
    parts = []
    cursor = 0
    for start, end, tag_label, matched, side in hits:
        parts.append(html.escape(text_str[cursor:start]))
        tip = html.escape(f"{'Issue' if side == 'det' else 'Strength' if side == 'del' else 'Tag'}: {tag_label}")
        side_class = f" ev-{side}" if side in ("det", "del") else ""
        parts.append(f'<span class="ev-highlight{side_class}" data-tag="{tip}">{html.escape(matched)}</span>')
        cursor = end
    parts.append(html.escape(text_str[cursor:]))
    return f"<div class='review-body'>{''.join(parts)}</div>"



def _build_evidence_lookup(processed_rows):
    """Build evidence lookup keyed by BOTH normalized DataFrame index AND review_id."""
    lookup = {}
    for rec in processed_rows or []:
        idx = _canonical_index_key(rec.get("idx", ""))
        rid = str(rec.get("review_id", rec.get("rid", ""))).strip()
        if not idx and not rid:
            continue
        entries = []
        for lab, evs in (rec.get("ev_det", {}) or {}).items():
            for e in (evs or []):
                if e and e.strip():
                    entries.append((e.strip(), lab, "det"))
        for lab, evs in (rec.get("ev_del", {}) or {}).items():
            for e in (evs or []):
                if e and e.strip():
                    entries.append((e.strip(), lab, "del"))
        if not entries:
            continue
        for key in [idx, rid]:
            if not key:
                continue
            lookup.setdefault(key, []).extend(entries)
    for key, values in list(lookup.items()):
        deduped = []
        seen = set()
        for item in values:
            norm = (str(item[0]).strip().lower(), str(item[1]).strip().lower(), str(item[2]).strip().lower())
            if norm in seen:
                continue
            seen.add(norm)
            deduped.append(item)
        lookup[key] = deduped
    return lookup



def _symptom_tags_html(det_tags, del_tags, *, ev_det=None, ev_del=None):
    if not det_tags and not del_tags:
        return ""
    ev_det = ev_det or {}
    ev_del = ev_del or {}
    def _chip(tag, color, ev_map):
        ev = ev_map.get(tag, [])
        tooltip = _esc(" | ".join(str(e)[:80] for e in ev)) if ev else "No evidence captured"
        ev_total_chars = sum(len(str(e)) for e in ev)
        if len(ev) >= 2 and ev_total_chars >= 30:
            conf_style = "opacity:1;border:1.5px solid;"
        elif len(ev) >= 1 and ev_total_chars >= 15:
            conf_style = "opacity:.88;border:1px solid;"
        else:
            conf_style = "opacity:.65;border:1px dashed;"
        ev_indicator = f"<span style='font-size:9px;opacity:.5;margin-left:2px;'>{len(ev)}ev</span>" if ev else "<span style='font-size:9px;opacity:.4;margin-left:2px;'>0ev</span>"
        return f"<span class='chip {color}' style='font-size:11px;padding:3px 8px;cursor:help;{conf_style}' title='{tooltip}'>{_esc(tag)}{ev_indicator}</span>"
    sym_html = "<div style='margin-top:9px;padding-top:9px;border-top:1px solid var(--border);display:flex;flex-direction:column;gap:6px;'>"
    if det_tags:
        det_chips = "".join(_chip(t, "red", ev_det) for t in det_tags)
        sym_html += f"<div style='display:flex;align-items:flex-start;gap:7px;flex-wrap:wrap;'><span style='font-size:10px;text-transform:uppercase;letter-spacing:.07em;color:var(--danger);font-weight:700;white-space:nowrap;padding-top:3px;'>Issues</span><div style='display:flex;gap:4px;flex-wrap:wrap;'>{det_chips}</div></div>"
    if del_tags:
        del_chips = "".join(_chip(t, "green", ev_del) for t in del_tags)
        sym_html += f"<div style='display:flex;align-items:flex-start;gap:7px;flex-wrap:wrap;'><span style='font-size:10px;text-transform:uppercase;letter-spacing:.07em;color:var(--success);font-weight:700;white-space:nowrap;padding-top:3px;'>Strengths</span><div style='display:flex;gap:4px;flex-wrap:wrap;'>{del_chips}</div></div>"
    sym_html += "</div>"
    return sym_html



def _review_card_sections(row):
    sections = []
    seen = set()
    section_map = [
        ("Review", row.get("review_text")),
        ("Pros", row.get("pros")),
        ("Cons", row.get("cons")),
        ("Headline", row.get("headline")),
        ("Body", row.get("body")),
        ("Comments", row.get("comments")),
        ("Reviewer notes", row.get("reviewer_comments")),
    ]
    fallback_text = _safe_text(row.get("title_and_text"))
    for label, raw in section_map:
        body = _safe_text(raw).strip()
        if not body:
            continue
        key = re.sub(r"\s+", " ", body).strip().lower()
        if key in seen:
            continue
        seen.add(key)
        sections.append((label, body))
    if not sections and fallback_text:
        sections.append(("Review", fallback_text))
    if not sections:
        sections.append(("Review", "—"))
    return sections



def _render_review_card(row, evidence_items=None):
    rating_val = _safe_int(row.get("rating"), 0) if pd.notna(row.get("rating")) else 0
    stars = "★" * max(0, min(rating_val, 5)) + "☆" * max(0, 5 - rating_val)
    title = _safe_text(row.get("title"), "No title") or "No title"
    meta_bits = [b for b in [_safe_text(row.get("submission_date")), _safe_text(row.get("content_locale")), _safe_text(row.get("retailer")), _safe_text(row.get("product_or_sku"))] if b]
    is_organic = not _safe_bool(row.get("incentivized_review"), False)
    status_chips = f"<span class='chip {'gray' if is_organic else 'yellow'}'>{'Organic' if is_organic else 'Incentivized'}</span>"
    rec = row.get("is_recommended")
    if not _is_missing(rec):
        status_chips += f"<span class='chip {'gray' if _safe_bool(rec, False) else 'red'}'>{'Recommended' if _safe_bool(rec, False) else 'Not recommended'}</span>"
    det_cols, del_cols = _symptom_col_lists_from_columns(row.index)
    det_tags = _collect_row_symptom_tags(row, det_cols)
    del_tags = _collect_row_symptom_tags(row, del_cols)
    sections = _review_card_sections(row)
    with st.container(border=True):
        top_cols = st.columns([5, 1.5])
        with top_cols[0]:
            st.markdown(f"<span style='color:#f59e0b;letter-spacing:-.01em;'>{stars}</span>&nbsp;<span style='font-size:12px;color:var(--slate-500);font-weight:600;'>{rating_val}/5</span>", unsafe_allow_html=True)
            st.markdown(f"<div style='font-weight:700;font-size:14.5px;color:var(--navy);margin:3px 0 2px;'>{_esc(title)}</div>", unsafe_allow_html=True)
            if meta_bits:
                st.markdown(f"<div style='font-size:12px;color:var(--slate-400);margin-bottom:4px;'>{' · '.join(_esc(b) for b in meta_bits)}</div>", unsafe_allow_html=True)
        with top_cols[1]:
            st.markdown(f"<div class='chip-wrap' style='justify-content:flex-end;gap:4px;flex-wrap:wrap;padding-top:2px;'>{status_chips}</div>", unsafe_allow_html=True)

        active_kw = str(st.session_state.get("rf_kw", "")).strip()
        matched_hits = 0
        for idx, (section_label, section_text) in enumerate(sections):
            if len(sections) > 1:
                margin_top = ".3rem" if idx == 0 else ".55rem"
                st.markdown(f"<div class='metric-label' style='margin:{margin_top} 0 .18rem;font-size:10.5px;'>{_esc(section_label)}</div>", unsafe_allow_html=True)
            if evidence_items:
                section_hits = _find_evidence_hits(section_text, evidence_items)
                matched_hits += len(section_hits)
                st.markdown(_highlight_evidence(section_text, evidence_items, hits=section_hits), unsafe_allow_html=True)
            elif active_kw:
                highlighted = _highlight_keywords_in_text(html.escape(section_text), active_kw.split())
                st.markdown(f"<div class='review-body'>{highlighted}</div>", unsafe_allow_html=True)
            else:
                st.markdown(f"<div class='review-body'>{html.escape(section_text)}</div>", unsafe_allow_html=True)

        if evidence_items:
            if matched_hits:
                st.caption("Evidence highlights now check the visible review fields the Symptomizer reads, including review text, pros/cons, and comments when available.")
            else:
                st.caption("Some evidence was captured as paraphrased or cross-field text, so not every tag snippet maps perfectly onto the visible review fields.")

        _ev_det_map, _ev_del_map = {}, {}
        try:
            _rid = _safe_text(row.get("review_id"))
            _oidx = _canonical_index_key(row.get("_orig_idx", getattr(row, "name", "")))
            for pr in (st.session_state.get("sym_processed_rows") or []):
                pr_idx = _canonical_index_key(pr.get("idx", ""))
                pr_rid = str(pr.get("review_id", pr.get("rid", "")))
                if pr_idx == _oidx or pr_rid == _rid or (pr_idx and pr_idx == _canonical_index_key(getattr(row, "name", ""))):
                    _ev_det_map = pr.get("ev_det", {})
                    _ev_del_map = pr.get("ev_del", {})
                    break
        except Exception:
            pass
        tag_html = _symptom_tags_html(det_tags, del_tags, ev_det=_ev_det_map, ev_del=_ev_del_map)
        if tag_html:
            st.markdown(tag_html, unsafe_allow_html=True)
        footer_bits = []
        rid = _safe_text(row.get("review_id"))
        if rid:
            footer_bits.append(f"<span style='font-size:11.5px;color:var(--slate-400);'>ID {_esc(rid)}</span>")
        loc = _safe_text(row.get("user_location"))
        if loc:
            footer_bits.append(f"<span style='font-size:11.5px;color:var(--slate-400);'>{_esc(loc)}</span>")
        if footer_bits:
            st.markdown(f"<div style='display:flex;align-items:center;gap:8px;flex-wrap:wrap;margin-top:8px;'>{' · '.join(footer_bits)}</div>", unsafe_allow_html=True)

def _render_ai_taxonomy_preview_table(preview_items, *, key_prefix, side_label):
    rows = []
    for item in preview_items or []:
        rows.append({
            "L1 Theme": _safe_text(item.get("l1_theme") or item.get("theme") or item.get("family")) or "Product Specific",
            "L2 Symptom": _safe_text(item.get("label")) or "—",
            "Bucket": _safe_text(item.get("bucket")) or "Product Specific",
            "Family": _safe_text(item.get("family")) or "—",
            "Review Hits": int(item.get("review_hits", 0) or 0),
            "Support %": round(float(item.get("support_ratio", 0.0) or 0.0) * 100.0, 1),
            "Aliases": ", ".join(list(item.get("aliases") or [])[:4]) or "—",
            "Example": (_safe_text((item.get("examples") or [""])[0]) or "—"),
            "Rationale": _safe_text(item.get("rationale")) or "—",
        })
    df = pd.DataFrame(rows)
    if df.empty:
        st.info(f"No {side_label.lower()} were generated for the current sample.")
        return
    st.markdown(_chip_html([
        (f"{df['L1 Theme'].nunique():,} themes", "blue"),
        (f"{df['L2 Symptom'].nunique():,} L2 symptoms", "indigo"),
        (f"{int((df['Bucket'] == 'Category Driver').sum())} category drivers", "gray"),
    ]), unsafe_allow_html=True)
    st.caption("This preview is already structured as L1 Themes → L2 Symptoms so it is easier to consolidate before you accept the taxonomy.")
    with st.expander("Preview table tools", expanded=False):
        c1, c2, c3 = st.columns([2.0, 1.2, 2.4])
        search = c1.text_input("Search preview", key=f"{key_prefix}_search", placeholder="Filter symptoms, themes, or aliases")
        bucket_filter = c2.multiselect("Buckets", options=sorted(df["Bucket"].dropna().unique().tolist()), default=sorted(df["Bucket"].dropna().unique().tolist()), key=f"{key_prefix}_buckets")
        visible_cols = c3.multiselect("Visible columns", options=df.columns.tolist(), default=[c for c in ["L1 Theme", "L2 Symptom", "Bucket", "Review Hits", "Aliases", "Example"] if c in df.columns], key=f"{key_prefix}_cols")
    if search:
        pattern = re.escape(search)
        mask = (
            df["L2 Symptom"].str.contains(pattern, case=False, na=False)
            | df["L1 Theme"].str.contains(pattern, case=False, na=False)
            | df["Aliases"].str.contains(pattern, case=False, na=False)
            | df["Family"].str.contains(pattern, case=False, na=False)
        )
        df = df[mask]
    if bucket_filter:
        df = df[df["Bucket"].isin(bucket_filter)]
    visible_cols = [c for c in (visible_cols or df.columns.tolist()) if c in df.columns] or df.columns.tolist()
    st.dataframe(df[visible_cols], use_container_width=True, hide_index=True, height=int(min(max(320, 36 * len(df) + 72), 700)))


def _taxonomy_preview_items_with_side(ai_result):
    items = []
    result = ai_result or {}
    for side_key, bucket_key in (("delighter", "preview_delighters"), ("detractor", "preview_detractors")):
        for raw in (result.get(bucket_key) or []):
            row = dict(raw)
            row["side"] = side_key
            items.append(row)
    return items



def _render_structured_taxonomy_table(rows, *, key_prefix, empty_label="No taxonomy rows to show."):
    if not rows:
        st.info(empty_label)
        return
    df = pd.DataFrame(rows)
    visible_defaults = [col for col in ["L1 Theme", "L2 Symptom", "Side", "Bucket", "Review Hits", "Support %", "Aliases"] if col in df.columns]
    with st.expander("Taxonomy table tools", expanded=False):
        c1, c2, c3 = st.columns([2.0, 1.35, 1.35])
        search = c1.text_input("Search taxonomy", key=f"{key_prefix}_search", placeholder="Filter by symptom, theme, alias, or rationale")
        side_opts = ["All"] + sorted(df["Side"].dropna().astype(str).unique().tolist()) if "Side" in df.columns else ["All"]
        side_choice = c2.selectbox("Side", options=side_opts, key=f"{key_prefix}_side")
        bucket_opts = sorted(df["Bucket"].dropna().astype(str).unique().tolist()) if "Bucket" in df.columns else []
        bucket_choice = c3.multiselect("Buckets", options=bucket_opts, default=bucket_opts, key=f"{key_prefix}_bucket")
        c4, c5, c6 = st.columns([1.35, 1.0, 2.65])
        theme_opts = sorted(df["L1 Theme"].dropna().astype(str).unique().tolist()) if "L1 Theme" in df.columns else []
        theme_choice = c4.multiselect("L1 themes", options=theme_opts, default=theme_opts, key=f"{key_prefix}_theme")
        row_options = [25, 50, 100, 250, "All"]
        row_choice = c5.selectbox("Rows", options=row_options, index=1, key=f"{key_prefix}_rows")
        visible_cols = c6.multiselect("Visible columns", options=df.columns.tolist(), default=visible_defaults or df.columns.tolist(), key=f"{key_prefix}_visible")
    filtered = df.copy()
    if search:
        pattern = re.escape(str(search).strip())
        searchable_cols = [col for col in ["L2 Symptom", "L1 Theme", "Aliases", "Rationale", "Example"] if col in filtered.columns]
        if searchable_cols:
            mask = pd.Series(False, index=filtered.index)
            for col in searchable_cols:
                mask = mask | filtered[col].astype(str).str.contains(pattern, case=False, na=False)
            filtered = filtered[mask]
    if side_choice != "All" and "Side" in filtered.columns:
        filtered = filtered[filtered["Side"] == side_choice]
    if bucket_choice and "Bucket" in filtered.columns:
        filtered = filtered[filtered["Bucket"].isin(bucket_choice)]
    if theme_choice and "L1 Theme" in filtered.columns:
        filtered = filtered[filtered["L1 Theme"].isin(theme_choice)]
    sort_cols = []
    sort_ascending = []
    for col, asc in (("Side", True), ("L1 Theme", True), ("Review Hits", False), ("L2 Symptom", True)):
        if col in filtered.columns:
            sort_cols.append(col)
            sort_ascending.append(asc)
    if sort_cols:
        filtered = filtered.sort_values(sort_cols, ascending=sort_ascending, kind="mergesort")
    if row_choice != "All":
        filtered = filtered.head(int(row_choice))
    visible_cols = [col for col in (visible_cols or visible_defaults or df.columns.tolist()) if col in filtered.columns] or filtered.columns.tolist()
    st.markdown(_chip_html([
        (f"{filtered['L1 Theme'].nunique():,} themes" if "L1 Theme" in filtered.columns else f"{len(filtered):,} rows", "blue"),
        (f"{filtered['L2 Symptom'].nunique():,} L2 symptoms" if "L2 Symptom" in filtered.columns else f"{len(filtered):,} rows", "indigo"),
        (f"{len(filtered):,} rows showing", "gray"),
    ]), unsafe_allow_html=True)
    if filtered.empty:
        st.info("No taxonomy rows match the current filters.")
        return
    st.dataframe(filtered[visible_cols], use_container_width=True, hide_index=True, height=int(min(max(360, 36 * len(filtered) + 72), 760)))



def _build_theme_impact_table(df_in, symptom_cols, theme_map, *, avg_rating, kind="detractors", shrink_k=3.0):
    long, symptomized_reviews = _prepare_symptom_long(df_in, symptom_cols)
    if long.empty:
        return pd.DataFrame(columns=["Item", "Mentions", "% Tagged Reviews", "Avg Star", "Avg Tags/Review", "Confidence %", "Net Hit", "Forecast Δ★", "Impact Score", "Severity Wt"])
    working = long.copy()
    working["symptom"] = working["symptom"].astype(str).str.title()
    working["Item"] = working["symptom"].map(lambda value: theme_map.get(str(value).title(), "Product Specific"))
    working = working[working["Item"].astype(str).str.strip() != ""]
    if working.empty:
        return pd.DataFrame(columns=["Item", "Mentions", "% Tagged Reviews", "Avg Star", "Avg Tags/Review", "Confidence %", "Net Hit", "Forecast Δ★", "Impact Score", "Severity Wt"])
    working["severity_wt"] = working["symptom"].map(lambda value: float(_label_severity_weight(value, kind=kind)))
    theme_review = working.groupby(["__row", "Item"], dropna=False).agg(
        star=("star", "first"),
        theme_l2_count=("symptom", "nunique"),
        theme_weight=("review_weight", "sum"),
        severity_wt=("severity_wt", "mean"),
    ).reset_index()
    baseline = float(avg_rating or 0)
    stars = pd.to_numeric(theme_review["star"], errors="coerce")
    sign = 1.0 if str(kind).lower().startswith("del") else -1.0
    if str(kind).lower().startswith("del"):
        gap = (stars - baseline).clip(lower=0)
    else:
        gap = (baseline - stars).clip(lower=0)
    theme_review["gap"] = gap.fillna(0.0)
    theme_review["attributed_gap"] = theme_review["theme_weight"].astype(float) * theme_review["gap"].astype(float)

    mention_reviews = theme_review.groupby("Item", dropna=False)["__row"].nunique().astype(int)
    avg_l2 = theme_review.groupby("Item", dropna=False)["theme_l2_count"].mean()
    avg_stars = theme_review.groupby("Item", dropna=False)["star"].mean()
    weighted_mentions = theme_review.groupby("Item", dropna=False)["theme_weight"].sum()
    severity = theme_review.groupby("Item", dropna=False)["severity_wt"].mean().fillna(1.0)
    raw_impact = theme_review.groupby("Item", dropna=False)["attributed_gap"].sum().astype(float) / float(max(len(df_in), 1))
    review_conf = mention_reviews.astype(float) / (mention_reviews.astype(float) + float(max(shrink_k, 0.1)))
    weight_conf = weighted_mentions.astype(float) / (weighted_mentions.astype(float) + float(max(shrink_k / 2.0, 0.5)))
    align_conf = _alignment_confidence(avg_stars, baseline, kind=kind).astype(float)
    confidence = np.sqrt(review_conf * weight_conf) * align_conf

    out = pd.DataFrame({
        "Item": [str(item) for item in mention_reviews.index.tolist()],
        "Mentions": mention_reviews.values.astype(int),
        "% Tagged Reviews": (mention_reviews.values / float(max(symptomized_reviews, 1)) * 100.0).round(1).astype(str) + "%",
        "Avg Star": [round(float(avg_stars[item]), 1) if item in avg_stars and not pd.isna(avg_stars[item]) else None for item in mention_reviews.index],
        "Avg Tags/Review": np.round(avg_l2.values.astype(float), 2),
        "Confidence %": (confidence.reindex(mention_reviews.index).fillna(0.0).clip(lower=0.0, upper=1.0) * 100.0).round(1).values,
        "Net Hit": (sign * raw_impact.reindex(mention_reviews.index).fillna(0.0)).round(3).values,
        "Forecast Δ★": (sign * raw_impact.reindex(mention_reviews.index).fillna(0.0) * confidence.reindex(mention_reviews.index).fillna(0.0)).round(3).values,
        "Severity Wt": severity.reindex(mention_reviews.index).round(2).values,
    })
    out["Impact Score"] = (pd.to_numeric(out["Forecast Δ★"], errors="coerce") * pd.to_numeric(out["Severity Wt"], errors="coerce")).round(3)
    out.attrs["symptomized_review_count"] = symptomized_reviews
    out.attrs["all_review_count"] = int(len(df_in))
    return out.sort_values(["Impact Score", "Mentions", "Item"], ascending=[True if str(kind).lower().startswith("det") else False, False, True], kind="mergesort", ignore_index=True)



def _render_symptomizer_taxonomy_workbench(*, processed_df, delighters, detractors, aliases, category, preview_items):
    taxonomy_rows = _build_structured_taxonomy_rows(
        delighters,
        detractors,
        aliases=aliases,
        category=category,
        preview_items=preview_items,
    )
    if not taxonomy_rows:
        return None
    merge_rows = _suggest_taxonomy_merges(taxonomy_rows)
    rows_df = pd.DataFrame(taxonomy_rows)
    l1_summary = pd.DataFrame()
    if not rows_df.empty:
        l1_summary = (
            rows_df.groupby(["Side", "L1 Theme"], dropna=False)
            .agg(
                L2_Symptoms=("L2 Symptom", "nunique"),
                Category_Drivers=("Bucket", lambda s: int((pd.Series(s).astype(str) == "Category Driver").sum())),
                Product_Specific=("Bucket", lambda s: int((pd.Series(s).astype(str) == "Product Specific").sum())),
                Universal_Neutral=("Bucket", lambda s: int((pd.Series(s).astype(str) == "Universal Neutral").sum())),
            )
            .reset_index()
            .rename(columns={"L1 Theme": "Item", "L2_Symptoms": "L2 Symptoms", "Category_Drivers": "Category Drivers", "Product_Specific": "Product Specific", "Universal_Neutral": "Universal Neutral"})
            .sort_values(["Side", "L2 Symptoms", "Item"], ascending=[True, False, True], ignore_index=True)
        )
    det_map = {row["L2 Symptom"]: row["L1 Theme"] for row in taxonomy_rows if row.get("side_key") == "detractor"}
    del_map = {row["L2 Symptom"]: row["L1 Theme"] for row in taxonomy_rows if row.get("side_key") == "delighter"}
    avg_rating = float(pd.to_numeric(processed_df.get("rating"), errors="coerce").mean() or 0.0) if processed_df is not None and not processed_df.empty else 0.0
    det_cols = [col for col in AI_DET_HEADERS if col in processed_df.columns]
    del_cols = [col for col in AI_DEL_HEADERS if col in processed_df.columns]
    det_l2 = _add_net_hit(analyze_symptoms_fast(processed_df, det_cols), avg_rating, total_reviews=len(processed_df), kind="detractors", detail_df=processed_df, symptom_cols=det_cols) if det_cols else pd.DataFrame()
    del_l2 = _add_net_hit(analyze_symptoms_fast(processed_df, del_cols), avg_rating, total_reviews=len(processed_df), kind="delighters", detail_df=processed_df, symptom_cols=del_cols) if del_cols else pd.DataFrame()
    if not det_l2.empty:
        det_l2["L1 Theme"] = det_l2["Item"].map(lambda value: det_map.get(str(value).title(), "Product Specific"))
    if not del_l2.empty:
        del_l2["L1 Theme"] = del_l2["Item"].map(lambda value: del_map.get(str(value).title(), "Product Specific"))
    det_l1 = _build_theme_impact_table(processed_df, det_cols, det_map, avg_rating=avg_rating, kind="detractors") if det_cols else pd.DataFrame()
    del_l1 = _build_theme_impact_table(processed_df, del_cols, del_map, avg_rating=avg_rating, kind="delighters") if del_cols else pd.DataFrame()

    st.markdown("### 4 · Structured taxonomy")
    st.markdown("<div class='section-sub'>The symptom catalog is organized as <b>L1 Themes</b> and <b>L2 Symptoms</b>. The default view now opens on the detailed L2 tables so you can inspect the exact issues and strengths first, then roll up to themes when needed.</div>", unsafe_allow_html=True)
    st.markdown(_chip_html([
        (f"{rows_df['L1 Theme'].nunique():,} L1 themes", "blue"),
        (f"{rows_df['L2 Symptom'].nunique():,} L2 symptoms", "indigo"),
        (f"{len(merge_rows):,} consolidation cues", "gray" if merge_rows else "green"),
        (f"{len(processed_df):,} processed reviews", "green"),
    ]), unsafe_allow_html=True)

    tabs = st.tabs(["🔴 Detractor impact", "🟢 Delighter impact"])
    with tabs[0]:
        st.caption("Default view opens on the L2 symptom table so the exact detractors are visible first.")
        theme_tabs = st.tabs(["L2 symptom table", "L1 theme rollup"])
        with theme_tabs[0]:
            _render_interactive_symptom_table(det_l2, key_prefix="sym_det_l2", empty_label="Detractors")
        with theme_tabs[1]:
            _render_interactive_symptom_table(det_l1, key_prefix="sym_det_l1", empty_label="Detractor themes")
    with tabs[1]:
        st.caption("Default view opens on the L2 symptom table so the exact delighters are visible first.")
        theme_tabs = st.tabs(["L2 symptom table", "L1 theme rollup"])
        with theme_tabs[0]:
            _render_interactive_symptom_table(del_l2, key_prefix="sym_del_l2", empty_label="Delighters")
        with theme_tabs[1]:
            _render_interactive_symptom_table(del_l1, key_prefix="sym_del_l1", empty_label="Delighter themes")

    return {
        "taxonomy_rows": taxonomy_rows,
        "merge_rows": merge_rows,
        "l1_summary": l1_summary,
        "rows_df": rows_df,
    }



def _render_symptomizer_taxonomy_housekeeping(taxonomy_meta):
    if not taxonomy_meta:
        return
    taxonomy_rows = taxonomy_meta.get("taxonomy_rows") or []
    merge_rows = taxonomy_meta.get("merge_rows") or []
    l1_summary = taxonomy_meta.get("l1_summary") if isinstance(taxonomy_meta.get("l1_summary"), pd.DataFrame) else pd.DataFrame()
    if not taxonomy_rows:
        return
    st.markdown("### 5 · Taxonomy map & consolidation")
    st.markdown("<div class='section-sub'>Reference views for the active taxonomy. These are separated from the main impact view so the Symptomizer results stay focused on action first.</div>", unsafe_allow_html=True)
    with st.expander("🗂 Taxonomy map", expanded=False):
        st.caption("Systematic taxonomy view: each L2 symptom rolls up into one L1 theme, with bucket, aliases, and sample support kept in one place.")
        if not l1_summary.empty:
            st.dataframe(l1_summary, use_container_width=True, hide_index=True, height=int(min(max(240, 34 * len(l1_summary) + 72), 520)))
        _render_structured_taxonomy_table(taxonomy_rows, key_prefix="sym_struct_taxonomy")
    with st.expander("🔗 Consolidation cues", expanded=False):
        if merge_rows:
            st.caption("These are not auto-merged. They are consolidation cues to help keep the taxonomy clean when two labels look like the same concept.")
            st.dataframe(pd.DataFrame(merge_rows), use_container_width=True, hide_index=True, height=int(min(max(240, 34 * len(merge_rows) + 72), 520)))
        else:
            st.success("No obvious duplicate labels detected in the current taxonomy. Canonical naming and alias normalization are already keeping it tight.")

# ═══════════════════════════════════════════════════════════════════════════════
#  TAB: DASHBOARD
# ═══════════════════════════════════════════════════════════════════════════════
def _top_theme_summary(df_in, cols, *, kind="detractors"):
    try:
        tbl = analyze_symptoms_fast(df_in, cols)
    except Exception:
        tbl = pd.DataFrame()
    if tbl is None or tbl.empty:
        return None, 0
    if "rating" in df_in.columns:
        tbl = _add_net_hit(tbl, float(_safe_mean(df_in["rating"]) or 0), total_reviews=len(df_in), kind=kind, detail_df=df_in, symptom_cols=cols)
    row = tbl.iloc[0]
    return _safe_text(row.get("Item")) or None, _safe_int(row.get("Mentions"), 0)


def _render_dashboard_snapshot(chart_df, overall_df=None):
    scope_df = chart_df if chart_df is not None else pd.DataFrame()
    if scope_df.empty:
        return
    m = _get_metrics(scope_df)
    od = overall_df if overall_df is not None else scope_df
    det_cols, del_cols = _get_symptom_col_lists(od)
    top_det, det_mentions = _top_theme_summary(scope_df, det_cols, kind="detractors")
    top_del, del_mentions = _top_theme_summary(scope_df, del_cols, kind="delighters")
    recommend_txt = _fmt_pct(m.get("recommend_rate")) if m.get("recommend_rate") is not None else "n/a"
    recommend_sub = "Reviews with recommendation signal" if m.get("recommend_rate") is not None else "No recommendation field available"
    organic_share = max(0.0, 1.0 - float(m.get("pct_incentivized") or 0.0))
    st.markdown(
        f"""
<div class="soft-panel">
  <div style="font-weight:800;color:var(--navy);margin-bottom:8px;">Executive snapshot</div>
  <div class="summary-grid">
    <div class="summary-item">
      <div class="label">Current scope</div>
      <div class="value">{len(scope_df):,} reviews</div>
      <div class="sub">{int((~scope_df['incentivized_review'].fillna(False)).sum()):,} organic in the current view.</div>
    </div>
    <div class="summary-item">
      <div class="label">Satisfaction</div>
      <div class="value">{_fmt_num(m.get('avg_rating'))} ★</div>
      <div class="sub">Recommend rate {recommend_txt}. {recommend_sub}.</div>
    </div>
    <div class="summary-item">
      <div class="label">Biggest risk</div>
      <div class="value">{_esc(top_det or 'Run Symptomizer')}</div>
      <div class="sub">{(str(det_mentions) + ' mentions in view') if top_det else 'Top risk themes will appear here once symptoms are tagged.'}</div>
    </div>
    <div class="summary-item">
      <div class="label">Quality of sample</div>
      <div class="value">{_fmt_pct(organic_share)} organic</div>
      <div class="sub">Low-star share {_fmt_pct(m.get('pct_low_star'))}. {(_esc(top_del) + ' leads positives.') if top_del else 'Top positive themes will appear here once symptoms are tagged.'}</div>
    </div>
  </div>
</div>
""",
        unsafe_allow_html=True,
    )


def _render_dashboard(filtered_df, overall_df=None):
    od = overall_df if overall_df is not None else filtered_df
    st.markdown("<div class='section-title'>Dashboard</div>", unsafe_allow_html=True)
    # ── Auto-generated key insights (no AI needed) ───────────────────────
    try:
        n_reviews = len(filtered_df)
        if n_reviews >= 5:
            avg_rating = pd.to_numeric(filtered_df.get("rating"), errors="coerce").mean()
            organic_mask = ~filtered_df.get("incentivized_review", pd.Series(False, index=filtered_df.index)).fillna(False)
            organic_avg = pd.to_numeric(filtered_df.loc[organic_mask, "rating"], errors="coerce").mean() if organic_mask.any() else avg_rating
            recent_30 = filtered_df[pd.to_datetime(filtered_df.get("submission_time"), errors="coerce", utc=True) >= (pd.Timestamp.now(tz="UTC") - pd.Timedelta(days=30))]
            recent_avg = pd.to_numeric(recent_30.get("rating"), errors="coerce").mean() if len(recent_30) >= 3 else None
            low_pct = (pd.to_numeric(filtered_df.get("rating"), errors="coerce") <= 2).mean()
            summary_pills = []
            if pd.notna(avg_rating):
                summary_pills.append(
                    f"<span class='dashboard-pill'><span class='meta'>Average rating</span><strong>{avg_rating:.2f}★</strong></span>"
                )
                summary_pills.append(
                    f"<span class='dashboard-pill'><span class='meta'>Reviews</span><strong>{n_reviews:,}</strong></span>"
                )
            if pd.notna(organic_avg) and organic_mask.sum() >= 3 and pd.notna(avg_rating):
                delta = organic_avg - avg_rating
                if abs(delta) >= 0.1:
                    direction = "Higher" if delta > 0 else "Lower"
                    summary_pills.append(
                        f"<span class='dashboard-pill'><span class='meta'>Organic</span><strong>{organic_avg:.2f}★</strong><span class='meta'>{direction} than overall</span></span>"
                    )
            if recent_avg is not None and pd.notna(recent_avg) and pd.notna(avg_rating):
                delta = recent_avg - avg_rating
                if abs(delta) >= 0.15:
                    trend_cls = "trend-up" if delta > 0 else "trend-down"
                    trend_label = "Trending up" if delta > 0 else "Trending down"
                    summary_pills.append(
                        f"<span class='dashboard-pill {trend_cls}'><span class='meta'>Last 30 days</span><strong>{recent_avg:.2f}★</strong><span class='meta'>{trend_label}</span></span>"
                    )
            if low_pct >= 0.15:
                summary_pills.append(
                    f"<span class='dashboard-pill warn'><span class='meta'>Low-star share</span><strong>{low_pct:.0%}</strong></span>"
                )
            if summary_pills:
                st.markdown(
                    "<div class='dashboard-brief'><div class='dashboard-brief-title'>At a glance</div><div class='dashboard-brief-row'>"
                    + "".join(summary_pills)
                    + "</div></div>",
                    unsafe_allow_html=True,
                )
    except Exception:
        pass
    # ── Symptomizer insights (if available) ───────────────────────────
    _sym_processed = st.session_state.get("sym_processed_rows") or []
    if _sym_processed and len(_sym_processed) >= 3:
        try:
            _det_freq = {}
            _del_freq = {}
            for _rec in _sym_processed:
                for _t in (_rec.get("wrote_dets") or []):
                    _det_freq[_t] = _det_freq.get(_t, 0) + 1
                for _t in (_rec.get("wrote_dels") or []):
                    _del_freq[_t] = _del_freq.get(_t, 0) + 1
            _n = max(len(_sym_processed), 1)
            _top_dets = sorted(_det_freq.items(), key=lambda x: -x[1])[:5]
            _top_dels = sorted(_del_freq.items(), key=lambda x: -x[1])[:5]
            if _top_dets or _top_dels:
                _max_count = max([c for _, c in (_top_dets + _top_dels)] or [1])
                dc1, dc2 = st.columns(2)
                with dc1:
                    bars_html = f"<div class='sym-insights-card'><div class='sym-insights-title' style='color:var(--danger);'>Top issues ({len(_sym_processed)} reviews tagged)</div>"
                    for label, count in _top_dets:
                        pct = count * 100 / _max_count
                        bars_html += f"<div class='sym-bar'><span class='sym-bar-label'>{_esc(label)}</span><div class='sym-bar-track'><div class='sym-bar-fill det' style='width:{pct:.0f}%'></div></div><span class='sym-bar-count'>{count}</span></div>"
                    if not _top_dets:
                        bars_html += "<div style='font-size:12px;color:var(--slate-400);'>No detractors tagged</div>"
                    bars_html += "</div>"
                    st.markdown(bars_html, unsafe_allow_html=True)
                with dc2:
                    bars_html = f"<div class='sym-insights-card'><div class='sym-insights-title' style='color:var(--success);'>Top strengths ({len(_sym_processed)} reviews tagged)</div>"
                    for label, count in _top_dels:
                        pct = count * 100 / _max_count
                        bars_html += f"<div class='sym-bar'><span class='sym-bar-label'>{_esc(label)}</span><div class='sym-bar-track'><div class='sym-bar-fill del' style='width:{pct:.0f}%'></div></div><span class='sym-bar-count'>{count}</span></div>"
                    if not _top_dels:
                        bars_html += "<div style='font-size:12px;color:var(--slate-400);'>No delighters tagged</div>"
                    bars_html += "</div>"
                    st.markdown(bars_html, unsafe_allow_html=True)
        except Exception:
            pass
    st.markdown("<div class='section-sub'>Start with the overall time trend and symptom tables. Additional charts are tucked into a cleaner secondary section below.</div>", unsafe_allow_html=True)
    sym_state = _detect_symptom_state(od)
    if sym_state == "none":
        st.markdown("""<div class="sym-state-banner" style="padding:1.5rem 1.8rem;text-align:left;display:flex;align-items:center;gap:16px;margin-bottom:.5rem;">
          <div style="font-size:2.2rem;flex-shrink:0;">💊</div>
          <div style="flex:1;">
            <div class="title" style="margin-bottom:4px;font-size:15px;">No symptoms tagged yet</div>
            <div class="sub" style="max-width:none;font-size:12.5px;">Run the Symptomizer to AI-tag delighters &amp; detractors — they'll surface here once complete.</div>
          </div>
        </div>""", unsafe_allow_html=True)
        if st.button("💊 Go to Symptomizer →", type="primary", key="dash_go_sym", use_container_width=False):
            st.session_state["workspace_tab_request"] = TAB_SYMPTOMIZER
            st.rerun()
        st.markdown("<div style='height:.5rem'></div>", unsafe_allow_html=True)

    chart_df = filtered_df.copy()
    if chart_df.empty:
        st.info("No reviews match the current scope.")
        return

    st.markdown("<div style='height:.3rem'></div>", unsafe_allow_html=True)
    _render_dashboard_snapshot(chart_df, od)
    _render_reviews_over_time_chart(chart_df)
    _render_source_rating_watch(chart_df)

    st.markdown("<div style='height:.75rem'></div>", unsafe_allow_html=True)
    _render_symptom_dashboard(chart_df, od)

    with st.expander("📊 Additional dashboard views", expanded=False):
        dash_tabs = st.tabs(["Rating mix", "Cohorts", "Sentiment", "Markets", "Review depth"])

        with dash_tabs[0]:
            rating_df = _rating_dist(chart_df)
            rating_df["rating_label"] = rating_df["rating"].map(lambda v: f"{int(v)}★")
            rating_df["count_pct_label"] = rating_df.apply(lambda r: f"{int(r['review_count']):,} · {_fmt_pct(r['share'])}", axis=1)
            with st.container(border=True):
                _render_chart_header("Rating distribution", "Volume and share by star rating for the current view.")
                fig = px.bar(
                    rating_df,
                    x="rating_label",
                    y="review_count",
                    text="count_pct_label",
                    category_orders={"rating_label": ["1★", "2★", "3★", "4★", "5★"]},
                    color="rating",
                    color_discrete_map={"1": "#ef4444", "2": "#f97316", "3": "#eab308", "4": "#84cc16", "5": "#22c55e"},
                    hover_data={"share": ":.1%", "review_count": True},
                )
                fig.update_traces(textposition="outside", cliponaxis=False, showlegend=False)
                fig.update_layout(title=None, margin=dict(l=20, r=18, t=18, b=20), xaxis_title="", yaxis_title="Reviews", height=330, plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)", font_family="Inter")
                fig = _sw_style_fig(fig)
                fig.update_layout(title=None, margin=dict(l=24, r=18, t=18, b=32))
                _show_plotly(fig)

        with dash_tabs[1]:
            cohort_df = _cohort_by_incentivized(chart_df)
            with st.container(border=True):
                _render_chart_header("Rating split: Organic vs Incentivized", "Compare cohort distributions without crowding the plot area.")
                if cohort_df.empty:
                    st.info("No cohort data.")
                else:
                    fig_c = px.bar(cohort_df, x="star", y="pct", color="cohort", barmode="group", labels={"star": "Star", "pct": "% of cohort", "cohort": "Cohort"}, color_discrete_map={"Organic": "#6366f1", "Incentivized": "#f59e0b"})
                    fig_c.update_layout(xaxis=dict(tickmode="array", tickvals=[1, 2, 3, 4, 5], ticktext=["1★", "2★", "3★", "4★", "5★"]), title=None, plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)", font_family="Inter", margin=dict(l=20, r=18, t=18, b=48), height=330)
                    fig_c.update_yaxes(ticksuffix="%")
                    fig_c = _sw_style_fig(fig_c)
                    fig_c.update_layout(title=None, legend=dict(orientation="h", y=-0.22, x=0, xanchor="left", yanchor="top", font=dict(size=11)), margin=dict(l=24, r=18, t=18, b=78))
                    _show_plotly(fig_c)

        with dash_tabs[2]:
            sb_df = _star_band_trend(chart_df)
            with st.container(border=True):
                _render_chart_header("Sentiment drift over time", "Track how low-star and high-star mix moves by month.")
                if sb_df.empty:
                    st.info("Insufficient date data for sentiment trend.")
                else:
                    fig_sb = go.Figure()
                    fig_sb.add_trace(go.Scatter(x=sb_df["month_start"], y=sb_df["pct_low"], name="% 1-2★", mode="lines+markers", line=dict(color="#ef4444", width=2), marker=dict(size=4), fill="tozeroy", fillcolor="rgba(239,68,68,0.08)"))
                    fig_sb.add_trace(go.Scatter(x=sb_df["month_start"], y=sb_df["pct_high"], name="% 4-5★", mode="lines+markers", line=dict(color="#22c55e", width=2), marker=dict(size=4)))
                    fig_sb.update_layout(title=None, hovermode="x unified", plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)", font_family="Inter", margin=dict(l=20, r=18, t=18, b=48), height=330)
                    fig_sb.update_yaxes(ticksuffix="%", title="% of monthly reviews")
                    fig_sb = _sw_style_fig(fig_sb)
                    fig_sb.update_layout(title=None, legend=dict(orientation="h", y=-0.22, x=0, xanchor="left", yanchor="top", font=dict(size=11)), margin=dict(l=24, r=18, t=18, b=78))
                    _show_plotly(fig_sb)

        with dash_tabs[3]:
            with st.container(border=True):
                _render_chart_header("Top markets by review volume", "Review count by locale, with average rating shown as scaled markers.")
                locale_df = _locale_breakdown(chart_df, top_n=None)
                if locale_df.empty:
                    st.info("No locale data.")
                else:
                    fig_loc = go.Figure()
                    fig_loc.add_trace(go.Bar(x=locale_df["count"], y=locale_df["content_locale"], orientation="h", name="Reviews", marker_color="#6366f1", opacity=0.82, hovertemplate="%{y}<br>%{x:,} reviews<extra></extra>"))
                    fig_loc.add_trace(go.Scatter(x=locale_df["avg_rating"] * locale_df["count"].max() / 5, y=locale_df["content_locale"], mode="markers", name="Avg ★ (scaled)", marker=dict(color=locale_df["avg_rating"], colorscale="RdYlGn", cmin=1, cmax=5, size=9, showscale=True, colorbar=dict(title="Avg ★", len=0.42, x=1.01)), hovertemplate="%{y}<br>Avg ★: %{text}<extra></extra>", text=[f"{v:.2f}" for v in locale_df["avg_rating"]]))
                    fig_loc.update_layout(title=None, height=max(320, 26 * len(locale_df) + 100), margin=dict(l=70, r=40, t=18, b=48), barmode="overlay", plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)", font_family="Inter", xaxis_title="Reviews", yaxis_title="")
                    fig_loc = _sw_style_fig(fig_loc)
                    fig_loc.update_layout(title=None, legend=dict(orientation="h", y=-0.18, x=0, xanchor="left", yanchor="top", font=dict(size=11)), margin=dict(l=74, r=58, t=18, b=84))
                    _show_plotly(fig_loc)
                st.markdown("<div style='height:.35rem'></div>", unsafe_allow_html=True)
                locs = _top_locations(chart_df, top_n=None)
                if not locs.empty:
                    st.markdown("**Top reviewer locations**")
                    locs_display = locs.copy()
                    locs_display["avg_rating"] = locs_display["avg_rating"].map(lambda v: f"{v:.2f}★" if pd.notna(v) else "—")
                    locs_display = locs_display.rename(columns={"user_location": "Location", "count": "Reviews", "avg_rating": "Avg ★"})
                    table_height = min(600, max(280, 35 * len(locs_display) + 40))
                    st.dataframe(locs_display[["Location", "Reviews", "Avg ★"]], use_container_width=True, hide_index=True, height=table_height)

        with dash_tabs[4]:
            with st.container(border=True):
                _render_chart_header("Review depth vs satisfaction", "Longer reviews often signal different satisfaction patterns and complexity.")
                len_df = _review_length_cohort(chart_df)
                if len_df.empty:
                    st.info("Insufficient data for review-length analysis.")
                else:
                    fig_len = go.Figure()
                    fig_len.add_trace(go.Bar(x=len_df["Length Quartile"], y=len_df["avg_rating"], text=[f"{v:.2f}★" for v in len_df["avg_rating"]], textposition="outside", marker_color=["#ef4444" if v < 3.5 else "#eab308" if v < 4.2 else "#22c55e" for v in len_df["avg_rating"]], hovertemplate="%{x}<br>Avg ★: %{y:.2f}<br>n=%{customdata}<extra></extra>", customdata=len_df["count"]))
                    fig_len.update_layout(title=None, yaxis_range=[1, 5.2], yaxis_title="Avg ★", xaxis_title="", plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)", font_family="Inter", margin=dict(l=20, r=18, t=18, b=24), height=330)
                    fig_len = _sw_style_fig(fig_len)
                    fig_len.update_layout(title=None, margin=dict(l=24, r=18, t=18, b=34))
                    _show_plotly(fig_len)

    # ── Symptomizer insights (visible when tags exist) ────────────────
    _sym_processed = st.session_state.get("sym_processed_rows") or []
    if _sym_processed and len(_sym_processed) >= 3:
        _det_freq = {}
        _del_freq = {}
        for rec in _sym_processed:
            for t in (rec.get("wrote_dets") or []):
                _det_freq[t] = _det_freq.get(t, 0) + 1
            for t in (rec.get("wrote_dels") or []):
                _del_freq[t] = _del_freq.get(t, 0) + 1
        _top_dets = sorted(_det_freq.items(), key=lambda x: -x[1])[:5]
        _top_dels = sorted(_del_freq.items(), key=lambda x: -x[1])[:5]
        if _top_dets or _top_dels:
            st.markdown("<div class='sym-insights-card'><div class='sym-insights-title'>Symptomizer insights · " + f"{len(_sym_processed):,} reviews tagged</div>", unsafe_allow_html=True)
            ic1, ic2 = st.columns(2)
            with ic1:
                for label, count in _top_dets:
                    pct = count / max(len(_sym_processed), 1) * 100
                    st.markdown(f"<div class='sym-bar'><div class='sym-bar-label'>{_esc(label)}</div><div class='sym-bar-track'><div class='sym-bar-fill det' style='width:{min(pct * 2, 100)}%;'></div></div><div class='sym-bar-count'>{pct:.0f}%</div></div>", unsafe_allow_html=True)
            with ic2:
                for label, count in _top_dels:
                    pct = count / max(len(_sym_processed), 1) * 100
                    st.markdown(f"<div class='sym-bar'><div class='sym-bar-label'>{_esc(label)}</div><div class='sym-bar-track'><div class='sym-bar-fill del' style='width:{min(pct * 2, 100)}%;'></div></div><div class='sym-bar-count'>{pct:.0f}%</div></div>", unsafe_allow_html=True)
            st.markdown("</div>", unsafe_allow_html=True)

# ═══════════════════════════════════════════════════════════════════════════════
#  TAB: REVIEW EXPLORER
# ═══════════════════════════════════════════════════════════════════════════════
def _set_review_explorer_page(page: int, page_count: int):
    target = max(1, min(int(page), max(1, int(page_count))))
    st.session_state["review_explorer_page"] = target
    st.session_state["re_page_input"] = target


def _render_review_explorer(*, summary, overall_df, filtered_df, prompt_artifacts, filter_description, active_items):
    st.markdown("<div class='section-title'>Review Explorer</div>", unsafe_allow_html=True)
    st.markdown(f"<div class='section-sub'>Showing <strong>{len(filtered_df):,}</strong> reviews · Use sidebar filters to narrow the stream.</div>", unsafe_allow_html=True)
    bundle, bundle_error = _safe_get_master_bundle(
        summary,
        filtered_df,
        prompt_artifacts,
        active_items=active_items,
        filter_description=filter_description,
        export_scope_label="Filtered current view" if active_items else "Current view (all reviews)",
        total_loaded_reviews=len(overall_df),
    )
    tc = st.columns([1.45, 1.3, 0.9, 1.15, 0.85])
    if bundle is not None:
        tc[0].download_button("⬇️ Download current view", data=bundle["excel_bytes"], file_name=bundle["excel_name"], mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True, key="explorer_dl")
    else:
        tc[0].button("⬇️ Download current view", use_container_width=True, disabled=True, key="explorer_dl_disabled")
    sort_mode = tc[1].selectbox("Sort", ["Newest", "Oldest", "Highest rating", "Lowest rating", "Longest", "Most tagged", "Least tagged"], key="review_explorer_sort")
    per_page = int(tc[2].selectbox("Per page", [10, 20, 30, 50], key="review_explorer_per_page"))
    tc[3].markdown("<div style='height:4px'></div>", unsafe_allow_html=True)
    show_ev = tc[3].toggle("Evidence highlights", value=True, key="re_show_highlights", help="Highlight Symptomizer evidence in yellow — hover to see the AI tag")
    if bundle_error is not None:
        st.warning(f"Current-view export is temporarily unavailable: {bundle_error}")
    ordered_df = _sort_reviews(filtered_df, sort_mode)
    # Preserve original DataFrame index as a column for evidence lookup
    if "_orig_idx" not in ordered_df.columns:
        ordered_df = ordered_df.assign(_orig_idx=ordered_df.index)
    ordered_df = ordered_df.reset_index(drop=True)
    if ordered_df.empty:
        st.info("No reviews match the current filters.")
        return
    page_count = max(1, math.ceil(len(ordered_df) / max(per_page, 1)))
    current_page = max(1, min(int(st.session_state.get("review_explorer_page", 1)), page_count))
    if "re_page_input" not in st.session_state:
        st.session_state["re_page_input"] = current_page
    start = (current_page - 1) * per_page
    page_df = ordered_df.iloc[start:start + per_page]
    ev_lookup = _build_evidence_lookup(st.session_state.get("sym_processed_rows") or [])
    for orig_idx, row in page_df.iterrows():
        _oidx = str(row.get("_orig_idx", orig_idx))
        _rid = str(row.get("review_id", ""))
        ev_items = (ev_lookup.get(_oidx) or ev_lookup.get(_rid) or ev_lookup.get(str(orig_idx))) if show_ev else None
        _render_review_card(row, evidence_items=ev_items)
        st.markdown("<div style='height:6px'></div>", unsafe_allow_html=True)
    with st.container(border=True):
        pc = st.columns([0.8, 0.8, 2.6, 0.9, 0.8, 0.8])
        pc[0].button("⏮", use_container_width=True, disabled=current_page <= 1, key="re_first", on_click=_set_review_explorer_page, args=(1, page_count))
        pc[1].button("‹", use_container_width=True, disabled=current_page <= 1, key="re_prev", on_click=_set_review_explorer_page, args=(current_page - 1, page_count))
        pc[2].markdown(f"<div class='compact-pager-status'>Page {current_page} of {page_count:,}<span class='compact-pager-sub'>{start + 1:,}–{min(start + per_page, len(ordered_df)):,} of {len(ordered_df):,} reviews</span></div>", unsafe_allow_html=True)
        go_page = int(pc[3].number_input("Go", min_value=1, max_value=page_count, value=int(st.session_state.get("re_page_input", current_page)), step=1, key="re_page_input", label_visibility="collapsed"))
        pc[4].button("›", use_container_width=True, disabled=current_page >= page_count, key="re_next", on_click=_set_review_explorer_page, args=(current_page + 1, page_count))
        pc[5].button("⏭", use_container_width=True, disabled=current_page >= page_count, key="re_last", on_click=_set_review_explorer_page, args=(page_count, page_count))
    if go_page != current_page:
        _set_review_explorer_page(go_page, page_count)
        st.rerun()

# ═══════════════════════════════════════════════════════════════════════════════
#  TAB: AI ANALYST
# ═══════════════════════════════════════════════════════════════════════════════
# ═══════════════════════════════════════════════════════════════════════════════
#  TAB: REVIEW PROMPT
# ═══════════════════════════════════════════════════════════════════════════════
def _render_review_prompt_tab(*, settings, overall_df, filtered_df, summary, filter_description):
    st.markdown("<div class='section-title'>Review Prompt</div>", unsafe_allow_html=True)
    st.markdown("<div class='section-sub'>Create row-level AI tags that become new review columns. Start with the default starter pack, then tailor the prompts to the questions ELT, Product, Quality, or Insights need answered repeatedly.</div>", unsafe_allow_html=True)
    with st.container(border=True):
        st.markdown("**Prompt library**")
        st.markdown("<div class='helper-chip-row'><span class='helper-chip'>Loudness</span><span class='helper-chip'>Reliability</span><span class='helper-chip'>Usage count</span><span class='helper-chip'>Safety</span><span class='helper-chip'>Ownership period</span></div>", unsafe_allow_html=True)
        sc = st.columns([1.2, 1.2, 1])
        if sc[0].button("Add starter pack", use_container_width=True, key="prompt_add"):
            new_rows = pd.DataFrame(REVIEW_PROMPT_STARTER_ROWS)
            existing = set(st.session_state["prompt_definitions_df"]["column_name"].astype(str))
            to_add = new_rows[~new_rows["column_name"].isin(existing)]
            if not to_add.empty:
                st.session_state["prompt_definitions_df"] = pd.concat([st.session_state["prompt_definitions_df"], to_add], ignore_index=True)
            st.rerun()
        if sc[1].button("Reset to starter", use_container_width=True, key="prompt_reset"):
            st.session_state["prompt_definitions_df"] = pd.DataFrame(REVIEW_PROMPT_STARTER_ROWS)
            st.rerun()
        if sc[2].button("Clear all", use_container_width=True, key="prompt_clear"):
            st.session_state["prompt_definitions_df"] = pd.DataFrame(columns=["column_name", "prompt", "labels"])
            st.rerun()
    st.markdown("#### Prompt definitions")
    edited_df = st.data_editor(
        st.session_state["prompt_definitions_df"],
        use_container_width=True,
        num_rows="dynamic",
        hide_index=True,
        key="prompt_def_editor",
        height=320,
        column_config={
            "column_name": st.column_config.TextColumn("Column name", width="medium"),
            "prompt": st.column_config.TextColumn("Prompt", width="large"),
            "labels": st.column_config.TextColumn("Labels (comma-separated)", width="large"),
        },
    )
    st.session_state["prompt_definitions_df"] = edited_df
    _pd_sig = edited_df.to_json() if not edited_df.empty else ""
    _cached_defs = st.session_state.get("_prompt_defs_cache", {})
    if _cached_defs.get("sig") == _pd_sig and _cached_defs.get("cols") == list(overall_df.columns):
        prompt_defs = _cached_defs["defs"]
        prompt_defs_error = _cached_defs.get("error")
    else:
        prompt_defs_error = None
        try:
            prompt_defs = _normalize_prompt_defs(edited_df, overall_df.columns)
        except ReviewDownloaderError as exc:
            prompt_defs = []
            prompt_defs_error = str(exc)
        st.session_state["_prompt_defs_cache"] = dict(sig=_pd_sig, cols=list(overall_df.columns), defs=prompt_defs, error=prompt_defs_error)
    if prompt_defs_error:
        st.error(prompt_defs_error)
    api_key = settings.get("api_key")
    client = _get_client()
    with st.container(border=True):
        sc = st.columns([1.25, 1, 1, 2.45])
        tagging_scope = sc[0].selectbox("Scope", ["Current filtered reviews", "All loaded reviews"], index=0, key="prompt_tagging_scope")
        scope_df = filtered_df if tagging_scope == "Current filtered reviews" else overall_df
        batch_size = int(st.session_state.get("sym_batch_size", 8))
        est = math.ceil(len(scope_df) / max(1, batch_size)) if len(scope_df) else 0
        sc[1].metric("Reviews", f"{len(scope_df):,}")
        sc[2].metric("Requests", f"{est:,}")
        sc[3].caption(f"Scope: {tagging_scope.lower()} · {filter_description}")
        run_disabled = (not api_key) or (not prompt_defs) or len(scope_df) == 0
        if st.button("▶️ Run Review Prompt", type="primary", use_container_width=True, disabled=run_disabled, key="prompt_run_btn"):
            overlay = _show_thinking("Classifying each review…")
            try:
                prd = _run_review_prompt_tagging(client=client, source_df=scope_df.reset_index(drop=True), prompt_defs=prompt_defs, chunk_size=batch_size)
                updated = _merge_prompt_results(overall_df, prd, prompt_defs)
                dataset = dict(st.session_state["analysis_dataset"])
                dataset["reviews_df"] = updated
                st.session_state["analysis_dataset"] = dataset
                summary_df = _summarize_prompt_results(prd, prompt_defs, source_df=scope_df)
                defsig = json.dumps([dict(col=p["column_name"], prompt=p["prompt"], labels=p["labels"]) for p in prompt_defs], sort_keys=True)
                st.session_state["prompt_run_artifacts"] = dict(definitions=prompt_defs, summary_df=summary_df, scope_label=tagging_scope, scope_filter_description=filter_description, scope_review_ids=list(prd["review_id"].astype(str)), definition_signature=defsig, review_count=len(prd), generated_utc=pd.Timestamp.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC"))
                st.session_state["master_export_bundle"] = None
                st.session_state["_prompt_bundle_ready"] = False
                st.session_state["prompt_run_notice"] = f"Finished tagging {len(prd):,} reviews."
            except Exception as exc:
                st.error(f"Review Prompt run failed: {exc}")
            finally:
                overlay.empty()
            st.rerun()
    notice = st.session_state.pop("prompt_run_notice", None)
    if notice:
        st.success(notice)
    pa = st.session_state.get("prompt_run_artifacts")
    if not pa:
        st.info("Run Review Prompt to generate AI columns.")
        return
    cur_sig = json.dumps([dict(col=p["column_name"], prompt=p["prompt"], labels=p["labels"]) for p in prompt_defs], sort_keys=True) if prompt_defs else ""
    if cur_sig != pa.get("definition_signature"):
        st.info("Prompt definitions changed — re-run to refresh.")
    updated_overall = st.session_state["analysis_dataset"]["reviews_df"]
    hc = st.columns([1.4, 1.4, 4])
    hc[2].caption(f"Run: {pa.get('generated_utc')} · Scope: {pa.get('scope_label')} · Reviews: {pa.get('review_count'):,}")
    if hc[0].button("🔄 Prepare download", use_container_width=True, key="prompt_prep_dl"):
        try:
            with st.spinner("Building export…"):
                _get_master_bundle(summary, updated_overall, pa)
            st.session_state["_prompt_bundle_ready"] = True
        except Exception as exc:
            st.session_state["_prompt_bundle_ready"] = False
            st.error(f"Could not build export: {exc}")
        st.rerun()
    bundle = st.session_state.get("master_export_bundle")
    dl_ready = bundle is not None
    hc[1].download_button("⬇️ Download tagged file", data=bundle["excel_bytes"] if dl_ready else b"", file_name=bundle["excel_name"] if dl_ready else "tagged.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", disabled=not dl_ready, key="prompt_dl_btn")
    if not dl_ready:
        st.caption("Click **Prepare download** first to build the export file.")
    plookup = {p["display_name"]: p for p in pa["definitions"]}
    pnames = list(plookup.keys())
    if not pnames:
        st.info("No prompt results yet.")
        return
    if st.session_state.get("prompt_result_view") not in pnames:
        st.session_state["prompt_result_view"] = pnames[0]
    sel = st.radio("Prompt result view", options=pnames, horizontal=True, key="prompt_result_view", label_visibility="collapsed")
    prompt = plookup[sel]
    pc_col = prompt["column_name"]
    rids = set(str(x) for x in pa.get("scope_review_ids", []))
    result_scope = updated_overall.loc[updated_overall["review_id"].astype(str).isin(rids)] if rids else updated_overall.iloc[0:0]
    lopts = [str(l) for l in pa["summary_df"][pa["summary_df"]["column_name"] == pc_col]["label"].tolist()]
    sel_labels = st.multiselect("Labels", options=lopts, default=lopts, key=f"plabels_{pc_col}")
    if pc_col in result_scope.columns and not result_scope.empty:
        _slab = result_scope[pc_col]
        _view = result_scope[_slab.isin(sel_labels)] if sel_labels else result_scope.iloc[0:0]
        _vc = _slab.value_counts() if not result_scope.empty else pd.Series(dtype=int)
        _ar = result_scope.groupby(pc_col)["rating"].mean() if "rating" in result_scope.columns and not result_scope.empty else pd.Series(dtype=float)
    else:
        _view = result_scope.iloc[0:0]
        _vc = pd.Series(dtype=int)
        _ar = pd.Series(dtype=float)
    total = max(len(result_scope), 1)
    ps_rows = [dict(label=l, review_count=int(_vc.get(l, 0)), share=_safe_pct(int(_vc.get(l, 0)), total), avg_rating=float(_ar[l]) if l in _ar.index and pd.notna(_ar[l]) else None) for l in prompt["labels"]]
    ps = pd.DataFrame(ps_rows)
    cc, tc_col = st.columns([1.45, 1.05])
    with cc:
        with st.container(border=True):
            if ps.empty or ps["review_count"].sum() == 0:
                st.info("No tagged reviews match current filters.")
            else:
                fig = px.pie(ps, names="label", values="review_count", hole=0.44, color_discrete_sequence=["#6366f1", "#10b981", "#f59e0b", "#ef4444", "#3b82f6", "#8b5cf6"])
                fig.update_layout(margin=dict(l=20, r=20, t=20, b=20), paper_bgcolor="rgba(0,0,0,0)", font_family="Inter")
                _show_plotly(fig)
    with tc_col:
        with st.container(border=True):
            st.markdown(f"**Column** `{pc_col}`")
            st.write(prompt["prompt"])
            if not ps.empty:
                ds = ps.copy()
                ds["avg_rating"] = ds["avg_rating"].map(lambda x: f"{x:.2f}★" if pd.notna(x) and x is not None else "—")
                ds["share"] = ds["share"].map(_fmt_pct)
                st.dataframe(ds[["label", "review_count", "avg_rating", "share"]], use_container_width=True, hide_index=True, height=240)
    prevcols = [c for c in ["review_id", "rating", "incentivized_review", "submission_time", "content_locale", "title", "review_text", pc_col] if c in _view.columns]
    st.markdown("**Tagged review preview**")
    st.dataframe(_view[prevcols].head(50), use_container_width=True, hide_index=True, height=300)

# ═══════════════════════════════════════════════════════════════════════════════
#  TAB: SYMPTOMIZER
# ═══════════════════════════════════════════════════════════════════════════════
def _render_symptomizer_tab(*, settings, overall_df, filtered_df, summary, filter_description):
    _apply_pending_symptomizer_ui_state()
    st.markdown("<div class='section-title'>Symptomizer</div>", unsafe_allow_html=True)
    st.markdown("<div class='section-sub'>Row-level AI tagging of delighters and detractors. Use this when you want a more structured theme layer for the Dashboard, Review Explorer, and downstream exports.</div>", unsafe_allow_html=True)
    with st.expander("How the Symptomizer works", expanded=False):
        st.markdown(
            """
**What it does**

The Symptomizer reads each review as a set of evidence-backed product claims, maps those claims to your active L1/L2 taxonomy, and writes structured detractor and delighter tags back into the workspace and export file.

**Advanced AI behavior inside this run**

- Rich taxonomy guidance can tell the model exactly when each label should be tagged as a detractor, a delighter, or skipped.
- Long reviews can route through a two-stage pipeline: extract claims first, then map them to the taxonomy, then cross-check with the single-pass result.
- Polarity audits look for negation, contrast clauses, and concern-then-relief language before a tag survives.
- Adaptive batching keeps short reviews fast while shrinking batch size for longer, denser reviews.
- Sparse-result follow-up and targeted verification re-check the results most likely to be wrong instead of blindly reprocessing everything.
- Alias learning and product knowledge enrichment use what the run learned to make future taxonomy passes sharper.

**Why a highlight can still miss sometimes**

The tagger reads more than the main review body — it may use pros, cons, comments, title/body variants, or other visible fields when available. Highlights work best when the stored evidence is verbatim. They can still miss when the captured snippet was paraphrased, heavily normalized, split across multiple fields, or pulled from a fuzzy evidence match instead of one exact span. This build now recovers more verbatim spans during evidence validation and checks a broader set of visible review fields to make highlighting much more consistent.
            """
        )
    client = _get_client()
    api_key = settings.get("api_key")
    sym_source = st.session_state.get("sym_symptoms_source", "none")
    if sym_source == "none":
        _try_load_symptoms_from_file()
    delighters, detractors, _ = _sync_symptom_catalog_session(default_source="built-in")
    include_universal_prev = bool(st.session_state.get("sym_include_universal_neutral", True))
    sym_source = st.session_state.get("sym_symptoms_source", "none")
    st.markdown("### 1 · Symptoms catalog")
    with st.expander("🧩 Universal Neutral Symptoms", expanded=False):
        include_universal = st.checkbox(
            "Include Universal Neutral Symptoms in the active taxonomy",
            value=include_universal_prev,
            key="sym_include_universal_neutral",
            help="Checked by default. These cross-product tags stay available for AI tagging, inline editing, and exports.",
        )
        st.caption("Turn this off to remove the universal neutral pack from the catalog, current AI tags, inline suggestions, and the export file. You can also extend the pack with your own custom cross-product symptoms below.")
        neutral_dels, neutral_dets = _universal_neutral_catalog()
        custom_neutral_dels, custom_neutral_dets = _custom_universal_catalog()
        u1, u2 = st.columns(2)
        with u1:
            st.markdown("**🟢 Delighters**")
            st.markdown(_chip_html([(lab, "green") for lab in neutral_dels]), unsafe_allow_html=True)
        with u2:
            st.markdown("**🔴 Detractors**")
            st.markdown(_chip_html([(lab, "red") for lab in neutral_dets]), unsafe_allow_html=True)
        st.markdown(_chip_html([
            (f"Built-in: {len(_UNIVERSAL_NEUTRAL_DELIGHTERS) + len(_UNIVERSAL_NEUTRAL_DETRACTORS)}", "gray"),
            (f"Custom: {len(custom_neutral_dels) + len(custom_neutral_dets)}", "indigo"),
        ]), unsafe_allow_html=True)
        with st.form("sym_custom_universal_form"):
            cu1, cu2 = st.columns(2)
            with cu1:
                st.markdown("**Custom universal delighters**")
                custom_del_text = st.text_area(
                    "Add custom universal delighters",
                    value="\n".join(custom_neutral_dels),
                    height=120,
                    key="sym_custom_universal_del_text",
                    placeholder="One per line or comma-separated, e.g. Comfortable, Attractive Design",
                    label_visibility="collapsed",
                )
            with cu2:
                st.markdown("**Custom universal detractors**")
                custom_det_text = st.text_area(
                    "Add custom universal detractors",
                    value="\n".join(custom_neutral_dets),
                    height=120,
                    key="sym_custom_universal_det_text",
                    placeholder="One per line or comma-separated, e.g. Uncomfortable, Poor Design",
                    label_visibility="collapsed",
                )
            cf1, cf2 = st.columns(2)
            save_custom_universal = cf1.form_submit_button("Save custom universal symptoms", use_container_width=True)
            clear_custom_universal = cf2.form_submit_button("Clear custom universal symptoms", use_container_width=True)
        if save_custom_universal:
            saved_custom_dels, saved_custom_dets = _save_custom_universal_catalog(
                delighters=_parse_manual_tag_entries(custom_del_text),
                detractors=_parse_manual_tag_entries(custom_det_text),
                merge=False,
            )
            st.session_state["sym_run_notice"] = f"Saved {len(saved_custom_dels)} custom universal delighters and {len(saved_custom_dets)} custom universal detractors."
            st.rerun()
        if clear_custom_universal:
            _save_custom_universal_catalog(delighters=[], detractors=[], merge=False)
            st.session_state["sym_run_notice"] = "Cleared custom universal symptoms for this workspace."
            st.rerun()
    include_universal = bool(st.session_state.get("sym_include_universal_neutral", True))
    if include_universal != include_universal_prev:
        _apply_universal_neutral_toggle(include_universal)
        st.rerun()
    st.markdown(_chip_html([
        (f"✓ {len(delighters)} delighters", "green"),
        (f"✓ {len(detractors)} detractors", "red"),
        ("Universal neutral: On" if include_universal else "Universal neutral: Off", "blue" if include_universal else "gray"),
        (f"Source: {sym_source}", "indigo"),
    ]), unsafe_allow_html=True)
    universal_del_set, universal_det_set = map(set, _universal_neutral_catalog())
    product_specific_dels = [label for label in delighters if label not in universal_del_set]
    product_specific_dets = [label for label in detractors if label not in universal_det_set]
    if not product_specific_dels and not product_specific_dets:
        if include_universal:
            st.info("Using only the Universal Neutral Symptoms pack right now. Add product-specific symptoms below for better recall and cleaner reporting.")
        else:
            st.warning("No active symptoms are configured. Add product-specific symptoms below or re-enable Universal Neutral Symptoms.")
    # ── Taxonomy status ─────────────────────────────────────────────────
    # has_taxonomy is True only when product-specific symptoms exist,
    # NOT when only universal neutrals are loaded (source: built-in/none)
    _sym_source = st.session_state.get("sym_symptoms_source", "none")
    _has_product_specific = bool(product_specific_dels or product_specific_dets)
    has_taxonomy = _has_product_specific and _sym_source not in ("none", "built-in", "")
    if has_taxonomy:
        st.markdown(f"<div class='helper-chip-row'><span class='helper-chip' style='background:rgba(5,150,105,.10);color:#059669;border-color:rgba(5,150,105,.25);'>✅ Taxonomy active — {len(st.session_state.get('sym_detractors', []))} det · {len(st.session_state.get('sym_delighters', []))} del · Source: {_sym_source}</span></div>", unsafe_allow_html=True)
    sym_tabs = st.tabs(["🚀  Generate taxonomy", "✏️  Manual entry", "📄  Upload workbook"])
    with sym_tabs[0]:
        if not api_key:
            st.warning("OpenAI API key required.")
        else:
            # ── Unified 3-step wizard ─────────────────────────────────────
            wizard_step = st.session_state.get("sym_wizard_step", 1)
            ai_result = st.session_state.get("sym_ai_build_result")
            _wizard_forced = st.session_state.pop("_sym_wizard_force_step", None)
            if _wizard_forced is not None:
                wizard_step = _wizard_forced
            elif ai_result and wizard_step < 2:
                wizard_step = 2
            else:
                # Only skip to step 3 when product-specific taxonomy exists
                if _has_product_specific and _sym_source not in ("none", "built-in", "") and wizard_step < 3 and not ai_result:
                    wizard_step = 3

            # Step indicator — progress bar style
            steps = [("1", "Generate", wizard_step >= 1), ("2", "Review", wizard_step >= 2), ("3", "Run", wizard_step >= 3)]
            step_html = "<div style='display:flex;align-items:center;gap:4px;margin:0 0 18px;'>"
            for i, (num, label, active) in enumerate(steps):
                is_current = (wizard_step == int(num))
                bg = "#6366f1" if is_current else ("rgba(99,102,241,.15)" if active else "var(--color-border-tertiary)")
                fg = "#fff" if is_current else ("#6366f1" if active else "var(--color-text-tertiary)")
                fw = "700" if is_current else "500"
                step_html += f"<span style='display:inline-flex;align-items:center;justify-content:center;min-width:28px;height:28px;border-radius:14px;background:{bg};color:{fg};font-size:12px;font-weight:{fw};padding:0 10px;'>{num}. {label}</span>"
                if num != "3":
                    bar_color = "#6366f1" if active and int(num) < wizard_step else "var(--color-border-tertiary)"
                    step_html += f"<div style='flex:1;height:2px;background:{bar_color};min-width:20px;'></div>"
            step_html += "</div>"
            st.markdown(step_html, unsafe_allow_html=True)

            # ── STEP 1: Product description + Generate ────────────────────
            if wizard_step == 1:
                _has_existing_knowledge = bool(st.session_state.get("sym_product_knowledge"))
                if _has_existing_knowledge:
                    st.caption("Rebuilding with existing product knowledge. Your previous product description and knowledge are preserved.")

                pdesc = st.text_area(
                    "Product description (optional — AI will draft from reviews if blank)",
                    value=st.session_state.get("sym_product_profile", ""),
                    placeholder="e.g. SharkNinja FlexStyle HD440 — air styling tool with 4 attachments for curl, smooth, volumize, and dry",
                    height=80,
                    key="sym_pdesc",
                )
                if not overall_df.empty and "review_text" in overall_df.columns:
                    max_samples = min(250, max(5, len(overall_df)))
                    sample_n = st.slider("Sample reviews for taxonomy generation", min_value=5, max_value=max_samples, value=min(50, max_samples), step=5, key="sym_sample_n")
                    st.caption(f"The AI will read {sample_n} of {len(overall_df):,} reviews to understand the product and build the taxonomy.")
                else:
                    sample_n = 50

                product_knowledge = _normalize_product_knowledge(st.session_state.get("sym_product_knowledge") or {})
                sample_reviews = _sample_reviews_for_symptomizer(overall_df, sample_n)

                # Verify client is initialized before showing generate button
                _gen_blocked = False
                if client is None:
                    st.error("OpenAI client could not be initialized. Verify your API key is correct in Settings → OpenAI API Key.")
                    _gen_blocked = True
                elif len(sample_reviews) == 0:
                    st.warning("No reviews available. Load reviews first, then generate the taxonomy.")
                    _gen_blocked = True

                generate_label = "🚀 Generate taxonomy" if pdesc.strip() else "🚀 Generate taxonomy from reviews"
                if st.button(generate_label, type="primary", use_container_width=True, key="sym_generate_taxonomy", disabled=_gen_blocked):
                    overlay = _show_thinking("Step 1/2 — Analyzing product and reviews…")
                    _gen_success = False
                    try:
                        # Step 1: Generate/update product knowledge (also drafts description if empty)
                        desc_result = _ai_generate_product_description(client=client, sample_reviews=sample_reviews, existing_description=pdesc.strip())
                        if desc_result.get("description"):
                            st.session_state["sym_product_profile"] = desc_result["description"]
                            pk = _normalize_product_knowledge(desc_result.get("product_knowledge") or product_knowledge)
                            st.session_state["sym_product_knowledge"] = pk
                            st.session_state["sym_product_profile_ai_note"] = desc_result.get("confidence_note", "")
                        else:
                            pk = product_knowledge
                        overlay.empty()
                        overlay = _show_thinking("Step 2/2 — Building symptom taxonomy…")
                        # Step 2: Generate taxonomy from knowledge + reviews
                        tax_result = _ai_build_symptom_list(
                            client=client,
                            product_description=st.session_state.get("sym_product_profile", pdesc),
                            sample_reviews=sample_reviews,
                            product_knowledge=pk,
                        )
                        if not tax_result or (not tax_result.get("delighters") and not tax_result.get("detractors")):
                            raise ValueError("AI returned empty taxonomy — try adding a product description or increasing sample size")
                        st.session_state["sym_ai_build_result"] = tax_result
                        st.session_state["sym_product_knowledge"] = _normalize_product_knowledge(tax_result.get("product_knowledge") or pk)
                        # NOTE: Auto-calibration removed from generate flow for speed.
                        # Users can validate with 🧪 button in step 2 or section 2.
                        st.session_state["sym_wizard_step"] = 2
                        _gen_success = True
                    except Exception as exc:
                        st.error(f"Taxonomy generation failed: {exc}")
                        _log.error("Taxonomy generation failed: %s", exc, exc_info=True)
                    finally:
                        overlay.empty()
                    if _gen_success:
                        st.rerun()

                # Show existing product knowledge if available
                ai_note = st.session_state.get("sym_product_profile_ai_note", "")
                if ai_note:
                    st.caption(ai_note)
                if _has_visible_product_knowledge(product_knowledge):
                    with st.expander("📚 Generated product knowledge", expanded=False):
                        _render_product_knowledge_panel(product_knowledge)

            # ── STEP 2: Review taxonomy + Approve ─────────────────────────
            elif wizard_step == 2 and ai_result:
                product_knowledge = _normalize_product_knowledge(st.session_state.get("sym_product_knowledge") or {})
                preview_dels = list(ai_result.get("preview_delighters") or [])
                preview_dets = list(ai_result.get("preview_detractors") or [])
                del_bucket_counts = Counter(str(item.get("bucket") or "Product Specific") for item in preview_dels)
                det_bucket_counts = Counter(str(item.get("bucket") or "Product Specific") for item in preview_dets)

                st.markdown("**Generated taxonomy — review before running:**")
                # Show calibration result if available
                cr = st.session_state.get("sym_calibration_result")
                if cr:
                    rec = cr.get("recommendation", "")
                    hit = cr.get("hit_rate", 0)
                    avg = cr.get("avg_tags_per_review", 0)
                    unu = len(cr.get("unused_detractors", [])) + len(cr.get("unused_delighters", []))
                    if rec == "ready":
                        st.success(f"✅ Calibration passed — {hit:.0%} hit rate, {avg:.1f} avg tags/review. {unu} unused labels.")
                    elif rec == "needs_tuning":
                        st.warning(f"⚠️ Calibration: {hit:.0%} hit rate — consider editing. {unu} unused labels.")
                    else:
                        st.error(f"🔴 Low coverage — {hit:.0%} hit rate. Review the taxonomy before running.")
                st.markdown(_chip_html([
                    (f"Category: {str(ai_result.get('category') or 'general').replace('_', ' ').title()}", "blue"),
                    (f"Delighters: {len(ai_result.get('delighters', []))}", "green"),
                    (f"Detractors: {len(ai_result.get('detractors', []))}", "red"),
                    (f"Category drivers: {del_bucket_counts.get('Category Driver', 0) + det_bucket_counts.get('Category Driver', 0)}", "indigo"),
                    (f"Product specific: {del_bucket_counts.get('Product Specific', 0) + det_bucket_counts.get('Product Specific', 0)}", "gray"),
                ]), unsafe_allow_html=True)
                # Show diff vs old taxonomy if rebuilding
                _old_tax = st.session_state.pop("_sym_old_taxonomy", None)
                if _old_tax:
                    new_det_count = len(ai_result.get("detractors", []))
                    new_del_count = len(ai_result.get("delighters", []))
                    old_det_count = _old_tax.get("det_count", 0)
                    old_del_count = _old_tax.get("del_count", 0)
                    new_det_set = set(str(l.get("label", l) if isinstance(l, dict) else l).lower() for l in ai_result.get("detractors", []))
                    new_del_set = set(str(l.get("label", l) if isinstance(l, dict) else l).lower() for l in ai_result.get("delighters", []))
                    old_det_set = set(str(l).lower() for l in _old_tax.get("det_labels", []))
                    old_del_set = set(str(l).lower() for l in _old_tax.get("del_labels", []))
                    added_det = len(new_det_set - old_det_set)
                    removed_det = len(old_det_set - new_det_set)
                    added_del = len(new_del_set - old_del_set)
                    removed_del = len(old_del_set - new_del_set)
                    if added_det or removed_det or added_del or removed_del:
                        diff_parts = []
                        if added_det: diff_parts.append(f"+{added_det} det")
                        if removed_det: diff_parts.append(f"-{removed_det} det")
                        if added_del: diff_parts.append(f"+{added_del} del")
                        if removed_del: diff_parts.append(f"-{removed_del} del")
                        st.caption(f"Changes vs previous taxonomy: {' · '.join(diff_parts)} (was {old_det_count} det / {old_del_count} del)")
                if ai_result.get("taxonomy_note"):
                    st.caption(str(ai_result.get("taxonomy_note")))

                # Product knowledge summary
                if _has_visible_product_knowledge(product_knowledge):
                    with st.expander("📚 Generated product knowledge", expanded=False):
                        _render_product_knowledge_panel(product_knowledge)

                # Taxonomy preview
                preview_tabs = st.tabs(["🟢 Delighters", "🔴 Detractors"])
                with preview_tabs[0]:
                    _render_ai_taxonomy_preview_table(preview_dels, key_prefix="sym_ai_preview_del", side_label="Delighters")
                with preview_tabs[1]:
                    _render_ai_taxonomy_preview_table(preview_dets, key_prefix="sym_ai_preview_det", side_label="Detractors")

                # Optional edit
                with st.expander("✏️ Edit before running", expanded=False):
                    r1, r2 = st.columns(2)
                    with r1:
                        st.markdown("🟢 Delighters")
                        st.text_area("Edit", value="\n".join(ai_result.get("delighters", [])), height=220, key="sym_ai_del_edit")
                    with r2:
                        st.markdown("🔴 Detractors")
                        st.text_area("Edit", value="\n".join(ai_result.get("detractors", [])), height=220, key="sym_ai_det_edit")

                # Review count + Approve & Run
                st.divider()
                rc1, rc2, rc3 = st.columns([1.5, 1, 1.5])
                colmap = _detect_sym_cols(overall_df)
                work = _detect_missing(overall_df, colmap)
                available = len(work[work["Needs_Symptomization"] | work["Needs_Delighters"] | work["Needs_Detractors"]])
                max_reviews = max(1, min(available, len(overall_df)))
                n_reviews = rc1.number_input("Reviews to symptomize", min_value=1, max_value=max_reviews, value=min(max_reviews, 100), step=10, key="sym_wizard_n_reviews")
                batch_size = int(rc2.number_input("Batch size", min_value=1, max_value=20, value=int(st.session_state.get("sym_batch_size", 8)), step=1, key="sym_wizard_batch_size"))
                rc3.metric("Available reviews", f"{available:,}")

                btn_cols = st.columns([1.5, 1, 1])
                if btn_cols[0].button("✅ Approve & run symptomizer", type="primary", use_container_width=True, key="sym_approve_and_run"):
                    # Accept taxonomy
                    accepted_dels, accepted_dets = _canonical_symptom_catalog(
                        _parse_manual_tag_entries(st.session_state.get("sym_ai_del_edit", "\n".join(ai_result.get("delighters", [])))),
                        _parse_manual_tag_entries(st.session_state.get("sym_ai_det_edit", "\n".join(ai_result.get("detractors", [])))),
                    )
                    include_neutral = bool(st.session_state.get("sym_include_universal_neutral", True))
                    final_dets, final_dels = _ensure_universal_taxonomy(accepted_dets, accepted_dels, include_universal_neutral=include_neutral)
                    st.session_state.update(
                        sym_delighters=final_dels,
                        sym_detractors=final_dets,
                        sym_aliases=_alias_map_for_catalog(final_dels, final_dets, extra_aliases=ai_result.get("aliases") or {}, existing_aliases=st.session_state.get("sym_aliases", {})),
                        sym_symptoms_source="ai",
                        sym_taxonomy_preview_items=_taxonomy_preview_items_with_side(ai_result),
                        sym_taxonomy_category=str(ai_result.get("category") or "general"),
                        sym_product_profile_ai_note=(ai_result.get("taxonomy_note") or st.session_state.get("sym_product_profile_ai_note", "")),
                        sym_wizard_step=3,
                        sym_wizard_auto_run=True,
                        _sym_run_n_reviews=int(n_reviews),
                        _sym_run_batch_size=int(batch_size),
                    )
                    if taxonomy_edited_in_wizard:
                        st.session_state["sym_qa_user_edited"] = True
                    st.session_state.pop("sym_ai_build_result", None)
                    st.rerun()
                if btn_cols[1].button("🔄 Regenerate", use_container_width=True, key="sym_regenerate"):
                    st.session_state.pop("sym_ai_build_result", None)
                    st.session_state["sym_wizard_step"] = 1
                    st.session_state["_sym_wizard_force_step"] = 1
                    st.rerun()
                if btn_cols[2].button("← Back", use_container_width=True, key="sym_wizard_back"):
                    st.session_state["sym_wizard_step"] = 1
                    st.session_state["_sym_wizard_force_step"] = 1
                    st.rerun()

            # ── STEP 3: Ready to run / already has taxonomy ───────────────
            elif wizard_step >= 3 or has_taxonomy:
                st.session_state["sym_wizard_step"] = 3
                st.caption(f"Product-specific taxonomy loaded ({len(st.session_state.get('sym_detractors', []))} det · {len(st.session_state.get('sym_delighters', []))} del). Scroll down to Section 2 to run the symptomizer, or rebuild here.")
                if st.button("🔄 Rebuild taxonomy from scratch", use_container_width=True, key="sym_rebuild_taxonomy"):
                    # Store old taxonomy counts for diff display in step 2
                    old_dets = list(st.session_state.get("sym_detractors") or [])
                    old_dels = list(st.session_state.get("sym_delighters") or [])
                    st.session_state["_sym_old_taxonomy"] = {"det_count": len(old_dets), "del_count": len(old_dels), "det_labels": old_dets[:30], "del_labels": old_dels[:30]}
                    st.session_state["sym_wizard_step"] = 1
                    st.session_state["_sym_wizard_force_step"] = 1
                    st.session_state.pop("sym_ai_build_result", None)
                    st.rerun()
                _sym_knowledge = _normalize_product_knowledge(st.session_state.get("sym_product_knowledge") or {})
                if _has_visible_product_knowledge(_sym_knowledge):
                    with st.expander("📚 Generated product knowledge", expanded=False):
                        _render_product_knowledge_panel(_sym_knowledge)

            else:
                # Fallback: no ai_result but wizard_step=2 — reset
                st.session_state["sym_wizard_step"] = 1
                st.rerun()
    with sym_tabs[1]:
        if include_universal:
            st.caption("Universal Neutral Symptoms are managed by the toggle above, so this editor is just for product-specific symptoms.")
        with st.form("sym_manual_catalog_form"):
            c1, c2 = st.columns(2)
            with c1:
                st.markdown("🟢 **Delighters**")
                del_text = st.text_area("One per line or comma-separated", value="\n".join(product_specific_dels), height=200, key="sym_del_manual")
            with c2:
                st.markdown("🔴 **Detractors**")
                det_text = st.text_area("One per line or comma-separated", value="\n".join(product_specific_dets), height=200, key="sym_det_manual")
            save_manual = st.form_submit_button("💾 Save symptoms", use_container_width=True)
        if save_manual:
            saved_dels, saved_dets = _canonical_symptom_catalog(_parse_manual_tag_entries(del_text), _parse_manual_tag_entries(det_text))
            st.session_state.update(sym_delighters=saved_dels, sym_detractors=saved_dets, sym_aliases=_alias_map_for_catalog(saved_dels, saved_dets, existing_aliases=st.session_state.get("sym_aliases", {})), sym_symptoms_source="manual", sym_taxonomy_preview_items=[])
            st.session_state["sym_qa_user_edited"] = True
            st.success("Saved.")
            st.rerun()
    with sym_tabs[2]:
        st.markdown("Upload an Excel workbook with a **Symptoms** sheet: columns Symptom, Type (Delighter/Detractor), optional Aliases. If there is no Symptoms sheet but the review tab already contains populated **Symptom 1–20** or **AI Symptom** columns, the local catalog will be inferred automatically.")
        sym_upload = st.file_uploader("Upload workbook", type=["xlsx", "xlsm"], key="sym_file_uploader")
        if sym_upload:
            raw = sym_upload.getvalue()
            st.session_state["_uploaded_raw_bytes"] = raw
            d, t, a = _get_symptom_whitelists(raw)
            if d or t:
                loaded_dels, loaded_dets = _canonical_symptom_catalog(d, t)
                st.session_state.update(sym_delighters=loaded_dels, sym_detractors=loaded_dets, sym_aliases=_alias_map_for_catalog(loaded_dels, loaded_dets, extra_aliases=a, existing_aliases=st.session_state.get("sym_aliases", {})), sym_symptoms_source="file", sym_taxonomy_preview_items=[], sym_taxonomy_category="general")
                st.session_state["sym_qa_user_edited"] = True
                st.success(f"Loaded {len(loaded_dels)} delighters and {len(loaded_dets)} detractors.")
                st.rerun()
            else:
                st.error("Could not find a usable Symptoms sheet or any populated Symptom / AI Symptom columns.")
    current_catalog_rows = _build_structured_taxonomy_rows(
        delighters,
        detractors,
        aliases=st.session_state.get("sym_aliases", {}),
        category=str(st.session_state.get("sym_taxonomy_category") or (st.session_state.get("sym_ai_build_result") or {}).get("category") or "general"),
        preview_items=list(st.session_state.get("sym_taxonomy_preview_items") or _taxonomy_preview_items_with_side(st.session_state.get("sym_ai_build_result") or {})),
    )
    with st.expander(f"🗂 Current L1/L2 taxonomy ({len(current_catalog_rows)})", expanded=False):
        st.caption("This is the active symptom list the Symptomizer will use right now, organized as L1 themes and L2 symptoms.")
        _render_structured_taxonomy_table(current_catalog_rows, key_prefix="sym_current_catalog")
    st.divider()
    st.markdown("### 2 · Configure and run")
    delighters = list(st.session_state.get("sym_delighters") or [])
    detractors = list(st.session_state.get("sym_detractors") or [])
    colmap = _detect_sym_cols(overall_df)
    work = _detect_missing(overall_df, colmap)
    st.markdown(
        f"""<div class="hero-grid" style="grid-template-columns:minmax(220px,280px);margin-top:0;margin-bottom:.7rem;">
      <div class="hero-stat accent"><div class="label">Total reviews</div><div class="value">{len(overall_df):,}</div></div>
    </div>""",
        unsafe_allow_html=True,
    )
    active_taxonomy_count = len(delighters) + len(detractors)
    with st.container(border=True):
        st.markdown(
            "<div class='builder-kicker'>Run setup</div><div class='builder-title'>Choose what to tag</div><div class='builder-sub'>Pick the review set, how many reviews to process, and batch size. Counts update automatically as you change the controls.</div>",
            unsafe_allow_html=True,
        )
        cfg = st.columns([1.45, 1.0, 0.95, 0.85, 0.95])
        scope_choice = cfg[0].selectbox("Review set", ["Missing both", "Any missing", "Current filtered reviews", "All loaded reviews"], key="sym_scope_choice")
        if scope_choice == "Missing both":
            target_df = work[(work["Needs_Delighters"]) & (work["Needs_Detractors"])]
        elif scope_choice == "Any missing":
            target_df = work[(work["Needs_Delighters"]) | (work["Needs_Detractors"])]
        elif scope_choice == "Current filtered reviews":
            fids = set(filtered_df["review_id"].astype(str))
            target_df = work[work["review_id"].astype(str).isin(fids)]
        else:
            target_df = work
        max_reviews_in_scope = max(1, len(target_df))
        default_n = int(st.session_state.get("sym_n_to_process", min(100, max_reviews_in_scope)))
        default_n = min(max(1, default_n), max_reviews_in_scope)
        st.session_state["sym_n_to_process"] = default_n
        n_to_process = cfg[1].number_input("Reviews to process", min_value=1, max_value=max_reviews_in_scope, value=default_n, step=1, key="sym_n_to_process")
        batch_size_default = int(st.session_state.get("sym_batch_size_run", st.session_state.get("sym_batch_size", 8)) or 8)
        batch_size_default = min(20, max(1, batch_size_default))
        batch_size = int(cfg[2].number_input("Batch size", min_value=1, max_value=20, value=batch_size_default, step=1, key="sym_batch_size_run"))
        est_batches = max(1, math.ceil(int(n_to_process) / batch_size)) if n_to_process else 0
        cfg[3].metric("In scope", f"{len(target_df):,}")
        cfg[4].metric("Batches", f"{est_batches:,}")
        st.caption(f"Model: {_shared_model()} · Review set: {scope_choice}")
        run_disabled = (not api_key) or (len(target_df) == 0) or (active_taxonomy_count == 0)
        if run_disabled and not api_key:
            st.warning("Add OPENAI_API_KEY to Streamlit secrets.")
        elif len(target_df) == 0:
            st.info("No reviews match the selected review set.")
        elif active_taxonomy_count == 0:
            st.info("Add at least one active symptom or re-enable Universal Neutral Symptoms before running the Symptomizer.")
        run_btn = st.button(f"▶️ Symptomize {min(int(n_to_process), len(target_df)):,} review(s)", type="primary", use_container_width=True, disabled=run_disabled, key="sym_run_btn")
        if not run_disabled and int(n_to_process) > 0:
            _est_batches = max(1, math.ceil(int(n_to_process) / max(batch_size, 1)))
            _est_secs = _est_batches * 2.5 / min(3, _est_batches)
            _est_label = f"~{_est_secs:.0f}s" if _est_secs < 60 else f"~{_est_secs / 60:.1f}min"
            st.markdown(f"<div class='time-estimate'>⏱ Estimated tagging time: {_est_label} for {int(n_to_process)} reviews ({_est_batches} batches × {batch_size}/batch)</div>", unsafe_allow_html=True)
    # ── Taxonomy overlap check ────────────────────────────────────────────
    if active_taxonomy_count >= 6 and (detractors and delighters):
        overlap_warnings = []
        det_stems = {label: _tokenize(label) for label in detractors}
        del_stems = {label: _tokenize(label) for label in delighters}
        for d_label, d_tokens in det_stems.items():
            for l_label, l_tokens in del_stems.items():
                if d_tokens and l_tokens:
                    shared = set(d_tokens) & set(l_tokens)
                    if len(shared) >= 2 and len(shared) / max(len(d_tokens), len(l_tokens)) > 0.5:
                        overlap_warnings.append(f"'{d_label}' ↔ '{l_label}' ({', '.join(shared)})")
        if overlap_warnings:
            with st.expander(f"⚠️ {len(overlap_warnings)} potential taxonomy overlaps", expanded=False):
                st.caption("These detractor/delighter pairs share significant keyword overlap and may cause the AI to confuse them:")
                for w in overlap_warnings[:8]:
                    st.markdown(f"- {w}")
    # Pre-resolve variables needed by both calibration and run
    profile = st.session_state.get("sym_product_profile", "")
    product_knowledge = _normalize_product_knowledge(st.session_state.get("sym_product_knowledge") or {})
    client = _get_client()
    if _HAS_SYMPTOMIZER_V3 and api_key and active_taxonomy_count > 0 and len(target_df) > 0:
        cc1, cc2 = st.columns([1, 3])
        if cc1.button("🧪 Validate taxonomy", use_container_width=True, key="sym_calibrate_btn", help="8-review calibration to check taxonomy fit."):
            with st.spinner("Calibrating…"):
                calib = _v3_calibration_preflight(client=client, sample_reviews=_sample_reviews_for_symptomizer(filtered_df, 8), allowed_detractors=list(detractors or []), allowed_delighters=list(delighters or []), product_profile=profile, chat_complete_fn=_chat_complete_with_fallback_models, safe_json_load_fn=_safe_json_load, model_fn=_shared_model, reasoning_fn=_shared_reasoning)
                st.session_state["sym_calibration_result"] = calib
        cr = st.session_state.get("sym_calibration_result")
        if cr:
            rec, hit, avg, unu = cr.get("recommendation",""), cr.get("hit_rate",0), cr.get("avg_tags_per_review",0), len(cr.get("unused_detractors",[]))+len(cr.get("unused_delighters",[]))
            if rec == "ready": cc2.success(f"✅ Taxonomy looks good — {hit:.0%} hit rate, {avg:.1f} avg tags. {unu} unused labels.")
            elif rec == "needs_tuning": cc2.warning(f"⚠️ Needs tuning — {hit:.0%} hit rate. Consider removing {unu} unused labels.")
            else: cc2.error(f"🔴 Low coverage — {hit:.0%} hit rate. Review taxonomy before full batch.")
            with st.expander("📊 Calibration details", expanded=False): st.json(cr)
    notice = st.session_state.pop("sym_run_notice", None)
    if notice:
        st.success(notice)
    # ── Auto-run from wizard step 2 "Approve & Run" ──────────────────
    wizard_auto_run = st.session_state.pop("sym_wizard_auto_run", False)
    if wizard_auto_run:
        n_to_process = st.session_state.pop("_sym_run_n_reviews", int(n_to_process))
        batch_size = st.session_state.pop("_sym_run_batch_size", batch_size)
        run_btn = True  # Force run
    if run_btn:
        prioritized = _prioritize_for_symptomization(target_df).head(int(n_to_process))
        rows_to_process = prioritized.copy()
        prog = st.progress(0.0, text="Preparing reviews…")
        status = st.empty()
        eta_box = st.empty()
        stats_box = st.empty()
        processed_local = []
        t0 = time.perf_counter()
        total_n = max(1, len(rows_to_process))
        done = 0
        failed_count = 0
        total_labels_written = 0
        updated_df = _ensure_ai_cols(overall_df.copy())
        profile = st.session_state.get("sym_product_profile", "")
        product_knowledge = _normalize_product_knowledge(st.session_state.get("sym_product_knowledge") or {})
        aliases = st.session_state.get("sym_aliases", {})
        rows_list = list(rows_to_process.iterrows())
        bidxs = list(range(0, len(rows_list), batch_size))
        empty_out = dict(dels=[], dets=[], ev_del={}, ev_det={}, unl_dels=[], unl_dets=[], safety="Not Mentioned", reliability="Not Mentioned", sessions="Unknown")
        _active_dets = list(detractors or [])
        _active_dels = list(delighters or [])
        _include_universal = bool(st.session_state.get("sym_include_universal_neutral", True))
        _ev_chars = int(st.session_state.get("sym_max_ev_chars", 120))
        _label_tracker = _v3_LabelTracker(_active_dets, _active_dels) if _HAS_SYMPTOMIZER_V3 else None
        if _HAS_SYMPTOMIZER_V3:
            avg_words = rows_to_process.get("review_length_words", pd.Series(50)).fillna(50).mean()
            if avg_words < 60 and batch_size < 12:
                batch_size = min(12, batch_size + 4)
            elif avg_words < 100 and batch_size < 10:
                batch_size = min(10, batch_size + 2)
            bidxs = list(range(0, len(rows_list), batch_size))
            if avg_words < 100:
                status.info(f"Avg {avg_words:.0f} words/review — using batch size {batch_size}.")

        _pre_category = st.session_state.get("sym_taxonomy_category", "")
        if _pre_category and _pre_category != "general":
            st.session_state["_sym_cached_category"] = _pre_category

        from concurrent.futures import ThreadPoolExecutor, as_completed
        _MAX_WORKERS = min(3, max(1, len(bidxs)))

        def _process_one_batch(bi_start_pair):
            bi, start = bi_start_pair
            batch = rows_list[start:start + batch_size]
            items = [
                dict(
                    idx=int(idx),
                    review=_symptomizer_review_text(row),
                    rating=row.get("rating"),
                    pros=_safe_text(row.get("pros")),
                    cons=_safe_text(row.get("cons")),
                )
                for idx, row in batch
            ]
            _adaptive_profile = profile
            if _label_tracker and bi > 3:
                hints = _label_tracker.get_prompt_hints()
                if hints:
                    _adaptive_profile = profile + "\n" + hints
            outs = {}
            _failed = 0
            if client:
                try:
                    outs = _call_symptomizer_batch(
                        client=client,
                        items=items,
                        allowed_delighters=_active_dels,
                        allowed_detractors=_active_dets,
                        product_profile=_adaptive_profile,
                        product_knowledge=product_knowledge,
                        max_ev_chars=_ev_chars,
                        aliases=aliases,
                        include_universal_neutral=_include_universal,
                    )
                except Exception as exc:
                    _log.warning("Batch %d failed (%s) — retrying individually", bi, exc)
                    _failed = len(items)
                    for it in items:
                        try:
                            single = _call_symptomizer_batch(
                                client=client,
                                items=[it],
                                allowed_delighters=_active_dels,
                                allowed_detractors=_active_dets,
                                product_profile=_adaptive_profile,
                                product_knowledge=product_knowledge,
                                max_ev_chars=_ev_chars,
                                aliases=aliases,
                                include_universal_neutral=_include_universal,
                            )
                            outs.update(single)
                            _failed -= 1
                        except Exception:
                            pass
            return bi, items, outs, _failed

        batch_inputs = list(enumerate(bidxs, 1))
        all_batch_results = []
        total_batches = max(1, len(batch_inputs))
        checkpoint_every = 2 if total_batches > 2 else 1

        if _MAX_WORKERS > 1 and len(bidxs) > 1:
            status.info(f"Tagging {total_n:,} reviews across {len(bidxs)} batches in adaptive waves ({_MAX_WORKERS} concurrent)…")
            for wave_start in range(0, len(batch_inputs), _MAX_WORKERS):
                wave = batch_inputs[wave_start:wave_start + _MAX_WORKERS]
                wave_results = []
                with ThreadPoolExecutor(max_workers=min(_MAX_WORKERS, len(wave))) as executor:
                    futures = {executor.submit(_process_one_batch, bp): bp for bp in wave}
                    for future in as_completed(futures):
                        wave_results.append(future.result())
                for result in sorted(wave_results, key=lambda x: x[0]):
                    all_batch_results.append(result)
                    if _label_tracker:
                        try:
                            _label_tracker.record_batch(result[2])
                        except Exception:
                            pass
                done_batches = len(all_batch_results)
                prep_progress = 0.65 * (done_batches / max(len(bidxs), 1))
                prog.progress(prep_progress, text=f"Batch {done_batches}/{len(bidxs)} complete")
        else:
            status.info(f"Tagging {total_n:,} reviews across {len(bidxs)} batch{'es' if len(bidxs) != 1 else ''}…")
            for bp in batch_inputs:
                result = _process_one_batch(bp)
                all_batch_results.append(result)
                if _label_tracker:
                    try:
                        _label_tracker.record_batch(result[2])
                    except Exception:
                        pass
                done_batches = len(all_batch_results)
                prep_progress = 0.65 * (done_batches / max(len(bidxs), 1))
                prog.progress(prep_progress, text=f"Batch {done_batches}/{len(bidxs)} complete")

        for bi, items, outs, batch_failed in sorted(all_batch_results, key=lambda x: x[0]):
            failed_count += batch_failed
            for it in items:
                idx = int(it["idx"])
                out = outs.get(idx, empty_out)
                dets_out = list(out.get("dets", []))[:10]
                dels_out = list(out.get("dels", []))[:10]
                updated_df = _write_ai_symptom_row(
                    updated_df,
                    idx,
                    dets=dets_out,
                    dels=dels_out,
                    safety=out.get("safety", "Not Mentioned"),
                    reliability=out.get("reliability", "Not Mentioned"),
                    sessions=out.get("sessions", "Unknown"),
                )
                actual_row = updated_df.loc[idx]
                final_dets = _normalize_tag_list(_collect_row_symptom_tags(actual_row, AI_DET_HEADERS))
                final_dels = _normalize_tag_list(_collect_row_symptom_tags(actual_row, AI_DEL_HEADERS))
                total_labels_written += len(dets_out or []) + len(dels_out or [])
                for lab in (out.get("unl_dels", []) or []):
                    _record_new_symptom_candidate(lab, idx=idx, side="delighter")
                for lab in (out.get("unl_dets", []) or []):
                    _record_new_symptom_candidate(lab, idx=idx, side="detractor")
                processed_local.append(
                    dict(
                        idx=idx,
                        review_id=str(actual_row.get("review_id", "")),
                        wrote_dets=final_dets,
                        wrote_dels=final_dels,
                        safety=out.get("safety", ""),
                        reliability=out.get("reliability", ""),
                        sessions=out.get("sessions", ""),
                        ev_det=out.get("ev_det", {}),
                        ev_del=out.get("ev_del", {}),
                        unl_dels=out.get("unl_dels", []),
                        unl_dets=out.get("unl_dets", []),
                    )
                )
                done += 1
            if bi % checkpoint_every == 0 or bi == total_batches:
                dataset_ck = dict(st.session_state["analysis_dataset"])
                dataset_ck["reviews_df"] = updated_df.copy()
                st.session_state["analysis_dataset"] = dataset_ck
                st.session_state["sym_processed_rows"] = list(processed_local)
                st.session_state["_sym_checkpoint_done"] = done
            elapsed = time.perf_counter() - t0
            rate = done / elapsed if elapsed > 0 else 0
            rem = (total_n - done) / rate if rate > 0 else 0
            apply_progress = 0.65 + (0.27 * (done / max(total_n, 1)))
            prog.progress(min(apply_progress, 0.92), text=f"{done}/{total_n} reviews tagged")
            eta_box.markdown(f"**Tagging speed:** {rate * 60:.1f} rev/min · **Tagging ETA:** ~{_fmt_secs(rem)}")
            avg_labels = total_labels_written / max(done, 1)
            stat_lines = [
                f"Labels written: <b>{total_labels_written}</b> · Avg per review: <b>{avg_labels:.1f}</b>"
                + (f" · ⚠️ {failed_count} failed" if failed_count > 0 else "")
            ]
            if _label_tracker:
                alerts = _label_tracker.check_alerts(min_reviews=batch_size * 3)
                for alert in alerts[:3]:
                    if alert["issue"] == "too_broad":
                        stat_lines.append(f"⚠️ '{_esc(alert['label'])}' is hitting {alert['pct']:.0f}% of reviews and may be too broad.")
                    elif alert["issue"] == "high_zero_rate":
                        stat_lines.append(f"⚠️ {alert['pct']:.0f}% of reviews returned zero tags — the taxonomy may need tuning.")
                    elif alert["issue"] == "zero_hits":
                        stat_lines.append(f"ℹ️ '{_esc(alert['label'])}' ({_esc(alert['side'])}) still has zero hits after {_label_tracker.total_reviews} reviews.")
                if bi > 3 and _label_tracker.get_prompt_hints():
                    stat_lines.append("🧠 Adaptive prompting is active and smoothing dominant or missing labels.")
            stats_box.markdown("<div class='status-note'>" + "<br>".join(stat_lines) + "</div>", unsafe_allow_html=True)
            if bi % 4 == 0 or bi == total_batches:
                gc.collect()

        def _needs_sparse_follow_up(rec):
            if not _HAS_SYMPTOMIZER_V3:
                return (not rec.get("wrote_dets")) or (not rec.get("wrote_dels"))
            if rec.get("idx") not in updated_df.index:
                return (not rec.get("wrote_dets")) or (not rec.get("wrote_dels"))
            source_row = updated_df.loc[rec.get("idx")]
            review_text = _symptomizer_review_text(source_row)
            needs_det, needs_del = _v3_gate_polarity(source_row.get("rating"), review_text)
            return (needs_det and not rec.get("wrote_dets")) or (needs_del and not rec.get("wrote_dels"))

        follow_up_candidates = sum(1 for rec in processed_local if _needs_sparse_follow_up(rec))
        status.info(
            "Tagging complete. Finalizing results"
            + (f" · focused follow-up is checking {follow_up_candidates:,} under-tagged review(s)" if follow_up_candidates else "")
            + "…"
        )
        eta_box.markdown(
            "<div class='status-note'>The last step saves results, runs a targeted sparse-result audit only on reviews that still look under-tagged, then refreshes the tables and export file. This stage can feel slower because it focuses on the hardest edge cases instead of the easy bulk tagging work.</div>",
            unsafe_allow_html=True,
        )
        prog.progress(0.95, text=f"{done}/{total_n} tagged · syncing and auditing edge cases")

        retry_changed = 0
        if _HAS_SYMPTOMIZER_V3 and client and processed_local:
            try:
                retry_t0 = time.perf_counter()
                retry_total = max(int(follow_up_candidates or 0), 0)

                def _retry_progress(done_retry, total_retry, phase):
                    total_retry = max(int(total_retry or 0), 0)
                    checked = min(max(int(done_retry or 0), 0), total_retry or max(int(done_retry or 0), 0))
                    elapsed_retry = time.perf_counter() - retry_t0
                    rate_retry = checked / elapsed_retry if elapsed_retry > 0 else 0.0
                    rem_retry = (max(total_retry - checked, 0) / rate_retry) if rate_retry > 0 and total_retry > 0 else 0.0
                    progress = 0.97 + (0.015 * (checked / max(total_retry, 1))) if total_retry else 0.982
                    msg = f"{done}/{total_n} tagged · auditing sparse results"
                    if total_retry:
                        msg += f" ({checked}/{total_retry})"
                    prog.progress(min(progress, 0.985), text=msg)
                    if total_retry:
                        eta_box.markdown(
                            f"**Sparse-result audit:** {checked}/{total_retry} review(s) checked · **Speed:** {rate_retry * 60:.1f} rev/min · **ETA:** ~{_fmt_secs(rem_retry)}"
                        )
                    elif phase == "queued":
                        eta_box.markdown(
                            "<div class='status-note'>No sparse-result follow-up was needed, so the app is moving straight into result sync and export prep.</div>",
                            unsafe_allow_html=True,
                        )

                _retry_progress(0, retry_total, "queued")
                all_items = {}
                for bs in bidxs:
                    for ri, rw in rows_list[bs:bs + batch_size]:
                        all_items[int(ri)] = dict(idx=int(ri), review=_symptomizer_review_text(rw), rating=rw.get("rating"))
                batch_res = {int(r["idx"]): r for r in processed_local}
                retry_res = _v3_retry_zero_tags(
                    client=client,
                    results=batch_res,
                    items=list(all_items.values()),
                    allowed_detractors=_active_dets,
                    allowed_delighters=_active_dels,
                    aliases=aliases,
                    max_ev_chars=_ev_chars,
                    chat_complete_fn=_chat_complete_with_fallback_models,
                    safe_json_load_fn=_safe_json_load,
                    model_fn=_shared_model,
                    reasoning_fn=_shared_reasoning,
                    max_workers=min(4, max(1, retry_total)),
                    progress_callback=_retry_progress,
                )
                for ri, nr in retry_res.items():
                    prev = batch_res.get(int(ri), {})
                    prev_dets = list(prev.get("wrote_dets") or [])
                    prev_dels = list(prev.get("wrote_dels") or [])
                    new_dets = list(nr.get("dets") or [])[:10]
                    new_dels = list(nr.get("dels") or [])[:10]
                    if new_dets != prev_dets or new_dels != prev_dels:
                        retry_changed += 1
                        if ri in updated_df.index:
                            updated_df = _write_ai_symptom_row(
                                updated_df,
                                ri,
                                dets=new_dets,
                                dels=new_dels,
                                safety=nr.get("safety"),
                                reliability=nr.get("reliability"),
                                sessions=nr.get("sessions"),
                            )
                        processed_local = _upsert_processed_symptom_record(
                            processed_local,
                            ri,
                            new_dets,
                            new_dels,
                            row_meta={
                                "AI Safety": nr.get("safety", prev.get("safety", "Not Mentioned")),
                                "AI Reliability": nr.get("reliability", prev.get("reliability", "Not Mentioned")),
                                "AI # of Sessions": nr.get("sessions", prev.get("sessions", "Unknown")),
                            },
                            ev_det=nr.get("ev_det"),
                            ev_del=nr.get("ev_del"),
                        )
                if retry_changed:
                    _log.info("Auto-retry recovered tags for %d review(s)", retry_changed)
                    status.info(f"Focused follow-up recovered tags for {retry_changed:,} review(s).")
            except Exception as retry_exc:
                _log.warning("Auto-retry failed: %s", retry_exc)

        prog.progress(0.985, text=f"{done}/{total_n} tagged · saving results")
        dataset = dict(st.session_state["analysis_dataset"])
        dataset["reviews_df"] = updated_df
        qa_baseline_map = _build_symptom_baseline_map(processed_local)
        qa_metrics = _compute_tag_edit_accuracy(qa_baseline_map, qa_baseline_map)
        final_label_count = sum(len(r.get("wrote_dets", [])) + len(r.get("wrote_dels", [])) for r in processed_local)
        elapsed_total = time.perf_counter() - t0
        st.session_state.update(
            analysis_dataset=dataset,
            sym_processed_rows=processed_local,
            master_export_bundle=None,
            sym_qa_baseline_map=qa_baseline_map,
            sym_qa_row_ids=list(qa_baseline_map.keys()),
            sym_qa_selected_row=(next(iter(qa_baseline_map.keys()), None)),
            sym_qa_accuracy=qa_metrics,
            sym_qa_user_edited=False,
            sym_qa_notice=None,
            sym_export_bytes=None,
        )
        st.session_state["sym_last_run_stats"] = {
            "reviews": done,
            "labels": final_label_count,
            "failed": failed_count,
            "avg_per_review": round(final_label_count / max(done, 1), 1),
            "elapsed_sec": round(elapsed_total, 1),
            "model": _shared_model(),
            "reasoning": _shared_reasoning(),
            "pipeline": (
                "v4 hybrid"
                if bool(st.session_state.get("sym_v4_pipeline", True))
                else ("staged" if bool(st.session_state.get("sym_staged_pipeline")) else "single-pass")
            ),
            "cache_stats": _v3_result_cache.stats if _HAS_SYMPTOMIZER_V3 else {},
            "retry_recovered": retry_changed,
        }

        try:
            enrichments = _enrich_product_knowledge_from_run(processed_local, min_mentions=max(3, done // 20))
            enrichment_count = sum(len(v) for v in enrichments.values())
            if enrichment_count > 0:
                st.session_state["sym_last_run_stats"]["knowledge_enriched"] = enrichment_count
                _log.info("Knowledge enrichment: %d new themes discovered", enrichment_count)
        except Exception as enrich_exc:
            _log.warning("Knowledge enrichment failed: %s", enrich_exc)
        try:
            aliases_learned = _auto_learn_aliases_from_run(processed_local)
            if aliases_learned > 0:
                st.session_state["sym_last_run_stats"]["aliases_learned"] = aliases_learned
        except Exception as alias_exc:
            _log.warning("Alias learning failed: %s", alias_exc)

        prog.progress(1.0, text=f"{done}/{total_n} tagged · complete")
        eta_box.markdown(
            "<div class='status-note'>Finished. Results below are already synced to the Dashboard and Review Explorer.</div>",
            unsafe_allow_html=True,
        )
        status.success(
            f"✅ Symptomized {done:,} reviews — {final_label_count} labels written ({final_label_count / max(done, 1):.1f} avg/review)."
            + (f" Focused follow-up recovered {retry_changed} review(s)." if retry_changed else "")
        )
        _log.info(
            "Symptomized %d reviews — %d labels written (%.1f avg)",
            done,
            final_label_count,
            final_label_count / max(done, 1),
        )
    st.divider()
    processed = st.session_state.get("sym_processed_rows") or []
    if not processed:
        st.info("Run the Symptomizer above to see results here.")
        return
    # ── Last run summary ─────────────────────────────────────────────────
    last_stats = st.session_state.get("sym_last_run_stats")
    if last_stats:
        sc1, sc2, sc3 = st.columns(3)
        sc1.metric("Reviews tagged", f"{last_stats.get('reviews', 0):,}")
        sc2.metric("Labels written", f"{last_stats.get('labels', 0):,} ({last_stats.get('avg_per_review', 0):.1f}/review)")
        pipeline_label = str(last_stats.get("pipeline") or ("staged" if last_stats.get("staged") else "single-pass")).title()
        sc3.metric("Speed", f"{last_stats.get('reviews', 0) / max(last_stats.get('elapsed_sec', 1), 0.1) * 60:.0f}/min · {pipeline_label}")
        # Sub-stats row: cache + knowledge enrichment
        sub_chips = []
        cache_s = last_stats.get("cache_stats", {})
        if cache_s.get("hits", 0) > 0:
            sub_chips.append((f"Cache: {cache_s['hits']} hits saved", "indigo"))
        rr = last_stats.get("retry_recovered", 0)
        if rr > 0:
            sub_chips.append((f"Follow-up recovered: {rr}", "green"))
        ke = last_stats.get("knowledge_enriched", 0)
        if ke > 0:
            sub_chips.append((f"Knowledge enriched: {ke} new themes", "green"))
        al = last_stats.get("aliases_learned", 0)
        if al > 0:
            sub_chips.append((f"Aliases learned: {al} consumer terms", "blue"))
        if sub_chips:
            st.markdown(_chip_html(sub_chips), unsafe_allow_html=True)
    st.markdown("### 3 · Results")
    total_tags = sum(len(r.get("wrote_dets", [])) + len(r.get("wrote_dels", [])) for r in processed)
    st.markdown(_chip_html([(f"{len(processed)} reviews tagged", "green"), (f"{total_tags} labels written", "indigo")]), unsafe_allow_html=True)

    # ── Tag distribution — visible without clicking ──────────────────
    if processed and total_tags > 0:
        det_freq = {}
        del_freq = {}
        zero_tag_count = 0
        for rec in processed:
            dets = rec.get("wrote_dets", [])
            dels = rec.get("wrote_dels", [])
            if not dets and not dels:
                zero_tag_count += 1
            for t in dets:
                det_freq[t] = det_freq.get(t, 0) + 1
            for t in dels:
                del_freq[t] = del_freq.get(t, 0) + 1
        top_dets = sorted(det_freq.items(), key=lambda x: -x[1])[:8]
        top_dels = sorted(del_freq.items(), key=lambda x: -x[1])[:8]

        det_col, del_col = st.columns(2)
        with det_col:
            if top_dets:
                st.markdown("<div style='font-size:12px;font-weight:700;color:var(--danger,#ef4444);text-transform:uppercase;letter-spacing:.05em;margin-bottom:6px;'>Top detractors</div>", unsafe_allow_html=True)
                for label, count in top_dets:
                    pct = count / max(len(processed), 1) * 100
                    bar_w = min(pct * 2, 100)
                    st.markdown(
                        f"<div style='display:flex;align-items:center;gap:8px;margin-bottom:3px;'>"
                        f"<div style='flex:1;font-size:12px;color:var(--color-text-primary);white-space:nowrap;overflow:hidden;text-overflow:ellipsis;'>{_esc(label)}</div>"
                        f"<div style='width:120px;height:14px;background:var(--color-border-tertiary,#eee);border-radius:7px;overflow:hidden;flex-shrink:0;'>"
                        f"<div style='width:{bar_w}%;height:100%;background:rgba(239,68,68,.7);border-radius:7px;'></div></div>"
                        f"<span style='font-size:11px;color:var(--color-text-secondary);min-width:36px;text-align:right;'>{count} ({pct:.0f}%)</span>"
                        f"</div>", unsafe_allow_html=True)
        with del_col:
            if top_dels:
                st.markdown("<div style='font-size:12px;font-weight:700;color:var(--success,#10b981);text-transform:uppercase;letter-spacing:.05em;margin-bottom:6px;'>Top delighters</div>", unsafe_allow_html=True)
                for label, count in top_dels:
                    pct = count / max(len(processed), 1) * 100
                    bar_w = min(pct * 2, 100)
                    st.markdown(
                        f"<div style='display:flex;align-items:center;gap:8px;margin-bottom:3px;'>"
                        f"<div style='flex:1;font-size:12px;color:var(--color-text-primary);white-space:nowrap;overflow:hidden;text-overflow:ellipsis;'>{_esc(label)}</div>"
                        f"<div style='width:120px;height:14px;background:var(--color-border-tertiary,#eee);border-radius:7px;overflow:hidden;flex-shrink:0;'>"
                        f"<div style='width:{bar_w}%;height:100%;background:rgba(16,185,129,.7);border-radius:7px;'></div></div>"
                        f"<span style='font-size:11px;color:var(--color-text-secondary);min-width:36px;text-align:right;'>{count} ({pct:.0f}%)</span>"
                        f"</div>", unsafe_allow_html=True)

        # Quick stats row
        _qs_chips = []
        if zero_tag_count > 0:
            _qs_chips.append((f"{zero_tag_count} reviews with 0 tags", "yellow"))
        _unl_dets = set()
        _unl_dels = set()
        for rec in processed:
            for u in (rec.get("unl_dets") or []):
                _unl_dets.add(str(u))
            for u in (rec.get("unl_dels") or []):
                _unl_dels.add(str(u))
        if _unl_dets or _unl_dels:
            _qs_chips.append((f"{len(_unl_dets) + len(_unl_dels)} unlisted themes discovered", "blue"))
        if _qs_chips:
            st.markdown(_chip_html(_qs_chips), unsafe_allow_html=True)

        # ── Quick AI insights from tagging results ────────────────────
        if top_dets or top_dels:
            if st.button("✨ Generate insights from tags", use_container_width=True, key="sym_quick_insights"):
                _qi_client = _get_client()
                if _qi_client:
                    with st.spinner("Generating insights…"):
                        _qi_det_str = ", ".join(f"{l} ({c}/{len(processed)}, {c/max(len(processed),1)*100:.0f}%)" for l, c in top_dets)
                        _qi_del_str = ", ".join(f"{l} ({c}/{len(processed)}, {c/max(len(processed),1)*100:.0f}%)" for l, c in top_dels)
                        _qi_zero_str = f"{zero_tag_count} reviews ({zero_tag_count/max(len(processed),1)*100:.0f}%) had zero tags." if zero_tag_count else ""
                        _qi_prompt = f"""Analyze these symptomizer results from {len(processed)} consumer product reviews.
Product: {st.session_state.get('sym_product_profile', 'Unknown product')}

Top detractors: {_qi_det_str}
Top delighters: {_qi_del_str}
{_qi_zero_str}

Write 3-5 bullet points of actionable insights. Be specific — reference the data. Focus on:
1. The #1 issue and how severe it is (what % of reviews mention it)
2. Any surprising patterns (e.g. a detractor that also appears in positive reviews)
3. What the data suggests the product team should prioritize
4. What's working well (top delighters) and how to protect those strengths
Keep it concise — each bullet should be 1-2 sentences max."""
                        try:
                            insights = _chat_complete_with_fallback_models(
                                _qi_client, model=_shared_model(), structured=False,
                                messages=[{"role": "user", "content": _qi_prompt}],
                                temperature=0.3, max_tokens=800, reasoning_effort=_shared_reasoning(),
                            )
                            st.session_state["sym_quick_insights"] = insights
                        except Exception as exc:
                            st.error(f"Insights generation failed: {exc}")
                    st.rerun()

            _cached_insights = st.session_state.get("sym_quick_insights")
            if _cached_insights:
                with st.container(border=True):
                    st.markdown(f"**✨ AI Insights**\n\n{_cached_insights}")

        if _unl_dets or _unl_dels:
            with st.expander(f"🆕 Unlisted themes ({len(_unl_dets)} det · {len(_unl_dels)} del)", expanded=False):
                uc1, uc2 = st.columns(2)
                if _unl_dets:
                    uc1.markdown("**Detractor candidates:** " + ", ".join(f"`{u}`" for u in sorted(_unl_dets)[:15]))
                if _unl_dels:
                    uc2.markdown("**Delighter candidates:** " + ", ".join(f"`{u}`" for u in sorted(_unl_dels)[:15]))
    if _HAS_SYMPTOMIZER_V3 and len(processed) >= 3:
        try:
            # Build items list from session data for polarity mismatch detection
            _audit_items = []
            _ds = st.session_state.get("analysis_dataset", {})
            _rdf = _ds.get("reviews_df") if isinstance(_ds, dict) else None
            if _rdf is not None and hasattr(_rdf, "loc"):
                for r in processed:
                    _ridx = r["idx"]
                    try:
                        _rrow = _rdf.loc[_ridx] if _ridx in _rdf.index else None
                        _audit_items.append({"idx": _ridx, "rating": _rrow.get("rating") if _rrow is not None else None})
                    except Exception:
                        _audit_items.append({"idx": _ridx, "rating": None})
            _aud = _v3_audit_distribution(
                {r["idx"]: {"dets":r.get("wrote_dets",[]),"dels":r.get("wrote_dels",[]),"ev_det":r.get("ev_det",{}),"ev_del":r.get("ev_del",{})} for r in processed},
                items=_audit_items or None,
            )
            _sing = _aud.get("singleton_detractors",[])+_aud.get("singleton_delighters",[])
            _dom = _aud.get("dominant_detractors",[])+_aud.get("dominant_delighters",[])
            _noev = _aud.get("zero_evidence_tags",[])
            _pol = _aud.get("polarity_mismatches",[])
            _shared = _aud.get("shared_evidence_labels",{})
            issue_count = len(_sing) + len(_dom) + len(_noev) + len(_pol) + len(_shared)
            if issue_count > 0:
                with st.expander(f"🔍 Tag quality audit — {len(_sing)} singletons · {len(_dom)} dominant · {len(_noev)} no-evidence" + (f" · {len(_pol)} polarity mismatches" if _pol else ""), expanded=False):
                    if _sing: st.caption("**Singletons** (appear once — possible hallucinations):"); st.markdown(", ".join(f"`{s}`" for s in _sing[:20]))
                    if _dom: st.caption("**Dominant** (>50% of reviews — may be too broad):"); st.markdown(", ".join(f"`{d}`" for d in _dom[:10]))
                    if _noev: st.caption("**No evidence** (tagged without supporting text):"); st.markdown(", ".join(f"`{e}`" for e in _noev[:15]))
                    if _pol:
                        st.caption(f"**Polarity mismatches** ({len(_pol)} reviews where star rating contradicts tags):")
                        for pm in _pol[:8]:
                            issue_desc = "5★ with only detractors" if pm["issue"] == "5star_only_detractors" else "1-2★ with only delighters"
                            tags = pm.get("dets", pm.get("dels", []))
                            st.markdown(f"Row {pm['idx']}: {issue_desc} — tags: {', '.join(f'`{t}`' for t in tags[:3])}")
                    if _shared:
                        st.caption(f"**Shared evidence** ({len(_shared)} evidence strings reused across 3+ labels — possible AI laziness):")
                        for ev, labels in list(_shared.items())[:5]:
                            st.markdown(f'"{ev[:50]}…" → {", ".join(f"`{l}`" for l in labels[:4])}')
        except Exception: pass
    # ── Phase 4: Taxonomy recommendations ────────────────────────────────
    if _HAS_SYMPTOMIZER_V3 and len(processed) >= 5:
        try:
            rec_data = {r["idx"]: {"dets": r.get("wrote_dets",[]), "dels": r.get("wrote_dels",[]),
                                    "ev_det": r.get("ev_det",{}), "ev_del": r.get("ev_del",{}),
                                    "unl_dets": r.get("unl_dets",[]), "unl_dels": r.get("unl_dels",[])} for r in processed}
            recs = _v3_generate_recommendations(rec_data,
                allowed_detractors=list(detractors or []), allowed_delighters=list(delighters or []))
            if recs:
                high_recs = [r for r in recs if r["priority"] == "high"]
                med_recs = [r for r in recs if r["priority"] == "medium"]
                with st.expander(f"💡 Taxonomy recommendations — {len(high_recs)} high priority · {len(med_recs)} medium", expanded=bool(high_recs)):
                    for rec in recs[:12]:
                        icon = {"promote": "⬆️", "split": "✂️", "merge": "🔗", "remove": "🗑️"}.get(rec["action"], "•")
                        badge_color = {"high": "rgba(220,38,38,.12)", "medium": "rgba(217,119,6,.12)"}.get(rec["priority"], "rgba(100,116,139,.1)")
                        badge_text = {"high": "#dc2626", "medium": "#d97706"}.get(rec["priority"], "#64748b")
                        st.markdown(
                            f"<div style='display:flex;align-items:flex-start;gap:8px;padding:6px 0;border-bottom:1px solid var(--border,#dde1e8);'>"
                            f"<span style='font-size:16px;flex-shrink:0;'>{icon}</span>"
                            f"<div style='flex:1;'>"
                            f"<span style='font-size:12.5px;font-weight:700;color:var(--navy);'>{rec['action'].title()}: {', '.join(rec['labels'][:3])}</span>"
                            f"<div style='font-size:11.5px;color:var(--slate-500);margin-top:2px;'>{rec['reason']}</div>"
                            f"</div>"
                            f"<span style='font-size:10px;padding:2px 6px;border-radius:99px;background:{badge_color};color:{badge_text};font-weight:700;text-transform:uppercase;'>{rec['priority']}</span>"
                            f"</div>",
                            unsafe_allow_html=True,
                        )
                    # One-click apply for promote recommendations
                    promote_recs = [r for r in recs if r["action"] == "promote"]
                    if promote_recs:
                        if st.button(f"⬆️ Promote {len(promote_recs)} suggested labels to catalog", key="sym_apply_promotions"):
                            for r in promote_recs:
                                for label in r["labels"]:
                                    side = r.get("side", "detractor")
                                    target_key = "sym_delighters" if side == "delighter" else "sym_detractors"
                                    current = list(st.session_state.get(target_key) or [])
                                    if label not in current:
                                        current.append(label)
                                        st.session_state[target_key] = current
                            st.session_state["sym_aliases"] = _alias_map_for_catalog(
                                st.session_state.get("sym_delighters", []),
                                st.session_state.get("sym_detractors", []))
                            st.session_state["sym_qa_user_edited"] = True
                            st.toast(f"Promoted {len(promote_recs)} labels to catalog", icon="⬆️")
                            st.rerun()
        except Exception:
            pass
    raw_cands = {k: v for k, v in (st.session_state.get("sym_new_candidates") or {}).items() if k.strip() and k.strip() not in (delighters + detractors)}
    det_candidate_rows = _candidate_rows_for_side(raw_cands, side="detractor") if raw_cands else []
    del_candidate_rows = _candidate_rows_for_side(raw_cands, side="delighter") if raw_cands else []
    total_candidate_labels = len({row["Label"] for row in det_candidate_rows + del_candidate_rows})
    if det_candidate_rows or del_candidate_rows:
        with st.expander(f"🟡 New symptom candidates ({total_candidate_labels})", expanded=False):
            st.caption("Separated by AI-suggested side so it is faster to add the right labels. You can also promote broad cross-product themes into the Universal Neutral pack right from here.")
            cand_tabs = st.tabs([f"🔴 Detractor candidates ({len(det_candidate_rows)})", f"🟢 Delighter candidates ({len(del_candidate_rows)})"])
            with cand_tabs[0]:
                if not det_candidate_rows:
                    st.info("No new detractor candidates in this run.")
                else:
                    det_df = pd.DataFrame(det_candidate_rows)
                    st.dataframe(det_df, use_container_width=True, hide_index=True, height=min(360, 48 + 35 * len(det_df)), column_config={"Count": st.column_config.NumberColumn(format="%d")})
                    det_lookup = {row["Label"]: row for row in det_candidate_rows}
                    with st.form("sym_new_det_form"):
                        det_pick = st.multiselect(
                            "Choose detractor candidates to add",
                            options=list(det_lookup.keys()),
                            format_func=lambda lab: f"{lab} · {det_lookup[lab].get('L1 Theme', 'Product Specific')} · {int(det_lookup[lab]['Count'])} mentions",
                            key="sym_new_det_pick",
                        )
                        dc1, dc2 = st.columns(2)
                        add_det = dc1.form_submit_button("Add selected → Detractors", use_container_width=True)
                        promote_det = dc2.form_submit_button("Promote selected → Universal Neutral Detractors", use_container_width=True)
                    if add_det and det_pick:
                        _, new_dets = _canonical_symptom_catalog(delighters, detractors + list(det_pick))
                        st.session_state["sym_detractors"] = new_dets
                        st.session_state["sym_aliases"] = _alias_map_for_catalog(st.session_state.get("sym_delighters") or delighters, new_dets, existing_aliases=st.session_state.get("sym_aliases", {}))
                        st.session_state["sym_qa_user_edited"] = True
                        st.success(f"Added {len(det_pick)} detractor candidate(s).")
                        st.rerun()
                    if promote_det and det_pick:
                        _promote_labels_to_custom_universal(det_pick, side="detractor")
                        st.success(f"Promoted {len(det_pick)} detractor candidate(s) to Universal Neutral Symptoms.")
                        st.rerun()
            with cand_tabs[1]:
                if not del_candidate_rows:
                    st.info("No new delighter candidates in this run.")
                else:
                    del_df = pd.DataFrame(del_candidate_rows)
                    st.dataframe(del_df, use_container_width=True, hide_index=True, height=min(360, 48 + 35 * len(del_df)), column_config={"Count": st.column_config.NumberColumn(format="%d")})
                    del_lookup = {row["Label"]: row for row in del_candidate_rows}
                    with st.form("sym_new_del_form"):
                        del_pick = st.multiselect(
                            "Choose delighter candidates to add",
                            options=list(del_lookup.keys()),
                            format_func=lambda lab: f"{lab} · {del_lookup[lab].get('L1 Theme', 'Product Specific')} · {int(del_lookup[lab]['Count'])} mentions",
                            key="sym_new_del_pick",
                        )
                        dc1, dc2 = st.columns(2)
                        add_del = dc1.form_submit_button("Add selected → Delighters", use_container_width=True)
                        promote_del = dc2.form_submit_button("Promote selected → Universal Neutral Delighters", use_container_width=True)
                    if add_del and del_pick:
                        new_dels, _ = _canonical_symptom_catalog(delighters + list(del_pick), detractors)
                        st.session_state["sym_delighters"] = new_dels
                        st.session_state["sym_aliases"] = _alias_map_for_catalog(new_dels, st.session_state.get("sym_detractors") or detractors, existing_aliases=st.session_state.get("sym_aliases", {}))
                        st.session_state["sym_qa_user_edited"] = True
                        st.success(f"Added {len(del_pick)} delighter candidate(s).")
                        st.rerun()
                    if promote_del and del_pick:
                        _promote_labels_to_custom_universal(del_pick, side="delighter")
                        st.success(f"Promoted {len(del_pick)} delighter candidate(s) to Universal Neutral Symptoms.")
                        st.rerun()
    updated_reviews = (st.session_state.get("analysis_dataset") or {}).get("reviews_df", overall_df)
    qa_row_ids = [str(rid) for rid in (st.session_state.get("sym_qa_row_ids") or []) if str(rid).strip()]
    if not qa_row_ids:
        qa_row_ids = [str(rec.get("idx")) for rec in processed if str(rec.get("idx", "")).strip()]
    if qa_row_ids and not st.session_state.get("sym_qa_baseline_map"):
        st.session_state["sym_qa_baseline_map"] = _build_symptom_baseline_map(processed)
    qa_row_ids = [rid for rid in qa_row_ids if rid in _collect_ai_tag_map(updated_reviews, row_ids=qa_row_ids)]
    if qa_row_ids:
        st.session_state["sym_qa_row_ids"] = qa_row_ids
    qa_metrics = _qa_accuracy_metrics(updated_reviews) if qa_row_ids else {}
    qa_notice = st.session_state.pop("sym_qa_notice", None)
    if qa_notice:
        st.success(qa_notice)
    if qa_metrics:
        qa_cols = st.columns(5)
        qa_cols[0].metric("Editable reviews", f"{len(qa_row_ids):,}")
        qa_cols[1].metric("Baseline tags", f"{qa_metrics.get('baseline_total_tags', 0):,}")
        qa_cols[2].metric("Added", f"{qa_metrics.get('added_tags', 0):,}")
        qa_cols[3].metric("Removed", f"{qa_metrics.get('removed_tags', 0):,}")
        qa_cols[4].metric("Accuracy", f"{qa_metrics.get('accuracy_pct', 100.0):.1f}%")
        goal_chip = "green" if float(qa_metrics.get("accuracy_pct", 100.0) or 0) >= 80 else "yellow"
        st.markdown(_chip_html([
            ("Goal: >80%", goal_chip),
            (f"Changed reviews: {qa_metrics.get('changed_reviews', 0)}", "indigo"),
            ("Inline edits save into export", "gray"),
        ]), unsafe_allow_html=True)
        with st.expander("How the accuracy % works", expanded=False):
            st.markdown(
                "Accuracy is calculated against the original AI baseline for this run: **Accuracy = 100 × (1 - (adds + removes) / baseline tags)**. "
                "It stays at 100% until someone makes a manual tag edit in this workspace. Example: 30 original tags, 2 additions, and 2 removals = 86.7% accuracy. "
                "The goal is to keep this above 80% while still correcting misses."
            )

    current_preview_items = list(st.session_state.get("sym_taxonomy_preview_items") or _taxonomy_preview_items_with_side(st.session_state.get("sym_ai_build_result") or {}))
    current_category = str(st.session_state.get("sym_taxonomy_category") or (st.session_state.get("sym_ai_build_result") or {}).get("category") or "general")
    processed_indices = []
    for rec in processed:
        try:
            processed_indices.append(int(rec.get("idx")))
        except Exception:
            continue
    processed_indices = [idx for idx in processed_indices if idx in getattr(updated_reviews, "index", [])]
    processed_df = updated_reviews.loc[processed_indices].copy() if processed_indices else pd.DataFrame(columns=updated_reviews.columns)
    taxonomy_meta = None
    if not processed_df.empty:
        taxonomy_meta = _render_symptomizer_taxonomy_workbench(
            processed_df=processed_df,
            delighters=delighters,
            detractors=detractors,
            aliases=st.session_state.get("sym_aliases", {}),
            category=current_category,
            preview_items=current_preview_items,
        )
        st.divider()
    evidence_lookup = _build_evidence_lookup(processed)
    log_options = [20, 50, 100, 250, "All"]
    if st.session_state.get("sym_review_log_limit") not in log_options:
        st.session_state["sym_review_log_limit"] = 50
    log_choice = st.selectbox("Review log size", options=log_options, key="sym_review_log_limit", help="Expand the result log beyond the old 20-review cap.")
    processed_pool = list(processed) if str(log_choice) == "All" else list(processed[-int(log_choice):])
    ordered_pool = list(reversed(processed_pool))
    page_size_options = [10, 20, 25, 50]
    if st.session_state.get("sym_review_page_size") not in page_size_options:
        st.session_state["sym_review_page_size"] = 20
    total_pages = max(1, int(math.ceil(len(ordered_pool) / float(max(int(st.session_state.get("sym_review_page_size") or 20), 1)))))
    current_page = int(st.session_state.get("sym_review_page") or 1)
    current_page = min(max(current_page, 1), total_pages)
    pg1, pg2, pg3 = st.columns([1.1, 1.0, 2.4])
    page_size = int(pg1.selectbox("Rows per page", options=page_size_options, index=page_size_options.index(int(st.session_state.get("sym_review_page_size") or 20)), key="sym_review_page_size"))
    total_pages = max(1, int(math.ceil(len(ordered_pool) / float(max(page_size, 1)))))
    current_page = min(max(current_page, 1), total_pages)
    st.session_state["sym_review_page"] = current_page
    page_num = int(pg2.number_input("Page", min_value=1, max_value=total_pages, value=current_page, step=1, key="sym_review_page"))
    pg3.caption(f"Rendering {min(page_size, len(ordered_pool)):,} review cards at a time keeps inline editing and taxonomy updates smoother on large runs.")
    page_start = max(0, (page_num - 1) * page_size)
    page_end = page_start + page_size
    processed_view = ordered_pool[page_start:page_end]
    with st.expander(f"📋 Review log — showing {len(processed_view)} of {len(processed_pool)} loaded from {len(processed)} processed", expanded=True):
        for rec in processed_view:
            idx = rec.get("idx", "?")
            row_key = _canonical_index_key(idx)
            head = f"Row {idx} — {len(rec.get('wrote_dets', []))} issues · {len(rec.get('wrote_dels', []))} strengths"
            with st.expander(head):
                row = None
                try:
                    row = updated_reviews.loc[int(float(idx))]
                    row_review_id = _safe_text(row.get("review_id"))
                    ev_items = evidence_lookup.get(row_key) or (evidence_lookup.get(row_review_id) if row_review_id else None)
                    _render_review_card(row, evidence_items=ev_items or None)
                except Exception:
                    try:
                        vb = str(overall_df.loc[int(float(idx)), "review_text"])[:800]
                        st.markdown(f"<div class='review-body'>{html.escape(vb)}</div>", unsafe_allow_html=True)
                    except Exception:
                        st.info("Could not load that review row.")
                st.markdown("<div class='chip-wrap' style='margin-top:8px;margin-bottom:4px;'>" + f"<span class='chip yellow'>Safety: {_esc(rec.get('safety', ''))}</span>" + f"<span class='chip indigo'>Reliability: {_esc(rec.get('reliability', ''))}</span>" + f"<span class='chip gray'>Sessions: {_esc(rec.get('sessions', ''))}</span>" + "</div>", unsafe_allow_html=True)

                if row is None:
                    continue

                baseline_map = st.session_state.get("sym_qa_baseline_map") or {}
                baseline_payload = baseline_map.get(row_key, {"detractors": [], "delighters": []})
                current_dets = _normalize_tag_list(_collect_row_symptom_tags(row, AI_DET_HEADERS))
                current_dels = _normalize_tag_list(_collect_row_symptom_tags(row, AI_DEL_HEADERS))
                suggestions = _build_inline_tag_suggestions(row, current_dets, current_dels, detractors, delighters)
                missing_dets = suggestions.get("missing_detractors", [])
                missing_dels = suggestions.get("missing_delighters", [])
                det_options = _normalize_tag_list(current_dets + detractors + missing_dets)
                del_options = _normalize_tag_list(current_dels + delighters + missing_dels)

                suggestion_chips = []
                if missing_dets:
                    suggestion_chips.extend((f"Missing detractor: {lab}", "red") for lab in missing_dets[:3])
                if missing_dels:
                    suggestion_chips.extend((f"Missing delighter: {lab}", "green") for lab in missing_dels[:3])

                current_issue_chips = [(label, "red") for label in current_dets[:8]]
                current_strength_chips = [(label, "green") for label in current_dels[:8]]
                baseline_issue_chips = [(label, "gray") for label in (baseline_payload.get("detractors", []) or [])[:8]]
                baseline_strength_chips = [(label, "gray") for label in (baseline_payload.get("delighters", []) or [])[:8]]

                with st.expander("Edit Taxonomy", expanded=False):
                    st.markdown(
                        "<div class='soft-panel' style='margin-top:.2rem;'><b>Fix tags right here</b> · Deselect a current tag to remove it, pick a missing symptom from the catalog, or type a brand-new symptom. This review updates the export file immediately after you save it.</div>",
                        unsafe_allow_html=True,
                    )
                    st.caption("Universal Neutral Symptoms are active here." if bool(st.session_state.get("sym_include_universal_neutral", True)) else "Universal Neutral Symptoms are turned off in this workspace, so this editor is only using product-specific tags.")

                    summary_cols = st.columns(2)
                    with summary_cols[0]:
                        st.markdown("**🔴 Issues**")
                        st.caption("Current saved tags")
                        if current_issue_chips:
                            st.markdown(_chip_html(current_issue_chips), unsafe_allow_html=True)
                        else:
                            st.caption("None")
                        st.caption("Original AI tags")
                        if baseline_issue_chips:
                            st.markdown(_chip_html(baseline_issue_chips), unsafe_allow_html=True)
                        else:
                            st.caption("None")
                    with summary_cols[1]:
                        st.markdown("**🟢 Strengths**")
                        st.caption("Current saved tags")
                        if current_strength_chips:
                            st.markdown(_chip_html(current_strength_chips), unsafe_allow_html=True)
                        else:
                            st.caption("None")
                        st.caption("Original AI tags")
                        if baseline_strength_chips:
                            st.markdown(_chip_html(baseline_strength_chips), unsafe_allow_html=True)
                        else:
                            st.caption("None")

                    if suggestion_chips:
                        st.markdown(_chip_html(suggestion_chips), unsafe_allow_html=True)
                    else:
                        st.caption("No obvious missing tags detected, but you can still add one manually below.")

                    suggested_det_text = ", ".join(missing_dets[:5]) if missing_dets else "No missing detractor suggestion"
                    suggested_del_text = ", ".join(missing_dels[:5]) if missing_dels else "No missing delighter suggestion"

                    with st.form(f"sym_inline_form_{row_key}"):
                        edit_cols = st.columns(2)
                        with edit_cols[0]:
                            st.caption(f"Suggested missing detractors: {suggested_det_text}")
                            chosen_dets = st.multiselect(
                                "Choose detractor symptoms",
                                options=det_options,
                                default=current_dets,
                                key=f"sym_inline_det_select_{row_key}",
                                help="Deselect a tag to remove it, or select a missing symptom to add it.",
                            )
                            new_det_text = st.text_input(
                                "Add a new detractor symptom",
                                placeholder="Comma-separated, e.g. Hard To Clean, Cheap Handle",
                                key=f"sym_inline_det_new_{row_key}",
                            )
                        with edit_cols[1]:
                            st.caption(f"Suggested missing delighters: {suggested_del_text}")
                            chosen_dels = st.multiselect(
                                "Choose delighter symptoms",
                                options=del_options,
                                default=current_dels,
                                key=f"sym_inline_del_select_{row_key}",
                                help="Deselect a tag to remove it, or select a missing symptom to add it.",
                            )
                            new_del_text = st.text_input(
                                "Add a new delighter symptom",
                                placeholder="Comma-separated, e.g. Quiet, Easy Cleanup",
                                key=f"sym_inline_del_new_{row_key}",
                            )

                        st.caption("Cleanest workflow: scan the review, make the edit here, then click Update tags + refresh accuracy.")
                        action_cols = st.columns([2.1, 1.25, 1.65])
                        apply_changes = action_cols[0].form_submit_button("✅ Update tags + refresh accuracy", type="primary", use_container_width=True)
                        reset_changes = action_cols[1].form_submit_button("↺ Reset to AI baseline", use_container_width=True)
                        action_cols[2].caption("Only the first 10 detractors and first 10 delighters are written to the export file for each review.")

                    if apply_changes:
                        new_dets = _parse_manual_tag_entries(new_det_text)
                        new_dels = _parse_manual_tag_entries(new_del_text)
                        ok, message = _save_inline_symptom_edit(
                            row_key,
                            row,
                            _normalize_tag_list(list(chosen_dets) + new_dets),
                            _normalize_tag_list(list(chosen_dels) + new_dels),
                            updated_reviews=updated_reviews,
                            processed_rows=processed,
                            detractors=detractors,
                            delighters=delighters,
                            notice_prefix="Updated",
                        )
                        if ok:
                            st.rerun()
                        else:
                            st.error(message)
                    if reset_changes:
                        ok, message = _save_inline_symptom_edit(
                            row_key,
                            row,
                            baseline_payload.get("detractors", []),
                            baseline_payload.get("delighters", []),
                            updated_reviews=updated_reviews,
                            processed_rows=processed,
                            detractors=detractors,
                            delighters=delighters,
                            notice_prefix="Reset",
                        )
                        if ok:
                            st.rerun()
                        else:
                            st.error(message)

    if taxonomy_meta:
        _render_symptomizer_taxonomy_housekeeping(taxonomy_meta)
        st.divider()

    ec1, ec2 = st.columns([1.5, 3])
    if ec1.button("🧾 Prepare export", use_container_width=True, key="sym_prep_export"):
        upd = st.session_state["analysis_dataset"]["reviews_df"]
        orig = st.session_state.get("_uploaded_raw_bytes")
        sym_bytes = _gen_symptomized_workbook(orig, upd) if orig else _build_master_excel(summary, upd)
        st.session_state["sym_export_bytes"] = sym_bytes
        st.success("Export prepared.")
    sym_bytes = st.session_state.get("sym_export_bytes")
    sym_export_name = f"{_safe_summary_product_slug(summary, updated_reviews, default='symptomized_reviews')}_Symptomized.xlsx"
    ec1.download_button("⬇️ Download symptomized file", data=sym_bytes or b"", file_name=sym_export_name, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", disabled=(sym_bytes is None), key="sym_dl")
    qa_export_metrics = st.session_state.get("sym_qa_accuracy") or {}
    if qa_export_metrics:
        ec2.caption(f"Writes the final AI symptom tags (including QA edits) plus AI Safety · Reliability · Sessions into existing AI columns when present, otherwise it appends safe new columns without overwriting local symptomization. Current QA accuracy: {float(qa_export_metrics.get('accuracy_pct', 100.0) or 0):.1f}%.")
    else:
        ec2.caption("Writes AI Symptom Detractor / Delighter columns plus AI Safety · Reliability · Sessions into existing AI columns when present, otherwise it appends safe new columns without overwriting local symptomization.")

# ── Social Listening (extracted to review_analyst/social_listening.py) ────
if not _HAS_SOCIAL_PKG:
    # Fallback stubs when package not available
    def _build_social_beta_query(raw_query): return raw_query
    def _social_demo_payload(): return {"posts":pd.DataFrame(),"detractors":[],"delighters":[],"viral":[],"top_comments":[],"compare":{},"metrics":{}}
    def _social_demo_query(p): return p
    def _social_demo_trend(s,e): return pd.DataFrame()
    def _render_social_metric_card(l,v,s,accent="indigo"): pass
    def _render_social_post_card(r,*,highlight=""): pass
    def _social_demo_answer(q): return "Social listening module not loaded."
    def _render_social_listening_tab():
        st.info("Social Listening module not available.")
else:
    from review_analyst.social_listening import (
        _build_social_beta_query, _social_demo_payload, _social_demo_query,
        _social_demo_trend, _render_social_metric_card, _render_social_post_card,
        _social_demo_answer, _render_social_listening_tab,
    )


def _render_lava_lamp_background():
    return


def main():
    _render_lava_lamp_background()
    st.markdown("""<div class='app-shell'>
      <div class='app-header'>
        <div class='app-brand'>
          <div class='app-logo'>✨</div>
          <div>
            <div class='app-title-row'>
              <div class='app-title'>StarWalk Review Analyst</div>
              <span class='beta-chip'>Beta</span>
            </div>
            <div class='app-subtitle'>Single-file Streamlit workspace for executive review, deep-dive exploration, Review Prompt, Symptomizer, and a placeholder Social Listening Beta route that works even before reviews exist.</div>
          </div>
        </div>
      </div>
    </div>""", unsafe_allow_html=True)

    dataset = st.session_state.get("analysis_dataset")
    if dataset:
        bc = st.columns([2.5, 1.05, 1.1, 0.55, 0.55])
        # Workspace name (editable)
        ws_name = st.session_state.get("workspace_name", dataset.get("source_label", "Untitled"))
        bc[0].caption(f"Active workspace · {dataset.get('source_type', '').title()} · {ws_name}")
        # Save button
        if _HAS_WORKSPACE_STORE:
            if bc[1].button("💾 Save workspace", use_container_width=True, key="ws_save_btn"):
                try:
                    summary = dataset.get("summary")
                    reviews_df = dataset.get("reviews_df", pd.DataFrame())
                    sym_state = {k: v for k, v in st.session_state.items() if k.startswith("sym_") and not k.startswith("_")}
                    ws_id = _ws_save(
                        workspace_name=ws_name,
                        source_type=dataset.get("source_type", "unknown"),
                        source_label=dataset.get("source_label", ""),
                        reviews_df=reviews_df,
                        dataset_payload={k: v for k, v in dataset.items() if k not in ("reviews_df",)},
                        state_payload=sym_state,
                        product_id=getattr(summary, "product_id", "") if summary else "",
                        product_url=getattr(summary, "product_url", "") if summary else "",
                        symptomized=bool(st.session_state.get("sym_processed_rows")),
                        workspace_id=st.session_state.get("workspace_id"),
                        allow_source_upsert=True,
                    )
                    st.session_state["workspace_id"] = ws_id
                    st.toast(f"Workspace saved: {ws_name}", icon="💾")
                except Exception as exc:
                    st.error(f"Save failed: {exc}")
            if bc[2].button("Save as new", use_container_width=True, key="ws_save_as_btn"):
                try:
                    summary = dataset.get("summary")
                    reviews_df = dataset.get("reviews_df", pd.DataFrame())
                    sym_state = {k: v for k, v in st.session_state.items() if k.startswith("sym_") and not k.startswith("_")}
                    new_name = ws_name if str(ws_name).strip().lower().endswith("copy") else f"{ws_name} copy"
                    ws_id = _ws_save(
                        workspace_name=new_name,
                        source_type=dataset.get("source_type", "unknown"),
                        source_label=dataset.get("source_label", ""),
                        reviews_df=reviews_df,
                        dataset_payload={k: v for k, v in dataset.items() if k not in ("reviews_df",)},
                        state_payload=sym_state,
                        product_id=getattr(summary, "product_id", "") if summary else "",
                        product_url=getattr(summary, "product_url", "") if summary else "",
                        symptomized=bool(st.session_state.get("sym_processed_rows")),
                        workspace_id=None,
                        allow_source_upsert=False,
                    )
                    st.session_state["workspace_id"] = ws_id
                    st.session_state["workspace_name"] = new_name
                    st.toast(f"Saved new workspace: {new_name}", icon="🧬")
                except Exception as exc:
                    st.error(f"Save as failed: {exc}")
        if bc[3].button("✏️", use_container_width=True, key="ws_rename_btn", help="Rename workspace"):
            st.session_state["_ws_show_rename"] = True
        if bc[4].button("Clear", use_container_width=True, key="ws_clear"):
            _reset_workspace_state(reset_source=True)
            st.rerun()
        # Rename dialog
        if st.session_state.get("_ws_show_rename"):
            rc1, rc2 = st.columns([3, 1])
            new_name = rc1.text_input("New workspace name", value=ws_name, key="_ws_rename_input")
            if rc2.button("Save name", key="_ws_rename_confirm"):
                st.session_state["workspace_name"] = new_name
                if _HAS_WORKSPACE_STORE and st.session_state.get("workspace_id"):
                    try:
                        _ws_rename(st.session_state["workspace_id"], new_name)
                    except Exception:
                        pass
                st.session_state["_ws_show_rename"] = False
                st.rerun()
        failures = dataset.get("source_failures") or []
        if failures:
            st.warning("Partial multi-link load: " + " | ".join(f"{u} → {err}" for u, err in failures[:3]))

    if st.session_state.get("workspace_source_mode") not in {SOURCE_MODE_URL, SOURCE_MODE_FILE}:
        st.session_state["workspace_source_mode"] = SOURCE_MODE_URL

    with st.expander("🧰 Build or switch workspace", expanded=(dataset is None and st.session_state.get("workspace_active_tab") != TAB_SOCIAL_LISTENING)):
        # ── Saved Workspaces ──────────────────────────────────────────────
        if _HAS_WORKSPACE_STORE:
            try:
                saved_count = _ws_count()
            except Exception:
                saved_count = 0
            if saved_count > 0:
                with st.container(border=True):
                    st.markdown(f"**💾 Saved Workspaces** ({saved_count})")
                    try:
                        saved_list = _ws_list(limit=20)
                    except Exception:
                        saved_list = []
                    if saved_list:
                        for ws in saved_list:
                            wc1, wc2, wc3 = st.columns([3.5, 0.8, 0.7])
                            avg_r = ws.get("avg_rating")
                            avg_str = f" · {avg_r:.1f}★" if avg_r else ""
                            sym_str = f" · {ws.get('symptomized_count', 0)} tagged" if ws.get("symptomized") else ""
                            wc1.markdown(f"**{_esc(ws.get('workspace_name', 'Untitled'))}** <span style='color:var(--slate-400);font-size:12px;'>{ws.get('review_count', 0):,} reviews{avg_str}{sym_str}</span>", unsafe_allow_html=True)
                            if wc2.button("Load", key=f"ws_load_{ws['workspace_id']}", use_container_width=True):
                                try:
                                    loaded = _ws_load(ws["workspace_id"])
                                    _ws_touch(ws["workspace_id"])
                                    _reset_review_filters()
                                    # Restore dataset
                                    restored_dataset = loaded.get("dataset_payload", {})
                                    restored_dataset["reviews_df"] = loaded.get("reviews_df", pd.DataFrame())
                                    st.session_state["analysis_dataset"] = restored_dataset
                                    st.session_state["workspace_id"] = ws["workspace_id"]
                                    st.session_state["workspace_name"] = ws.get("workspace_name", "Untitled")
                                    # Restore symptomizer state
                                    for k, v in (loaded.get("state_payload") or {}).items():
                                        if k.startswith("sym_"):
                                            st.session_state[k] = v
                                    st.session_state["workspace_active_tab"] = TAB_DASHBOARD
                                    st.toast(f"Loaded: {ws.get('workspace_name', 'Untitled')}", icon="📂")
                                    st.rerun()
                                except Exception as exc:
                                    st.error(f"Load failed: {exc}")
                            if wc3.button("🗑️", key=f"ws_del_{ws['workspace_id']}", use_container_width=True, help="Delete"):
                                try:
                                    _ws_delete(ws["workspace_id"])
                                    st.toast(f"Deleted: {ws.get('workspace_name', '')}", icon="🗑️")
                                    st.rerun()
                                except Exception as exc:
                                    st.error(f"Delete failed: {exc}")
                    else:
                        st.caption("No saved workspaces yet. Build one below and save it.")
                st.markdown("<div style='height:.5rem'></div>", unsafe_allow_html=True)
        st.markdown("<div class='builder-card'><div class='builder-kicker'>Get started</div><div class='builder-title'>Build a review workspace</div><div class='builder-sub'>Paste a supported product page or review API URL, or upload a review export file. Start with the Dashboard for an executive summary, then move into the other tabs for deeper work.</div><div class='helper-chip-row'><span class='helper-chip'>Shark/Ninja US</span><span class='helper-chip'>Shark/Ninja UK/EU</span><span class='helper-chip'>Costco</span><span class='helper-chip'>Sephora</span><span class='helper-chip'>Ulta</span><span class='helper-chip'>Hoka</span><span class='helper-chip'>CurrentBody</span><span class='helper-chip'>Okendo API</span><span class='helper-chip'>Bazaarvoice API</span><span class='helper-chip'>PowerReviews API</span></div></div>", unsafe_allow_html=True)
        with st.container(border=True):
            st.markdown(
                "<div class='builder-kicker'>Source setup</div><div class='builder-title' style='font-size:16px;'>Choose how to load reviews</div><div class='builder-sub' style='margin-bottom:.7rem;'>Pick a link-based workspace or an uploaded file flow. The controls below stay grouped so the source choice, mode, and primary action feel like one connected setup.</div>",
                unsafe_allow_html=True,
            )
            src_label_col, src_control_col = st.columns([1.0, 2.4])
            with src_label_col:
                st.markdown("<div class='metric-label' style='margin-top:.35rem;'>Workspace source</div>", unsafe_allow_html=True)
            with src_control_col:
                source_mode = st.radio(
                    "Workspace source",
                    [SOURCE_MODE_URL, SOURCE_MODE_FILE],
                    horizontal=True,
                    key="workspace_source_mode",
                    label_visibility="collapsed",
                )
            if source_mode == SOURCE_MODE_URL:
                mode_label_col, mode_control_col = st.columns([1.0, 2.4])
                with mode_label_col:
                    st.markdown("<div class='metric-label' style='margin-top:.2rem;'>Link mode</div>", unsafe_allow_html=True)
                with mode_control_col:
                    url_mode = st.radio(
                        "Link mode",
                        ["Single link", "Multiple links"],
                        horizontal=True,
                        key="workspace_url_entry_mode",
                        label_visibility="collapsed",
                    )
                if url_mode == "Single link":
                    st.text_input(
                        "Product or review URL",
                        key="workspace_product_url",
                        placeholder="Paste a product page or direct review endpoint",
                        label_visibility="collapsed",
                    )
                    st.caption("Fastest path: paste a retailer product page or a direct Bazaarvoice / PowerReviews / Okendo review endpoint.")
                    if st.button("Build review workspace", type="primary", key="ws_build_url", use_container_width=True):
                        try:
                            nd = _load_product_reviews_dispatch(st.session_state.get("workspace_product_url", DEFAULT_PRODUCT_URL))
                            _reset_review_filters()
                            st.session_state.update(analysis_dataset=nd, chat_messages=[], master_export_bundle=None, prompt_run_artifacts=None, sym_processed_rows=[], sym_new_candidates={}, sym_symptoms_source="none", sym_delighters=[], sym_detractors=[], sym_custom_universal_delighters=[], sym_custom_universal_detractors=[], sym_aliases={}, sym_taxonomy_preview_items=[], sym_taxonomy_category="general", sym_qa_baseline_map={}, sym_qa_accuracy={}, sym_qa_user_edited=False, sym_qa_row_ids=[], sym_qa_selected_row=None, sym_qa_notice=None, sym_product_profile_ai_note="", sym_product_knowledge={}, workspace_active_tab=TAB_DASHBOARD, workspace_tab_request=None, ai_include_references=False, ot_show_volume=False, _uploaded_raw_bytes=None, sym_export_bytes=None)
                            with st.spinner("Auto-discovering product profile…"):
                                _auto_discover_product(nd)
                            st.rerun()
                        except requests.HTTPError as exc:
                            st.error(f"HTTP error: {exc}")
                        except ReviewDownloaderError as exc:
                            st.error(str(exc))
                        except Exception as exc:
                            st.error(str(exc))
                else:
                    st.text_area(
                        "Product or review URLs",
                        key="workspace_product_urls_bulk",
                        height=150,
                        placeholder="Paste one product page or review endpoint per line\nhttps://www.costco.com/...\nhttps://www.sephora.com/...\nhttps://www.hoka.com/...",
                        label_visibility="collapsed",
                    )
                    bulk_urls = _parse_bulk_product_urls(st.session_state.get("workspace_product_urls_bulk", ""))
                    if bulk_urls:
                        preview = ", ".join((_strip_www(urlparse(u).netloc) or u) for u in bulk_urls[:4])
                        if len(bulk_urls) > 4:
                            preview += f" +{len(bulk_urls) - 4}"
                        st.caption(f"Ready to load {len(bulk_urls)} link(s) · {preview}")
                    else:
                        st.caption("Paste one product page or direct review endpoint per line to build a combined workspace.")
                    if st.button("Build combined workspace", type="primary", key="ws_build_url_multi", use_container_width=True):
                        try:
                            nd = _load_multiple_product_reviews_dispatch(bulk_urls)
                            _reset_review_filters()
                            st.session_state.update(analysis_dataset=nd, chat_messages=[], master_export_bundle=None, prompt_run_artifacts=None, sym_processed_rows=[], sym_new_candidates={}, sym_symptoms_source="none", sym_delighters=[], sym_detractors=[], sym_custom_universal_delighters=[], sym_custom_universal_detractors=[], sym_aliases={}, sym_taxonomy_preview_items=[], sym_taxonomy_category="general", sym_qa_baseline_map={}, sym_qa_accuracy={}, sym_qa_user_edited=False, sym_qa_row_ids=[], sym_qa_selected_row=None, sym_qa_notice=None, sym_product_profile_ai_note="", sym_product_knowledge={}, workspace_active_tab=TAB_DASHBOARD, workspace_tab_request=None, ai_include_references=False, ot_show_volume=False, _uploaded_raw_bytes=None, sym_export_bytes=None)
                            with st.spinner("Auto-discovering product profile…"):
                                _auto_discover_product(nd)
                            st.rerun()
                        except requests.HTTPError as exc:
                            st.error(f"HTTP error: {exc}")
                        except ReviewDownloaderError as exc:
                            st.error(str(exc))
                        except Exception as exc:
                            st.error(str(exc))
            else:
                st.markdown("<div class='metric-label' style='margin:.25rem 0 .3rem;'>Uploaded review file</div>", unsafe_allow_html=True)
                uploader_key = f"workspace_files_{int(st.session_state.get('workspace_file_uploader_nonce', 0))}"
                uploaded_files = st.file_uploader(
                    "Upload review export files",
                    type=["csv", "xlsx", "xls"],
                    accept_multiple_files=True,
                    help="Supports Axion-style exports and similar CSV/XLSX review files.",
                    key=uploader_key,
                    label_visibility="collapsed",
                )
                include_local_symptomization = st.checkbox(
                    "Include local symptomization if detected",
                    value=bool(st.session_state.get("workspace_include_local_symptomization", True)),
                    key="workspace_include_local_symptomization",
                    help="Preserves populated Symptom 1–20 / AI Symptom columns from uploaded files so the Dashboard, Review Explorer, filters, and Symptomizer can use them immediately.",
                )
                st.caption("Mapped columns include Event Id, Base SKU, Review Text, Rating, Opened date, Seeded Flag, and Retailer when available. Turn on local symptomization to keep existing Symptom 1–20 / AI Symptom tags and infer the symptom catalog automatically.")
                if st.button("Build review workspace from file", type="primary", key="ws_build_file", use_container_width=True):
                    try:
                        nd = _load_uploaded_files_dispatch(uploaded_files or [], include_local_symptomization=include_local_symptomization)
                        _reset_review_filters()
                        raw_bytes = None
                        if uploaded_files and len(uploaded_files) == 1:
                            fname = getattr(uploaded_files[0], "name", "")
                            if fname.lower().endswith((".xlsx", ".xlsm")):
                                raw_bytes = uploaded_files[0].getvalue()
                        local_delighters, local_detractors = ([], [])
                        if include_local_symptomization:
                            local_delighters, local_detractors = _local_symptom_catalog(nd["reviews_df"])
                        has_local_catalog = bool(local_delighters or local_detractors)
                        st.session_state.update(
                            analysis_dataset=nd,
                            chat_messages=[],
                            master_export_bundle=None,
                            prompt_run_artifacts=None,
                            sym_processed_rows=[],
                            sym_new_candidates={},
                            sym_symptoms_source=("local upload" if has_local_catalog else "none"),
                            sym_delighters=(local_delighters if has_local_catalog else []),
                            sym_detractors=(local_detractors if has_local_catalog else []),
                            sym_aliases=_alias_map_for_catalog((local_delighters if has_local_catalog else []), (local_detractors if has_local_catalog else [])),
                            sym_taxonomy_preview_items=[],
                            sym_taxonomy_category="general",
                            sym_qa_baseline_map={},
                            sym_qa_accuracy={},
                            sym_qa_user_edited=False,
                            sym_qa_row_ids=[],
                            sym_qa_selected_row=None,
                            sym_qa_notice=None,
                            sym_product_profile_ai_note="",
                            sym_product_knowledge={},
                            workspace_active_tab=TAB_DASHBOARD,
                            workspace_tab_request=None,
                            ai_include_references=False,
                            ot_show_volume=False,
                            _uploaded_raw_bytes=raw_bytes,
                            sym_export_bytes=None,
                        )
                        st.rerun()
                    except ReviewDownloaderError as exc:
                        st.error(str(exc))
                    except Exception as exc:
                        st.exception(exc)

    dataset = st.session_state.get("analysis_dataset")
    settings = _render_sidebar(dataset["reviews_df"] if dataset else None)
    if not dataset:
        active_tab = st.session_state.get("workspace_active_tab", TAB_DASHBOARD)
        if active_tab == TAB_SOCIAL_LISTENING:
            st.markdown("""<div class='soft-panel' style='margin-top:.55rem;'>
              <b>Social-only beta mode</b> · No uploaded review file is required here. The placeholder social experience below previews the future Meltwater-powered workflow before Bazaarvoice / PowerReviews / uploads are brought in.
            </div>""", unsafe_allow_html=True)
            _render_social_listening_tab()
            return
        st.markdown("""<div class='empty-state-card'>
          <div style="font-size:2.5rem;margin-bottom:.75rem;">📊</div>
          <div class='empty-state-title' style="font-size:16px;">No workspace loaded</div>
          <div class='empty-state-sub'>Build a workspace above to unlock the Dashboard, Review Explorer, AI Analyst, Review Prompt, and Symptomizer. Or skip reviews entirely and open <b>Social Listening Beta</b> from the sidebar to explore the placeholder Meltwater + AI social workflow.</div>
        </div>""", unsafe_allow_html=True)
        if st.button("📣 Open Social Listening Beta", type="primary", key="empty_state_open_social"):
            st.session_state["workspace_active_tab"] = TAB_SOCIAL_LISTENING
            st.rerun()
        return

    summary = dataset["summary"]
    overall_df = dataset["reviews_df"]
    source_type = dataset.get("source_type", "bazaarvoice")
    source_label = dataset.get("source_label", "")
    filter_state = settings["review_filters"]
    filtered_df = filter_state["filtered_df"]
    filter_description = filter_state["description"]
    new_filter_sig = json.dumps(filter_state["active_items"], default=str)
    if st.session_state.get("review_filter_signature") != new_filter_sig:
        st.session_state["review_filter_signature"] = new_filter_sig
        st.session_state["review_explorer_page"] = 1

    _render_workspace_header(summary, overall_df, filtered_df, st.session_state.get("prompt_run_artifacts"), source_type=source_type, source_label=source_label, filter_description=filter_description, active_items=filter_state["active_items"])
    _render_top_metrics(overall_df, filtered_df)
    _render_active_filter_summary(filter_state, overall_df)

    pending_tab = st.session_state.pop("workspace_tab_request", None)
    if pending_tab in WORKSPACE_TABS:
        st.session_state["workspace_active_tab"] = pending_tab
    elif st.session_state.get("workspace_active_tab") not in WORKSPACE_TABS:
        st.session_state["workspace_active_tab"] = TAB_DASHBOARD
    active_tab = _render_workspace_nav()
    common = dict(settings=settings, overall_df=overall_df, filtered_df=filtered_df, summary=summary, filter_description=filter_description)

    def _safe_render(fn, *args, **kwargs):
        """Error boundary: catch crashes in individual tabs without taking down the app."""
        try:
            fn(*args, **kwargs)
        except Exception as exc:
            st.error(f"This tab encountered an error: {exc}")
            _log.error("Tab render error in %s: %s", fn.__name__, exc, exc_info=True)
            with st.expander("Technical details", expanded=False):
                import traceback
                st.code(traceback.format_exc(), language="text")

    if active_tab == TAB_DASHBOARD:
        _safe_render(_render_dashboard, filtered_df, overall_df)
    elif active_tab == TAB_REVIEW_EXPLORER:
        _safe_render(_render_review_explorer, summary=summary, overall_df=overall_df, filtered_df=filtered_df, prompt_artifacts=st.session_state.get("prompt_run_artifacts"), filter_description=filter_description, active_items=filter_state["active_items"])
    elif active_tab == TAB_REVIEW_PROMPT:
        _safe_render(_render_review_prompt_tab, **common)
    elif active_tab == TAB_SYMPTOMIZER:
        _safe_render(_render_symptomizer_tab, **common)
    elif active_tab == TAB_SOCIAL_LISTENING:
        _safe_render(_render_social_listening_tab)

    # ── Persistent AI chat bar (bottom of every page) ─────────────────
    _render_bottom_chat_bar(settings=settings, overall_df=overall_df, filtered_df=filtered_df, summary=summary, filter_description=filter_description)


# ═══════════════════════════════════════════════════════════════════════════════
#  BOTTOM CHAT BAR — persistent AI assistant on every page
# ═══════════════════════════════════════════════════════════════════════════════
def _render_bottom_chat_bar(*, settings, overall_df, filtered_df, summary, filter_description):
    """Persistent AI chat bar at the bottom of every page."""
    chat_messages = st.session_state.get("chat_messages") or []
    active_tab = st.session_state.get("workspace_active_tab") or TAB_DASHBOARD
    view_reviews = len(filtered_df) if isinstance(filtered_df, pd.DataFrame) else 0
    total_reviews = len(overall_df) if isinstance(overall_df, pd.DataFrame) else 0
    has_symptoms = bool(st.session_state.get("sym_processed_rows"))
    product_label = _product_name(summary, overall_df if isinstance(overall_df, pd.DataFrame) else pd.DataFrame())

    focus_options = [
        "General",
        "Action plan",
        "Root cause",
        "Consumer language",
        "Product / CX opportunities",
    ]
    focus_version_key = "_bottom_chat_focus_default_v3"
    if st.session_state.get(focus_version_key) != 3 or st.session_state.get("bottom_chat_focus") not in focus_options:
        st.session_state["bottom_chat_focus"] = "General"
        st.session_state[focus_version_key] = 3
    answer_styles = ["Actionable", "Balanced", "Deep dive"]
    if st.session_state.get("bottom_chat_answer_style") not in answer_styles:
        st.session_state["bottom_chat_answer_style"] = "Balanced"

    with st.container(border=True):
        top_left, top_right = st.columns([4.2, 1.0])
        with top_left:
            st.markdown("### Ask AI about this product or workspace")
            st.caption(
                "Grounded in the active filters, current symptom tags, product knowledge, and recent review evidence. "
                "Leave AI mode on General for open-ended questions, or switch to a guided mode when you want a sharper action-plan, root-cause, language, or opportunity frame."
            )
        with top_right:
            if st.button("Clear chat", key="bottom_chat_clear", use_container_width=True, disabled=not chat_messages):
                st.session_state["chat_messages"] = []
                st.rerun()

        st.markdown(
            _chip_html(
                [
                    (f"Product: {product_label}", "blue"),
                    (f"View: {view_reviews:,} reviews", "indigo"),
                    (f"Loaded: {total_reviews:,}", "gray"),
                    (f"Symptomizer: {'On' if has_symptoms else 'Off'}", "green" if has_symptoms else "gray"),
                    (f"Tab: {str(active_tab).replace('_', ' ').title()}", "gray"),
                ]
            ),
            unsafe_allow_html=True,
        )
        if filter_description:
            st.caption(f"Current filters: {filter_description}")

        ctl1, ctl2 = st.columns([1.25, 1.0])
        focus = ctl1.selectbox(
            "AI mode",
            options=focus_options,
            key="bottom_chat_focus",
            help="General is the default for freeform Q&A. Switch modes only when you want the answer intentionally framed as an action plan, root-cause readout, consumer-language summary, or opportunity scan.",
        )
        answer_style = ctl2.selectbox(
            "Answer style",
            options=answer_styles,
            key="bottom_chat_answer_style",
            help="Controls how concise or detailed the answer should be.",
        )

        if chat_messages:
            latest_messages = chat_messages[-2:] if len(chat_messages) >= 2 else chat_messages[-1:]
            for msg in latest_messages:
                with st.chat_message(msg.get("role", "user")):
                    st.markdown(msg.get("content", ""), unsafe_allow_html=True)
            older_messages = chat_messages[:-2]
            if older_messages:
                with st.expander(f"Earlier messages ({len(older_messages)})", expanded=False):
                    for msg in older_messages[-8:]:
                        with st.chat_message(msg.get("role", "user")):
                            st.markdown(msg.get("content", ""), unsafe_allow_html=True)
        else:
            st.markdown(
                "<div class='soft-panel' style='margin:.35rem 0 .8rem;'><b>Best uses</b> · Ask for the biggest complaints in this slice, what to fix first, where the rating drag is concentrated, what consumers actually praise, or how to translate the current view into product/CX actions.</div>",
                unsafe_allow_html=True,
            )
            suggestions = []
            if has_symptoms:
                suggestions.extend([
                    "What is the general story in this view?",
                    "Which detractors deserve action first in this view?",
                    "What are the likely root causes behind the top detractors?",
                    "What do the strongest delighters suggest we should protect?",
                ])
            elif view_reviews > 0:
                suggestions.extend([
                    "What is the general story in this filtered view?",
                    "Summarize the biggest complaints in this filtered view",
                    "What are the clearest consumer themes here?",
                    "What separates the low-star reviews from the high-star reviews?",
                ])
            else:
                suggestions = [
                    "How should I use this workspace?",
                    "Give me the general story here",
                    "What should I look for first?",
                    "How does the Symptomizer help here?",
                ]
            rows = [suggestions[i:i + 2] for i in range(0, min(len(suggestions), 4), 2)]
            for ridx, row in enumerate(rows):
                cols = st.columns(len(row))
                for cidx, (col, suggestion) in enumerate(zip(cols, row)):
                    if col.button(suggestion, key=f"chat_suggest_{ridx}_{cidx}", use_container_width=True):
                        st.session_state["_chat_pending_prompt"] = suggestion
                        st.rerun()

    pending_prompt = st.session_state.pop("_chat_pending_prompt", None)
    prompt = st.chat_input("Ask AI about this product or workspace...", key="bottom_chat_input")
    if pending_prompt:
        prompt = pending_prompt

    if prompt:
        client = _get_client()
        api_key = settings.get("api_key")
        if not client or not api_key:
            st.error("OpenAI API key required. Add it in Settings -> OpenAI API Key.")
            return

        persona_map = {
            "General": None,
            "Action plan": "Product Development",
            "Root cause": "Quality Engineer",
            "Consumer language": "Consumer Insights",
            "Product / CX opportunities": "Product Development",
        }
        focus_instructions = {
            "General": "Answer the question directly using the current workspace and filtered view. Stay flexible and only impose structure when it helps the user.",
            "Action plan": "Prioritize what should happen now, next, and later. Be specific about product, CX, ops, or messaging actions.",
            "Root cause": "Explain the most likely failure modes, workflow friction, or taxonomy-level causes behind the patterns in this view. Separate confirmed evidence from inference.",
            "Consumer language": "Summarize the actual consumer language, emotional tone, and repeated phrasing. Highlight what customers care about and how they describe it.",
            "Product / CX opportunities": "Convert the current view into opportunity areas for product, support, onboarding, merchandising, retention, or messaging.",
        }
        response_scaffolds = {
            "General": "",
            "Action plan": "Structure the answer as: What matters most; Prioritized actions (Now / Next / Later); Risks or dependencies; What to monitor next.",
            "Root cause": "Structure the answer as: What is happening; Most likely root causes; Evidence supporting each cause; What is still uncertain; What to verify next.",
            "Consumer language": "Structure the answer as: What customers keep saying; Emotional tone; Repeated phrases or ideas; Messaging implications; Quotes or examples only when they add value.",
            "Product / CX opportunities": "Structure the answer as: Biggest opportunity areas; Product opportunities; CX/onboarding/support opportunities; Quick wins; Longer-term bets.",
        }
        style_instructions = {
            "Actionable": "Be concise, decisive, and practical. Favor ranked recommendations over long narrative.",
            "Balanced": "Balance diagnosis with recommendations. Include enough explanation to make the actions credible.",
            "Deep dive": "Go deeper on tradeoffs, evidence, uncertainty, and second-order implications."
        }
        general_target_words = {"Actionable": 220, "Balanced": 340, "Deep dive": 520}
        guided_target_words = {"Actionable": 260, "Balanced": 430, "Deep dive": 650}
        target_words = (general_target_words if focus == "General" else guided_target_words).get(answer_style, 340 if focus == "General" else 430)
        grounding_note = (
            "Ground the answer in the current filtered view and current workspace. Quantify where possible, separate evidence from inference, avoid generic advice, and answer simple questions plainly when the data supports them. "
            "Use named symptom themes, concrete consumer language, and likely owner-level actions when they are relevant to the question."
        )
        prompt_parts = [
            prompt,
            f"Mode: {focus}.",
            f"Answer style: {answer_style}.",
            f"Special instruction: {focus_instructions.get(focus, '')}",
            f"Style instruction: {style_instructions.get(answer_style, '')}",
            grounding_note,
        ]
        if focus != "General" and response_scaffolds.get(focus):
            prompt_parts.insert(4, f"Response format: {response_scaffolds.get(focus, '')}")
        else:
            prompt_parts.insert(4, "Response format: Freeform. Answer naturally and only add sections when they improve clarity.")
        enriched_prompt = "\n\n".join(part for part in prompt_parts if part)

        st.session_state.setdefault("chat_messages", []).append({"role": "user", "content": prompt})
        with st.spinner("Thinking through the current view…"):
            try:
                answer = _call_analyst(
                    question=enriched_prompt,
                    overall_df=overall_df,
                    filtered_df=filtered_df,
                    summary=summary,
                    filter_description=filter_description,
                    chat_history=chat_messages,
                    persona_name=persona_map.get(focus),
                    target_words=target_words,
                    include_references=bool(st.session_state.get("ai_include_references", False)),
                    freeform_mode=(focus == "General"),
                )
                st.session_state["chat_messages"].append({"role": "assistant", "content": answer})
            except Exception as exc:
                st.session_state["chat_messages"].append({"role": "assistant", "content": f"Error: {exc}"})
        st.rerun()


if __name__ == "__main__":
    main()

