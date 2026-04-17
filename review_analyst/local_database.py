from __future__ import annotations

import csv
import hashlib
import json
import os
import pickle
import re
import sqlite3
import time
from difflib import SequenceMatcher
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Callable, Dict, Iterable, Iterator, List, Optional, Sequence, Tuple

import pandas as pd
from openpyxl import load_workbook

from .models import ReviewBatchSummary
from .normalization import (
    REVIEW_ID_ALIASES,
    REVIEW_TEXT_ALIASES,
    TITLE_ALIASES,
    DATE_ALIASES,
    RATING_ALIASES,
    _score_uploaded_sheet,
    finalize_df,
    normalize_uploaded_df,
)

_LOCAL_DB_ENV = "STARWALK_LOCAL_REVIEW_DB_ROOT"
_DB_FILENAME = "central_review_database.sqlite3"
_SNAPSHOT_FILENAME = "central_review_database_snapshot.xlsx"
_DEFAULT_CHUNK_SIZE = 50000
_WORKSPACE_CACHE_SUBDIR = "workspace_cache"
_MAX_WORKSPACE_CACHE_FILES = 24
_SEARCH_RESULT_SCAN_LIMIT = 15000
_REQUIRED_READY_TABLES = {"reviews_enriched", "sku_catalog", "base_model_directory", "product_directory"}
_WORKSPACE_EXCLUDED_COLUMNS = {"raw_json", "context_data_json", "badges"}
_SUPPORTED_TABULAR_SUFFIXES = {".csv", ".xlsx", ".xls", ".xlsm"}
_DYSON_SUBDIR = "dyson"

_MANIFEST_SQL = """
CREATE TABLE IF NOT EXISTS import_manifest (
    manifest_id INTEGER PRIMARY KEY AUTOINCREMENT,
    synced_at TEXT NOT NULL,
    review_file_name TEXT,
    review_file_path TEXT,
    review_file_size INTEGER,
    review_file_mtime REAL,
    mapping_file_name TEXT,
    mapping_file_path TEXT,
    mapping_file_size INTEGER,
    mapping_file_mtime REAL,
    review_count INTEGER NOT NULL DEFAULT 0,
    mapped_review_count INTEGER NOT NULL DEFAULT 0,
    unmatched_review_count INTEGER NOT NULL DEFAULT 0,
    base_model_count INTEGER NOT NULL DEFAULT 0,
    distinct_product_count INTEGER NOT NULL DEFAULT 0,
    notes_json TEXT NOT NULL DEFAULT '{}'
)
"""

_REVIEW_BASE_COLUMNS = [
    "review_id",
    "product_id",
    "base_sku",
    "sku_item",
    "product_or_sku",
    "original_product_name",
    "title",
    "review_text",
    "rating",
    "is_recommended",
    "content_locale",
    "submission_time",
    "submission_date",
    "submission_month",
    "incentivized_review",
    "is_syndicated",
    "photos_count",
    "photo_urls",
    "title_and_text",
    "retailer",
    "post_link",
    "age_group",
    "user_nickname",
    "user_location",
    "total_positive_feedback_count",
    "source_system",
    "source_file",
]

_CATALOG_COLUMNS = [
    "sku",
    "sku_key",
    "master_item",
    "master_item_key",
    "base_model_number",
    "mapped_brand",
    "mapped_category",
    "mapped_subcategory",
    "mapped_subsub_category",
    "mapped_region",
    "item_status",
    "lifecycle_phase",
    "description",
    "user_item_type",
    "item_class",
    "item_long_description",
    "ean",
    "upc",
]

_ENRICHED_REVIEW_COLUMNS = [
    "base_model_number",
    "master_item",
    "catalog_match_type",
    "catalog_match_key",
    "mapped_sku",
    "mapped_brand",
    "mapped_category",
    "mapped_subcategory",
    "mapped_subsub_category",
    "mapped_region",
    "mapped_item_status",
    "mapped_lifecycle_phase",
    "mapped_description",
    "mapped_user_item_type",
    "mapped_item_class",
    "mapped_item_long_description",
    "mapped_ean",
    "mapped_upc",
    "base_model_detected",
]


def _utcnow_iso() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def _safe_text(value: Any) -> str:
    return str(value or "").strip()


def _dedupe_preserve_order(values: Iterable[str]) -> List[str]:
    seen: set[str] = set()
    ordered: List[str] = []
    for raw in values:
        text = _safe_text(raw)
        if not text or text in seen:
            continue
        seen.add(text)
        ordered.append(text)
    return ordered


def _normalize_selection_values(value: Any) -> List[str]:
    if value is None:
        return []
    if isinstance(value, (list, tuple, set)):
        items = list(value)
    else:
        text = _safe_text(value)
        items = [text] if text else []
    return _dedupe_preserve_order(_safe_text(item) for item in items)


def _selection_payload(value: Any) -> Any:
    normalized = _normalize_selection_values(value)
    if not normalized:
        return []
    if len(normalized) == 1:
        return normalized[0]
    return normalized


def _primary_selection_value(value: Any) -> str:
    normalized = _normalize_selection_values(value)
    return normalized[0] if normalized else ""


def _normalize_key(value: Any) -> str:
    return _safe_text(value).upper()


def _normalize_series(series: pd.Series) -> pd.Series:
    return series.fillna("").astype(str).str.strip().str.upper()


def _file_info(path: Optional[Path], *, relative_to: Optional[Path] = None) -> Dict[str, Any]:
    if path is None or not Path(path).exists():
        return {
            "path": "",
            "name": "",
            "relative_path": "",
            "size": 0,
            "mtime": 0.0,
            "modified_at": "",
        }
    resolved = Path(path).resolve()
    stat = resolved.stat()
    relative_path = ""
    if relative_to is not None:
        try:
            relative_path = str(resolved.relative_to(Path(relative_to).resolve()))
        except Exception:
            relative_path = resolved.name
    return {
        "path": str(resolved),
        "name": resolved.name,
        "relative_path": relative_path,
        "size": int(stat.st_size),
        "mtime": float(stat.st_mtime),
        "modified_at": datetime.fromtimestamp(stat.st_mtime, tz=timezone.utc).replace(microsecond=0).isoformat(),
    }


def default_local_review_db_root() -> str:
    raw = _safe_text(os.environ.get(_LOCAL_DB_ENV))
    if raw:
        path = Path(raw).expanduser().resolve()
    else:
        path = (Path(__file__).resolve().parents[1] / ".starwalk_data" / "local_review_database").resolve()
    path.mkdir(parents=True, exist_ok=True)
    return str(path)


def ensure_local_review_db_dirs(root: Optional[str] = None) -> Dict[str, str]:
    root_path = Path(root or default_local_review_db_root()).expanduser().resolve()
    incoming_path = root_path / "incoming"
    reviews_path = incoming_path / "reviews"
    mapping_path = incoming_path / "sku_mapping"
    dyson_reviews_path = reviews_path / _DYSON_SUBDIR
    dyson_mapping_path = mapping_path / _DYSON_SUBDIR
    exports_path = root_path / "exports"
    cache_path = root_path / "cache"
    workspace_cache_path = cache_path / _WORKSPACE_CACHE_SUBDIR
    for path in [root_path, incoming_path, reviews_path, mapping_path, dyson_reviews_path, dyson_mapping_path, exports_path, cache_path, workspace_cache_path]:
        path.mkdir(parents=True, exist_ok=True)
    return {
        "root": str(root_path),
        "incoming_dir": str(incoming_path),
        "reviews_dir": str(reviews_path),
        "mapping_dir": str(mapping_path),
        "dyson_reviews_dir": str(dyson_reviews_path),
        "dyson_mapping_dir": str(dyson_mapping_path),
        "exports_dir": str(exports_path),
        "cache_dir": str(cache_path),
        "workspace_cache_dir": str(workspace_cache_path),
        "db_path": str((root_path / _DB_FILENAME).resolve()),
        "snapshot_path": str((exports_path / _SNAPSHOT_FILENAME).resolve()),
    }


def _candidate_files(folder: Path, *, recursive: bool = False) -> List[Path]:
    if not folder.exists():
        return []
    iterator = folder.rglob("*") if recursive else folder.iterdir()
    candidates = [path for path in iterator if path.is_file() and path.suffix.lower() in _SUPPORTED_TABULAR_SUFFIXES]
    return sorted(candidates, key=lambda p: (p.stat().st_mtime, p.name.lower()), reverse=True)


def _dedupe_paths(paths: Iterable[Path]) -> List[Path]:
    seen: set[str] = set()
    ordered: List[Path] = []
    for raw_path in paths:
        try:
            resolved = Path(raw_path).resolve()
        except Exception:
            resolved = Path(raw_path)
        key = str(resolved)
        if key in seen:
            continue
        seen.add(key)
        ordered.append(resolved)
    return ordered


def _source_file_infos(paths: Sequence[Path], *, relative_to: Path) -> List[Dict[str, Any]]:
    return [_file_info(path, relative_to=relative_to) for path in paths if path is not None and Path(path).exists()]


def _source_signature(file_infos: Sequence[Dict[str, Any]]) -> str:
    payload = [
        {
            "relative_path": _safe_text(info.get("relative_path")),
            "name": _safe_text(info.get("name")),
            "size": int(info.get("size") or 0),
            "mtime": float(info.get("mtime") or 0.0),
        }
        for info in file_infos
    ]
    payload.sort(key=lambda item: (item.get("relative_path", ""), item.get("name", "")))
    return hashlib.sha1(json.dumps(payload, sort_keys=True, ensure_ascii=False).encode("utf-8")).hexdigest()


def _primary_standard_files(folder: Path) -> List[Path]:
    files = []
    for path in _candidate_files(folder):
        lower_name = path.name.lower()
        if lower_name.startswith(f"{_DYSON_SUBDIR}__") or _DYSON_SUBDIR in lower_name:
            continue
        files.append(path)
    return files


def _dyson_source_files(folder: Path) -> List[Path]:
    dyson_dir = folder / _DYSON_SUBDIR
    nested = _candidate_files(dyson_dir, recursive=True)
    root_matches = [
        path
        for path in _candidate_files(folder)
        if path.name.lower().startswith(f"{_DYSON_SUBDIR}__") or _DYSON_SUBDIR in path.name.lower()
    ]
    return _dedupe_paths([*nested, *root_matches])


def discover_latest_local_sources(root: Optional[str] = None) -> Dict[str, Any]:
    dirs = ensure_local_review_db_dirs(root)
    reviews_dir = Path(dirs["reviews_dir"])
    mapping_dir = Path(dirs["mapping_dir"])
    primary_review_candidates = _primary_standard_files(reviews_dir)
    primary_mapping_candidates = _primary_standard_files(mapping_dir)
    dyson_review_paths = _dyson_source_files(reviews_dir)
    dyson_mapping_paths = _dyson_source_files(mapping_dir)
    review_paths = _dedupe_paths([*(primary_review_candidates[:1]), *dyson_review_paths])
    mapping_paths = _dedupe_paths([*(primary_mapping_candidates[:1]), *dyson_mapping_paths])
    review_file = primary_review_candidates[0] if primary_review_candidates else (review_paths[0] if review_paths else None)
    mapping_file = primary_mapping_candidates[0] if primary_mapping_candidates else (mapping_paths[0] if mapping_paths else None)
    review_files = _source_file_infos(review_paths, relative_to=Path(dirs["incoming_dir"]))
    mapping_files = _source_file_infos(mapping_paths, relative_to=Path(dirs["incoming_dir"]))
    dyson_review_files = _source_file_infos(dyson_review_paths, relative_to=Path(dirs["incoming_dir"]))
    dyson_mapping_files = _source_file_infos(dyson_mapping_paths, relative_to=Path(dirs["incoming_dir"]))
    return {
        **dirs,
        "review_file": _file_info(review_file, relative_to=Path(dirs["incoming_dir"])),
        "mapping_file": _file_info(mapping_file, relative_to=Path(dirs["incoming_dir"])),
        "review_files": review_files,
        "mapping_files": mapping_files,
        "dyson_review_files": dyson_review_files,
        "dyson_mapping_files": dyson_mapping_files,
        "review_source_count": len(review_files),
        "mapping_source_count": len(mapping_files),
        "dyson_review_count": len(dyson_review_files),
        "dyson_mapping_count": len(dyson_mapping_files),
        "source_signature": _source_signature([*review_files, *mapping_files]),
        "has_inputs": bool(review_paths and mapping_paths),
    }


def _connect(db_path: str) -> sqlite3.Connection:
    conn = sqlite3.connect(db_path)
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA synchronous=NORMAL")
    conn.execute("PRAGMA busy_timeout=5000")
    conn.execute("PRAGMA temp_store=MEMORY")
    conn.execute("PRAGMA wal_autocheckpoint=1000")
    conn.execute("PRAGMA cache_size=-131072")
    conn.execute("PRAGMA cache_spill=FALSE")
    try:
        conn.execute("PRAGMA mmap_size=1073741824")
    except Exception:
        pass
    conn.execute(_MANIFEST_SQL)
    return conn


def _latest_manifest(conn: sqlite3.Connection) -> Dict[str, Any]:
    conn.row_factory = sqlite3.Row
    row = conn.execute(
        "SELECT * FROM import_manifest ORDER BY manifest_id DESC LIMIT 1"
    ).fetchone()
    if row is None:
        return {}
    result = dict(row)
    notes_raw = _safe_text(result.get("notes_json"))
    try:
        result["notes"] = json.loads(notes_raw) if notes_raw else {}
    except Exception:
        result["notes"] = {}
    return result


def _manifest_matches_sources(manifest: Dict[str, Any], sources: Dict[str, Any]) -> bool:
    if not manifest:
        return False
    notes = dict(manifest.get("notes") or {})
    manifest_signature = _safe_text(notes.get("source_signature"))
    source_signature = _safe_text(sources.get("source_signature"))
    if manifest_signature and source_signature:
        return manifest_signature == source_signature
    if source_signature and (int(sources.get("review_source_count") or 0) > 1 or int(sources.get("mapping_source_count") or 0) > 1):
        return False
    review = sources.get("review_file") or {}
    mapping = sources.get("mapping_file") or {}
    return all(
        [
            _safe_text(manifest.get("review_file_name")) == _safe_text(review.get("name")),
            int(manifest.get("review_file_size") or 0) == int(review.get("size") or 0),
            float(manifest.get("review_file_mtime") or 0.0) == float(review.get("mtime") or 0.0),
            _safe_text(manifest.get("mapping_file_name")) == _safe_text(mapping.get("name")),
            int(manifest.get("mapping_file_size") or 0) == int(mapping.get("size") or 0),
            float(manifest.get("mapping_file_mtime") or 0.0) == float(mapping.get("mtime") or 0.0),
        ]
    )


def get_local_review_db_status(root: Optional[str] = None) -> Dict[str, Any]:
    sources = discover_latest_local_sources(root)
    db_path = sources["db_path"]
    manifest: Dict[str, Any] = {}
    tables: List[str] = []
    if Path(db_path).exists():
        with _connect(db_path) as conn:
            manifest = _latest_manifest(conn)
            tables = [str(row[0]) for row in conn.execute("SELECT name FROM sqlite_master WHERE type='table'").fetchall()]
    status = dict(sources)
    ready_tables = _REQUIRED_READY_TABLES.issubset(set(tables or []))
    status.update(
        {
            "db_exists": Path(db_path).exists(),
            "manifest": manifest,
            "tables": tables,
            "needs_sync": bool(sources.get("has_inputs")) and (not manifest or not ready_tables or not _manifest_matches_sources(manifest, sources)),
            "is_ready": bool(Path(db_path).exists()) and bool(manifest) and ready_tables,
        }
    )
    return status


def _select_best_excel_sheet(workbook_path: Path) -> str:
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    best_name = wb.sheetnames[0] if wb.sheetnames else ""
    best_score = -1
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = ws.iter_rows(min_row=1, max_row=1, values_only=True)
        header_row = next(rows, None) or []
        score = _score_uploaded_sheet(list(header_row))
        if score > best_score:
            best_score = score
            best_name = sheet_name
    return best_name


def _iter_csv_chunks(path: Path, *, chunk_size: int = _DEFAULT_CHUNK_SIZE) -> Iterator[pd.DataFrame]:
    last_error: Optional[Exception] = None
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            with open(path, "r", newline="", encoding=encoding) as handle:
                reader = csv.DictReader(handle)
                rows: List[Dict[str, Any]] = []
                rows_seen = 0
                for row in reader:
                    rows.append(row)
                    if len(rows) >= chunk_size:
                        chunk = pd.DataFrame(rows)
                        chunk.attrs["chunk_row_offset"] = rows_seen
                        yield chunk
                        rows_seen += len(rows)
                        rows = []
                if rows:
                    chunk = pd.DataFrame(rows)
                    chunk.attrs["chunk_row_offset"] = rows_seen
                    yield chunk
            return
        except UnicodeDecodeError as exc:
            last_error = exc
            continue
    if last_error is not None:
        raise last_error


def _iter_excel_chunks(path: Path, *, chunk_size: int = _DEFAULT_CHUNK_SIZE) -> Iterator[pd.DataFrame]:
    sheet_name = _select_best_excel_sheet(path)
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    row_iter = ws.iter_rows(values_only=True)
    header_row = next(row_iter, None)
    if header_row is None:
        return
    headers = [str(value).strip() if value is not None else "" for value in header_row]
    batch: List[Dict[str, Any]] = []
    rows_seen = 0
    for raw_row in row_iter:
        values = list(raw_row)
        payload = {headers[idx]: values[idx] if idx < len(values) else None for idx in range(len(headers))}
        batch.append(payload)
        if len(batch) >= chunk_size:
            chunk = pd.DataFrame(batch)
            chunk.attrs["chunk_row_offset"] = rows_seen
            yield chunk
            rows_seen += len(batch)
            batch = []
    if batch:
        chunk = pd.DataFrame(batch)
        chunk.attrs["chunk_row_offset"] = rows_seen
        yield chunk


def _iter_tabular_chunks(path: str, *, chunk_size: int = _DEFAULT_CHUNK_SIZE) -> Iterator[pd.DataFrame]:
    file_path = Path(path)
    suffix = file_path.suffix.lower()
    if suffix == ".csv":
        yield from _iter_csv_chunks(file_path, chunk_size=chunk_size)
        return
    if suffix in {".xlsx", ".xlsm"}:
        yield from _iter_excel_chunks(file_path, chunk_size=chunk_size)
        return
    if suffix == ".xls":
        raw = pd.read_excel(file_path)
        if raw.empty:
            return
        for start in range(0, len(raw), chunk_size):
            chunk = raw.iloc[start:start + chunk_size].copy()
            chunk.attrs["chunk_row_offset"] = start
            yield chunk
        return
    raise ValueError(f"Unsupported review file type: {file_path.name}")


def _read_tabular_file(path: str) -> pd.DataFrame:
    file_path = Path(path)
    suffix = file_path.suffix.lower()
    if suffix == ".csv":
        last_error: Optional[Exception] = None
        for encoding in ("utf-8-sig", "utf-8", "latin-1"):
            try:
                return pd.read_csv(file_path, encoding=encoding)
            except UnicodeDecodeError as exc:
                last_error = exc
                continue
        if last_error is not None:
            raise last_error
        return pd.DataFrame()
    if suffix in {".xlsx", ".xlsm"}:
        sheet_name = _select_best_excel_sheet(file_path)
        return pd.read_excel(file_path, sheet_name=sheet_name)
    if suffix == ".xls":
        return pd.read_excel(file_path)
    raise ValueError(f"Unsupported tabular file type: {file_path.name}")


def _read_mapping_file(path: str) -> pd.DataFrame:
    file_path = Path(path)
    suffix = file_path.suffix.lower()
    if suffix == ".csv":
        return _read_tabular_file(path)
    if suffix in {".xlsx", ".xlsm", ".xls"}:
        return pd.read_excel(file_path)
    raise ValueError(f"Unsupported mapping file type: {file_path.name}")


def load_master_sku_catalog(path: str) -> pd.DataFrame:
    raw = _read_mapping_file(path)
    raw.columns = [str(col).strip() for col in raw.columns]
    catalog = pd.DataFrame()
    catalog["sku"] = raw.get("SKU", pd.Series(dtype="object")).astype("string").fillna("").str.strip()
    catalog["sku_key"] = catalog["sku"].str.upper()
    catalog["master_item"] = raw.get("Master Item", pd.Series(dtype="object")).astype("string").fillna("").str.strip()
    catalog["master_item_key"] = catalog["master_item"].str.upper()
    catalog["base_model_number"] = catalog["master_item"].where(catalog["master_item"].ne(""), catalog["sku"])
    catalog["mapped_brand"] = raw.get("Brand", pd.Series(dtype="object")).astype("string").fillna("").str.strip().replace({"": pd.NA})
    catalog["mapped_category"] = raw.get("Category", pd.Series(dtype="object")).astype("string").fillna("").str.strip().replace({"": pd.NA})
    catalog["mapped_subcategory"] = raw.get("SubCategory", pd.Series(dtype="object")).astype("string").fillna("").str.strip().replace({"": pd.NA})
    catalog["mapped_subsub_category"] = raw.get("SubsubCategory", pd.Series(dtype="object")).astype("string").fillna("").str.strip().replace({"": pd.NA})
    catalog["mapped_region"] = raw.get("REGION", pd.Series(dtype="object")).astype("string").fillna("").str.strip().replace({"": pd.NA})
    catalog["item_status"] = raw.get("Item Status", pd.Series(dtype="object")).astype("string").fillna("").str.strip().replace({"": pd.NA})
    catalog["lifecycle_phase"] = raw.get("Lifecyle Phase", pd.Series(dtype="object")).astype("string").fillna("").str.strip().replace({"": pd.NA})
    catalog["description"] = raw.get("Description", pd.Series(dtype="object")).astype("string").fillna("").str.strip().replace({"": pd.NA})
    catalog["user_item_type"] = raw.get("User Item Type", pd.Series(dtype="object")).astype("string").fillna("").str.strip().replace({"": pd.NA})
    catalog["item_class"] = raw.get("Item Class", pd.Series(dtype="object")).astype("string").fillna("").str.strip().replace({"": pd.NA})
    catalog["item_long_description"] = raw.get("Item Long Description", pd.Series(dtype="object")).astype("string").fillna("").str.strip().replace({"": pd.NA})
    catalog["ean"] = raw.get("EAN", pd.Series(dtype="object")).astype("string").fillna("").str.strip().replace({"": pd.NA})
    catalog["upc"] = raw.get("UPC", pd.Series(dtype="object")).astype("string").fillna("").str.strip().replace({"": pd.NA})
    catalog = catalog[_CATALOG_COLUMNS].copy()
    catalog = catalog[(catalog["sku_key"].ne("")) | (catalog["master_item_key"].ne(""))].reset_index(drop=True)
    return catalog


def load_dyson_product_catalog(path: str) -> pd.DataFrame:
    raw = _read_mapping_file(path)
    raw.columns = [str(col).strip() for col in raw.columns]
    catalog = pd.DataFrame()
    catalog["sku"] = raw.get("PID", raw.get("Product ID", pd.Series(dtype="object"))).astype("string").fillna("").str.strip()
    catalog["sku_key"] = catalog["sku"].str.upper()
    catalog["master_item"] = raw.get("Model Code", raw.get("Master Item", pd.Series(dtype="object"))).astype("string").fillna("").str.strip()
    catalog["master_item_key"] = catalog["master_item"].str.upper()
    catalog["base_model_number"] = catalog["master_item"].where(catalog["master_item"].ne(""), catalog["sku"])
    catalog["mapped_brand"] = raw.get("Brand", pd.Series(["Dyson"] * len(raw), index=raw.index)).astype("string").fillna("Dyson").str.strip().replace({"": "Dyson"})
    catalog["mapped_category"] = raw.get("Category", pd.Series(dtype="object")).astype("string").fillna("").str.strip().replace({"": pd.NA})
    catalog["mapped_subcategory"] = raw.get("Sub-Category", raw.get("SubCategory", pd.Series(dtype="object"))).astype("string").fillna("").str.strip().replace({"": pd.NA})
    catalog["mapped_subsub_category"] = raw.get("SubsubCategory", pd.Series(dtype="object")).astype("string").fillna("").str.strip().replace({"": pd.NA})
    catalog["mapped_region"] = raw.get("Region", raw.get("REGION", pd.Series(dtype="object"))).astype("string").fillna("").str.strip().replace({"": pd.NA})
    catalog["item_status"] = raw.get("Item Status", pd.Series(dtype="object")).astype("string").fillna("").str.strip().replace({"": pd.NA})
    catalog["lifecycle_phase"] = raw.get("Lifecycle Phase", raw.get("Lifecyle Phase", pd.Series(dtype="object"))).astype("string").fillna("").str.strip().replace({"": pd.NA})
    product_name = raw.get("Product Name", pd.Series(dtype="object")).astype("string").fillna("").str.strip()
    color_variant = raw.get("Color Variant", pd.Series(dtype="object")).astype("string").fillna("").str.strip()
    notes = raw.get("Notes", pd.Series(dtype="object")).astype("string").fillna("").str.strip()
    catalog["description"] = product_name.replace({"": pd.NA})
    catalog["user_item_type"] = raw.get("Product Type", pd.Series(dtype="object")).astype("string").fillna("").str.strip().replace({"": pd.NA})
    catalog["item_class"] = raw.get("Sub-Category", raw.get("SubCategory", pd.Series(dtype="object"))).astype("string").fillna("").str.strip().replace({"": pd.NA})
    detail_series = product_name.where(product_name.ne(""), catalog["base_model_number"].astype("string"))
    detail_series = detail_series.astype("string").fillna("").str.strip()
    detail_series = detail_series.where(color_variant.eq(""), detail_series + " · " + color_variant)
    detail_series = detail_series.where(notes.eq(""), detail_series + " · " + notes)
    catalog["item_long_description"] = detail_series.replace({"": pd.NA})
    catalog["ean"] = raw.get("EAN", pd.Series(dtype="object")).astype("string").fillna("").str.strip().replace({"": pd.NA})
    catalog["upc"] = raw.get("UPC", pd.Series(dtype="object")).astype("string").fillna("").str.strip().replace({"": pd.NA})
    catalog = catalog[_CATALOG_COLUMNS].copy()
    catalog = catalog[(catalog["sku_key"].ne("")) | (catalog["master_item_key"].ne(""))].reset_index(drop=True)
    return catalog


def _combine_catalog_frames(frames: Sequence[pd.DataFrame]) -> pd.DataFrame:
    valid_frames = [frame.copy() for frame in frames if isinstance(frame, pd.DataFrame) and not frame.empty]
    if not valid_frames:
        return pd.DataFrame(columns=_CATALOG_COLUMNS)
    combined = pd.concat(valid_frames, ignore_index=True)
    for col in _CATALOG_COLUMNS:
        if col not in combined.columns:
            combined[col] = pd.NA
    combined = combined[_CATALOG_COLUMNS].copy()
    combined["sku_key"] = combined["sku_key"].astype("string").fillna("").str.strip().str.upper()
    combined["master_item_key"] = combined["master_item_key"].astype("string").fillna("").str.strip().str.upper()
    combined = combined.drop_duplicates(subset=["sku_key", "master_item_key", "mapped_brand", "mapped_category", "mapped_subcategory"], keep="first")
    combined = combined.reset_index(drop=True)
    return combined


def _dyson_pid_from_filename(path: Path) -> str:
    stem = path.stem.strip()
    lower = stem.lower()
    if lower.startswith(f"{_DYSON_SUBDIR}__"):
        return stem.split("__", 1)[1].strip().lower()
    return stem.strip().lower()


def _dyson_catalog_by_pid(catalog_df: pd.DataFrame) -> Dict[str, Dict[str, Any]]:
    if catalog_df.empty:
        return {}
    lookup: Dict[str, Dict[str, Any]] = {}
    for _, row in catalog_df.iterrows():
        pid = _safe_text(row.get("sku")).lower()
        if not pid or pid in lookup:
            continue
        lookup[pid] = row.to_dict()
    return lookup


def normalize_dyson_review_chunk(raw_chunk: pd.DataFrame, *, source_path: Path, catalog_row: Optional[Dict[str, Any]] = None) -> pd.DataFrame:
    working = raw_chunk.copy()
    working.columns = [str(col).strip() for col in working.columns]
    pid = _dyson_pid_from_filename(source_path)
    catalog_row = dict(catalog_row or {})
    base_model = _safe_text(catalog_row.get("base_model_number") or catalog_row.get("master_item") or pid)
    product_name = _safe_text(catalog_row.get("description") or catalog_row.get("item_long_description") or pid)
    brand = _safe_text(catalog_row.get("mapped_brand") or "Dyson") or "Dyson"
    review_offset = 0
    if not working.empty:
        review_offset = int(getattr(raw_chunk, "attrs", {}).get("chunk_row_offset", 0) or 0)
    if "Product ID" not in working.columns:
        working["Product ID"] = pid
    if "Base SKU" not in working.columns:
        working["Base SKU"] = base_model
    if "SKU Item" not in working.columns:
        working["SKU Item"] = pid
    if "Product Name" not in working.columns:
        working["Product Name"] = product_name
    if "Brand" not in working.columns:
        working["Brand"] = brand
    if "Retailer" not in working.columns:
        working["Retailer"] = "Dyson"
    if "Source" not in working.columns:
        working["Source"] = "Dyson"
    if "Moderation Status" not in working.columns:
        working["Moderation Status"] = "APPROVED"
    if "Review ID" not in working.columns:
        working["Review ID"] = [f"dyson__{pid}__{review_offset + idx + 1}" for idx in range(len(working))]
    normalized = normalize_uploaded_df(working, source_name=source_path.name, include_local_symptomization=False)
    normalized["product_id"] = normalized["product_id"].astype("string").fillna("").str.strip().replace({"": pid})
    normalized["base_sku"] = normalized["base_sku"].astype("string").fillna("").str.strip().replace({"": base_model})
    normalized["sku_item"] = normalized["sku_item"].astype("string").fillna("").str.strip().replace({"": pid})
    normalized["original_product_name"] = normalized["original_product_name"].astype("string").fillna("").str.strip().replace({"": product_name})
    normalized["brand_raw"] = normalized.get("brand_raw", pd.Series(pd.NA, index=normalized.index)).astype("string").fillna("").str.strip().replace({"": brand})
    normalized["retailer"] = normalized["retailer"].astype("string").fillna("").str.strip().replace({"": "Dyson"})
    normalized["source_system"] = "Dyson Uploaded File"
    normalized["source_file"] = source_path.name
    return finalize_df(normalized)


def _lookup_frame(catalog: pd.DataFrame, key_col: str) -> pd.DataFrame:
    frame = catalog[catalog[key_col].fillna("").astype(str).str.strip().ne("")].copy()
    if frame.empty:
        return frame
    return frame.drop_duplicates(subset=[key_col], keep="first").set_index(key_col)


def enrich_reviews_with_catalog(reviews_df: pd.DataFrame, catalog_df: pd.DataFrame) -> pd.DataFrame:
    working = reviews_df.copy()
    sku_lookup = _lookup_frame(catalog_df, "sku_key")
    master_lookup = _lookup_frame(catalog_df, "master_item_key")

    for col in _ENRICHED_REVIEW_COLUMNS:
        if col not in working.columns:
            working[col] = pd.NA

    product_key = _normalize_series(working.get("product_id", pd.Series(index=working.index, dtype="object")))
    base_key = _normalize_series(working.get("base_sku", pd.Series(index=working.index, dtype="object")))
    sku_item_key = _normalize_series(working.get("sku_item", pd.Series(index=working.index, dtype="object")))

    mappings = [
        ("product_to_sku", product_key, sku_lookup),
        ("sku_item_to_sku", sku_item_key, sku_lookup),
        ("base_to_sku", base_key, sku_lookup),
        ("base_to_master", base_key, master_lookup),
        ("product_to_master", product_key, master_lookup),
        ("sku_item_to_master", sku_item_key, master_lookup),
    ]
    field_map = {
        "base_model_number": "base_model_number",
        "master_item": "master_item",
        "mapped_sku": "sku",
        "mapped_brand": "mapped_brand",
        "mapped_category": "mapped_category",
        "mapped_subcategory": "mapped_subcategory",
        "mapped_subsub_category": "mapped_subsub_category",
        "mapped_region": "mapped_region",
        "mapped_item_status": "item_status",
        "mapped_lifecycle_phase": "lifecycle_phase",
        "mapped_description": "description",
        "mapped_user_item_type": "user_item_type",
        "mapped_item_class": "item_class",
        "mapped_item_long_description": "item_long_description",
        "mapped_ean": "ean",
        "mapped_upc": "upc",
    }

    assigned = pd.Series(False, index=working.index)
    for match_type, key_series, lookup in mappings:
        if lookup.empty:
            continue
        hit = (~assigned) & key_series.ne("") & key_series.isin(lookup.index)
        if not bool(hit.any()):
            continue
        keys = key_series.loc[hit]
        matched = lookup.reindex(keys)
        working.loc[hit, "catalog_match_type"] = match_type
        working.loc[hit, "catalog_match_key"] = keys.values
        for target_col, source_col in field_map.items():
            if source_col in matched.columns:
                working.loc[hit, target_col] = matched[source_col].values
        assigned |= hit

    base_fallback = working.get("base_sku", pd.Series(index=working.index, dtype="object")).astype("string").fillna("").str.strip()
    product_fallback = working.get("product_id", pd.Series(index=working.index, dtype="object")).astype("string").fillna("").str.strip()
    working["base_model_number"] = working["base_model_number"].astype("string").fillna("").str.strip()
    working["base_model_number"] = working["base_model_number"].where(working["base_model_number"].ne(""), base_fallback)
    working["base_model_number"] = working["base_model_number"].where(working["base_model_number"].ne(""), product_fallback)
    working["base_model_number"] = working["base_model_number"].replace({"": pd.NA})
    working["master_item"] = working["master_item"].astype("string").fillna("").str.strip()
    working["master_item"] = working["master_item"].where(working["master_item"].ne(""), working["base_model_number"].astype("string"))
    working["master_item"] = working["master_item"].replace({"": pd.NA})
    working["base_model_detected"] = working["base_model_number"].astype("string").fillna("").str.strip().ne("")
    return working


def _maybe_report(progress_callback: Optional[Callable[..., None]], **kwargs: Any) -> None:
    if callable(progress_callback):
        try:
            progress_callback(**kwargs)
        except TypeError:
            try:
                progress_callback(kwargs.get("progress"), kwargs.get("title", ""), kwargs.get("detail", ""))
            except Exception:
                return
        except Exception:
            return


def _sanitize_sql_label(value: str) -> str:
    return '"' + str(value).replace('"', '""') + '"'


def _ensure_table_has_columns(conn: sqlite3.Connection, table_name: str, columns: Sequence[str]) -> None:
    existing = set(_table_columns(conn, table_name))
    for column in columns:
        label = str(column)
        if label in existing:
            continue
        conn.execute(f"ALTER TABLE {_sanitize_sql_label(table_name)} ADD COLUMN {_sanitize_sql_label(label)}")
        existing.add(label)


def _append_value_filter(where_parts: List[str], params: Dict[str, Any], column: str, values: Any, *, param_prefix: str) -> None:
    normalized = _normalize_selection_values(values)
    if not normalized:
        return
    placeholders: List[str] = []
    for idx, value in enumerate(normalized):
        key = f"{param_prefix}_{idx}"
        params[key] = value
        placeholders.append(f":{key}")
    if len(placeholders) == 1:
        where_parts.append(f"{column} = {placeholders[0]}")
    else:
        where_parts.append(f"{column} IN ({', '.join(placeholders)})")


def _workspace_projection_columns(conn: sqlite3.Connection) -> List[str]:
    columns = _table_columns(conn, "reviews_enriched")
    projected = [col for col in columns if col not in _WORKSPACE_EXCLUDED_COLUMNS]
    return projected or columns


def _create_indexes(conn: sqlite3.Connection) -> None:
    index_sql = [
        "CREATE INDEX IF NOT EXISTS idx_reviews_review_id ON reviews_enriched(review_id)",
        "CREATE INDEX IF NOT EXISTS idx_reviews_base_model ON reviews_enriched(base_model_number)",
        "CREATE INDEX IF NOT EXISTS idx_reviews_product_id ON reviews_enriched(product_id)",
        "CREATE INDEX IF NOT EXISTS idx_reviews_mapped_brand ON reviews_enriched(mapped_brand)",
        "CREATE INDEX IF NOT EXISTS idx_reviews_mapped_category ON reviews_enriched(mapped_category)",
        "CREATE INDEX IF NOT EXISTS idx_reviews_mapped_subcategory ON reviews_enriched(mapped_subcategory)",
        "CREATE INDEX IF NOT EXISTS idx_reviews_submission_date ON reviews_enriched(submission_date)",
        "CREATE INDEX IF NOT EXISTS idx_reviews_submission_time ON reviews_enriched(submission_time DESC, review_id DESC)",
        "CREATE INDEX IF NOT EXISTS idx_reviews_base_model_time ON reviews_enriched(base_model_number, submission_time DESC)",
        "CREATE INDEX IF NOT EXISTS idx_reviews_product_time ON reviews_enriched(product_id, submission_time DESC)",
        "CREATE INDEX IF NOT EXISTS idx_reviews_brand_category_time ON reviews_enriched(mapped_brand, mapped_category, mapped_subcategory, submission_time DESC)",
        "CREATE INDEX IF NOT EXISTS idx_reviews_selection_combo_time ON reviews_enriched(mapped_brand, mapped_category, mapped_subcategory, base_model_number, product_id, submission_time DESC)",
        "CREATE INDEX IF NOT EXISTS idx_reviews_brand_country_time ON reviews_enriched(mapped_brand, country, submission_time DESC)",
        "CREATE INDEX IF NOT EXISTS idx_reviews_moderation_bucket ON reviews_enriched(moderation_bucket)",
        "CREATE INDEX IF NOT EXISTS idx_reviews_moderation_incent_time ON reviews_enriched(moderation_bucket, incentivized_review, submission_time DESC)",
        "CREATE INDEX IF NOT EXISTS idx_reviews_review_origin_group ON reviews_enriched(review_origin_group)",
        "CREATE INDEX IF NOT EXISTS idx_reviews_country ON reviews_enriched(country)",
        "CREATE INDEX IF NOT EXISTS idx_catalog_sku_key ON sku_catalog(sku_key)",
        "CREATE INDEX IF NOT EXISTS idx_catalog_master_key ON sku_catalog(master_item_key)",
        "CREATE INDEX IF NOT EXISTS idx_base_model_directory_base_model ON base_model_directory(base_model_number)",
        "CREATE INDEX IF NOT EXISTS idx_base_model_directory_filters ON base_model_directory(mapped_brand, mapped_category, mapped_subcategory, base_model_number)",
        "CREATE INDEX IF NOT EXISTS idx_product_directory_base_model ON product_directory(base_model_number)",
        "CREATE INDEX IF NOT EXISTS idx_product_directory_product_id ON product_directory(product_id)",
        "CREATE INDEX IF NOT EXISTS idx_product_directory_filters ON product_directory(mapped_brand, mapped_category, mapped_subcategory, base_model_number, product_id)",
        "CREATE INDEX IF NOT EXISTS idx_product_directory_name ON product_directory(original_product_name)",
    ]
    for sql in index_sql:
        conn.execute(sql)



def _optimize_database(conn: sqlite3.Connection) -> None:
    try:
        conn.execute("ANALYZE")
    except Exception:
        pass
    try:
        conn.execute("PRAGMA optimize")
    except Exception:
        pass


def _rebuild_directory_tables(conn: sqlite3.Connection) -> None:
    conn.execute("DROP TABLE IF EXISTS base_model_directory")
    conn.execute(
        """
        CREATE TABLE base_model_directory AS
        SELECT
            COALESCE(NULLIF(base_model_number, ''), NULLIF(base_sku, ''), NULLIF(product_id, ''), 'UNMAPPED') AS base_model_number,
            COALESCE(NULLIF(mapped_brand, ''), '') AS mapped_brand,
            COALESCE(NULLIF(mapped_category, ''), '') AS mapped_category,
            COALESCE(NULLIF(mapped_subcategory, ''), '') AS mapped_subcategory,
            COALESCE(NULLIF(mapped_subsub_category, ''), '') AS mapped_subsub_category,
            COUNT(*) AS review_count,
            COUNT(DISTINCT review_id) AS distinct_review_count,
            COUNT(DISTINCT product_id) AS distinct_product_count,
            ROUND(AVG(CAST(rating AS REAL)), 3) AS avg_rating,
            MIN(submission_date) AS first_review_date,
            MAX(submission_date) AS latest_review_date
        FROM reviews_enriched
        GROUP BY 1, 2, 3, 4, 5
        """
    )
    conn.execute("DROP TABLE IF EXISTS product_directory")
    conn.execute(
        """
        CREATE TABLE product_directory AS
        SELECT
            COALESCE(NULLIF(base_model_number, ''), NULLIF(base_sku, ''), NULLIF(product_id, ''), 'UNMAPPED') AS base_model_number,
            COALESCE(NULLIF(product_id, ''), NULLIF(product_or_sku, ''), 'UNKNOWN_PRODUCT') AS product_id,
            COALESCE(NULLIF(original_product_name, ''), '') AS original_product_name,
            COALESCE(NULLIF(mapped_brand, ''), '') AS mapped_brand,
            COALESCE(NULLIF(mapped_category, ''), '') AS mapped_category,
            COALESCE(NULLIF(mapped_subcategory, ''), '') AS mapped_subcategory,
            COALESCE(NULLIF(mapped_subsub_category, ''), '') AS mapped_subsub_category,
            COUNT(*) AS review_count,
            COUNT(DISTINCT review_id) AS distinct_review_count,
            ROUND(AVG(CAST(rating AS REAL)), 3) AS avg_rating,
            MIN(submission_date) AS first_review_date,
            MAX(submission_date) AS latest_review_date
        FROM reviews_enriched
        GROUP BY 1, 2, 3, 4, 5, 6, 7
        """
    )


def _selection_label(*, base_model_number: Any = None, mapped_brand: Any = None, mapped_category: Any = None, mapped_subcategory: Any = None, product_id: Any = None) -> str:
    def _label(values: Any) -> str:
        normalized = _normalize_selection_values(values)
        if not normalized:
            return ""
        if len(normalized) <= 2:
            return ", ".join(normalized)
        return f"{normalized[0]} +{len(normalized) - 1} more"

    parts = []
    if _label(base_model_number):
        parts.append(_label(base_model_number))
    if _label(product_id):
        parts.append(_label(product_id))
    if _label(mapped_brand):
        parts.append(_label(mapped_brand))
    if _label(mapped_category):
        parts.append(_label(mapped_category))
    if _label(mapped_subcategory):
        parts.append(_label(mapped_subcategory))
    return " · ".join(parts) if parts else "Central review database"


def _manifest_token(status: Dict[str, Any]) -> str:
    manifest = dict(status.get("manifest") or {})
    review = dict(status.get("review_file") or {})
    mapping = dict(status.get("mapping_file") or {})
    payload = {
        "manifest_id": manifest.get("manifest_id"),
        "review_name": review.get("name"),
        "review_size": review.get("size"),
        "review_mtime": review.get("mtime"),
        "mapping_name": mapping.get("name"),
        "mapping_size": mapping.get("size"),
        "mapping_mtime": mapping.get("mtime"),
    }
    return hashlib.sha1(json.dumps(payload, sort_keys=True, ensure_ascii=False).encode("utf-8")).hexdigest()



def _workspace_cache_key(
    *,
    status: Dict[str, Any],
    base_model_number: Any = None,
    mapped_brand: Any = None,
    mapped_category: Any = None,
    mapped_subcategory: Any = None,
    product_id: Any = None,
    limit_rows: Optional[int] = None,
) -> str:
    payload = {
        "manifest_token": _manifest_token(status),
        "selection": {
            "base_model_number": _selection_payload(base_model_number),
            "mapped_brand": _selection_payload(mapped_brand),
            "mapped_category": _selection_payload(mapped_category),
            "mapped_subcategory": _selection_payload(mapped_subcategory),
            "product_id": _selection_payload(product_id),
            "limit_rows": int(limit_rows) if limit_rows is not None else None,
        },
    }
    return hashlib.sha1(json.dumps(payload, sort_keys=True, ensure_ascii=False).encode("utf-8")).hexdigest()



def _workspace_cache_path(
    status: Dict[str, Any],
    *,
    base_model_number: Any = None,
    mapped_brand: Any = None,
    mapped_category: Any = None,
    mapped_subcategory: Any = None,
    product_id: Any = None,
    limit_rows: Optional[int] = None,
) -> Path:
    dirs = ensure_local_review_db_dirs(status.get("root") or None)
    cache_dir = Path(dirs["workspace_cache_dir"]).expanduser().resolve()
    cache_key = _workspace_cache_key(
        status=status,
        base_model_number=base_model_number,
        mapped_brand=mapped_brand,
        mapped_category=mapped_category,
        mapped_subcategory=mapped_subcategory,
        product_id=product_id,
        limit_rows=limit_rows,
    )
    return cache_dir / f"{cache_key}.pkl"



def _prune_workspace_cache(cache_dir: Path, *, keep: int = _MAX_WORKSPACE_CACHE_FILES) -> None:
    try:
        files = [path for path in cache_dir.iterdir() if path.is_file() and path.suffix.lower() == ".pkl"]
    except Exception:
        return
    if len(files) <= keep:
        return
    for stale in sorted(files, key=lambda path: path.stat().st_mtime, reverse=True)[keep:]:
        try:
            stale.unlink(missing_ok=True)
        except Exception:
            continue



def _load_cached_workspace_dataset(cache_path: Path) -> Optional[Dict[str, Any]]:
    if not cache_path.exists():
        return None
    try:
        with open(cache_path, "rb") as handle:
            payload = pickle.load(handle)
    except Exception:
        return None
    if not isinstance(payload, dict) or not isinstance(payload.get("reviews_df"), pd.DataFrame):
        return None
    payload.setdefault("selection", {})
    payload.setdefault("load_meta", {})
    payload["selection"]["cache_hit"] = True
    payload["load_meta"]["cache_hit"] = True
    payload["load_meta"]["cache_path"] = str(cache_path)
    return payload



def _save_cached_workspace_dataset(cache_path: Path, dataset: Dict[str, Any]) -> None:
    try:
        cache_path.parent.mkdir(parents=True, exist_ok=True)
        with open(cache_path, "wb") as handle:
            pickle.dump(dataset, handle, protocol=pickle.HIGHEST_PROTOCOL)
        _prune_workspace_cache(cache_path.parent)
    except Exception:
        return



def _normalize_search_text(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", _safe_text(value).lower())



def _tokenize_search_text(value: Any) -> List[str]:
    return [token for token in re.split(r"[^a-z0-9]+", _safe_text(value).lower()) if token]



def _entity_candidate_fields(row: Dict[str, Any]) -> List[Tuple[str, str]]:
    return [
        ("base_model_number", _safe_text(row.get("base_model_number"))),
        ("product_id", _safe_text(row.get("product_id"))),
        ("original_product_name", _safe_text(row.get("original_product_name"))),
        ("mapped_brand", _safe_text(row.get("mapped_brand"))),
        ("mapped_category", _safe_text(row.get("mapped_category"))),
        ("mapped_subcategory", _safe_text(row.get("mapped_subcategory"))),
    ]



def _entity_search_score(row: Dict[str, Any], *, query_text: str, query_key: str, query_tokens: Sequence[str]) -> float:
    if not query_text:
        return 0.0
    score = 0.0
    review_count = float(row.get("review_count") or 0.0)
    code_like = bool(re.search(r"[a-z].*\d|\d.*[a-z]", query_text))
    for field_name, raw_value in _entity_candidate_fields(row):
        value = _safe_text(raw_value)
        if not value:
            continue
        lower = value.lower()
        key = _normalize_search_text(value)
        if lower == query_text:
            score = max(score, 420.0 if field_name in {"base_model_number", "product_id"} else 360.0)
        if key and key == query_key:
            score = max(score, 430.0 if field_name in {"base_model_number", "product_id"} else 370.0)
        if lower.startswith(query_text):
            score = max(score, 300.0 if field_name in {"base_model_number", "product_id"} else 245.0)
        if key and query_key and key.startswith(query_key):
            score = max(score, 310.0 if field_name in {"base_model_number", "product_id"} else 255.0)
        if query_text in lower:
            score = max(score, 255.0 if field_name in {"base_model_number", "product_id"} else 205.0)
        if key and query_key and query_key in key:
            score = max(score, 265.0 if field_name in {"base_model_number", "product_id"} else 215.0)
        if query_tokens:
            value_tokens = set(_tokenize_search_text(value))
            overlap = len(set(query_tokens) & value_tokens)
            if overlap:
                score += overlap * (26.0 if field_name in {"base_model_number", "product_id"} else 16.0)
        if key and query_key:
            ratio = SequenceMatcher(None, query_key, key).ratio()
            if ratio >= 0.72:
                score = max(score, 120.0 + ratio * (150.0 if field_name in {"base_model_number", "product_id"} else 120.0))
            elif ratio >= 0.58:
                score = max(score, 70.0 + ratio * (100.0 if field_name in {"original_product_name", "mapped_category", "mapped_subcategory"} else 80.0))
    if code_like and _safe_text(row.get("entity_type")) == "Base model":
        score += 18.0
    score += min(review_count / 500.0, 18.0)
    return score



def _rank_entity_search_results(rows: Sequence[Dict[str, Any]], *, query: str, limit: int) -> List[Dict[str, Any]]:
    query_text = _safe_text(query).lower()
    query_key = _normalize_search_text(query)
    query_tokens = _tokenize_search_text(query)
    ranked: List[Dict[str, Any]] = []
    seen: set[Tuple[str, str, str]] = set()
    for row in rows:
        item = dict(row)
        item["score"] = _entity_search_score(item, query_text=query_text, query_key=query_key, query_tokens=query_tokens)
        entity_key = (
            _safe_text(item.get("entity_type")),
            _safe_text(item.get("base_model_number")),
            _safe_text(item.get("product_id")),
        )
        if entity_key in seen:
            continue
        seen.add(entity_key)
        ranked.append(item)
    ranked = [row for row in ranked if float(row.get("score") or 0.0) > 0 or not query_text]
    ranked.sort(key=lambda row: (-float(row.get("score") or 0.0), -float(row.get("review_count") or 0.0), _safe_text(row.get("base_model_number")).lower(), _safe_text(row.get("product_id")).lower()))
    return ranked[: max(1, int(limit or 25))]


def _status_paths(items: Sequence[Dict[str, Any]]) -> List[Path]:
    paths: List[Path] = []
    for item in items or []:
        path_text = _safe_text((item or {}).get("path"))
        if not path_text:
            continue
        path = Path(path_text)
        if path.exists():
            paths.append(path)
    return _dedupe_paths(paths)


def _is_dyson_source_file(path: Path, *, relative_path: str = "") -> bool:
    rel = _safe_text(relative_path).replace("\\", "/").lower()
    name = path.name.lower()
    parts = {part.lower() for part in path.parts}
    if f"/{_DYSON_SUBDIR}/" in rel or rel.startswith(f"{_DYSON_SUBDIR}/"):
        return True
    if name.startswith(f"{_DYSON_SUBDIR}__") or _DYSON_SUBDIR in name:
        return True
    return _DYSON_SUBDIR in parts


def sync_local_review_database(
    root: Optional[str] = None,
    *,
    force: bool = False,
    chunk_size: int = _DEFAULT_CHUNK_SIZE,
    progress_callback: Optional[Callable[..., None]] = None,
    export_snapshot: bool = False,
) -> Dict[str, Any]:
    status = discover_latest_local_sources(root)
    review_infos = list(status.get("review_files") or [])
    mapping_infos = list(status.get("mapping_files") or [])
    review_paths = _status_paths(review_infos)
    mapping_paths = _status_paths(mapping_infos)
    dyson_review_paths = _status_paths(status.get("dyson_review_files") or [])
    dyson_mapping_paths = _status_paths(status.get("dyson_mapping_files") or [])
    if not review_paths or not mapping_paths:
        raise ValueError("Add at least one review export into incoming/reviews and at least one SKU mapping workbook into incoming/sku_mapping first.")
    if dyson_review_paths and not dyson_mapping_paths:
        raise ValueError("Dyson review files were detected. Add the Dyson product mapping workbook into incoming/sku_mapping/dyson before syncing.")

    db_path = status["db_path"]
    with _connect(db_path) as conn:
        manifest = _latest_manifest(conn)
        if not force and Path(db_path).exists() and manifest and _manifest_matches_sources(manifest, status):
            result = get_local_review_db_status(root)
            result["skipped"] = True
            result["message"] = "The local review database is already in sync with the latest files."
            return result

    standard_catalog_frames: List[pd.DataFrame] = []
    dyson_catalog_frames: List[pd.DataFrame] = []
    total_mapping_sources = max(len(mapping_infos), 1)
    for mapping_index, info in enumerate(mapping_infos, start=1):
        mapping_path = Path(info.get("path") or "")
        if not mapping_path.exists():
            continue
        is_dyson = _is_dyson_source_file(mapping_path, relative_path=_safe_text(info.get("relative_path")))
        progress = 0.03 + (0.09 * (mapping_index / total_mapping_sources))
        title = "Reading Dyson mapping" if is_dyson else "Reading master SKU mapping"
        detail = f"Loading {mapping_path.name} ({mapping_index}/{total_mapping_sources}) and preparing the product catalog."
        _maybe_report(progress_callback, progress=progress, title=title, detail=detail)
        frame = load_dyson_product_catalog(str(mapping_path)) if is_dyson else load_master_sku_catalog(str(mapping_path))
        if is_dyson:
            dyson_catalog_frames.append(frame)
        else:
            standard_catalog_frames.append(frame)

    catalog_df = _combine_catalog_frames([*standard_catalog_frames, *dyson_catalog_frames])
    dyson_catalog_lookup = _dyson_catalog_by_pid(_combine_catalog_frames(dyson_catalog_frames))

    with _connect(db_path) as conn:
        conn.execute("DROP TABLE IF EXISTS sku_catalog")
        conn.execute("DROP TABLE IF EXISTS reviews_enriched")
        conn.execute("DROP TABLE IF EXISTS base_model_directory")
        conn.execute("DROP TABLE IF EXISTS product_directory")
        catalog_df.to_sql("sku_catalog", conn, if_exists="replace", index=False)

        total_reviews = 0
        mapped_reviews = 0
        first_chunk = True
        total_review_sources = max(len(review_infos), 1)
        total_chunks_loaded = 0
        for source_index, info in enumerate(review_infos, start=1):
            review_path = Path(info.get("path") or "")
            if not review_path.exists():
                continue
            is_dyson = _is_dyson_source_file(review_path, relative_path=_safe_text(info.get("relative_path")))
            source_label = review_path.name
            source_window_start = 0.12 + 0.58 * ((source_index - 1) / total_review_sources)
            source_window_end = 0.12 + 0.58 * (source_index / total_review_sources)
            ingest_label = "Ingesting Dyson reviews" if is_dyson else "Ingesting reviews"
            _maybe_report(progress_callback, progress=source_window_start, title=ingest_label, detail=f"Streaming {source_label} ({source_index}/{total_review_sources}) into the local SQLite database.")
            dyson_catalog_row = dyson_catalog_lookup.get(_dyson_pid_from_filename(review_path), {}) if is_dyson else {}
            for source_chunk_index, raw_chunk in enumerate(_iter_tabular_chunks(str(review_path), chunk_size=chunk_size), start=1):
                if is_dyson:
                    normalized = normalize_dyson_review_chunk(raw_chunk, source_path=review_path, catalog_row=dyson_catalog_row)
                else:
                    normalized = normalize_uploaded_df(raw_chunk, source_name=source_label, include_local_symptomization=True)
                enriched = enrich_reviews_with_catalog(normalized, catalog_df)
                total_reviews += len(enriched)
                mapped_reviews += int(enriched["catalog_match_type"].notna().sum()) if "catalog_match_type" in enriched.columns else 0
                if not first_chunk:
                    _ensure_table_has_columns(conn, "reviews_enriched", list(enriched.columns))
                enriched.to_sql("reviews_enriched", conn, if_exists="replace" if first_chunk else "append", index=False, chunksize=20000)
                first_chunk = False
                total_chunks_loaded += 1
                chunk_fraction = source_chunk_index / max(source_chunk_index + 1, 1)
                progress = source_window_start + ((source_window_end - source_window_start) * min(chunk_fraction, 0.92))
                _maybe_report(
                    progress_callback,
                    progress=progress,
                    title=ingest_label,
                    detail=f"Loaded {total_reviews:,} review rows so far across {total_chunks_loaded:,} chunk(s). Current source: {source_label}.",
                )

        if first_chunk:
            empty = finalize_df(pd.DataFrame())
            empty.to_sql("reviews_enriched", conn, if_exists="replace", index=False)

        _maybe_report(progress_callback, progress=0.78, title="Building directories", detail="Creating base-model and product directories for fast filtering.")
        _rebuild_directory_tables(conn)
        _create_indexes(conn)
        _maybe_report(progress_callback, progress=0.9, title="Optimizing SQLite", detail="Refreshing indexes and query statistics so future loads, Action Center views, and search stay fast.")
        _optimize_database(conn)

        counts = conn.execute(
            "SELECT COUNT(*) AS review_count, COUNT(DISTINCT product_id) AS distinct_product_count, COUNT(DISTINCT COALESCE(NULLIF(base_model_number, ''), NULLIF(base_sku, ''), NULLIF(product_id, ''), 'UNMAPPED')) AS base_model_count, COUNT(CASE WHEN catalog_match_type IS NOT NULL AND TRIM(CAST(catalog_match_type AS TEXT)) <> '' THEN 1 END) AS mapped_review_count FROM reviews_enriched"
        ).fetchone()
        review_count = int(counts[0] or 0)
        distinct_product_count = int(counts[1] or 0)
        base_model_count = int(counts[2] or 0)
        mapped_review_count = int(counts[3] or 0)
        unmatched_review_count = max(review_count - mapped_review_count, 0)

        conn.execute(
            """
            INSERT INTO import_manifest (
                synced_at,
                review_file_name,
                review_file_path,
                review_file_size,
                review_file_mtime,
                mapping_file_name,
                mapping_file_path,
                mapping_file_size,
                mapping_file_mtime,
                review_count,
                mapped_review_count,
                unmatched_review_count,
                base_model_count,
                distinct_product_count,
                notes_json
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                _utcnow_iso(),
                status["review_file"]["name"],
                status["review_file"]["path"],
                int(status["review_file"]["size"] or 0),
                float(status["review_file"]["mtime"] or 0.0),
                status["mapping_file"]["name"],
                status["mapping_file"]["path"],
                int(status["mapping_file"]["size"] or 0),
                float(status["mapping_file"]["mtime"] or 0.0),
                review_count,
                mapped_review_count,
                unmatched_review_count,
                base_model_count,
                distinct_product_count,
                json.dumps({
                    "catalog_rows": int(len(catalog_df)),
                    "chunk_size": int(chunk_size),
                    "source_signature": _safe_text(status.get("source_signature")),
                    "review_source_count": int(status.get("review_source_count") or 0),
                    "mapping_source_count": int(status.get("mapping_source_count") or 0),
                    "dyson_review_count": int(status.get("dyson_review_count") or 0),
                    "dyson_mapping_count": int(status.get("dyson_mapping_count") or 0),
                    "review_files": [
                        {
                            "name": _safe_text(item.get("name")),
                            "relative_path": _safe_text(item.get("relative_path")),
                            "size": int(item.get("size") or 0),
                            "mtime": float(item.get("mtime") or 0.0),
                        }
                        for item in review_infos
                    ],
                    "mapping_files": [
                        {
                            "name": _safe_text(item.get("name")),
                            "relative_path": _safe_text(item.get("relative_path")),
                            "size": int(item.get("size") or 0),
                            "mtime": float(item.get("mtime") or 0.0),
                        }
                        for item in mapping_infos
                    ],
                }, ensure_ascii=False, sort_keys=True),
            ),
        )

    if export_snapshot:
        try:
            export_local_review_database_snapshot(root)
        except Exception:
            pass

    source_suffix = ""
    if int(status.get("dyson_review_count") or 0) > 0:
        source_suffix = f" including {int(status.get('dyson_review_count') or 0):,} Dyson review file(s)"
    _maybe_report(progress_callback, progress=1.0, title="Local database ready", detail=f"Synced {review_count:,} reviews across {base_model_count:,} base models{source_suffix}.")
    result = get_local_review_db_status(root)
    result["message"] = f"Synced {review_count:,} reviews into the local review database{source_suffix}."
    result["skipped"] = False
    return result


def _selection_sql(
    *,
    base_model_number: Any = None,
    mapped_brand: Any = None,
    mapped_category: Any = None,
    mapped_subcategory: Any = None,
    product_id: Any = None,
) -> Tuple[str, Dict[str, Any]]:
    where_parts: List[str] = []
    params: Dict[str, Any] = {}
    _append_value_filter(where_parts, params, "base_model_number", base_model_number, param_prefix="base_model_number")
    _append_value_filter(where_parts, params, "mapped_brand", mapped_brand, param_prefix="mapped_brand")
    _append_value_filter(where_parts, params, "mapped_category", mapped_category, param_prefix="mapped_category")
    _append_value_filter(where_parts, params, "mapped_subcategory", mapped_subcategory, param_prefix="mapped_subcategory")
    _append_value_filter(where_parts, params, "product_id", product_id, param_prefix="product_id")
    where_sql = " AND ".join(where_parts)
    return where_sql, params


def _directory_option_values(
    conn: sqlite3.Connection,
    *,
    column: str,
    mapped_brand: Any = None,
    mapped_category: Any = None,
    mapped_subcategory: Any = None,
    base_model_number: Any = None,
    product_id: Any = None,
    exclude_field: Optional[str] = None,
    limit: Optional[int] = None,
) -> List[str]:
    where_parts = [f"{column} IS NOT NULL", f"TRIM({column}) <> ''"]
    params: Dict[str, Any] = {}
    filter_items = [
        ("mapped_brand", mapped_brand),
        ("mapped_category", mapped_category),
        ("mapped_subcategory", mapped_subcategory),
        ("base_model_number", base_model_number),
        ("product_id", product_id),
    ]
    for field_name, values in filter_items:
        if field_name == exclude_field:
            continue
        _append_value_filter(where_parts, params, field_name, values, param_prefix=field_name)
    sql = f"SELECT {column} AS option_value, SUM(review_count) AS total_reviews FROM product_directory"
    if where_parts:
        sql += " WHERE " + " AND ".join(where_parts)
    sql += f" GROUP BY {column} ORDER BY total_reviews DESC, {column}"
    if limit is not None and int(limit) > 0:
        sql += f" LIMIT {int(limit)}"
    return [str(row[0]) for row in conn.execute(sql, params).fetchall()]


def get_local_review_db_filter_options(
    root: Optional[str] = None,
    *,
    mapped_brand: Any = None,
    mapped_category: Any = None,
    mapped_subcategory: Any = None,
    base_model_number: Any = None,
    product_id: Any = None,
) -> Dict[str, List[str]]:
    status = get_local_review_db_status(root)
    if not status.get("is_ready"):
        return {"mapped_brand": [], "mapped_category": [], "mapped_subcategory": [], "base_model_number": [], "product_id": []}
    db_path = status["db_path"]
    with _connect(db_path) as conn:
        try:
            brands = _directory_option_values(
                conn,
                column="mapped_brand",
                mapped_brand=mapped_brand,
                mapped_category=mapped_category,
                mapped_subcategory=mapped_subcategory,
                base_model_number=base_model_number,
                product_id=product_id,
                exclude_field="mapped_brand",
                limit=5000,
            )
            categories = _directory_option_values(
                conn,
                column="mapped_category",
                mapped_brand=mapped_brand,
                mapped_category=mapped_category,
                mapped_subcategory=mapped_subcategory,
                base_model_number=base_model_number,
                product_id=product_id,
                exclude_field="mapped_category",
                limit=5000,
            )
            subcategories = _directory_option_values(
                conn,
                column="mapped_subcategory",
                mapped_brand=mapped_brand,
                mapped_category=mapped_category,
                mapped_subcategory=mapped_subcategory,
                base_model_number=base_model_number,
                product_id=product_id,
                exclude_field="mapped_subcategory",
                limit=5000,
            )
            base_models = _directory_option_values(
                conn,
                column="base_model_number",
                mapped_brand=mapped_brand,
                mapped_category=mapped_category,
                mapped_subcategory=mapped_subcategory,
                base_model_number=base_model_number,
                product_id=product_id,
                exclude_field="base_model_number",
                limit=5000,
            )
            products = _directory_option_values(
                conn,
                column="product_id",
                mapped_brand=mapped_brand,
                mapped_category=mapped_category,
                mapped_subcategory=mapped_subcategory,
                base_model_number=base_model_number,
                product_id=product_id,
                exclude_field="product_id",
                limit=5000,
            )
        except sqlite3.OperationalError:
            return {"mapped_brand": [], "mapped_category": [], "mapped_subcategory": [], "base_model_number": [], "product_id": []}
    return {
        "mapped_brand": brands,
        "mapped_category": categories,
        "mapped_subcategory": subcategories,
        "base_model_number": base_models,
        "product_id": products,
    }


def count_local_review_db_selection(
    root: Optional[str] = None,
    *,
    base_model_number: Any = None,
    mapped_brand: Any = None,
    mapped_category: Any = None,
    mapped_subcategory: Any = None,
    product_id: Any = None,
) -> int:
    status = get_local_review_db_status(root)
    if not status.get("is_ready"):
        return 0
    where_sql, params = _selection_sql(
        base_model_number=base_model_number,
        mapped_brand=mapped_brand,
        mapped_category=mapped_category,
        mapped_subcategory=mapped_subcategory,
        product_id=product_id,
    )
    sql = "SELECT COALESCE(SUM(review_count), 0) FROM product_directory" + (f" WHERE {where_sql}" if where_sql else "")
    with _connect(status["db_path"]) as conn:
        try:
            value = conn.execute(sql, params).fetchone()[0]
        except sqlite3.OperationalError:
            fallback_sql = "SELECT COUNT(*) FROM reviews_enriched" + (f" WHERE {where_sql}" if where_sql else "")
            value = conn.execute(fallback_sql, params).fetchone()[0]
    return int(value or 0)


def load_local_review_workspace(
    root: Optional[str] = None,
    *,
    base_model_number: Any = None,
    mapped_brand: Any = None,
    mapped_category: Any = None,
    mapped_subcategory: Any = None,
    product_id: Any = None,
    limit_rows: Optional[int] = None,
) -> Dict[str, Any]:
    status = get_local_review_db_status(root)
    if not status.get("is_ready"):
        raise ValueError("Sync the local review database first.")

    load_started = time.perf_counter()
    cache_path = _workspace_cache_path(
        status,
        base_model_number=base_model_number,
        mapped_brand=mapped_brand,
        mapped_category=mapped_category,
        mapped_subcategory=mapped_subcategory,
        product_id=product_id,
        limit_rows=limit_rows,
    )
    cached_dataset = _load_cached_workspace_dataset(cache_path)
    if cached_dataset is not None:
        selection = dict(cached_dataset.get("selection") or {})
        selection["cache_hit"] = True
        selection["load_seconds"] = round(time.perf_counter() - load_started, 4)
        load_meta = dict(cached_dataset.get("load_meta") or {})
        load_meta.update(
            {
                "cache_hit": True,
                "cache_path": str(cache_path),
                "load_strategy": "selection_cache",
                "load_seconds": round(time.perf_counter() - load_started, 4),
            }
        )
        cached_dataset["selection"] = selection
        cached_dataset["load_meta"] = load_meta
        return cached_dataset

    where_sql, params = _selection_sql(
        base_model_number=base_model_number,
        mapped_brand=mapped_brand,
        mapped_category=mapped_category,
        mapped_subcategory=mapped_subcategory,
        product_id=product_id,
    )
    count_sql = "SELECT COUNT(*) FROM reviews_enriched" + (f" WHERE {where_sql}" if where_sql else "")
    with _connect(status["db_path"]) as conn:
        try:
            total_count = int((conn.execute("SELECT COALESCE(SUM(review_count), 0) FROM product_directory" + (f" WHERE {where_sql}" if where_sql else ""), params).fetchone() or [0])[0] or 0)
        except sqlite3.OperationalError:
            total_count = 0
        if total_count <= 0:
            total_count = int(conn.execute(count_sql, params).fetchone()[0] or 0)
        if total_count <= 0:
            raise ValueError("No reviews match the current local database selection.")
        projected_columns = _workspace_projection_columns(conn)
        select_sql = "SELECT " + ", ".join(_sanitize_sql_label(col) for col in projected_columns) + " FROM reviews_enriched" + (f" WHERE {where_sql}" if where_sql else "") + " ORDER BY submission_time DESC, review_id DESC"
        if limit_rows is not None and int(limit_rows) > 0:
            select_sql += f" LIMIT {int(limit_rows)}"
        df = pd.read_sql_query(select_sql, conn, params=params)
    df = finalize_df(df)
    label = _selection_label(
        base_model_number=base_model_number,
        mapped_brand=mapped_brand,
        mapped_category=mapped_category,
        mapped_subcategory=mapped_subcategory,
        product_id=product_id,
    )
    summary = ReviewBatchSummary(
        product_url="",
        product_id=_primary_selection_value(base_model_number) or _primary_selection_value(product_id) or _primary_selection_value(mapped_category) or "LOCAL_REVIEW_DATABASE",
        total_reviews=total_count,
        page_size=max(len(df), 1),
        requests_needed=0,
        reviews_downloaded=len(df),
    )
    source_signature = json.dumps(
        {
            "db_path": status["db_path"],
            "selection": {
                "base_model_number": _selection_payload(base_model_number),
                "mapped_brand": _selection_payload(mapped_brand),
                "mapped_category": _selection_payload(mapped_category),
                "mapped_subcategory": _selection_payload(mapped_subcategory),
                "product_id": _selection_payload(product_id),
                "limit_rows": int(limit_rows) if limit_rows is not None else None,
            },
            "manifest_id": (status.get("manifest") or {}).get("manifest_id"),
        },
        sort_keys=True,
        ensure_ascii=False,
    )
    load_seconds = round(time.perf_counter() - load_started, 4)
    dataset = {
        "summary": summary,
        "reviews_df": df,
        "source_type": "local_database",
        "source_label": f"Local DB · {label}",
        "source_root": status["root"],
        "source_signature": source_signature,
        "selection": {
            "base_model_number": _selection_payload(base_model_number),
            "mapped_brand": _selection_payload(mapped_brand),
            "mapped_category": _selection_payload(mapped_category),
            "mapped_subcategory": _selection_payload(mapped_subcategory),
            "product_id": _selection_payload(product_id),
            "limit_rows": int(limit_rows) if limit_rows is not None else None,
            "total_count": total_count,
            "cache_hit": False,
            "load_seconds": load_seconds,
        },
        "load_meta": {
            "cache_hit": False,
            "cache_path": str(cache_path),
            "load_strategy": "sqlite_read",
            "load_seconds": load_seconds,
        },
    }
    _save_cached_workspace_dataset(cache_path, dataset)
    return dataset


def _table_columns(conn: sqlite3.Connection, table_name: str) -> List[str]:
    return [str(row[1]) for row in conn.execute(f"PRAGMA table_info({_sanitize_sql_label(table_name)})").fetchall()]


def load_local_review_analytics_frame(
    root: Optional[str] = None,
    *,
    columns: Optional[Sequence[str]] = None,
    base_model_number: Any = None,
    mapped_brand: Any = None,
    mapped_category: Any = None,
    mapped_subcategory: Any = None,
    product_id: Any = None,
    moderation_buckets: Optional[Sequence[str]] = None,
    organic_only: bool = False,
    brand_values: Optional[Sequence[str]] = None,
    country_values: Optional[Sequence[str]] = None,
    limit_rows: Optional[int] = None,
) -> pd.DataFrame:
    status = get_local_review_db_status(root)
    if not status.get("is_ready"):
        raise ValueError("Sync the local review database first.")

    with _connect(status["db_path"]) as conn:
        available = set(_table_columns(conn, "reviews_enriched"))
        selected = [str(col) for col in (columns or []) if str(col) in available]
        if not selected:
            selected = [col for col in [
                "review_id", "submission_time", "submission_date", "rating", "incentivized_review",
                "review_origin_group", "review_acquisition_channel", "moderation_bucket", "country",
                "content_locale", "mapped_brand", "mapped_category", "mapped_subcategory",
                "base_model_number", "product_id", "original_product_name",
            ] if col in available]
        if not selected:
            return pd.DataFrame()

        where_sql, params = _selection_sql(
            base_model_number=base_model_number,
            mapped_brand=mapped_brand,
            mapped_category=mapped_category,
            mapped_subcategory=mapped_subcategory,
            product_id=product_id,
        )
        where_parts = [where_sql] if where_sql else []
        bind: Dict[str, Any] = dict(params)

        mod_values = [_safe_text(value) for value in (moderation_buckets or []) if _safe_text(value)]
        if mod_values and "moderation_bucket" in available:
            placeholders = []
            for idx, value in enumerate(mod_values):
                key = f"mod_{idx}"
                bind[key] = value
                placeholders.append(f":{key}")
            where_parts.append(f"moderation_bucket IN ({', '.join(placeholders)})")

        brand_filters = [_safe_text(value) for value in (brand_values or []) if _safe_text(value)]
        if brand_filters and "mapped_brand" in available:
            placeholders = []
            for idx, value in enumerate(brand_filters):
                key = f"brand_{idx}"
                bind[key] = value
                placeholders.append(f":{key}")
            where_parts.append(f"mapped_brand IN ({', '.join(placeholders)})")

        country_filters = [_safe_text(value) for value in (country_values or []) if _safe_text(value)]
        if country_filters and "country" in available:
            placeholders = []
            for idx, value in enumerate(country_filters):
                key = f"country_{idx}"
                bind[key] = value
                placeholders.append(f":{key}")
            where_parts.append(f"country IN ({', '.join(placeholders)})")

        if organic_only and "incentivized_review" in available:
            where_parts.append("COALESCE(CAST(incentivized_review AS INTEGER), 0) = 0")

        sql = "SELECT " + ", ".join(_sanitize_sql_label(col) for col in selected) + " FROM reviews_enriched"
        if where_parts:
            sql += " WHERE " + " AND ".join(part for part in where_parts if part)
        if "submission_time" in available:
            sql += " ORDER BY submission_time DESC, review_id DESC"
        elif "submission_date" in available:
            sql += " ORDER BY submission_date DESC, review_id DESC"
        if limit_rows is not None and int(limit_rows) > 0:
            sql += f" LIMIT {int(limit_rows)}"
        df = pd.read_sql_query(sql, conn, params=bind)

    if "submission_time" in df.columns:
        df["submission_time"] = pd.to_datetime(df["submission_time"], errors="coerce")
    if "submission_date" in df.columns:
        df["submission_date"] = pd.to_datetime(df["submission_date"], errors="coerce").dt.date
    if "rating" in df.columns:
        df["rating"] = pd.to_numeric(df["rating"], errors="coerce")
    if "incentivized_review" in df.columns:
        df["incentivized_review"] = pd.Series(df["incentivized_review"]).astype("boolean").fillna(False).astype(bool)
    return df


def search_local_review_entities(root: Optional[str] = None, *, query: str = "", limit: int = 25) -> List[Dict[str, Any]]:
    status = get_local_review_db_status(root)
    if not status.get("is_ready"):
        return []
    needle = _safe_text(query)
    take = max(1, int(limit or 25))

    with _connect(status["db_path"]) as conn:
        if not needle:
            rows = conn.execute(
                """
                SELECT
                    'Base model' AS entity_type,
                    base_model_number,
                    '' AS product_id,
                    '' AS original_product_name,
                    mapped_brand,
                    mapped_category,
                    mapped_subcategory,
                    review_count,
                    avg_rating,
                    0 AS score
                FROM base_model_directory
                ORDER BY review_count DESC, base_model_number
                LIMIT ?
                """,
                (take,),
            ).fetchall()
            keys = [
                "entity_type", "base_model_number", "product_id", "original_product_name", "mapped_brand",
                "mapped_category", "mapped_subcategory", "review_count", "avg_rating", "score",
            ]
            return [dict(zip(keys, row)) for row in rows]

        base_rows = conn.execute(
            """
            SELECT
                'Base model' AS entity_type,
                base_model_number,
                '' AS product_id,
                '' AS original_product_name,
                mapped_brand,
                mapped_category,
                mapped_subcategory,
                review_count,
                avg_rating
            FROM base_model_directory
            ORDER BY review_count DESC, base_model_number
            LIMIT ?
            """,
            (_SEARCH_RESULT_SCAN_LIMIT,),
        ).fetchall()
        product_rows = conn.execute(
            """
            SELECT
                'Product' AS entity_type,
                base_model_number,
                product_id,
                original_product_name,
                mapped_brand,
                mapped_category,
                mapped_subcategory,
                review_count,
                avg_rating
            FROM product_directory
            ORDER BY review_count DESC, base_model_number, product_id
            LIMIT ?
            """,
            (_SEARCH_RESULT_SCAN_LIMIT,),
        ).fetchall()

    keys = [
        "entity_type", "base_model_number", "product_id", "original_product_name", "mapped_brand",
        "mapped_category", "mapped_subcategory", "review_count", "avg_rating",
    ]
    rows: List[Dict[str, Any]] = [dict(zip(keys, row)) for row in list(base_rows) + list(product_rows)]
    return _rank_entity_search_results(rows, query=needle, limit=take)


def export_local_review_database_snapshot(root: Optional[str] = None, *, destination: Optional[str] = None) -> str:
    status = get_local_review_db_status(root)
    if not status.get("is_ready"):
        raise ValueError("Sync the local review database first.")
    out_path = Path(destination or status["snapshot_path"]).expanduser().resolve()
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with _connect(status["db_path"]) as conn, pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        manifest = pd.DataFrame([status.get("manifest") or {}])
        if not manifest.empty:
            manifest.to_excel(writer, sheet_name="Manifest", index=False)
        pd.read_sql_query("SELECT * FROM base_model_directory ORDER BY review_count DESC, base_model_number", conn).to_excel(writer, sheet_name="Base Models", index=False)
        pd.read_sql_query("SELECT * FROM product_directory ORDER BY review_count DESC, base_model_number, product_id", conn).to_excel(writer, sheet_name="Products", index=False)
        pd.read_sql_query("SELECT * FROM sku_catalog ORDER BY base_model_number, sku", conn).to_excel(writer, sheet_name="SKU Catalog", index=False)
    return str(out_path)
