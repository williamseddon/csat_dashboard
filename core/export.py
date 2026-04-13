"""Export builders: Excel workbook generation and formatting

Extracted from app.py. Uses _app() for cross-module access.
"""
from __future__ import annotations
import io, json, sys
from typing import Any, Dict, List, Optional
import pandas as pd
import streamlit as st
try:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
except ImportError:
    pass

NON_VALUES = {"", "NA", "N/A", "NONE", "NULL", "NAN", "<NA>", "NOT MENTIONED"}
DET_INDEXES = [column_index_from_string(c) for c in DET_LETTERS]
DEL_INDEXES = [column_index_from_string(c) for c in DEL_LETTERS]
META_INDEXES = {name: column_index_from_string(col) for name,col in META_ORDER}
UPLOAD_REVIEW_ID_ALIASES = ["Event Id", "Event ID", "Review ID", "Review Id", "Id", "review_id"]
AI_DEL_HEADERS = [f"AI Symptom Delighter {i}" for i in range(1,11)]
AI_DET_HEADERS = [f"AI Symptom Detractor {i}" for i in range(1,11)]

def _app():
    """Lazy access to app module namespace."""
    return sys.modules.get('__main__', sys.modules.get('app'))

def _esc(text):
    return html.escape(str(text or "")) if 'html' in dir() else str(text or "")

def _safe_text(value, default=""):
    if value is None: return default
    s = str(value).strip()
    return s if s else default

def _normalize_tag_list(tags):
    if not tags: return []
    seen = set()
    out = []
    for t in tags:
        s = str(t).strip()
        if not s or s.upper() in NON_VALUES: continue
        key = s.lower()
        if key not in seen: seen.add(key); out.append(s)
    return out


def _autosize_cell_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, (pd.Series, pd.DataFrame, pd.Index, list, tuple, set, dict)):
        try:
            return json.dumps(value, ensure_ascii=False, default=str)
        except Exception:
            return str(value)
    try:
        missing = pd.isna(value)
    except Exception:
        missing = False
    if isinstance(missing, bool) and missing:
        return ""
    return str(value)



def _autosize_ws(ws, df):
    ws.freeze_panes = "A2"
    if df is None or not isinstance(df, pd.DataFrame) or df.empty and len(df.columns) == 0:
        return
    for idx, col in enumerate(list(df.columns), 1):
        try:
            series = df.iloc[:, idx - 1]
        except Exception:
            series = pd.Series(dtype="object")
        if isinstance(series, pd.DataFrame):
            values = []
            for nested_idx in range(series.shape[1]):
                values.extend(series.iloc[:, nested_idx].head(250).tolist())
        else:
            values = series.head(250).tolist()
        max_len = max([len(str(col))] + [len(_autosize_cell_text(v)) for v in values] + [0])
        ws.column_dimensions[get_column_letter(idx)].width = min(max_len + 2, 48)



def _filter_criteria_df(active_items=None, *, filter_description="", export_scope_label="", total_loaded_reviews=None, export_review_count=None):
    rows = [
        {"field": "Export scope", "value": export_scope_label or "Current view"},
        {"field": "Export review count", "value": export_review_count if export_review_count is not None else ""},
        {"field": "Total loaded reviews", "value": total_loaded_reviews if total_loaded_reviews is not None else ""},
        {"field": "Filter description", "value": filter_description or "No active filters"},
    ]
    if active_items:
        for label, value in active_items:
            rows.append({"field": label, "value": value})
    else:
        rows.append({"field": "Filters", "value": "No active filters"})
    return pd.DataFrame(rows)



def _build_master_excel(summary, reviews_df, *, prompt_defs=None, prompt_summary_df=None, prompt_scope="", active_items=None, filter_description="", export_scope_label="Current view", total_loaded_reviews=None):
    metrics = _get_metrics(reviews_df)
    try:
        rd = _app()._rating_dist(reviews_df)
        md = _monthly_trend(reviews_df)
    except Exception:
        rd = pd.DataFrame()
        md = pd.DataFrame()
    summary_df = pd.DataFrame([dict(
        product_name=_product_name(summary, reviews_df),
        product_id=summary.product_id,
        product_url=summary.product_url,
        reviews_downloaded=summary.reviews_downloaded,
        export_review_count=len(reviews_df),
        total_loaded_reviews=total_loaded_reviews if total_loaded_reviews is not None else len(reviews_df),
        export_scope=export_scope_label,
        filter_description=filter_description or "No active filters",
        avg_rating=metrics.get("avg_rating"),
        avg_rating_non_incentivized=metrics.get("avg_rating_non_incentivized"),
        pct_low_star=metrics.get("pct_low_star"),
        pct_incentivized=metrics.get("pct_incentivized"),
        generated_utc=pd.Timestamp.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC"),
    )])
    priority_cols = ["review_id", "product_id", "rating", "incentivized_review", "is_recommended", "submission_time", "content_locale", "title", "review_text"]
    pc = [p["column_name"] for p in (prompt_defs or []) if p["column_name"] in reviews_df.columns]
    ordered = [c for c in priority_cols + pc if c in reviews_df.columns]
    remaining = [c for c in reviews_df.columns if c not in ordered]
    exp_reviews = reviews_df[ordered + remaining]
    filter_df = _filter_criteria_df(active_items, filter_description=filter_description, export_scope_label=export_scope_label, total_loaded_reviews=total_loaded_reviews, export_review_count=len(reviews_df))
    out = io.BytesIO()
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        sheets = {"Summary": summary_df, "FilterCriteria": filter_df, "Reviews": exp_reviews, "RatingDistribution": rd, "ReviewVolume": md}
        if prompt_defs:
            sheets["ReviewPromptDefinitions"] = pd.DataFrame([dict(column_name=p["column_name"], display_name=p["display_name"], prompt=p["prompt"], labels=", ".join(p["labels"]), scope=prompt_scope) for p in prompt_defs])
        if prompt_summary_df is not None and not prompt_summary_df.empty:
            sheets["ReviewPromptSummary"] = prompt_summary_df
        for sname, df_ in sheets.items():
            if df_ is None or df_.empty:
                continue
            df_.to_excel(writer, sheet_name=sname, index=False)
            _autosize_ws(writer.sheets[sname], df_)
    out.seek(0)
    return out.getvalue()



def _get_master_bundle(summary, reviews_df, prompt_artifacts, *, active_items=None, filter_description="", export_scope_label="Current view", total_loaded_reviews=None):
    pd_ = (prompt_artifacts or {}).get("definitions") or []
    psd = (prompt_artifacts or {}).get("summary_df")
    ps = (prompt_artifacts or {}).get("scope_label", "")
    key = json.dumps(dict(
        pid=summary.product_id,
        n=len(reviews_df),
        cols=sorted(str(c) for c in reviews_df.columns),
        psig=(prompt_artifacts or {}).get("definition_signature"),
        filters=list(active_items or []),
        filter_description=filter_description or "",
        export_scope_label=export_scope_label or "",
        total_loaded_reviews=total_loaded_reviews,
    ), sort_keys=True)
    b = st.session_state.get("master_export_bundle")
    if b and b.get("key") == key:
        return b
    xlsx = _build_master_excel(
        summary,
        reviews_df,
        prompt_defs=pd_,
        prompt_summary_df=psd,
        prompt_scope=ps,
        active_items=active_items,
        filter_description=filter_description,
        export_scope_label=export_scope_label,
        total_loaded_reviews=total_loaded_reviews,
    )
    ts = pd.Timestamp.utcnow().strftime("%Y%m%d_%H%M%S")
    scope_slug = "filtered_view" if active_items else "review_workspace"
    b = dict(key=key, excel_bytes=xlsx, excel_name=f"{summary.product_id}_{scope_slug}_{ts}.xlsx")
    st.session_state["master_export_bundle"] = b
    return b



def _safe_get_master_bundle(summary, reviews_df, prompt_artifacts, **kwargs):
    try:
        return _get_master_bundle(summary, reviews_df, prompt_artifacts, **kwargs), None
    except Exception as exc:
        return None, exc

# ═══════════════════════════════════════════════════════════════════════════════
#  SYMPTOMIZER HELPERS
# ═══════════════════════════════════════════════════════════════════════════════



def _gen_symptomized_workbook(original_bytes, updated_df):
    wb = load_workbook(io.BytesIO(original_bytes))
    sheet_name = _app()._best_uploaded_excel_sheet_name(original_bytes)
    if not sheet_name or sheet_name not in wb.sheetnames:
        sheet_name = "Star Walk scrubbed verbatims" if "Star Walk scrubbed verbatims" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]
    df2 = _ensure_ai_cols(updated_df.copy())

    def _canon_review_key(value):
        if value is None or (isinstance(value, float) and pd.isna(value)):
            return ""
        if isinstance(value, (int, np.integer)):
            return str(int(value))
        if isinstance(value, (float, np.floating)) and not pd.isna(value) and float(value).is_integer():
            return str(int(value))
        text = str(value).strip()
        if re.fullmatch(r"-?\d+\.0+", text):
            text = text.split(".", 1)[0]
        return text

    def _header_lookup(sheet):
        out = {}
        for col_idx in range(1, sheet.max_column + 1):
            header = _safe_text(sheet.cell(row=1, column=col_idx).value).strip()
            if header and header.lower() not in out:
                out[header.lower()] = col_idx
        return out

    header_map = _header_lookup(ws)

    def _ensure_ai_column(header_name, *, fallback_index=None):
        lower = header_name.lower()
        if lower in header_map:
            return header_map[lower]
        if fallback_index is not None:
            existing = _safe_text(ws.cell(row=1, column=fallback_index).value).strip()
            if not existing or existing.lower() == lower:
                ws.cell(row=1, column=fallback_index, value=header_name)
                header_map[lower] = fallback_index
                return fallback_index
        new_idx = max(ws.max_column, 0) + 1
        ws.cell(row=1, column=new_idx, value=header_name)
        header_map[lower] = new_idx
        return new_idx

    det_write_cols = [
        _ensure_ai_column(header, fallback_index=DET_INDEXES[j - 1])
        for j, header in enumerate(AI_DET_HEADERS, start=1)
    ]
    del_write_cols = [
        _ensure_ai_column(header, fallback_index=DEL_INDEXES[j - 1])
        for j, header in enumerate(AI_DEL_HEADERS, start=1)
    ]
    meta_write_cols = {
        "AI Safety": _ensure_ai_column("AI Safety", fallback_index=META_INDEXES["Safety"]),
        "AI Reliability": _ensure_ai_column("AI Reliability", fallback_index=META_INDEXES["Reliability"]),
        "AI # of Sessions": _ensure_ai_column("AI # of Sessions", fallback_index=META_INDEXES["# of Sessions"]),
    }

    review_id_col = next((header_map[a.lower()] for a in UPLOAD_REVIEW_ID_ALIASES if a.lower() in header_map), None)
    worksheet_rows_by_review_id = {}
    if review_id_col is not None:
        for row_idx in range(2, ws.max_row + 1):
            key = _canon_review_key(ws.cell(row=row_idx, column=review_id_col).value)
            if key and key not in worksheet_rows_by_review_id:
                worksheet_rows_by_review_id[key] = row_idx

    fg = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fr = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    fy = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    fb = PatternFill(start_color="CFE2F3", end_color="CFE2F3", fill_type="solid")
    fp = PatternFill(start_color="EAD1DC", end_color="EAD1DC", fill_type="solid")

    def _write_cell(row_idx, col_idx, value, fill):
        clean = None if (pd.isna(value) or str(value).strip() == "") else str(value).strip()
        cell = ws.cell(row=row_idx, column=col_idx, value=clean)
        cell.fill = fill if clean else PatternFill()

    for seq_row_idx, (_, row) in enumerate(df2.iterrows(), start=2):
        target_row = None
        if review_id_col is not None:
            review_key = _canon_review_key(row.get("review_id"))
            if review_key and review_key in worksheet_rows_by_review_id:
                target_row = worksheet_rows_by_review_id[review_key]
            elif review_key:
                target_row = ws.max_row + 1
                worksheet_rows_by_review_id[review_key] = target_row
                ws.cell(row=target_row, column=review_id_col, value=review_key)
        if target_row is None:
            target_row = seq_row_idx
        for j, ci in enumerate(det_write_cols, 1):
            _write_cell(target_row, ci, row.get(f"AI Symptom Detractor {j}"), fr)
        for j, ci in enumerate(del_write_cols, 1):
            _write_cell(target_row, ci, row.get(f"AI Symptom Delighter {j}"), fg)
        _write_cell(target_row, meta_write_cols["AI Safety"], row.get("AI Safety"), fy)
        _write_cell(target_row, meta_write_cols["AI Reliability"], row.get("AI Reliability"), fb)
        _write_cell(target_row, meta_write_cols["AI # of Sessions"], row.get("AI # of Sessions"), fp)
    for c in det_write_cols + del_write_cols + list(meta_write_cols.values()):
        try:
            ws.column_dimensions[get_column_letter(c)].width = 28
        except Exception:
            pass
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


