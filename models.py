import io
from typing import Dict, Any

import pandas as pd


def coerce_to_dataframe(obj: Any) -> pd.DataFrame:
    if obj is None:
        return pd.DataFrame()

    if isinstance(obj, pd.DataFrame):
        return obj.copy()

    if isinstance(obj, pd.Series):
        return obj.to_frame().reset_index(drop=True)

    if isinstance(obj, dict):
        try:
            return pd.DataFrame([obj])
        except Exception:
            return pd.DataFrame({"value": [str(obj)]})

    if isinstance(obj, (list, tuple)):
        if len(obj) == 0:
            return pd.DataFrame()
        try:
            return pd.DataFrame(obj)
        except Exception:
            return pd.DataFrame({"value": [str(x) for x in obj]})

    return pd.DataFrame({"value": [str(obj)]})


def safe_len(value: Any) -> int:
    if value is None:
        return 0
    try:
        if pd.isna(value):
            return 0
    except Exception:
        pass
    try:
        return len(str(value))
    except Exception:
        return 0


def get_column_display_width(series: pd.Series, header: Any, min_width: int = 8, max_width: int = 60) -> int:
    header_len = safe_len(header)

    try:
        values = series.tolist()
    except Exception:
        try:
            values = list(series.astype("object"))
        except Exception:
            values = []

    value_max = max((safe_len(v) for v in values), default=0)
    return min(max(max(header_len, value_max) + 2, min_width), max_width)


def autosize_worksheet_xlsxwriter(
    worksheet,
    df_: pd.DataFrame,
    max_width: int = 60,
    min_width: int = 8,
) -> None:
    for idx, col in enumerate(df_.columns):
        series = df_.iloc[:, idx] if isinstance(df_, pd.DataFrame) and df_.shape[1] > idx else pd.Series(dtype="object")
        width = get_column_display_width(series, col, min_width=min_width, max_width=max_width)
        worksheet.set_column(idx, idx, width)


def autosize_worksheet_openpyxl(
    worksheet,
    df_: pd.DataFrame,
    max_width: int = 60,
    min_width: int = 8,
) -> None:
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Alignment, Font, PatternFill

    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    header_font = Font(bold=True)
    wrap_alignment = Alignment(wrap_text=True, vertical="top")

    for col_idx, col_name in enumerate(df_.columns, start=1):
        col_letter = get_column_letter(col_idx)
        series = df_.iloc[:, col_idx - 1] if isinstance(df_, pd.DataFrame) and df_.shape[1] >= col_idx else pd.Series(dtype="object")
        width = get_column_display_width(series, col_name, min_width=min_width, max_width=max_width)
        worksheet.column_dimensions[col_letter].width = width

        header_cell = worksheet.cell(row=1, column=col_idx)
        header_cell.fill = header_fill
        header_cell.font = header_font
        header_cell.alignment = wrap_alignment

        col_lower = str(col_name).lower()
        if any(token in col_lower for token in ["review", "text", "comment", "summary", "reason", "details"]):
            worksheet.column_dimensions[col_letter].width = min(max(40, min_width), max_width)
            for row_idx in range(2, worksheet.max_row + 1):
                worksheet.cell(row=row_idx, column=col_idx).alignment = wrap_alignment

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions


def build_master_excel(
    summary,
    overall_df,
    extra_sheets: Dict[str, Any] | None = None,
) -> bytes:
    output = io.BytesIO()

    sheets = {
        "Summary": summary,
        "All Reviews": overall_df,
    }
    if extra_sheets:
        sheets.update(extra_sheets)

    engine = "openpyxl"
    try:
        import xlsxwriter  # noqa: F401
        engine = "xlsxwriter"
    except ImportError:
        pass

    with pd.ExcelWriter(output, engine=engine) as writer:
        for raw_sheet_name, df_ in sheets.items():
            export_df = coerce_to_dataframe(df_)

            if export_df.empty and len(export_df.columns) == 0:
                export_df = pd.DataFrame({"message": ["No data available"]})

            sheet_name = str(raw_sheet_name)[:31]
            export_df.to_excel(writer, index=False, sheet_name=sheet_name)

            if engine == "xlsxwriter":
                workbook = writer.book
                worksheet = writer.sheets[sheet_name]

                header_fmt = workbook.add_format({
                    "bold": True,
                    "text_wrap": True,
                    "valign": "top",
                    "border": 1,
                    "bg_color": "#D9EAF7",
                })
                wrap_fmt = workbook.add_format({
                    "text_wrap": True,
                    "valign": "top",
                })

                rows, cols = export_df.shape
                worksheet.freeze_panes(1, 0)
                if cols > 0:
                    worksheet.autofilter(0, 0, max(rows, 1), cols - 1)

                for col_idx, col_name in enumerate(export_df.columns):
                    worksheet.write(0, col_idx, str(col_name), header_fmt)

                for col_idx, col_name in enumerate(export_df.columns):
                    col_lower = str(col_name).lower()
                    if any(token in col_lower for token in ["review", "text", "comment", "summary", "reason", "details"]):
                        worksheet.set_column(col_idx, col_idx, 40, wrap_fmt)

                autosize_worksheet_xlsxwriter(worksheet, export_df)

            else:
                worksheet = writer.sheets[sheet_name]
                autosize_worksheet_openpyxl(worksheet, export_df)

    output.seek(0)
    return output.getvalue()
